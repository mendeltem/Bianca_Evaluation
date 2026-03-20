#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Severity Level Definitions (Phase I + Phase II-A + LOCATE pool)
===============================================================
1. Updates severity_level in-place in source Excel files using
   fixed Decision Tree cutoffs (Challenge, N=40).
2. Generates summary Excel with severity distributions.
3. Exports tables as JPG images.

Severity cutoffs derived from Decision Tree classification
(Challenge dataset, N=40):
  LOW:    ROI_Volume <= 6.96 mL
  MIDDLE: 6.96 < ROI_Volume <= 27.40 mL
  HIGH:   ROI_Volume > 27.40 mL
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

SEVERITY_CUTOFF_LOW_MID = 6.96
SEVERITY_CUTOFF_MID_HIGH = 27.40

PHASE1_XLSX = os.path.join(SCRIPT_DIR, "Phase_1", "LOCATE_SET", "bianca_pool_wihtouth_ge.xlsx")
PHASE2A_XLSX = os.path.join(SCRIPT_DIR, "Phase_2_a", "RESULTS", "LOCATE_Results_Metrics_DICE_ONLY.xlsx")
LOCATE_POOL_XLSX = os.path.join(SCRIPT_DIR, "Phase_1", "LOCATE_SET", "locate_pool.xlsx")
OUTPUT_XLSX = os.path.join(SCRIPT_DIR, "plots", "2_severity_definitions.xlsx")

INPLACE_UPDATE_FILES = [PHASE1_XLSX, LOCATE_POOL_XLSX]


def assign_severity(roi_vol):
    if pd.isna(roi_vol):
        return np.nan
    if roi_vol <= SEVERITY_CUTOFF_LOW_MID:
        return "low"
    elif roi_vol <= SEVERITY_CUTOFF_MID_HIGH:
        return "middle"
    else:
        return "high"


# =============================================================
# TABLE TO IMAGE (JPG)
# =============================================================

def table_to_image(df, filepath, col_widths=None):
    """Render a DataFrame as a publication-ready table image."""
    n_rows, n_cols = df.shape

    col_w = col_widths if col_widths else [max(len(str(c)), max(len(str(v)) for v in df[c])) * 0.11 + 0.1 for c in df.columns]
    fig_w = sum(col_w)
    row_h = 0.28
    fig_h = (n_rows + 1) * row_h

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    table = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc='center',
        loc='center',
    )
    table.auto_set_font_size(False)

    for j in range(n_cols):
        cell = table[0, j]
        cell.set_facecolor('#2F5496')
        cell.set_text_props(color='white', fontweight='bold',
                            fontsize=11, fontfamily='sans-serif')
        cell.set_edgecolor('#1a3a6e')
        cell.set_linewidth(1.5)
        cell.set_height(1.0 / (n_rows + 1))

    for i in range(1, n_rows + 1):
        for j in range(n_cols):
            cell = table[i, j]
            cell.set_facecolor('white')
            cell.set_edgecolor('black')
            cell.set_linewidth(0.8)
            cell.set_height(1.0 / (n_rows + 1))
            cell.set_text_props(fontsize=11, fontfamily='sans-serif')

    for j, w in enumerate(col_w):
        for i in range(n_rows + 1):
            table[i, j].set_width(w / fig_w)

    fig.patch.set_facecolor('#FEFEFE')
    fig.savefig(filepath, dpi=300, bbox_inches='tight', pad_inches=0,
                facecolor='#FEFEFE', edgecolor='none')
    plt.show()
    plt.close(fig)

    from PIL import Image, ImageChops
    img = Image.open(filepath).convert('RGB')
    bg = Image.new('RGB', img.size, (254, 254, 254))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        cropped = img.crop(bbox)
        arr = np.array(cropped)
        mask = (arr[:, :, 0] == 254) & (arr[:, :, 1] == 254) & (arr[:, :, 2] == 254)
        arr[mask] = [255, 255, 255]
        Image.fromarray(arr).save(filepath, quality=95)

    print(f"  Saved: {os.path.basename(filepath)}")


# =============================================================
# HELPERS
# =============================================================

header_fill = PatternFill('solid', fgColor='2F5496')
header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
meta_font = Font(italic=True, size=9, name='Arial', color='555555')
data_font = Font(size=9, name='Arial')
bold_font = Font(bold=True, size=9, name='Arial')
border_thin = Border(bottom=Side(style='thin', color='D9D9D9'))
ctr = Alignment(horizontal='center', vertical='center')
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)


def write_meta(ws, meta_lines, max_col=10):
    for i, line in enumerate(meta_lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = meta_font
        cell.alignment = left_wrap
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=max_col)
    return len(meta_lines)


def write_header(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = ctr


def auto_width(ws, max_col, header_row, max_row):
    for col in range(1, max_col + 1):
        mx = max(len(str(ws.cell(row=r, column=col).value or ''))
                 for r in range(header_row, max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 35)


def build_severity_df(df, cohort_label):
    """Build a flat DataFrame of severity stats for one cohort."""
    if 'severity_level' not in df.columns or 'ROI_Volume' not in df.columns:
        return pd.DataFrame()

    sub = df[['severity_level', 'ROI_Volume']].dropna()
    rows = []

    all_v = sub['ROI_Volume']
    rows.append({
        'Cohort': cohort_label,
        'Severity': 'Overall',
        'n': len(sub),
        'Range (mL)': f"{all_v.min():.2f} \u2013 {all_v.max():.2f}",
        'Median': f"{all_v.median():.2f}",
        'IQR Low': f"{all_v.quantile(0.25):.2f}",
        'IQR High': f"{all_v.quantile(0.75):.2f}",
    })

    for sev in ['low', 'middle', 'high']:
        vals = sub[sub['severity_level'] == sev]['ROI_Volume']
        if len(vals) == 0:
            continue
        rows.append({
            'Cohort': '',
            'Severity': sev.capitalize(),
            'n': len(vals),
            'Range (mL)': f"{vals.min():.2f} \u2013 {vals.max():.2f}",
            'Median': f"{vals.median():.2f}",
            'IQR Low': f"{vals.quantile(0.25):.2f}",
            'IQR High': f"{vals.quantile(0.75):.2f}",
        })

    return pd.DataFrame(rows)


def write_severity_block(ws, start_row, df, cohort_label):
    """Write severity stats for one cohort to Excel."""
    if 'severity_level' not in df.columns or 'ROI_Volume' not in df.columns:
        ws.cell(row=start_row, column=1,
                value=f"{cohort_label}: severity_level or ROI_Volume not available").font = meta_font
        return start_row + 1

    sub = df[['severity_level', 'ROI_Volume']].dropna()
    all_v = sub['ROI_Volume']
    row = start_row

    ws.cell(row=row, column=1, value=cohort_label).font = bold_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    ws.cell(row=row, column=1, value='Overall').font = bold_font
    ws.cell(row=row, column=2, value=len(sub))
    ws.cell(row=row, column=3, value=f"{all_v.min():.2f} \u2013 {all_v.max():.2f}")
    ws.cell(row=row, column=4, value=f"{all_v.median():.2f}")
    ws.cell(row=row, column=5, value=f"{all_v.quantile(0.25):.2f}")
    ws.cell(row=row, column=6, value=f"{all_v.quantile(0.75):.2f}")
    ws.cell(row=row, column=7, value=f"\u2264 {SEVERITY_CUTOFF_LOW_MID} / > {SEVERITY_CUTOFF_MID_HIGH}")
    row += 1

    for sev in ['low', 'middle', 'high']:
        vals = sub[sub['severity_level'] == sev]['ROI_Volume']
        if len(vals) == 0:
            continue
        ws.cell(row=row, column=1, value=sev.capitalize()).font = data_font
        ws.cell(row=row, column=2, value=len(vals))
        ws.cell(row=row, column=3, value=f"{vals.min():.2f} \u2013 {vals.max():.2f}")
        ws.cell(row=row, column=4, value=f"{vals.median():.2f}")
        ws.cell(row=row, column=5, value=f"{vals.quantile(0.25):.2f}")
        ws.cell(row=row, column=6, value=f"{vals.quantile(0.75):.2f}")
        ws.cell(row=row, column=7, value='')
        row += 1

    for r in range(start_row + 1, row):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.alignment = ctr
            cell.border = border_thin

    return row + 1


# =========================================================
# IN-PLACE SEVERITY UPDATE
# =========================================================
print("\n--- In-place severity_level updates ---")
for fpath in INPLACE_UPDATE_FILES:
    if not os.path.isfile(fpath):
        print(f"  SKIP (not found): {fpath}")
        continue
    df_tmp = pd.read_excel(fpath)
    if 'ROI_Volume' not in df_tmp.columns:
        print(f"  SKIP (no ROI_Volume): {fpath}")
        continue
    old_counts = df_tmp['severity_level'].value_counts().to_dict() if 'severity_level' in df_tmp.columns else {}
    df_tmp['severity_level'] = df_tmp['ROI_Volume'].apply(assign_severity)
    new_counts = df_tmp['severity_level'].value_counts().to_dict()
    df_tmp.to_excel(fpath, index=False)
    print(f"  UPDATED: {os.path.basename(fpath)} (n={len(df_tmp)})")
    if old_counts != new_counts:
        print(f"    before: {old_counts}")
        print(f"    after:  {new_counts}")
    else:
        print(f"    no change: {new_counts}")


# =========================================================
# LOAD DATA FOR REPORT
# =========================================================
print("\nLoading Phase I pool ...")
df_p1 = pd.read_excel(PHASE1_XLSX)
df_p1['severity_level'] = df_p1['ROI_Volume'].apply(assign_severity)
print(f"  Phase I: {len(df_p1)} subjects")

df_locate = None
if os.path.isfile(LOCATE_POOL_XLSX):
    print("Loading LOCATE pool ...")
    df_locate = pd.read_excel(LOCATE_POOL_XLSX)
    df_locate['severity_level'] = df_locate['ROI_Volume'].apply(assign_severity)
    print(f"  LOCATE pool: {len(df_locate)} subjects")

print("Loading Phase II-A ...")
df_p2a = pd.read_excel(PHASE2A_XLSX)
if "subject_with_mask" in df_p2a.columns:
    df_p2a = df_p2a[df_p2a["subject_with_mask"] == 1].copy()
if "subject" in df_p2a.columns:
    df_p2a = df_p2a.drop_duplicates(subset="subject", keep="first")
df_p2a['severity_level'] = df_p2a['ROI_Volume'].apply(assign_severity)
print(f"  Phase II-A: {len(df_p2a)} subjects")


# =========================================================
# BUILD EXCEL
# =========================================================
wb = Workbook()
ws1 = wb.active
ws1.title = 'Severity Definitions'

meta1 = [
    "WMH severity level definitions based on Decision Tree classification (Challenge, N=40).",
    f"Cutoffs: LOW \u2264 {SEVERITY_CUTOFF_LOW_MID} mL | MIDDLE {SEVERITY_CUTOFF_LOW_MID}\u2013{SEVERITY_CUTOFF_MID_HIGH} mL | HIGH > {SEVERITY_CUTOFF_MID_HIGH} mL.",
    "ROI_Volume = total WMH volume from manual segmentation.",
]
n_meta = write_meta(ws1, meta1, 7)

headers = ['Cohort / Severity', 'n', 'Range (mL)', 'Median',
           'IQR Low', 'IQR High', 'Cutoffs (mL)']
hr = n_meta + 1
write_header(ws1, hr, headers)

next_row = write_severity_block(ws1, hr + 1, df_p1,
    f"Phase I (n={len(df_p1)}): BeLOVE + Challenge, excl. GE")
if df_locate is not None:
    next_row = write_severity_block(ws1, next_row, df_locate,
        f"LOCATE pool (n={len(df_locate)}): held-out training set")
next_row = write_severity_block(ws1, next_row, df_p2a,
    f"Phase II-A (n={len(df_p2a)}): BeLOVE")
auto_width(ws1, 7, hr, next_row)

# Scanner x Severity sheets
datasets = [("Phase I", df_p1), ("Phase II-A", df_p2a)]
if df_locate is not None:
    datasets.insert(1, ("LOCATE pool", df_locate))
for label, df in datasets:
    if 'scanner' in df.columns and 'severity_level' in df.columns:
        ws = wb.create_sheet(f'Scanner x Severity ({label})')
        ct = pd.crosstab(df['scanner'], df['severity_level'], margins=True)
        meta = [f"Scanner \u00d7 WMH severity cross-tabulation ({label}, n={len(df)})."]
        n_m = write_meta(ws, meta, ct.shape[1] + 1)
        hr_s = n_m + 1
        write_header(ws, hr_s, ['Scanner'] + [str(c) for c in ct.columns])
        for r_idx, (idx_val, row_data) in enumerate(ct.iterrows(), hr_s + 1):
            ws.cell(row=r_idx, column=1, value=str(idx_val)).font = data_font
            for c_idx, val in enumerate(row_data, 2):
                ws.cell(row=r_idx, column=c_idx, value=int(val)).font = data_font
                ws.cell(row=r_idx, column=c_idx).alignment = ctr
                ws.cell(row=r_idx, column=c_idx).border = border_thin
        auto_width(ws, ct.shape[1] + 1, hr_s, hr_s + len(ct))

wb.save(OUTPUT_XLSX)
print(f"\n\u2705 {OUTPUT_XLSX}")


# =========================================================
# JPG EXPORTS
# =========================================================
print("\n--- Exporting table images ---")
plots_dir = os.path.join(SCRIPT_DIR, "plots")
os.makedirs(plots_dir, exist_ok=True)

# Build flat DataFrames for JPG export
all_severity_rows = []
cohorts = [
    (f"Phase I (n={len(df_p1)})", df_p1),
    (f"Phase II-A (n={len(df_p2a)})", df_p2a),
]
if df_locate is not None:
    cohorts.insert(1, (f"LOCATE pool (n={len(df_locate)})", df_locate))

for label, df in cohorts:
    sev_df = build_severity_df(df, label)
    if len(sev_df) > 0:
        all_severity_rows.append(sev_df)

if all_severity_rows:
    combined_sev = pd.concat(all_severity_rows, ignore_index=True)
    table_to_image(combined_sev, os.path.join(plots_dir, "severity_definitions.jpg"))

# Scanner x Severity JPGs
for label, df in cohorts:
    if 'scanner' in df.columns and 'severity_level' in df.columns:
        ct = pd.crosstab(df['scanner'], df['severity_level'], margins=True)
        ct_reset = ct.reset_index().rename(columns={'index': 'Scanner'})
        ct_reset.columns = [str(c) for c in ct_reset.columns]
        safe_label = label.split('(')[0].strip().replace(' ', '_')
        table_to_image(ct_reset, os.path.join(plots_dir, f"scanner_severity_{safe_label}.jpg"))

print("\nDone.")