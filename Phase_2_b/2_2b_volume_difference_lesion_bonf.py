#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Volume Difference Analysis  BIANCA WMH Revision (Phase II-B, n=211)
==================================================
Two preprocessing conditions: Non Removed | Removed
Three WMH compartments: Total | Deep | Periventricular
Stratified by lesion type with Bonferroni correction.

Single pairwise comparison (Wilcoxon signed-rank, paired):
  1. Non Removed vs Removed

Rationale: Phase II-A demonstrated that removed and inpainted conditions
yield equivalent segmentation accuracy (all Cliff's delta < 0.005) and
equivalent WMH volumes (convergence comparison: all p_bonf = 1.0,
all |delta| negligible). Therefore, Phase II-B volume analysis uses
the non_removed vs removed comparison only.

Effect sizes: Cliff's Delta + bootstrap 95% CI (1000 iterations)
  Thresholds (Hess & Kromrey, 2004):
    negligible |delta| < 0.147
    small      0.147 <= |delta| < 0.33
    medium     0.33 <= |delta| < 0.474
    large      |delta| >= 0.474

Multiple testing: Bonferroni correction per comparison pair.
  Family structure: Each pairwise comparison forms one Bonferroni family
  with k = n_lesion_types x n_compartments tests (5 types x 3 compartments
  = 15 tests, corrected alpha = 0.0033).
  ICH (n=12) included in main analysis (unlike Phase II-A where n=1).


Key results
-----------
Non Removed vs Removed:
  Infarcts (n=96) and mixed lesions (n=25) showed significant volume
  differences after Bonferroni correction (k=15, alpha_corr=0.0033).
  Infarcts: Total/Peri/Deep all p<0.001***, delta -0.027 to -0.033.
  Mixed: Total p<0.001***, Deep p=0.004**, Peri p=0.003**.
  Lacunes showed marginal significance for Periventricular only
  (p_bonf=0.018*). Infratentorial strokes and ICH were non-significant
  (all p_bonf >= 0.47). Removed consistently yielded higher WMH volumes
  across most lesion types and compartments. All Cliff's Delta negligible
  (max |delta| = 0.056 for ICH Peri, n=12; max |delta| = 0.048 for
  mixed, n=25; max |delta| = 0.033 for infarcts, n=96). Median absolute
  differences: infarcts 0.04-0.14 mL, mixed 0.07-0.29 mL, lacunes
  0.02-0.06 mL, infra 0.01-0.05 mL, ICH 0.07-0.19 mL.

Overall interpretation:
  Volume differences between preprocessing conditions are driven
  primarily by infarcts and mixed lesions, consistent with the
  mechanistic explanation that larger lesions produce greater intensity
  histogram distortion. Lacunes show only marginal periventricular
  effects; infratentorial strokes and ICH show no significant effects.
  Phase II-A convergence analysis confirmed that zero-filling and
  NAWM-based inpainting produce equivalent results, isolating lesion
  removal as the active preprocessing step. Despite statistical
  significance for infarcts and mixed lesions, all effect sizes remain
  negligible (max |delta| = 0.048 for mixed, 0.033 for infarcts),
  confirming that absolute volume differences do not reach a magnitude
  warranting concern for group-level analyses.


Bonferroni family structure
---------------------------
  Single comparison (Non Removed vs Removed) forms one family:
    k = 5 lesion types x 3 compartments = 15 tests
    Corrected alpha = 0.05 / 15 = 0.0033
  ICH (n=12) included in main analysis.


Revision context
----------------
  R1 Comment 1 / R5 Comment 9: zero-filling vs inpainting
  R5 Comment 4: systematic bias despite negligible effect size


Paper changes
-------------
  Section 2.8 (Statistical Analysis): Bonferroni-corrected Wilcoxon
  Section 3.2: Volume difference results by lesion type
  Table X: Formatted volume difference table


Response to Reviewers
---------------------
  R1 Comment 1 / R5 Comment 9: "Volume differences were quantified
    between non_removed and removed conditions stratified by lesion type
    and WMH compartment (Bonferroni-corrected, k=15, alpha=0.0033).
    Infarcts (n=96) and mixed lesions (n=25) showed significant effects,
    consistent with larger lesions producing greater intensity histogram
    distortion. Lacunes showed marginal significance for periventricular
    WMH only (p_bonf=0.018). All Cliff's Delta remained negligible
    (max |delta| = 0.048). This single comparison was chosen because
    Phase II-A convergence analysis confirmed removed and inpainted
    conditions yield equivalent WMH volumes (all p_bonf = 1.0)."

  R5 Comment 4: "Despite statistical significance for infarcts and
    mixed lesions, absolute volume differences remained small (median
    0.04-0.29 mL) with negligible effect sizes (max |delta| = 0.048).
    The systematic, size-dependent nature of these differences supports
    reporting them as a methodological consideration rather than a
    source of meaningful bias in group-level analyses."


Outputs
-------
  volume_diff_raw_{cond1}_vs_{cond2}.xlsx
    -> Sheet 'Results': Bonferroni-corrected pairwise tests (all lesion types incl. ICH)
  volume_diff_formatted_{cond1}_vs_{cond2}.xlsx
    -> Sheet 'Results': Publication-ready table (all lesion types incl. ICH)
"""

import os
import warnings
from datetime import datetime
import numpy as np
import pandas as pd
from scipy.stats import wilcoxon
from cliffs_delta import cliffs_delta as _cliffs_delta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_ALL.xlsx")
PLOT_DIR = os.path.join(SCRIPT_DIR, "plots", "2b_VolumeDifference")
N_BOOT = 1000
BOOT_SEED = 42

# Minimum N for Wilcoxon (below this, result is flagged)
MIN_N_WILCOXON = 10

LESION_ORDER = ["infra", "lacune", "infarct", "mixed", "ICH"]
# Phase II-B: ICH n=12, included in main analysis (unlike Phase II-A where n=1)
LESION_DISPLAY = {
    "infra":   "infratentorial strokes",
    "lacune":  "lacunes",
    "infarct": "infarcts",
    "mixed":   "mixed (infarcts+lacunes)",
    "ICH":     "intracranial hemorrhage",
}

# Condition suffixes in column names → display labels
COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
    "filled":      "Inpainted",
}

# WMH compartments: (prefix, display name)
COMPARTMENTS = [
    ("WMH",     "Total"),
    ("deepWMH", "Deep"),
    ("perWMH",  "Periventricular"),
]

# Pairwise comparisons: (cond1, cond2)
COMPARISONS = [
    ("non_removed", "removed"),
]

# Cliff's Delta thresholds: Hess & Kromrey (2004)
# Using library defaults: negligible < 0.147, small < 0.33, medium < 0.474, large >= 0.474
# Consistent with cliffs_delta library and all other analysis scripts.
DELTA_THRESHOLD = 0.147  # boundary for negligible → small


# ─────────────────────────────────────────────
# STATISTICS
# ─────────────────────────────────────────────

def cliffs_delta_wrapper(x, y):
    """Cliff's Delta using library with default thresholds (Hess & Kromrey, 2004)."""
    x, y = list(x), list(y)
    if len(x) == 0 or len(y) == 0:
        return np.nan, "N/A"
    d, size = _cliffs_delta(x, y)
    return round(d, 4), size


def bootstrap_cliffs_delta_ci(x, y, n_boot=N_BOOT, alpha=0.05, seed=BOOT_SEED):
    """Bootstrap 95% CI for Cliff's Delta using library."""
    rng = np.random.default_rng(seed)
    x, y = np.array(x), np.array(y)
    if len(x) < 2 or len(y) < 2:
        return np.nan, np.nan
    deltas = []
    for _ in range(n_boot):
        x_b = rng.choice(x, size=len(x), replace=True)
        y_b = rng.choice(y, size=len(y), replace=True)
        d, _ = _cliffs_delta(list(x_b), list(y_b))
        deltas.append(d)
    lo = np.percentile(deltas, 100 * alpha / 2)
    hi = np.percentile(deltas, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def determine_higher(v1, v2, l1, l2):
    """Determine which condition has higher median values."""
    m1, m2 = v1.median(), v2.median()
    if abs(m1 - m2) < 1e-10:
        return "Equal"
    return l1 if m1 > m2 else l2


# ─────────────────────────────────────────────
# MAIN ANALYSIS
# ─────────────────────────────────────────────

def run_analysis(df, comparison, tag="", lesion_order=None):
    """
    Run volume difference analysis for one pairwise comparison
    across all lesion types and compartments.

    Bonferroni family: one family per comparison pair, with
    k = n_lesion_types × n_compartments tests.

    Returns a DataFrame of raw results (before Bonferroni).
    """
    if lesion_order is None:
        lesion_order = LESION_ORDER
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
    print(f"\n{'='*60}")
    print(f"  {tag}{l1} vs {l2}")
    print(f"{'='*60}")

    rows = []
    for lesion in lesion_order:
        ld = df[df["lesion_type"] == lesion]
        if len(ld) == 0:
            print(f"  Warning: no data for '{lesion}'")
            continue

        print(f"  {lesion} (n={len(ld)})")

        for prefix, comp_name in COMPARTMENTS:
            col1 = f"{prefix}_{cond1}_volume_ml"
            col2 = f"{prefix}_{cond2}_volume_ml"

            if col1 not in df.columns or col2 not in df.columns:
                print(f"    Missing column: {col1} or {col2}")
                continue

            v1 = ld[col1].dropna()
            v2 = ld[col2].dropna()
            # Align indices
            common = v1.index.intersection(v2.index)
            v1, v2 = v1.loc[common], v2.loc[common]

            if len(v1) == 0:
                continue

            # Flag small subgroups
            n_subj = len(v1)
            small_n_flag = n_subj < MIN_N_WILCOXON
            if small_n_flag:
                print(f"    WARNING: {lesion}/{comp_name} n={n_subj} < {MIN_N_WILCOXON}, "
                      f"Wilcoxon results should be interpreted with caution")

            diff = (v1 - v2).abs()

            # Wilcoxon signed-rank
            try:
                w_stat, p_val = wilcoxon(v1, v2)
            except Exception:
                w_stat, p_val = np.nan, np.nan

            # Cliff's Delta + bootstrap CI
            cd, cd_label = cliffs_delta_wrapper(v1.values, v2.values)
            ci_lo, ci_hi = bootstrap_cliffs_delta_ci(v1.values, v2.values)

            # Directionality
            higher = determine_higher(v1, v2, l1, l2)

            rows.append({
                "Lesion Type":  lesion,
                "Compartment":  comp_name,
                "N":            n_subj,
                "Small N":      small_n_flag,
                f"{l1} Median (ml)":  round(v1.median(), 2),
                f"{l1} Q1 (ml)":      round(v1.quantile(0.25), 2),
                f"{l1} Q3 (ml)":      round(v1.quantile(0.75), 2),
                f"{l2} Median (ml)":  round(v2.median(), 2),
                f"{l2} Q1 (ml)":      round(v2.quantile(0.25), 2),
                f"{l2} Q3 (ml)":      round(v2.quantile(0.75), 2),
                "Abs Diff Median (ml)": round(diff.median(), 2),
                "Abs Diff Q1 (ml)":     round(diff.quantile(0.25), 2),
                "Abs Diff Q3 (ml)":     round(diff.quantile(0.75), 2),
                "Higher":         higher,
                "W-Statistic":    round(w_stat, 2) if not np.isnan(w_stat) else np.nan,
                "P-Value":        p_val,
                "Cliff's Delta":  cd,
                "CD CI lo":       ci_lo,
                "CD CI hi":       ci_hi,
                "Effect Size":    cd_label,
            })

    return pd.DataFrame(rows)


def apply_bonferroni(results_df):
    """
    Apply Bonferroni correction within one comparison family.

    Family = all lesion types × compartments for a single pairwise
    comparison (e.g., Non Removed vs Removed). This follows the same
    family structure as Bland-Altman and test condition scripts.
    """
    n_tests = len(results_df)
    if n_tests == 0:
        return results_df

    results_df = results_df.copy()
    results_df["P-Value (Bonferroni)"] = (results_df["P-Value"] * n_tests).clip(upper=1.0)

    results_df["P-Value (Bonf)-Display"] = results_df["P-Value (Bonferroni)"].apply(
        lambda x: "<0.001" if x < 0.001 else f"{x:.3f}"
    )

    def sig_label(p):
        if p < 0.001: return "***"
        if p < 0.01:  return "**"
        if p < 0.05:  return "*"
        return "ns"

    results_df["Significance"] = results_df["P-Value (Bonferroni)"].apply(sig_label)

    # Flag meaningful effects (|delta| >= 0.147)
    results_df["Meaningful (|delta|>=0.147)"] = (
        results_df["Cliff's Delta"].abs() >= DELTA_THRESHOLD
    )

    print(f"  Bonferroni family: {n_tests} tests, corrected alpha = {0.05/n_tests:.6f}")
    return results_df


# ─────────────────────────────────────────────
# FORMATTED TABLE (publication-ready)
# ─────────────────────────────────────────────

def format_results_table(results_df, l1, l2, lesion_order=None):
    """
    Create a publication-ready table grouped by lesion type.
    Columns: WMH | Cond1 (median, IQR) | Cond2 (median, IQR) |
             Diff (median, IQR) | Higher | p (Bonf) |
             Cliff's Delta [95% CI] | Effect size
    """
    if lesion_order is None:
        lesion_order = LESION_ORDER
    formatted_rows = []

    for lesion in lesion_order:
        ld = results_df[results_df["Lesion Type"] == lesion]
        if len(ld) == 0:
            continue

        n_val = ld["N"].iloc[0]
        display = LESION_DISPLAY.get(lesion, lesion)
        small_n = ld["Small N"].iloc[0]
        n_label = f"{display} n={n_val}" + (" (caution: small N)" if small_n else "")

        # Header row
        formatted_rows.append({
            "WMH": n_label,
            f"{l1} (mL)\n(median, IQR)": "",
            f"{l2} (mL)\n(median, IQR)": "",
            "Diff (mL)\n(median, IQR)": "",
            "Higher": "",
            "p (Bonf)": "",
            "Cliff's Delta [95% CI]": "",
            "Effect size": "",
        })

        # Compartment order: Total, Peri, Deep
        comp_order = ["Total", "Periventricular", "Deep"]
        comp_short = {"Total": "Total", "Periventricular": "Peri", "Deep": "Deep"}

        for comp in comp_order:
            cd = ld[ld["Compartment"] == comp]
            if len(cd) == 0:
                continue
            r = cd.iloc[0]

            c1_text = f"{r[f'{l1} Median (ml)']}\n({r[f'{l1} Q1 (ml)']}-{r[f'{l1} Q3 (ml)']})"
            c2_text = f"{r[f'{l2} Median (ml)']}\n({r[f'{l2} Q1 (ml)']}-{r[f'{l2} Q3 (ml)']})"
            diff_text = f"{r['Abs Diff Median (ml)']}\n({r['Abs Diff Q1 (ml)']}-{r['Abs Diff Q3 (ml)']})"

            p_text = f"{r['P-Value (Bonf)-Display']} {r['Significance']}"
            cd_val = r["Cliff's Delta"]
            cd_text = f"{cd_val:.2f} [{r['CD CI lo']}, {r['CD CI hi']}]"

            formatted_rows.append({
                "WMH": comp_short[comp],
                f"{l1} (mL)\n(median, IQR)": c1_text,
                f"{l2} (mL)\n(median, IQR)": c2_text,
                "Diff (mL)\n(median, IQR)": diff_text,
                "Higher": r["Higher"],
                "p (Bonf)": p_text,
                "Cliff's Delta [95% CI]": cd_text,
                "Effect size": r["Effect Size"],
            })

    return pd.DataFrame(formatted_rows)


# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────

def save_formatted_excel(df, filepath, metadata_lines, n_data_cols=8, ich_df=None):
    """Save with professional formatting matching BA script style.
    Optionally adds a second sheet for ICH (n=1) data."""
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        if ich_df is not None and len(ich_df) > 0:
            ich_df.to_excel(writer, index=False, sheet_name="ICH (n=1)")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Results"]

    # ── Insert metadata header rows ──
    n_meta = len(metadata_lines)
    ws.insert_rows(1, n_meta)
    meta_font = Font(italic=True, size=9, name="Arial", color="555555")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    max_col_letter = openpyxl.utils.get_column_letter(ws.max_column)

    for i, line in enumerate(metadata_lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = meta_font
        cell.alignment = left_wrap
        ws.merge_cells(f"A{i}:{max_col_letter}{i}")

    # ── Header row styling ──
    header_row = n_meta + 1
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # ── Data styling ──
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(size=9, name="Arial")

    col_widths = [32, 20, 20, 20, 14, 14, 24, 15]
    for i, w in enumerate(col_widths[:ws.max_column], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > header_row:
                cell.font = data_font
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal="center")

    ws.row_dimensions[header_row].height = 35
    for r in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 40

    # ── Merge lesion-type header rows ──
    for r in range(header_row + 1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "")
        if any(kw in val.lower() for kw in
               ["strokes", "lacunes", "infarcts", "mixed", "hemorrhage"]):
            ws.merge_cells(f"A{r}:{max_col_letter}{r}")
            ws.cell(r, 1).alignment = Alignment(horizontal="center",
                                                 vertical="center")
            ws.cell(r, 1).font = Font(bold=True, size=10, name="Arial")

    # ── Highlight significant rows (green) ──
    sig_fill = PatternFill("solid", fgColor="E2EFDA")
    p_col = None
    for c in range(1, ws.max_column + 1):
        if "p (" in str(ws.cell(header_row, c).value or "").lower():
            p_col = c
            break
    if p_col:
        for r in range(header_row + 1, ws.max_row + 1):
            val = str(ws.cell(r, p_col).value or "")
            if "*" in val:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = sig_fill

    # ── Format ICH sheet (minimal, descriptive only) ──
    if "ICH (n=1)" in wb.sheetnames:
        ws_ich = wb["ICH (n=1)"]
        ich_meta = [
            "ICH (intracranial hemorrhage, n=1): reported separately.",
            "Statistics not interpretable with n=1. Descriptive values only.",
        ]
        n_ich_meta = len(ich_meta)
        ws_ich.insert_rows(1, n_ich_meta)
        for i, line in enumerate(ich_meta, start=1):
            cell = ws_ich.cell(row=i, column=1, value=line)
            cell.font = meta_font
            cell.alignment = left_wrap
            ich_max_col = openpyxl.utils.get_column_letter(ws_ich.max_column)
            ws_ich.merge_cells(f"A{i}:{ich_max_col}{i}")

        ich_header_row = n_ich_meta + 1
        for col in range(1, ws_ich.max_column + 1):
            cell = ws_ich.cell(row=ich_header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for i, w in enumerate(col_widths[:ws_ich.max_column], 1):
            ws_ich.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(filepath)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading data …")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    # Phase II-B: BeLOVE subjects only (exclude Challenge subjects)
    df = df[df["subject"].str.startswith("sub-")].copy()

    # Exclude sub-027: fsl_anat preprocessing failed
    df = df[df["subject"] != "sub-027"].copy()

    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")

    print(f"Subjects: {len(df)}")
    print(f"Lesion types:\n{df['lesion_type'].value_counts()}")

    os.makedirs(PLOT_DIR, exist_ok=True)

    for cond1, cond2 in COMPARISONS:
        l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]

        # ── Main analysis (all lesion types incl. ICH) ──
        raw = run_analysis(df, (cond1, cond2), tag="", lesion_order=LESION_ORDER)
        corrected = apply_bonferroni(raw)

        # ── Save raw results ──
        raw_path = os.path.join(PLOT_DIR, f"volume_diff_raw_{cond1}_vs_{cond2}.xlsx")
        corrected.to_excel(raw_path, sheet_name="Results", index=False)
        print(f"  Raw saved: {raw_path}")

        # ── Metadata for formatted table ──
        n_tests = len(corrected)
        metadata = [
            f"Volume Difference Analysis (Phase II-B): {l1} vs {l2}",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"Script: 2_2b_volume_difference_lesion_bonf.py",
            f"N subjects: {len(df)}  |  "
            f"Bonferroni family: {n_tests} tests, "
            f"corrected alpha = {0.05/n_tests:.4f}" if n_tests > 0 else "",
            f"Effect size: Cliff's Delta with bootstrap 95% CI ({N_BOOT} iterations).",
            f"Thresholds (Hess & Kromrey, 2004): negligible |delta| < 0.147, "
            f"small 0.147-0.33, medium 0.33-0.474, large >= 0.474.",
        ]

        # ── Formatted table ──
        fmt = format_results_table(corrected, l1, l2, lesion_order=LESION_ORDER)

        fmt_path = os.path.join(PLOT_DIR, f"volume_diff_formatted_{cond1}_vs_{cond2}.xlsx")
        save_formatted_excel(fmt, fmt_path, metadata_lines=metadata)
        print(f"  Formatted saved: {fmt_path}")

    print("\nDone.")