#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase II-B: Correlation Analysis -- Stroke Lesion Volume vs WMH Volume Difference
=================================================================================
Spearman rank correlations between stroke lesion volume and WMH volume
difference (non_removed vs removed), stratified by lesion type.
Three WMH compartments analyzed separately: Total, Deep, Periventricular.

Both trimmed (<=90th percentile of stroke lesion volume) and full-sample
correlations are reported. Trimmed values minimize outlier influence
(Wilcox, 2012; Pernet et al., 2012); full-sample values are provided
for transparency (R5 Comment 15).

Design
------
  Single comparison: Non Removed vs Removed.
  Phase II-A demonstrated that removed and inpainted conditions yield
  equivalent WMH volumes (convergence: all p_bonf = 1.0, all |delta|
  negligible). Phase II-B therefore uses only the non_removed vs removed
  comparison.

  X-axis: stroke lesion volume (mL)
  Y-axis: signed WMH volume difference (removed - non_removed, mL)
  Positive values = removal yields higher WMH volume.

  Three WMH compartments: Total, Deep, Periventricular.
  Stratified by lesion type: infra, lacune, infarct, mixed, ICH.

  Trimming: exclude subjects above 90th percentile of stroke lesion
  volume within each lesion type (Wilcox, 2012).

  Bootstrap 95% CIs (1000 iterations, seed 42) on both trimmed and
  full samples (Efron & Tibshirani, 1993).

  Bonferroni correction: one family for the comparison,
  k = 5 lesion types x 3 compartments = 15 tests,
  corrected alpha = 0.05 / 15 = 0.0033.

  Subgroups with n < 10 flagged with caution note (R5 Comment 12).

Statistical note
----------------
  Spearman rank correlation is reported throughout. No trend lines are
  plotted, as Spearman assumes monotonic but not necessarily linear
  relationships. The rho values in annotation boxes communicate
  correlation strength.


Key results
-----------
  [To be filled after first run]


Revision context
----------------
  R5 Comment 15: Trimmed + full correlations for transparency.
  R5 Comment 12: Small subgroups (n<10) flagged.
  R5 Comment 5:  Bonferroni correction per comparison family.
  R1 Comment 1 / R5 Comment 9: Convergence justifies single comparison.

Paper changes
-------------
  Section 3.2.3: Size-dependent scaling reported with Phase II-B
    sample (n=210) alongside Phase II-A results.
  Supplemental Table: Combined trimmed + full Spearman correlations
    for all lesion types and WMH compartments.

Response to Reviewers
---------------------
  [To be drafted after results]

References
----------
  Efron, B., Tibshirani, R.J., 1993. An Introduction to the Bootstrap.
    Chapman and Hall, New York.
  Pernet, C.R., Wilcox, R., Rousselet, G.A., 2012. Robust Correlation
    Analyses. Front. Psychol. 3, 606.
  Wilcox, R.R., 2012. Introduction to Robust Estimation and Hypothesis
    Testing, 3rd ed. Academic Press, Amsterdam.

Outputs
-------
  correlation_results_phase2b.xlsx
    -> Sheet 'Publication': Trimmed + full side-by-side
    -> Sheet 'Detailed': All numerical values
  correlation_plot_{compartment}_{cond1}_vs_{cond2}.png
"""

import os
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
from scipy.stats import spearmanr
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

warnings.filterwarnings("ignore")


# =============================================================================
# CONFIG
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_ALL.xlsx")
PLOT_DIR = os.path.join(SCRIPT_DIR, "plots", "8_2b_CorrelationAnalysis")

TRIM_PERCENTILE = 90
N_BOOTSTRAP = 1000
BOOT_SEED = 42
MIN_N_RELIABLE = 10  # below this, flag as unreliable (R5 #12)

LESION_ORDER = ["infra", "lacune", "infarct", "mixed", "ICH"]
LESION_DISPLAY = {
    "infra":   "Infratentorial Strokes",
    "lacune":  "Lacunes",
    "infarct": "Ischemic Infarcts",
    "mixed":   "Mixed (Infarcts + Lacunes)",
    "ICH":     "Intracranial Hemorrhage",
}

COMPARTMENTS = [
    ("WMH",     "Total WMH"),
    ("deepWMH", "Deep WMH"),
    ("perWMH",  "Periventricular WMH"),
]
COMP_SHORT = {
    "Total WMH":            "Total",
    "Deep WMH":             "Deep",
    "Periventricular WMH":  "Peri",
}

COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
}

COMPARISONS = [
    ("non_removed", "removed"),
]

# Bonferroni: k = 5 lesion types x 3 compartments = 15
K_BONF = len(LESION_ORDER) * len(COMPARTMENTS)
ALPHA_ADJ = 0.05 / K_BONF


# =============================================================================
# STATISTICS
# =============================================================================

def spearman_with_bootstrap_ci(x, y, n_boot=N_BOOTSTRAP, seed=BOOT_SEED):
    """
    Spearman correlation with bootstrap 95% CI.
    Returns dict with rho, p, ci_lower, ci_upper, n.
    """
    x, y = np.asarray(x, dtype=float), np.asarray(y, dtype=float)
    valid = ~(np.isnan(x) | np.isnan(y))
    x, y = x[valid], y[valid]
    n = len(x)

    if n < 4:
        return {"n": n, "rho": np.nan, "p": np.nan,
                "ci_lower": np.nan, "ci_upper": np.nan}

    rho, p = spearmanr(x, y)

    rng = np.random.default_rng(seed)
    boot_rhos = np.empty(n_boot)
    for i in range(n_boot):
        idx = rng.choice(n, size=n, replace=True)
        r, _ = spearmanr(x[idx], y[idx])
        boot_rhos[i] = r

    ci_lo = np.nanpercentile(boot_rhos, 2.5)
    ci_hi = np.nanpercentile(boot_rhos, 97.5)

    return {"n": n, "rho": rho, "p": p,
            "ci_lower": ci_lo, "ci_upper": ci_hi}


def compute_trimmed_and_full(x_all, y_all, trim_pct=TRIM_PERCENTILE):
    """
    Compute both full-sample and trimmed Spearman correlations.
    Trimming: exclude subjects above trim_pct percentile of x.
    Returns (full_stats, trimmed_stats, trim_info).
    """
    x_all, y_all = np.asarray(x_all, dtype=float), np.asarray(y_all, dtype=float)
    valid = ~(np.isnan(x_all) | np.isnan(y_all))
    x_clean, y_clean = x_all[valid], y_all[valid]

    full_stats = spearman_with_bootstrap_ci(x_clean, y_clean)

    if len(x_clean) < 4:
        trimmed_stats = {"n": 0, "rho": np.nan, "p": np.nan,
                         "ci_lower": np.nan, "ci_upper": np.nan}
        trim_info = {"threshold_ml": np.nan, "n_excluded": 0}
    else:
        threshold = np.percentile(x_clean, trim_pct)
        mask = x_clean <= threshold
        x_trim, y_trim = x_clean[mask], y_clean[mask]
        trimmed_stats = spearman_with_bootstrap_ci(x_trim, y_trim)
        trim_info = {
            "threshold_ml": round(threshold, 2),
            "n_excluded": int((~mask).sum()),
        }

    return full_stats, trimmed_stats, trim_info


# =============================================================================
# ANALYSIS
# =============================================================================

def run_correlation_analysis(df, comparison):
    """
    For one pairwise comparison, compute Spearman correlations between
    stroke lesion volume and WMH volume difference, stratified by
    lesion type and WMH compartment.

    Returns list of result dicts.
    """
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
    lesion_vol_col = "infarct_volume_ml"

    results = []
    for lesion in LESION_ORDER:
        ld = df[df["lesion_type"] == lesion].copy()
        n_subj = len(ld)
        if n_subj == 0:
            continue

        small_n = n_subj < MIN_N_RELIABLE
        if small_n:
            print(f"    WARNING: {lesion} n={n_subj} < {MIN_N_RELIABLE}, "
                  f"interpret with caution (R5 #12)")

        for prefix, comp_name in COMPARTMENTS:
            vol_col1 = f"{prefix}_{cond1}_volume_ml"
            vol_col2 = f"{prefix}_{cond2}_volume_ml"

            if vol_col1 not in df.columns or vol_col2 not in df.columns:
                print(f"  ERROR: Missing columns {vol_col1} or {vol_col2}")
                continue

            x = ld[lesion_vol_col].values
            y = (ld[vol_col2] - ld[vol_col1]).values

            full_stats, trim_stats, trim_info = compute_trimmed_and_full(x, y)

            results.append({
                "lesion_type": lesion,
                "compartment": comp_name,
                "comparison": f"{l1} vs {l2}",
                "n_full": full_stats["n"],
                "rho_full": full_stats["rho"],
                "p_full": full_stats["p"],
                "ci_lower_full": full_stats["ci_lower"],
                "ci_upper_full": full_stats["ci_upper"],
                "n_trimmed": trim_stats["n"],
                "rho_trimmed": trim_stats["rho"],
                "p_trimmed": trim_stats["p"],
                "ci_lower_trimmed": trim_stats["ci_lower"],
                "ci_upper_trimmed": trim_stats["ci_upper"],
                "trim_threshold_ml": trim_info["threshold_ml"],
                "n_excluded": trim_info["n_excluded"],
                "small_n": small_n,
            })

    return results


def apply_bonferroni(results, k=K_BONF):
    """
    Apply Bonferroni correction across all tests in the family.
    k = 5 lesion types x 3 compartments = 15.
    """
    alpha_adj = 0.05 / k

    for r in results:
        p_f = r["p_full"]
        p_t = r["p_trimmed"]
        r["p_full_bonf"] = min(p_f * k, 1.0) if not np.isnan(p_f) else np.nan
        r["p_trimmed_bonf"] = min(p_t * k, 1.0) if not np.isnan(p_t) else np.nan
        r["k_bonf"] = k
        r["alpha_adj"] = alpha_adj

    return results


# =============================================================================
# FORMATTING HELPERS
# =============================================================================

def fmt_p(p):
    """Format p-value for publication."""
    if np.isnan(p):
        return "-"
    if p < 0.001:
        return "<0.001"
    if p < 0.01:
        return f"{p:.3f}"
    return f"{p:.2f}"


def fmt_rho(rho):
    """Format correlation coefficient."""
    if np.isnan(rho):
        return "-"
    return f"{rho:.2f}"


def fmt_ci(lo, hi):
    """Format 95% CI."""
    if np.isnan(lo) or np.isnan(hi):
        return "-"
    return f"[{lo:.2f}, {hi:.2f}]"


def fmt_rho_ci(rho, ci_lo, ci_hi):
    """Compact rho [CI] notation for publication tables."""
    if np.isnan(rho):
        return "-"
    ci = fmt_ci(ci_lo, ci_hi)
    return f"{rho:.2f} {ci}"


# =============================================================================
# TABLE GENERATION
# =============================================================================

def generate_publication_table(results):
    """
    Combined Supplemental table: grouped by lesion type, then compartment.
    Columns: Lesion Type | WMH | n_trim | rho [CI] trim |
    p_bonf trim | n_full | rho [CI] full | p_bonf full
    """
    rows = []
    for lesion in LESION_ORDER:
        lr = [r for r in results if r["lesion_type"] == lesion]
        if not lr:
            continue

        display = LESION_DISPLAY.get(lesion, lesion)
        n_subj = lr[0]["n_full"]
        caution = " (dagger)" if lr[0]["small_n"] else ""

        for r in lr:
            comp_short = COMP_SHORT.get(r["compartment"], r["compartment"])
            row = {
                "Lesion Type": f"{display} (n={n_subj}){caution}",
                "WMH": comp_short,
                "n (trim)": r["n_trimmed"],
                "rho [95% CI] (trim)": fmt_rho_ci(
                    r["rho_trimmed"], r["ci_lower_trimmed"],
                    r["ci_upper_trimmed"]),
                "p_Bonf (trim)": fmt_p(r["p_trimmed_bonf"]),
                "n (full)": r["n_full"],
                "rho [95% CI] (full)": fmt_rho_ci(
                    r["rho_full"], r["ci_lower_full"],
                    r["ci_upper_full"]),
                "p_Bonf (full)": fmt_p(r["p_full_bonf"]),
            }
            rows.append(row)

    return pd.DataFrame(rows)


def generate_detailed_table(results):
    """Detailed table with all raw numerical values."""
    rows = []
    for r in results:
        row = {
            "Lesion Type": r["lesion_type"],
            "Compartment": r["compartment"],
            "Comparison": r["comparison"],
            "n_full": r["n_full"],
            "rho_full": round(r["rho_full"], 4) if not np.isnan(r["rho_full"]) else None,
            "p_full": r["p_full"],
            "p_full_bonf": r["p_full_bonf"],
            "CI_lower_full": round(r["ci_lower_full"], 4) if not np.isnan(r["ci_lower_full"]) else None,
            "CI_upper_full": round(r["ci_upper_full"], 4) if not np.isnan(r["ci_upper_full"]) else None,
            "n_trimmed": r["n_trimmed"],
            "rho_trimmed": round(r["rho_trimmed"], 4) if not np.isnan(r["rho_trimmed"]) else None,
            "p_trimmed": r["p_trimmed"],
            "p_trimmed_bonf": r["p_trimmed_bonf"],
            "CI_lower_trimmed": round(r["ci_lower_trimmed"], 4) if not np.isnan(r["ci_lower_trimmed"]) else None,
            "CI_upper_trimmed": round(r["ci_upper_trimmed"], 4) if not np.isnan(r["ci_upper_trimmed"]) else None,
            "trim_threshold_ml": r["trim_threshold_ml"],
            "n_excluded": r["n_excluded"],
            "k_bonf": r["k_bonf"],
            "alpha_adj": r["alpha_adj"],
            "small_n_flag": r["small_n"],
        }
        rows.append(row)
    return pd.DataFrame(rows)


# =============================================================================
# PLOTTING
# =============================================================================

def plot_correlation_by_compartment(df, comparison, results, save_dir):
    """
    For each WMH compartment, create a multi-panel scatter plot
    stratified by lesion type.

    Uniform gray color (no Dice criterion in Phase II-B).
    """
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
    lesion_vol_col = "infarct_volume_ml"

    SCATTER_COLOR = "#5B8DB8"

    for prefix, comp_name in COMPARTMENTS:
        vol_col1 = f"{prefix}_{cond1}_volume_ml"
        vol_col2 = f"{prefix}_{cond2}_volume_ml"

        available = [lt for lt in LESION_ORDER
                     if lt in df["lesion_type"].unique()]
        n_plots = len(available)
        if n_plots == 0:
            continue

        n_cols = 3
        n_rows = int(np.ceil(n_plots / n_cols))
        fig, axes = plt.subplots(n_rows, n_cols,
                                 figsize=(6 * n_cols, 5 * n_rows))
        axes = np.atleast_1d(axes).flatten()

        res_lookup = {(r["lesion_type"], r["compartment"]): r
                      for r in results}

        for idx, lesion in enumerate(available):
            ax = axes[idx]
            ld = df[df["lesion_type"] == lesion].copy()

            x = ld[lesion_vol_col].values
            y = (ld[vol_col2] - ld[vol_col1]).values
            valid = ~(np.isnan(x) | np.isnan(y))

            ax.scatter(x[valid], y[valid], c=SCATTER_COLOR, alpha=0.6,
                       s=40, edgecolors="white", linewidth=0.5, zorder=2)

            ax.axhline(y=0, color="gray", linestyle=":", linewidth=1,
                       alpha=0.5)

            r = res_lookup.get((lesion, comp_name))
            if r:
                caution = ("\n(n<10, interpret with caution)" 
                           if r["small_n"] else "")
                lines = [
                    f"Trimmed: rho = {fmt_rho(r['rho_trimmed'])}, "
                    f"p = {fmt_p(r['p_trimmed_bonf'])}, "
                    f"n = {r['n_trimmed']}",
                    f"95% CI: {fmt_ci(r['ci_lower_trimmed'], r['ci_upper_trimmed'])}",
                    f"Full: rho = {fmt_rho(r['rho_full'])}, "
                    f"p = {fmt_p(r['p_full_bonf'])}, "
                    f"n = {r['n_full']}",
                    f"95% CI: {fmt_ci(r['ci_lower_full'], r['ci_upper_full'])}",
                ]
                stats_text = "\n".join(lines) + caution

                ax.text(0.02, 0.98, stats_text, transform=ax.transAxes,
                        fontsize=7.5, verticalalignment="top",
                        horizontalalignment="left",
                        bbox=dict(boxstyle="round,pad=0.4",
                                  facecolor="white", alpha=0.9,
                                  edgecolor="gray"))

            ax.set_xlabel("Stroke Lesion Volume (mL)", fontsize=10)
            ax.set_ylabel(f"WMH Volume Diff (mL)\n({l2} - {l1})",
                          fontsize=10)
            ax.set_title(LESION_DISPLAY.get(lesion, lesion),
                         fontsize=11, fontweight="bold")
            ax.grid(True, alpha=0.3)

        for idx in range(n_plots, len(axes)):
            axes[idx].set_visible(False)

        comp_short = COMP_SHORT.get(comp_name, comp_name)
        plt.suptitle(
            f"Correlation: Stroke Lesion Volume vs {comp_name} "
            f"Volume Difference\n{l1} vs {l2} (Phase II-B, n={len(df)})",
            fontsize=13, fontweight="bold", y=1.01)
        plt.tight_layout()

        fname = (f"correlation_plot_{comp_short.lower()}_"
                 f"{cond1}_vs_{cond2}.png")
        save_path = os.path.join(save_dir, fname)
        plt.savefig(save_path, dpi=300, bbox_inches="tight",
                    facecolor="white")
        print(f"  Plot saved: {save_path}")
        plt.close(fig)


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def save_formatted_excel(pub_table, detail_table, filepath, metadata_lines):
    """Save publication + detailed tables with professional formatting."""
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        pub_table.to_excel(writer, sheet_name="Publication",
                           index=False, startrow=len(metadata_lines) + 1)
        detail_table.to_excel(writer, sheet_name="Detailed", index=False)

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Publication"]

    # ── Metadata header ──
    n_meta = len(metadata_lines)
    meta_font = Font(italic=True, size=9, name="Arial", color="555555")
    left_wrap = Alignment(horizontal="left", vertical="center",
                          wrap_text=True)

    for i in range(n_meta):
        row_idx = i + 1
        cell = ws.cell(row=row_idx, column=1, value=metadata_lines[i])
        cell.font = meta_font
        cell.alignment = left_wrap
        max_col = get_column_letter(ws.max_column)
        ws.merge_cells(f"A{row_idx}:{max_col}{row_idx}")

    # ── Header row styling ──
    header_row = n_meta + 1
    header_font = Font(bold=True, size=10, name="Arial")
    center = Alignment(horizontal="center", vertical="center",
                       wrap_text=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row + 1, column=col)
        if cell.value:
            cell.font = header_font
            cell.alignment = center

    # ── Column widths ──
    for col in ws.columns:
        real_cells = [c for c in col if not isinstance(c, MergedCell)]
        if not real_cells:
            continue
        max_len = max((len(str(c.value or "")) for c in real_cells),
                      default=10)
        ws.column_dimensions[real_cells[0].column_letter].width = min(
            max_len + 3, 30)

    wb.save(filepath)
    print(f"  Excel saved: {filepath}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("  Phase II-B Correlation: Stroke Lesion Volume vs WMH Volume Diff")
    print("=" * 70)

    # ── Load data ──
    print(f"\nLoading: {XLSX_PATH}")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    # Phase II-B: BeLOVE subjects only
    df = df[df["subject"].str.startswith("sub-")].copy()
    # Exclude sub-027: fsl_anat preprocessing failed
    df = df[df["subject"] != "sub-027"].copy()

    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")

    # Filter: subjects with stroke lesion volume > 0
    if "infarct_volume_ml" not in df.columns:
        raise SystemExit("Column 'infarct_volume_ml' not found.")
    n_before = len(df)
    # df = df[df["infarct_volume_ml"].notna()
    #         & (df["infarct_volume_ml"] > 0)].copy()
    
    
    
    n_after = len(df)
    if n_before != n_after:
        print(f"  Excluded {n_before - n_after} subjects with "
              f"infarct_volume_ml = 0 or NaN")

    print(f"Subjects: {len(df)}")
    print(f"Lesion types:\n{df['lesion_type'].value_counts()}")

    os.makedirs(PLOT_DIR, exist_ok=True)

    # ── Run analysis ──
    for cond1, cond2 in COMPARISONS:
        l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
        print(f"\n{'=' * 60}")
        print(f"  {l1} vs {l2}")
        print(f"{'=' * 60}")

        results = run_correlation_analysis(df, (cond1, cond2))
        results = apply_bonferroni(results, k=K_BONF)

        # ── Print summary ──
        print(f"\n  {'Lesion':<12} {'WMH':<6} {'n_t':>4} {'rho_t':>6} "
              f"{'p_bonf_t':>10} {'n_f':>4} {'rho_f':>6} {'p_bonf_f':>10}")
        print(f"  {'-' * 64}")
        for r in results:
            print(f"  {r['lesion_type']:<12} "
                  f"{COMP_SHORT.get(r['compartment'], '?'):<6} "
                  f"{r['n_trimmed']:>4} "
                  f"{fmt_rho(r['rho_trimmed']):>6} "
                  f"{fmt_p(r['p_trimmed_bonf']):>10} "
                  f"{r['n_full']:>4} "
                  f"{fmt_rho(r['rho_full']):>6} "
                  f"{fmt_p(r['p_full_bonf']):>10}")

        # ── Tables ──
        pub_table = generate_publication_table(results)
        detail_table = generate_detailed_table(results)

        # ── Excel ──
        metadata = [
            "Supplemental Table: Spearman Correlations -- Stroke Lesion "
            "Volume vs WMH Volume Difference (Phase II-B)",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"Script: 8_2b_correlation_lesion_volume.py",
            f"N subjects: {len(df)} (BeLOVE, sub-027 excluded, "
            f"infarct_volume > 0)",
            f"Comparison: {l1} vs {l2}",
            f"Bonferroni family: k={K_BONF} tests "
            f"(5 lesion types x 3 compartments), "
            f"corrected alpha = {ALPHA_ADJ:.4f}",
            f"Trimming: <=90th percentile of stroke lesion volume "
            f"(Wilcox, 2012).",
            f"Bootstrap 95% CIs: {N_BOOTSTRAP} iterations "
            f"(Efron & Tibshirani, 1993).",
            f"(dagger) = n<{MIN_N_RELIABLE}, interpret with caution.",
        ]

        xlsx_path = os.path.join(PLOT_DIR,
                                 "correlation_results_phase2b.xlsx")
        save_formatted_excel(pub_table, detail_table, xlsx_path, metadata)

        # ── Plots ──
        plot_correlation_by_compartment(df, (cond1, cond2), results,
                                       PLOT_DIR)

    print("\nDone.")