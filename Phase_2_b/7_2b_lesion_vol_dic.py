#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lesion Volume Dichotomization (2 mL)  BIANCA WMH Revision (Phase II-B, n=211)
=============================================================================
Two preprocessing conditions: Non Removed | Removed
Three WMH compartments: Total | Deep | Periventricular
Dichotomized by stroke lesion volume (threshold=2 mL).

Single pairwise comparison (Wilcoxon signed-rank, paired):
  1. Non Removed vs Removed

Rationale: Phase II-A demonstrated that removed and inpainted conditions
yield equivalent segmentation accuracy (all Cliff's delta < 0.005) and
equivalent WMH volumes (convergence comparison: all p_bonf = 1.0,
all |delta| negligible). Therefore, Phase II-B volume analysis uses
the non_removed vs removed comparison only.

Within the comparison: Small (<2 mL) vs Large (>=2 mL),
for Total / Deep / Periventricular WMH volumes.

Effect sizes: Cliff's Delta with bootstrap 95% CI (1000 iterations).
  Thresholds (Hess & Kromrey, 2004):
    negligible |delta| < 0.147
    small      0.147 <= |delta| < 0.33
    medium     0.33 <= |delta| < 0.474
    large      |delta| >= 0.474

Multiple testing: Bonferroni correction.
  Family structure: Single comparison forms one Bonferroni family
  with k = 2 groups x 3 compartments = 6 tests.
  Corrected alpha = 0.05 / 6 = 0.0083.

Note: The 2 mL threshold was chosen a priori based on lesion volume
distributions (approximate median) and is presented as a complementary
descriptive analysis alongside the primary continuous correlation
analysis (Section 3.2.3). It is not intended to replace continuous
modelling of the size-effect relationship.


Key results
-----------
Non Removed vs Removed, dichotomized at 2 mL (n=210):
  Clear size-dependent pattern confirmed:

  Small lesions (<2 mL, n=105):
    Minimal differences: Total -0.01 mL, Deep 0.00 mL, Peri -0.01 mL
    Peri p=0.003**, Total p=0.013*, Deep ns (p=0.315)
    All Cliff's Delta negligible (max |delta| = 0.005)

  Large lesions (>=2 mL, n=105):
    Consistently significant: Total -0.29 mL, Deep -0.09 mL, Peri -0.17 mL
    All p<0.001***
    All Cliff's Delta negligible (max |delta| = 0.040)
    Removed always higher

  The ~29x difference in median volume effect (0.29 vs 0.01 mL for Total)
  confirms the size-dependent scaling pattern from the continuous
  correlation analysis. However, even in the large lesion group, all
  effect sizes remain negligible (max |delta| = 0.040).

  Note: N=210 (not 211) because one subject had infarct_volume_ml = 0
  or missing and was excluded from the dichotomization.


Bonferroni family structure
---------------------------
  Single comparison (Non Removed vs Removed) forms one family:
    k = 2 size groups x 3 compartments = 6 tests
    Corrected alpha = 0.05 / 6 = 0.0083


Revision context
----------------
  R5 Comment 4: Do results differ by stroke lesion size?
  R5 Comment 6: Dichotomization at 2 mL -- justification required.


Paper changes
-------------
  Section 2.8: Dichotomization methodology
  Section 3.3: Size-dependent results (Phase II-B)
  Supplemental Table: Volume difference by lesion size group


Response to Reviewers
---------------------
  R5 Comment 4: "Lesion volume dichotomization (threshold 2 mL) in
    Phase II-B (n=210, Small n=105, Large n=105) revealed a clear
    size-dependent pattern: subjects with small stroke lesions showed
    minimal WMH volume differences (median 0.00-0.01 mL, marginal
    significance after Bonferroni correction), whereas subjects with
    large lesions showed consistently significant differences (median
    0.09-0.29 mL, all p<0.001). However, all effect sizes remained
    negligible (Cliff's delta <= 0.040), confirming that while larger
    lesions produce a statistically detectable effect, its magnitude
    does not reach a level warranting concern for group-level analyses."

  R5 Comment 6: "We acknowledge the reviewer's concern regarding
    dichotomization. The 2 mL threshold was chosen a priori based on
    the lesion volume distribution and is presented as a complementary
    descriptive analysis to illustrate the size-dependent pattern. The
    primary analysis of size effects relies on continuous Spearman
    correlations (Section 3.2.3), which preserve the full distributional
    information."


Outputs
-------
  2ml_raw_non_removed_vs_removed.xlsx
    -> Bonferroni-corrected pairwise tests
  2ml_formatted_non_removed_vs_removed.xlsx
    -> Publication-ready table with metadata header
"""

import os
import warnings
from datetime import datetime
import numpy as np
import pandas as pd
from scipy.stats import wilcoxon
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cliffs_delta import cliffs_delta as _cliffs_delta

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_ALL.xlsx")
PLOT_DIR = os.path.join(SCRIPT_DIR, "plots", "2b_LesionVolume_Dichotomized")
VOLUME_THRESHOLD = 2  # mL

# Bootstrap parameters
N_BOOT = 1000
BOOT_SEED = 42

# Cliff's Delta threshold: negligible → small boundary (Hess & Kromrey, 2004)
DELTA_THRESHOLD = 0.147

COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
}

COMPARISONS = [
    ("non_removed", "removed"),
]

COMPARTMENTS = [
    ("WMH",     "Total WMH"),
    ("deepWMH", "Deep WMH"),
    ("perWMH",  "Periventricular WMH"),
]

COMP_SHORT = {
    "Total WMH": "Total",
    "Periventricular WMH": "Peri",
    "Deep WMH": "Deep",
}

# ─────────────────────────────────────────────
# STATISTICS
# ─────────────────────────────────────────────

def cliffs_delta_wrapper(x, y):
    """Cliff's Delta using cliffs_delta library."""
    x, y = list(x), list(y)
    if len(x) == 0 or len(y) == 0:
        return np.nan, "N/A"
    d, size = _cliffs_delta(x, y)
    return round(d, 4), size


def bootstrap_cliffs_delta_ci(x, y, n_boot=N_BOOT, alpha=0.05, seed=BOOT_SEED):
    """Bootstrap 95% CI for Cliff's Delta (1000 iterations)."""
    rng = np.random.default_rng(seed)
    x, y = np.array(x), np.array(y)
    if len(x) < 2 or len(y) < 2:
        return np.nan, np.nan
    deltas = []
    for _ in range(n_boot):
        x_b = rng.choice(x, size=len(x), replace=True)
        y_b = rng.choice(y, size=len(y), replace=True)
        d, _ = _cliffs_delta(list(x_b), list(y_b))
        deltas.append(d)
    lo = np.percentile(deltas, 100 * alpha / 2)
    hi = np.percentile(deltas, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def determine_higher(v1, v2, l1, l2):
    """Determine which condition has higher values based on paired differences."""
    diff_median = (v1 - v2).median()
    if abs(diff_median) < 1e-10:
        return "Equal"
    return l1 if diff_median > 0 else l2


def format_p(p):
    if pd.isna(p):
        return "N/A"
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def sig_label(p):
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    return "ns"


# ─────────────────────────────────────────────
# ANALYSIS
# ─────────────────────────────────────────────

def run_volume_dichotomy_analysis(df, comparison):
    """Run volume difference analysis for one comparison, split by lesion size."""
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]

    print(f"\n{'='*60}")
    print(f"  {l1} vs {l2}")
    print(f"{'='*60}")

    # Dichotomize
    df_small = df[df["infarct_volume_ml"] < VOLUME_THRESHOLD].copy()
    df_large = df[df["infarct_volume_ml"] >= VOLUME_THRESHOLD].copy()

    print(f"  Small (<{VOLUME_THRESHOLD} mL): n={len(df_small)}, "
          f"Large (>={VOLUME_THRESHOLD} mL): n={len(df_large)}")

    rows = []
    for group_label, group_df in [("Small", df_small), ("Large", df_large)]:
        for prefix, wmh_name in COMPARTMENTS:
            col1 = f"{prefix}_{cond1}_volume_ml"
            col2 = f"{prefix}_{cond2}_volume_ml"

            if col1 not in df.columns or col2 not in df.columns:
                continue

            v1 = group_df[col1].dropna()
            v2 = group_df[col2].dropna()
            common = v1.index.intersection(v2.index)
            v1, v2 = v1.loc[common], v2.loc[common]

            if len(v1) < 3:
                continue

            diff = v1 - v2

            # Wilcoxon signed-rank (paired)
            try:
                _, p_val = wilcoxon(v1, v2)
            except Exception:
                p_val = np.nan

            # Cliff's Delta + bootstrap CI
            cd, cd_label = cliffs_delta_wrapper(v1.values, v2.values)
            ci_lo, ci_hi = bootstrap_cliffs_delta_ci(v1.values, v2.values)

            # Higher condition
            higher = determine_higher(v1, v2, l1, l2)

            rows.append({
                "Group":     group_label,
                "WMH_Type":  wmh_name,
                "N":         len(v1),
                f"{l1} Median (ml)":  round(v1.median(), 2),
                f"{l1} Q1 (ml)":      round(v1.quantile(0.25), 2),
                f"{l1} Q3 (ml)":      round(v1.quantile(0.75), 2),
                f"{l2} Median (ml)":  round(v2.median(), 2),
                f"{l2} Q1 (ml)":      round(v2.quantile(0.25), 2),
                f"{l2} Q3 (ml)":      round(v2.quantile(0.75), 2),
                "Diff Median (ml)":   round(diff.median(), 2),
                "Diff Q1 (ml)":       round(diff.quantile(0.25), 2),
                "Diff Q3 (ml)":       round(diff.quantile(0.75), 2),
                "Higher":             higher,
                "P-Value":            p_val,
                "Cliff's Delta":      round(cd, 4) if not np.isnan(cd) else np.nan,
                "CI Lower":           ci_lo,
                "CI Upper":           ci_hi,
                "Effect Size":        cd_label,
            })

    return pd.DataFrame(rows)


def apply_bonferroni(results_df):
    """Apply Bonferroni correction (6 tests)."""
    n_tests = len(results_df)
    if n_tests == 0:
        return results_df

    results_df = results_df.copy()
    results_df["P-Value (Bonferroni)"] = (results_df["P-Value"] * n_tests).clip(upper=1.0)
    results_df["P-Value (Bonf)-Display"] = results_df["P-Value (Bonferroni)"].apply(format_p)
    results_df["Significance"] = results_df["P-Value (Bonferroni)"].apply(sig_label)

    print(f"  Bonferroni: {n_tests} tests, corrected alpha = {0.05/n_tests:.6f}")
    return results_df


# ─────────────────────────────────────────────
# FORMATTED TABLE
# ─────────────────────────────────────────────

def format_results_table(results_df, l1, l2):
    """Publication-ready table grouped by lesion size group."""
    formatted_rows = []

    for group_label in ["Small", "Large"]:
        gd = results_df[results_df["Group"] == group_label]
        if len(gd) == 0:
            continue

        n_val = gd["N"].iloc[0]
        if group_label == "Small":
            header = f"Low volume <{VOLUME_THRESHOLD} mL  (n={n_val})"
        else:
            header = f"High volume >={VOLUME_THRESHOLD} mL  (n={n_val})"

        formatted_rows.append({
            "WMH": header,
            f"{l1} (mL)\n(median, IQR)": "",
            f"{l2} (mL)\n(median, IQR)": "",
            "Diff (mL)\n(median, IQR)": "",
            "Higher": "",
            "p (Bonf)": "",
            "Cliff's Delta\n[95% CI]": "",
            "Effect size": "",
        })

        for _, wmh_name in COMPARTMENTS:
            cd = gd[gd["WMH_Type"] == wmh_name]
            if len(cd) == 0:
                continue
            r = cd.iloc[0]

            c1_text = f"{r[f'{l1} Median (ml)']} ({r[f'{l1} Q1 (ml)']}-{r[f'{l1} Q3 (ml)']})"
            c2_text = f"{r[f'{l2} Median (ml)']} ({r[f'{l2} Q1 (ml)']}-{r[f'{l2} Q3 (ml)']})"
            diff_text = f"{r['Diff Median (ml)']} ({r['Diff Q1 (ml)']}-{r['Diff Q3 (ml)']})"

            p_text = f"{r['P-Value (Bonf)-Display']} {r['Significance']}"

            cd_val = r["Cliff's Delta"]
            ci_lo = r["CI Lower"]
            ci_hi = r["CI Upper"]
            if not pd.isna(cd_val):
                cd_str = f"{cd_val:.2f} [{ci_lo:.2f}, {ci_hi:.2f}]"
            else:
                cd_str = "-"

            formatted_rows.append({
                "WMH": COMP_SHORT.get(wmh_name, wmh_name),
                f"{l1} (mL)\n(median, IQR)": c1_text,
                f"{l2} (mL)\n(median, IQR)": c2_text,
                "Diff (mL)\n(median, IQR)": diff_text,
                "Higher": r["Higher"],
                "p (Bonf)": p_text,
                "Cliff's Delta\n[95% CI]": cd_str,
                "Effect size": r["Effect Size"],
            })

    return pd.DataFrame(formatted_rows)


# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────

def save_formatted_excel(df, filepath, comparison_label, n_subjects, n_tests):
    """Professional formatting with metadata header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Metadata header
    meta_lines = [
        f"Lesion Volume Dichotomization (Phase II-B): {comparison_label}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Script: 4_2b_lesion_volume_dichotomized.py",
        f"N subjects: {n_subjects}  |  Threshold: {VOLUME_THRESHOLD} mL  |  "
        f"Bonferroni: {n_tests} tests, corrected alpha = {0.05/n_tests:.4f}",
        f"Effect size: Cliff's Delta with bootstrap 95% CI ({N_BOOT} iterations).",
        f"Thresholds (Hess & Kromrey, 2004): negligible |delta| < 0.147, "
        f"small 0.147-0.33, medium 0.33-0.474, large >= 0.474.",
    ]
    for i, line in enumerate(meta_lines, 1):
        ws.cell(i, 1, line)
        ws.merge_cells(start_row=i, start_column=1,
                        end_row=i, end_column=len(df.columns))
        ws.cell(i, 1).font = Font(italic=True, size=10, color="444444")
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)

    # Column headers
    header_row = len(meta_lines) + 1
    for c, col_name in enumerate(df.columns, 1):
        cell = ws.cell(header_row, c, col_name)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="center")

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), header_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                        horizontal="center")

    # Borders
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=header_row,
                             max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Column widths
    col_widths = [22, 20, 20, 20, 14, 14, 20, 14]
    for i, w in enumerate(col_widths[:ws.max_column], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[header_row].height = 35
    for r in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 40

    # Merge volume group header rows
    max_col_letter = get_column_letter(ws.max_column)
    for r in range(header_row + 1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or "").lower()
        if "volume" in val:
            ws.merge_cells(f"A{r}:{max_col_letter}{r}")
            ws.cell(r, 1).alignment = Alignment(horizontal="center",
                                                  vertical="center")
            ws.cell(r, 1).font = Font(bold=True)

    # Highlight significant rows (green)
    sig_fill = PatternFill("solid", fgColor="E2EFDA")
    p_col = None
    for c in range(1, ws.max_column + 1):
        if "p (" in str(ws.cell(header_row, c).value or "").lower():
            p_col = c
            break
    if p_col:
        for r in range(header_row + 1, ws.max_row + 1):
            val = str(ws.cell(r, p_col).value or "")
            if "*" in val:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = sig_fill

    wb.save(filepath)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading data ...")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    print(f"Total subjects loaded: {len(df)}")

    # Phase II-B: BeLOVE subjects only (exclude Challenge subjects)
    df = df[df["subject"].str.startswith("sub-")].copy()

    # Exclude sub-027: fsl_anat preprocessing failed
    df = df[df["subject"] != "sub-027"].copy()

    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")
        
        
    #here we have 211 : subject

    # Filter to subjects with stroke lesions (infarct_volume_ml > 0)
    # if "infarct_volume_ml" in df.columns:
    #     n_before = len(df)
    #     df = df[df["infarct_volume_ml"].notna() & (df["infarct_volume_ml"] > 0)].copy()
    #     print(f"Subjects with stroke lesions: {len(df)} (filtered from {n_before})")
    # else:
    #     raise SystemExit("Column 'infarct_volume_ml' not found -- cannot dichotomize.")





    n_subjects = len(df)
    print(f"Subjects for dichotomization: {n_subjects}")
    print(f"Volume threshold: {VOLUME_THRESHOLD} mL")
    print(f"Small (<{VOLUME_THRESHOLD} mL): {(df['infarct_volume_ml'] < VOLUME_THRESHOLD).sum()}, "
          f"Large (>={VOLUME_THRESHOLD} mL): {(df['infarct_volume_ml'] >= VOLUME_THRESHOLD).sum()}")

    os.makedirs(PLOT_DIR, exist_ok=True)

    for cond1, cond2 in COMPARISONS:
        l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
        comp_key = f"{cond1}_vs_{cond2}"
        comp_label = f"{l1} vs {l2}"

        # Analysis
        raw = run_volume_dichotomy_analysis(df, (cond1, cond2))
        corrected = apply_bonferroni(raw)

        # Raw Excel
        raw_path = os.path.join(PLOT_DIR, f"2ml_raw_{comp_key}.xlsx")
        corrected.to_excel(raw_path, index=False)
        print(f"  Raw saved: {raw_path}")

        # Formatted table
        fmt = format_results_table(corrected, l1, l2)
        n_tests = len(raw)
        fmt_path = os.path.join(PLOT_DIR, f"2ml_formatted_{comp_key}.xlsx")
        save_formatted_excel(fmt, fmt_path, comp_label, n_subjects, n_tests)
        print(f"  Formatted saved: {fmt_path}")

        # Print summary
        print(f"\n  --- {comp_label} ---")
        for group in ["Small", "Large"]:
            gd = corrected[corrected["Group"] == group]
            if len(gd) == 0:
                continue
            print(f"  {group} (<{VOLUME_THRESHOLD} mL):" if group == "Small"
                  else f"  {group} (>={VOLUME_THRESHOLD} mL):")
            for _, r in gd.iterrows():
                sig = r["Significance"]
                ci_str = f"[{r['CI Lower']:.2f}, {r['CI Upper']:.2f}]" if not pd.isna(r["CI Lower"]) else ""
                cd_val = r["Cliff's Delta"]
                higher_val = r["Higher"]
                print(f"    {r['WMH_Type']:25s}  diff={r['Diff Median (ml)']:+.2f} mL  "
                      f"p={r['P-Value (Bonf)-Display']:>7s} {sig:3s}  "
                      f"delta={cd_val:+.2f} {ci_str}  "
                      f"Higher={higher_val}")

    print("\nDone.")