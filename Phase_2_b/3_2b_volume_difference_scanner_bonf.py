#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Volume Difference Analysis by Scanner  BIANCA WMH Revision (Phase II-B, n=211)
===========================================================================
Two preprocessing conditions: Non Removed | Removed
Three WMH compartments: Total | Deep | Periventricular
Stratified by scanner type with Bonferroni correction.

Single pairwise comparison (Wilcoxon signed-rank, paired):
  1. Non Removed vs Removed

Rationale: Phase II-A demonstrated that removed and inpainted conditions
yield equivalent segmentation accuracy (all Cliff's delta < 0.005) and
equivalent WMH volumes (convergence comparison: all p_bonf = 1.0,
all |delta| negligible). Therefore, Phase II-B volume analysis uses
the non_removed vs removed comparison only.

Effect sizes: Cliff's Delta + bootstrap 95% CI (1000 iterations)
  Thresholds (Hess & Kromrey, 2004):
    negligible |delta| < 0.147
    small      0.147 <= |delta| < 0.33
    medium     0.33 <= |delta| < 0.474
    large      |delta| >= 0.474

Multiple testing: Bonferroni correction.
  Family structure: Single comparison forms one Bonferroni family
  with k = 3 scanners x 3 compartments = 9 tests.
  Corrected alpha = 0.05 / 9 = 0.0056.


Key results
-----------
Non Removed vs Removed by Scanner (n=211):
  All scanner types showed significant volume differences after
  Bonferroni correction (k=9, alpha=0.0056), with Removed consistently
  yielding higher WMH volumes across all scanners and compartments.

  Prisma_fit (n=124):
    Total 0.11 mL, Peri 0.07 mL, Deep 0.03 mL (all p<0.001***)
    Cliff's Delta: -0.02 to -0.01, all negligible

  Tim Trio (n=51):
    Total 0.09 mL, Peri 0.08 mL, Deep 0.03 mL
    Total/Peri p<0.001***, Deep p=0.032*
    Cliff's Delta: -0.02 to -0.01, all negligible

  Philips (n=36):
    Total 0.18 mL, Peri 0.09 mL, Deep 0.07 mL (all p<0.001***)
    Cliff's Delta: -0.06 to -0.05, all negligible
    Largest absolute differences and effect sizes among scanners

  All Cliff's Delta negligible (max |delta| = 0.06 for Philips Total).
  Philips shows ~2x larger median differences than Siemens scanners,
  consistent with scanner type being the second-ranked predictor in
  SHAP analysis (Phase II-A). However, Philips subjects also had
  higher baseline WMH volumes (median 24.12 mL vs 14.67-15.58 mL),
  so absolute differences may partly reflect the size-dependent
  scaling observed in the lesion-type analysis.


Bonferroni family structure
---------------------------
  Single comparison (Non Removed vs Removed) forms one family:
    k = 3 scanners x 3 compartments = 9 tests
    Corrected alpha = 0.05 / 9 = 0.0056


Revision context
----------------
  R5 Comment 10: Scanner-specific effects on WMH volume differences
  R5 Comment 11: Philips n=8 vs Siemens n=78 in Phase II-A;
    Phase II-B provides better scanner balance (Philips n=36, 17.1%)


Paper changes
-------------
  Section 3.3: Scanner-stratified volume differences (Phase II-B)
  Table/Supplemental: Scanner-stratified volume difference table


Response to Reviewers
---------------------
  R5 Comment 10/11: "Scanner-stratified volume analysis in Phase II-B
    (n=211) with improved scanner balance (Philips 17.1% vs 9.3% in
    Phase II-A) confirmed that preprocessing effects are consistent
    across all three scanner types. Philips showed the largest absolute
    differences (Total: 0.18 mL, delta=-0.06) compared to Prisma_fit
    (0.11 mL, delta=-0.02) and Tim Trio (0.09 mL, delta=-0.02),
    but all effect sizes remained negligible (max |delta| = 0.06).
    The larger Philips differences may partly reflect higher baseline
    WMH burden in this subgroup (median 24.12 mL vs 14.67-15.58 mL)."


Outputs
-------
  volume_diff_scanner_raw_non_removed_vs_removed.xlsx
    -> Sheet 'Results': Bonferroni-corrected pairwise tests
  volume_diff_scanner_formatted_non_removed_vs_removed.xlsx
    -> Publication-ready table with metadata header
  volume_diff_scanner_non_removed_vs_removed.png
    -> Boxplot of absolute volume differences by scanner
"""

import os
import warnings
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import wilcoxon
from cliffs_delta import cliffs_delta as _cliffs_delta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_ALL.xlsx")
PLOT_DIR = os.path.join(SCRIPT_DIR, "plots", "2b_VolumeDifference_Scanner")
N_BOOT = 1000
BOOT_SEED = 42

SCANNER_ORDER = ["Prisma_fit", "Tim Trio", "Philips"]

COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
    "filled":      "Inpainted",
}

COMPARTMENTS = [
    ("WMH",     "Total"),
    ("deepWMH", "Deep"),
    ("perWMH",  "Periventricular"),
]

COMPARISONS = [
    ("non_removed", "removed"),
]

# ─────────────────────────────────────────────
# STATISTICS
# ─────────────────────────────────────────────

def cliffs_delta(x, y):
    x, y = np.asarray(x), np.asarray(y)
    if len(x) == 0 or len(y) == 0:
        return np.nan, "N/A"
    d, _ = _cliffs_delta(x, y)
    ad = abs(d)
    label = ("negligible" if ad < 0.147 else
             "small" if ad < 0.33 else
             "medium" if ad < 0.474 else "large")
    return round(d, 4), label


def bootstrap_ci(x, y, n_boot=N_BOOT, seed=BOOT_SEED):
    rng = np.random.default_rng(seed)
    boot = [cliffs_delta(rng.choice(x, len(x), replace=True),
                         rng.choice(y, len(y), replace=True))[0]
            for _ in range(n_boot)]
    return round(np.percentile(boot, 2.5), 4), round(np.percentile(boot, 97.5), 4)


# ─────────────────────────────────────────────
# ANALYSIS
# ─────────────────────────────────────────────

def run_analysis(df, comparison):
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
    print(f"\n{'='*60}")
    print(f"  {l1} vs {l2}")
    print(f"{'='*60}")

    rows = []
    for scanner in SCANNER_ORDER:
        sd = df[df["scanner"] == scanner]
        if len(sd) == 0:
            continue
        print(f"  {scanner} (n={len(sd)})")

        for prefix, comp_name in COMPARTMENTS:
            col1 = f"{prefix}_{cond1}_volume_ml"
            col2 = f"{prefix}_{cond2}_volume_ml"
            if col1 not in df.columns or col2 not in df.columns:
                continue

            v1 = sd[col1].dropna()
            v2 = sd[col2].dropna()
            common = v1.index.intersection(v2.index)
            v1, v2 = v1.loc[common], v2.loc[common]
            if len(v1) == 0:
                continue

            diff = (v1 - v2).abs()

            try:
                w_stat, p_val = wilcoxon(v1, v2)
            except Exception:
                w_stat, p_val = np.nan, np.nan

            cd, cd_label = cliffs_delta(v1.values, v2.values)
            ci_lo, ci_hi = bootstrap_ci(v1.values, v2.values)

            higher = l2 if cd < 0 else l1 if cd > 0 else "No difference"

            rows.append({
                "Scanner":        scanner,
                "Compartment":    comp_name,
                "N":              len(v1),
                f"{l1} Median (ml)":  round(v1.median(), 2),
                f"{l1} Q1 (ml)":      round(v1.quantile(0.25), 2),
                f"{l1} Q3 (ml)":      round(v1.quantile(0.75), 2),
                f"{l2} Median (ml)":  round(v2.median(), 2),
                f"{l2} Q1 (ml)":      round(v2.quantile(0.25), 2),
                f"{l2} Q3 (ml)":      round(v2.quantile(0.75), 2),
                "Abs Diff Median (ml)": round(diff.median(), 2),
                "Abs Diff Q1 (ml)":     round(diff.quantile(0.25), 2),
                "Abs Diff Q3 (ml)":     round(diff.quantile(0.75), 2),
                "Higher":         higher,
                "W-Statistic":    round(w_stat, 2) if not np.isnan(w_stat) else np.nan,
                "P-Value":        p_val,
                "Cliff's Delta":  cd,
                "CD CI lo":       ci_lo,
                "CD CI hi":       ci_hi,
                "Effect Size":    cd_label,
            })

    return pd.DataFrame(rows)


def apply_bonferroni(results_df):
    n_tests = len(results_df)
    if n_tests == 0:
        return results_df
    results_df = results_df.copy()
    results_df["P-Value (Bonferroni)"] = (results_df["P-Value"] * n_tests).clip(upper=1.0)
    results_df["P-Value (Bonf)-Display"] = results_df["P-Value (Bonferroni)"].apply(
        lambda x: "<0.001" if x < 0.001 else f"{x:.3f}")

    def sig_label(p):
        if p < 0.001: return "***"
        if p < 0.01:  return "**"
        if p < 0.05:  return "*"
        return "ns"

    results_df["Significance"] = results_df["P-Value (Bonferroni)"].apply(sig_label)
    print(f"  Bonferroni: {n_tests} tests, corrected alpha = {0.05/n_tests:.6f}")
    return results_df


# ─────────────────────────────────────────────
# FORMATTED TABLE
# ─────────────────────────────────────────────

def format_results_table(results_df, l1, l2):
    formatted_rows = []
    for scanner in SCANNER_ORDER:
        sd = results_df[results_df["Scanner"] == scanner]
        if len(sd) == 0:
            continue
        n_val = sd["N"].iloc[0]

        formatted_rows.append({
            "WMH": f"{scanner} n={n_val}",
            f"{l1} (mL)\n(median, IQR)": "",
            f"{l2} (mL)\n(median, IQR)": "",
            "Diff (mL)\n(median, IQR)": "",
            "Higher": "",
            "p (Bonf)": "",
            "Cliff's Delta [95% CI]": "",
            "Effect size": "",
        })

        for comp in ["Total", "Periventricular", "Deep"]:
            cd = sd[sd["Compartment"] == comp]
            if len(cd) == 0:
                continue
            r = cd.iloc[0]
            c1_text = f"{r[f'{l1} Median (ml)']}\n({r[f'{l1} Q1 (ml)']}-{r[f'{l1} Q3 (ml)']})"
            c2_text = f"{r[f'{l2} Median (ml)']}\n({r[f'{l2} Q1 (ml)']}-{r[f'{l2} Q3 (ml)']})"
            diff_text = f"{r['Abs Diff Median (ml)']}\n({r['Abs Diff Q1 (ml)']}-{r['Abs Diff Q3 (ml)']})"
            p_text = f"{r['P-Value (Bonf)-Display']} {r['Significance']}"
            cd_text = f"{r['Cliff\'s Delta']:.2f} [{r['CD CI lo']}, {r['CD CI hi']}]"

            formatted_rows.append({
                "WMH": {"Total": "Total", "Periventricular": "Peri", "Deep": "Deep"}[comp],
                f"{l1} (mL)\n(median, IQR)": c1_text,
                f"{l2} (mL)\n(median, IQR)": c2_text,
                "Diff (mL)\n(median, IQR)": diff_text,
                "Higher": r["Higher"],
                "p (Bonf)": p_text,
                "Cliff's Delta [95% CI]": cd_text,
                "Effect size": r["Effect Size"],
            })

    return pd.DataFrame(formatted_rows)


# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────

def save_formatted_excel(df, filepath, metadata_lines):
    df.to_excel(filepath, index=False, sheet_name="Results", startrow=len(metadata_lines))
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Insert metadata
    for i, line in enumerate(metadata_lines, 1):
        ws.cell(i, 1, line)
        ws.merge_cells(start_row=i, start_column=1,
                       end_row=i, end_column=ws.max_column)
        ws.cell(i, 1).font = Font(italic=True, size=10, color="444444")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    col_widths = [30, 20, 20, 20, 16, 14, 24, 15]
    for i, w in enumerate(col_widths[:ws.max_column], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    header_row = len(metadata_lines) + 1
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="center")
            if cell.row == header_row:
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(wrap_text=True, vertical="center",
                                           horizontal="center")
    ws.row_dimensions[header_row].height = 35
    for r in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 40

    max_col_letter = chr(64 + ws.max_column)
    for r in range(header_row + 1, ws.max_row + 1):
        val = str(ws[f"A{r}"].value or "").lower()
        if any(kw in val for kw in ["prisma", "tim trio", "philips"]):
            ws.merge_cells(f"A{r}:{max_col_letter}{r}")
            ws[f"A{r}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"A{r}"].font = Font(bold=True)

    sig_fill = PatternFill("solid", fgColor="E2EFDA")
    p_col = None
    for c in range(1, ws.max_column + 1):
        if "p (" in str(ws.cell(header_row, c).value or "").lower():
            p_col = c
            break
    if p_col:
        for r in range(header_row + 1, ws.max_row + 1):
            val = str(ws.cell(r, p_col).value or "")
            if "*" in val:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = sig_fill
    wb.save(filepath)


# ─────────────────────────────────────────────
# VISUALIZATION
# ─────────────────────────────────────────────

def create_boxplot(df, cond1, cond2, output_path):
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
    plot_data = []
    for prefix, comp_name in COMPARTMENTS:
        col1 = f"{prefix}_{cond1}_volume_ml"
        col2 = f"{prefix}_{cond2}_volume_ml"
        if col1 not in df.columns or col2 not in df.columns:
            continue
        tmp = df[["scanner"]].copy()
        tmp["absolute_difference"] = (df[col1] - df[col2]).abs()
        tmp["volume_type"] = comp_name + " WMH"
        plot_data.append(tmp)

    melted = pd.concat(plot_data, ignore_index=True).dropna()
    palette = {
        "Total WMH":            "#2ca02c",
        "Periventricular WMH":  "#1f77b4",
        "Deep WMH":             "#ff7f0e",
    }

    fig, ax = plt.subplots(figsize=(12, 9))
    sns.boxplot(x="scanner", y="absolute_difference", hue="volume_type",
                data=melted, palette=palette, showfliers=True,
                order=SCANNER_ORDER, ax=ax)
    ax.set_xlabel("Scanner Type", fontsize=12, fontweight="bold")
    ax.set_ylabel(f"Absolute Volume Difference\n|{l1} - {l2}| (ml)",
                  fontsize=12, fontweight="bold")
    ax.set_title(f"WMH Volume Differences by Scanner Type (Phase II-B)\n{l1} vs {l2}",
                 fontsize=14, fontweight="bold", pad=20)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha="right")
    ax.legend(title="Volume Type", loc="center left",
              bbox_to_anchor=(1, 0.5), title_fontsize=11, fontsize=10)
    ax.grid(True, alpha=0.3, axis="y")
    plt.tight_layout()
    fig.savefig(output_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Plot saved: {output_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading data …")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    # Phase II-B: BeLOVE subjects only (exclude Challenge subjects)
    df = df[df["subject"].str.startswith("sub-")].copy()

    # Exclude sub-027: fsl_anat preprocessing failed
    df = df[df["subject"] != "sub-027"].copy()

    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")

    print(f"Subjects: {len(df)}")
    print(f"Scanners:\n{df['scanner'].value_counts()}")

    os.makedirs(PLOT_DIR, exist_ok=True)

    for cond1, cond2 in COMPARISONS:
        l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]

        raw = run_analysis(df, (cond1, cond2))
        corrected = apply_bonferroni(raw)

        # ── Save raw results ──
        raw_path = os.path.join(PLOT_DIR, f"volume_diff_scanner_raw_{cond1}_vs_{cond2}.xlsx")
        corrected.to_excel(raw_path, index=False)
        print(f"  Raw saved: {raw_path}")

        # ── Metadata for formatted table ──
        n_tests = len(corrected)
        metadata = [
            f"Volume Difference Analysis by Scanner (Phase II-B): {l1} vs {l2}",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
            f"Script: 3_2b_volume_difference_scanner_bonf.py",
            f"N subjects: {len(df)}  |  "
            f"Bonferroni family: {n_tests} tests, "
            f"corrected alpha = {0.05/n_tests:.4f}" if n_tests > 0 else "",
            f"Effect size: Cliff's Delta with bootstrap 95% CI ({N_BOOT} iterations).",
            f"Thresholds (Hess & Kromrey, 2004): negligible |delta| < 0.147, "
            f"small 0.147-0.33, medium 0.33-0.474, large >= 0.474.",
        ]

        # ── Formatted table ──
        fmt = format_results_table(corrected, l1, l2)
        fmt_path = os.path.join(PLOT_DIR, f"volume_diff_scanner_formatted_{cond1}_vs_{cond2}.xlsx")
        save_formatted_excel(fmt, fmt_path, metadata_lines=metadata)
        print(f"  Formatted saved: {fmt_path}")

        # ── Boxplot ──
        plot_path = os.path.join(PLOT_DIR, f"volume_diff_scanner_{cond1}_vs_{cond2}.png")
        create_boxplot(df, cond1, cond2, plot_path)

    print("\nDone.")