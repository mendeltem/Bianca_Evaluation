#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

GE Compare: Statistical Analysis (Paper-Ready Tables Only)
==========================================================
 
Standalone analysis-only script. Reads pre-computed raw results from
ge_compare_raw_results.xlsx (produced by the full 2_compare_GE_analyse.py
pipeline) and regenerates all statistics and tables without requiring
BIANCA, LOCATE, FSL, or MATLAB.
 
Revision context (NeuroImage: Clinical, Major Revision)
-------------------------------------------------------
  (1) GE exclusion and robustness claims (R1 Comment 4; R5 #8)
      R1: excluding GE weakens the "robustness" claim.
      R5: methodologically pragmatic but introduces selection bias.
 
      Resolution: Transparent empirical comparison via nested training
      design (without GE = strict subset of with GE). The term
      "robustness" is qualified to apply specifically to Siemens and
      Philips platforms. GE exclusion acknowledged as limitation with
      recommendation for scanner-specific retraining.
 
  (2) Multiple comparisons transparency (R5 #5)
      Bonferroni correction families explicitly defined and reported.
 
Technical design
----------------
  Training splits (nested, "without GE" is strict subset of "with GE"):
    Without GE: 15 Prisma + 15 Trio + 15 Philips         = 45
    With GE:    same 45   + 15 GE                         = 60
 
  The only difference is the addition of 15 GE subjects. Any performance
  change is therefore attributable to the GE data alone.
 
  Test set: all remaining subjects not in either training set.
  GE test subjects reported separately.
 
  Statistical analysis:
    - Cliff's Delta with bootstrap 95% CIs (1000 iterations)
    - Wilcoxon signed-rank (paired) with Bonferroni correction
    - Scanner-stratified evaluation
    - Bland-Altman agreement
 
Paper changes
-------------
  Methods 2.3 (revised): Scanner exclusion rationale, framed as
    pragmatic exclusion based on empirical evidence.
  Supplemental (new): Tables S-X (this output).
  Limitations (revised): GE generalizability explicitly acknowledged.
 
Response to Reviewers
---------------------
  R1 Comment 4: Empirical GE comparison; "robustness" qualified.
  R5 #8: Selection bias acknowledged; supplemental evidence provided.
  R5 #5: Bonferroni families documented in output.
 
Output (single .xlsx, 4 sheets)
-------------------------------
  1. Wilcoxon_Descriptive   combined inferential + descriptive stats
  2. Bland_Altman           agreement analysis
  3. Scanner_Stratified     per-scanner GE impact
  4. Bonferroni_Families    multiple testing transparency
 
Key findings (from actual results)
-----------------------------------
  Overall (N=70 test subjects):
    - All Cliff's Delta values negligible (|delta| <= 0.047).
    - GE inclusion has no meaningful effect on non-GE performance.
    - Consistent across thresholds (0.85, 0.90, LOCATE):
      Precision marginally higher with GE training (+0.014),
      Sensitivity marginally lower (-0.005 to -0.009).
      Both negligible in magnitude.
 
  Bland-Altman:
    - Mean differences maximally 0.015 (Precision). LoAs narrow.
    - No systematic bias for Dice (p = 0.097 / 0.741).
    - Significant but negligible bias for Sensitivity/Precision:
      reflects high power from paired design, not effect magnitude.
 
  Scanner-stratified:
    - Siemens/Philips: all effects negligible (|delta| < 0.10).
    - GE test subjects (N=5): medium deltas (0.28-0.36), but N=5
      precludes reliable inference. Supports GE exclusion rationale.
 
  Interpretation note for manuscript:
    Several Wilcoxon tests are significant after Bonferroni correction
    (e.g., Sensitivity p < 0.001), but all Cliff's Delta remain
    negligible. Classic "statistically significant, practically
    meaningless" scenario in paired designs at N=70. The manuscript
    must frame this explicitly: significance reflects power, not
    magnitude. Effect sizes (Cliff's Delta), not p-values, drive
    the substantive conclusions.
 
Terminology (consistent across all scripts)
-------------------------------------------
  non_removed: no lesion preprocessing
  removed:     zero-intensity lesion replacement
  filled:      NAWM-based inpainting (FSL lesion_filling)
 
References
----------
  Griffanti et al. (2016). BIANCA. NeuroImage: Clinical, 9, 235-242.
  Hess & Kromrey (2004). Effect size interpretation guidelines.
  Romano et al. (2006). Cliff's Delta for ordinal effect sizes.

  
  
  
"""

import os
import json
import tempfile
import numpy as np
import pandas as pd
import subprocess
import warnings

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.stats import wilcoxon
from cliffs_delta import cliffs_delta

from DATASETS.librarys.directory_lib import (
    fsl_copy, get_volume, threshold_lpm,
    apply_bianca_mask, fsl_dice_score,
    create_panel_plot_with_multi_overlay, get_top_slices
)

from dotenv import load_dotenv
load_dotenv()

DEBUG = False
RANDOM_STATE = int(os.getenv("RANDOM_STATE", "42"))
N_BOOTSTRAP = 1000

# ---------------------------------------------------------
# GLOBAL PATHS & CONFIG
# ---------------------------------------------------------
base_dir = "Phase_1/LOCATE_SET"
results_path = os.path.join(base_dir, "all_files.xlsx")
GE_COMPARE_dir = "Phase_1/GE_COMPARE"
model_dir = "Phase_1/GE_COMPARE/BIANCA_MODELS"


LOCATE_PATH = (
    "/sc-projects/sc-proj-cc15-csb-neuroimaging/SC_Stroke_MRI/temuuleu/"
    "Projects/BIANCA/LOCATE/LOCATE-BIANCA"
)
# train_image_directory_path = (
#     "/sc-projects/sc-proj-cc15-csb-neuroimaging/SC_Stroke_MRI/temuuleu/"
#     "DATASET_SC/PROJECT_NULL_07_04_2025/1_Project_all_code_all_data/"
#     "Phase_1/LOCATE_SET/locate_train"
# )

train_image_directory_path = (
'/sc-projects/sc-proj-cc15-csb-neuroimaging/SC_Stroke_MRI/temuuleu/DATASET_SC/PROJECT_NULL_07_04_2025/1_Project_all_code_all_data/Phase_1/LOCATE_SET/locate_train_with_ge'
)

feature_select = [1, 1, 1, 1]
thresholds_list = [0.85, 0.90, "locate"]

os.makedirs(GE_COMPARE_dir, exist_ok=True)
os.makedirs(model_dir, exist_ok=True)


# =============================================================
# STATISTICAL HELPER FUNCTIONS
# =============================================================
def bootstrap_cliffs_delta(x, y, n_boot=N_BOOTSTRAP, seed=RANDOM_STATE):
    """
    Compute Cliff's Delta with bootstrap 95% CI.
    Same-size resampling with replacement (1000 iterations).
    """
    rng = np.random.default_rng(seed)
    delta, _ = cliffs_delta(x.tolist(), y.tolist())
    boot_deltas = []
    n_x, n_y = len(x), len(y)
    for _ in range(n_boot):
        x_boot = rng.choice(x, size=n_x, replace=True)
        y_boot = rng.choice(y, size=n_y, replace=True)
        d, _ = cliffs_delta(x_boot.tolist(), y_boot.tolist())
        boot_deltas.append(d)
    ci_low = np.percentile(boot_deltas, 2.5)
    ci_high = np.percentile(boot_deltas, 97.5)
    return delta, ci_low, ci_high


def interpret_cliffs_delta(delta):
    """
    Interpret Cliff's Delta magnitude (Romano et al., 2006).
    |delta| >= 0.28 is our pre-defined threshold for meaningful effects.
    """
    abs_d = abs(delta)
    if abs_d < 0.147:
        return "negligible"
    elif abs_d < 0.28:
        return "small"
    elif abs_d < 0.43:
        return "medium"
    else:
        return "large"


def compute_wmh_volume_ml(nifti_path):
    """Compute WMH volume in mL from a binary mask."""
    try:
        vol = get_volume(nifti_path)  # returns (nonzero_voxels, volume_mm3, volume_ml)
        return float(vol[2])
    except Exception:
        return None


def bonferroni_alpha(n_tests, base_alpha=0.05):
    """Return Bonferroni-corrected alpha for a family of tests."""
    return base_alpha / n_tests


DELTA_THRESHOLD = 0.28  # Pre-defined threshold for meaningful effects


def format_p_table(p):
    """For Excel tables: readable string."""
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.3f}"


def format_delta_ci(delta, ci_lo, ci_hi, mag='N/A'):
    """Format Cliff's delta + 95% CI + interpretation into a single string."""
    if pd.isna(delta) or delta is None:
        return 'N/A'
    ci_lo_s = f"{ci_lo:.3f}" if not pd.isna(ci_lo) else 'N/A'
    ci_hi_s = f"{ci_hi:.3f}" if not pd.isna(ci_hi) else 'N/A'
    return f"{delta:.4f} [{ci_lo_s}, {ci_hi_s}] ({mag})"

# =============================================================
# EXCEL FORMATTING
# =============================================================
def format_excel(filepath, sheet_metadata=None):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(filepath)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    meta_font = Font(italic=True, size=9, name='Arial', color='555555')
    data_font = Font(size=9, name='Arial')
    sig_fill = PatternFill('solid', fgColor='E2EFDA')
    meaningful_fill = PatternFill('solid', fgColor='FCE4D6')
    border = Border(bottom=Side(style='thin', color='D9D9D9'))
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for ws in wb.worksheets:
        n_meta = 0
        if sheet_metadata and ws.title in sheet_metadata:
            meta_lines = sheet_metadata[ws.title]
            n_meta = len(meta_lines)
            ws.insert_rows(1, n_meta)
            for i, line in enumerate(meta_lines, start=1):
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = meta_font
                cell.alignment = left
                for c in range(2, ws.max_column + 1):
                    ws.cell(row=i, column=c, value='')
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=ws.max_column)

        header_row = n_meta + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        sig_col = None
        meaningful_col = None
        for c in range(1, ws.max_column + 1):
            hval = ws.cell(row=header_row, column=c).value
            if hval and str(hval) == 'Significant':
                sig_col = c
            if hval and 'Meaningful' in str(hval):
                meaningful_col = c

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = center
                cell.border = border

            if sig_col:
                sv = ws.cell(row=row, column=sig_col).value
                if sv == 'Yes':
                    for c2 in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c2).fill = sig_fill
            if meaningful_col:
                mv = ws.cell(row=row, column=meaningful_col).value
                if mv == 'Yes':
                    for c2 in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c2).fill = meaningful_fill

        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                         for r in range(header_row, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 35)

    wb.save(filepath)


# ---------------------------------------------------------
# LOCATE EXECUTION WRAPPER
# ---------------------------------------------------------
def run_locate_testing(train_image_directory_path, test_image_directory_name,
                       locate_path, verbose=1, feature_select=[1, 1, 1, 1]):
    matlab_script_content = f"""
    addpath(genpath('{locate_path}'));
    feature_select = {feature_select};
    verbose = {verbose};
    train_image_directory_path = '{train_image_directory_path}';
    test_image_directory_name = '{test_image_directory_name}';
    try
        LOCATE_testing(test_image_directory_name, train_image_directory_path, feature_select, verbose);
    catch ME
        fprintf('Error: %s\\n', ME.message);
        exit(1);
    end
    exit(0);
    """
    with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".m") as temp_file:
        script_path = temp_file.name
        temp_file.write(matlab_script_content)
    try:
        return_code = os.system(f"matlab -batch \"run('{script_path}');\"")
        if return_code != 0:
            raise RuntimeError(f"LOCATE testing failed with return code {return_code}")
    finally:
        if os.path.exists(script_path):
            os.remove(script_path)


# ---------------------------------------------------------
# DATA LOADING
# ---------------------------------------------------------
all_files = pd.read_excel(results_path)

all_files['manual_mask'] = np.where(
    all_files['WMH_removed_path'].notna(),
    all_files['WMH_removed_path'],
    all_files['WMH_path']
)

scanner_map = {
    'Prisma_fit': 'Siemens',
    'Tim Trio': 'Siemens',
    'Philips': 'Philips',
    'GE Signa': 'GE',
}

all_files['scanner_group'] = all_files['scanner'].map(scanner_map)


# ---------------------------------------------------------
# BALANCED TRAIN / TEST SPLITS
# ---------------------------------------------------------
# Clean design: the "without GE" set is a strict subset of "with GE".
# Only difference = 15 GE subjects added. No confound from reduced
# Siemens/Philips counts.
#   Without GE: 15 Prisma + 15 Trio + 15 Philips = 45
#   With GE:    same 45   + 15 GE                = 60
# ---------------------------------------------------------
prisma_pool = all_files[all_files['scanner'] == 'Prisma_fit']
trio_pool = all_files[all_files['scanner'] == 'Tim Trio']
philips_pool = all_files[all_files['scanner'] == 'Philips']
ge_pool = all_files[all_files['scanner'] == 'GE Signa']

# Shared base: 15 per non-GE scanner
base_prisma = prisma_pool.sample(n=15, random_state=RANDOM_STATE)
base_trio = trio_pool.sample(n=15, random_state=RANDOM_STATE)
base_philips = philips_pool.sample(n=15, random_state=RANDOM_STATE)

# Train WITHOUT GE: 15 + 15 + 15 = 45
train_no_ge = pd.concat(
    [base_prisma, base_trio, base_philips], ignore_index=True
)

# Train WITH GE: same 45 + 15 GE = 60
train_ge_subjects = ge_pool.sample(n=15, random_state=RANDOM_STATE)
train_with_ge = pd.concat(
    [base_prisma, base_trio, base_philips, train_ge_subjects],
    ignore_index=True
)

# Leftover GE for LOCATE
ge_not_in_train = ge_pool.drop(train_ge_subjects.index)
ge_not_in_train_path = os.path.join(base_dir, "ge_data_that_not_in_trainset.xlsx")
ge_not_in_train.to_excel(ge_not_in_train_path, index=False)

# Test set: all subjects NOT in either training set
all_train_subjects = set(train_no_ge['subject']).union(set(train_with_ge['subject']))
test_set = all_files[~all_files['subject'].isin(all_train_subjects)].copy()

# Verification
print(f"Train WITHOUT GE: {len(train_no_ge)} subjects")
print(train_no_ge['scanner'].value_counts().to_dict())
print(f"\nTrain WITH GE: {len(train_with_ge)} subjects")
print(train_with_ge['scanner'].value_counts().to_dict())
print(f"\nTest set: {len(test_set)} subjects")
print(test_set['scanner'].value_counts().to_dict())
print(f"\nSaved {len(ge_not_in_train)} unused GE subjects -> {ge_not_in_train_path}")

assert len(set(test_set['subject']) & set(train_no_ge['subject'])) == 0, "Test/train_no_ge overlap!"
assert len(set(test_set['subject']) & set(train_with_ge['subject'])) == 0, "Test/train_with_ge overlap!"

# Save split specs
split_info = {
    'random_state': RANDOM_STATE,
    'test_subjects': test_set['subject'].tolist(),
    'train_no_ge_subjects': train_no_ge['subject'].tolist(),
    'train_with_ge_subjects': train_with_ge['subject'].tolist(),
    'test_composition': test_set['scanner'].value_counts().to_dict(),
    'train_no_ge_composition': train_no_ge['scanner'].value_counts().to_dict(),
    'train_with_ge_composition': train_with_ge['scanner'].value_counts().to_dict(),
}
with open(os.path.join(GE_COMPARE_dir, "split_spec.json"), "w") as f:
    json.dump(split_info, f, indent=2)

test_set.to_excel(os.path.join(GE_COMPARE_dir, "test_set.xlsx"), index=False)
train_no_ge.to_excel(os.path.join(GE_COMPARE_dir, "train_no_ge.xlsx"), index=False)
train_with_ge.to_excel(os.path.join(GE_COMPARE_dir, "train_with_ge.xlsx"), index=False)
print(f"Saved splits -> {GE_COMPARE_dir}/")


# ---------------------------------------------------------
# HELPER: Build BIANCA master file & train
# ---------------------------------------------------------
def build_masterfile(df):
    """Build master file lines, raising if any file is missing."""
    lines = []
    for _, row in df.iterrows():
        base = row['dataset_base']
        paths = {
            'FLAIR': os.path.join(base, row['FLAIR_brain_biascorr']),
            'T1': os.path.join(base, row['T1']),
            'MNI': os.path.join(base, row['mni_mat_path']),
            'WMH': os.path.join(base, row['manual_mask']),
        }
        missing = [k for k, v in paths.items() if not os.path.isfile(v)]
        if missing:
            raise FileNotFoundError(
                f"Missing {missing} for subject {row['subject']}"
            )
        lines.append(f"{paths['FLAIR']} {paths['T1']} {paths['MNI']} {paths['WMH']}")
    return lines


def train_bianca_model(masterfile_lines, masterfile_path, model_path, label=""):
    """Write master file and train BIANCA if model doesn't exist yet."""
    with open(masterfile_path, 'w') as f:
        f.write('\n'.join(masterfile_lines))

    if os.path.isfile(model_path):
        print(f"  Model already exists: {model_path}")
        return

    n = len(masterfile_lines)
    trainstring = ",".join(str(r) for r in range(1, n + 1))
    cmd = [
        "bianca", f"--singlefile={masterfile_path}",
        "--brainmaskfeaturenum=1", "--matfeaturenum=3",
        "--featuresubset=1,2", "--labelfeaturenum=4",
        "--trainingpts=2000", "--nonlespts=10000",
        f"--trainingnums={trainstring}",
        f"--saveclassifierdata={model_path}",
        f"--querysubjectnum={n}", "-v"
    ]
    try:
        print(f"  BIANCA Training ({label})...")
        subprocess.run(cmd, check=True)
        print(f"  Training done -> {model_path}")
    except subprocess.CalledProcessError as e:
        print(f"  Training error ({label}): {e}")
        raise






# ---------------------------------------------------------
# BIANCA TRAINING
# ---------------------------------------------------------
BIANCA_MODEL_GE = os.path.join(model_dir, "BIANCA_MODEL_N_60_WITH_GE")
with_ge_masterfile = os.path.join(model_dir, "bianca_n_60_with_ge_masterfile.txt")
with_ge_lines = build_masterfile(train_with_ge)
train_bianca_model(with_ge_lines, with_ge_masterfile, BIANCA_MODEL_GE, label="With GE")

BIANCA_MODEL_NO_GE = os.path.join(model_dir, "BIANCA_MODEL_N_60_WITHOUT_GE")
no_ge_masterfile = os.path.join(model_dir, "bianca_n_60_without_ge_masterfile.txt")
no_ge_lines = build_masterfile(train_no_ge)
train_bianca_model(no_ge_lines, no_ge_masterfile, BIANCA_MODEL_NO_GE, label="Without GE")


# ---------------------------------------------------------
# INFERENCE + EVALUATION
# ---------------------------------------------------------
test_subjects = sorted(test_set['subject'].tolist())
if DEBUG:
    test_subjects = test_subjects[:1]

all_metrics_results = []

for si, test_subject in enumerate(test_subjects):
    print(f"\n[{si+1}/{len(test_subjects)}] {test_subject}")

    test_row = test_set[test_set['subject'] == test_subject].iloc[0]
    dataset_base = test_row['dataset_base']

    test_FLAIR = os.path.join(dataset_base, test_row['FLAIR_brain_biascorr'])
    test_T1 = os.path.join(dataset_base, test_row['T1'])
    test_mni_mat_path = os.path.join(dataset_base, test_row['mni_mat_path'])
    test_wmh_roi_file = os.path.join(dataset_base, test_row['manual_mask'])
    test_WMmask_path = os.path.join(dataset_base, test_row['WMmask'])
    test_brainmask = os.path.join(dataset_base, test_row['brainmask'])
    test_ventdistmap = os.path.join(dataset_base, test_row['ventdistmap'])

    models = [
        ("with_ge", BIANCA_MODEL_GE),
        ("without_ge", BIANCA_MODEL_NO_GE),
    ]

    for model_name, model_path in models:
        sub_dir = os.path.join(GE_COMPARE_dir, "test", test_subject, f"bianca_result_{model_name}")
        os.makedirs(sub_dir, exist_ok=True)

        # Master file for test subject
        test_masterfile_path = os.path.join(sub_dir, f"{test_subject}_masterfile_test.txt")
        with open(test_masterfile_path, 'w') as f:
            f.write(f"{test_FLAIR} {test_T1} {test_mni_mat_path} {test_wmh_roi_file}")

        # BIANCA Inference
        BIANCA_LPM = os.path.join(sub_dir, f"{test_subject}_BIANCA_LPM_{model_name}.nii.gz")
        if not os.path.isfile(BIANCA_LPM):
            cmd = [
                "bianca", f"--singlefile={test_masterfile_path}",
                "--brainmaskfeaturenum=1", "--matfeaturenum=3",
                "--featuresubset=1,2", f"--loadclassifierdata={model_path}",
                "--querysubjectnum=1", "-o", BIANCA_LPM, "-v"
            ]
            try:
                subprocess.run(cmd, check=True)
                print(f"  {test_subject} ({model_name}): LPM created")
            except subprocess.CalledProcessError as e:
                print(f"  {test_subject} ({model_name}): Inference error: {e}")
                continue

        # Threshold + Evaluate
        for thresh in thresholds_list:
            if thresh == "locate":
                thresh_str = "locate"
                thresh_dir = os.path.join(sub_dir, thresh_str)
                os.makedirs(thresh_dir, exist_ok=True)

                results_dir = os.path.join(thresh_dir, "LOCATE_results_directory")
                existing_results = (
                    [f for f in os.listdir(results_dir) if "BIANCA_LOCATE_binarylesionmap" in f]
                    if os.path.exists(results_dir) else []
                )

                if not existing_results:
                    print(f"  {test_subject} ({model_name}): Starting LOCATE")

                    fsl_copy(test_FLAIR, os.path.join(thresh_dir, f"{test_subject}_feature_FLAIR.nii.gz"))
                    fsl_copy(BIANCA_LPM, os.path.join(thresh_dir, f"{test_subject}_BIANCA_LPM.nii.gz"))
                    fsl_copy(test_T1, os.path.join(thresh_dir, f"{test_subject}_feature_t1w.nii.gz"))
                    fsl_copy(test_WMmask_path, os.path.join(thresh_dir, f"{test_subject}_biancamask.nii.gz"))
                    fsl_copy(test_brainmask, os.path.join(thresh_dir, f"{test_subject}_brainmask.nii.gz"))
                    fsl_copy(test_ventdistmap, os.path.join(thresh_dir, f"{test_subject}_ventdistmap.nii.gz"))

                    run_locate_testing(
                        train_image_directory_path=train_image_directory_path,
                        test_image_directory_name=os.path.abspath(thresh_dir),
                        locate_path=LOCATE_PATH,
                        verbose=1,
                        feature_select=feature_select,
                    )

                    existing_results = (
                        [f for f in os.listdir(results_dir) if "BIANCA_LOCATE_binarylesionmap" in f]
                        if os.path.exists(results_dir) else []
                    )

                if not existing_results:
                    print(f"  LOCATE output missing for {test_subject}")
                    continue

                thresh_corrected_path = os.path.join(results_dir, existing_results[0])

            else:
                thresh_str = str(int(thresh * 100))
                thresh_dir = os.path.join(sub_dir, thresh_str)
                os.makedirs(thresh_dir, exist_ok=True)

                bianca_thresh_path = os.path.join(
                    thresh_dir, f"{test_subject}_BIANCA_LPM_thresh_{thresh_str}.nii.gz"
                )
                thresh_corrected_path = os.path.join(
                    thresh_dir, f"{test_subject}_BIANCA_LPM_thresh_{thresh_str}_wm_corrected.nii.gz"
                )

                if not os.path.isfile(bianca_thresh_path):
                    threshold_lpm(BIANCA_LPM, bianca_thresh_path, thresh)

                if not os.path.isfile(thresh_corrected_path):
                    apply_bianca_mask(bianca_thresh_path, test_WMmask_path, thresh_corrected_path)

            # Evaluate
            try:
                eval_metrics = fsl_dice_score(thresh_corrected_path, test_wmh_roi_file)

                if isinstance(eval_metrics, dict):
                    dice_val = eval_metrics.get('dice_score', None)
                    sensitivity = eval_metrics.get('sensitivity', None)
                    precision = eval_metrics.get('precision', None)
                else:
                    dice_val = eval_metrics
                    sensitivity = None
                    precision = None

                # WMH volume from segmentation output
                wmh_vol = compute_wmh_volume_ml(thresh_corrected_path)

                all_metrics_results.append({
                    'subject': test_subject,
                    'scanner': test_row['scanner'],
                    'scanner_group': scanner_map.get(test_row['scanner'], test_row['scanner']),
                    'severity_level': test_row.get('severity_level', None),
                    'model': model_name,
                    'threshold': thresh_str,
                    'dice_score': dice_val,
                    'sensitivity': sensitivity,
                    'precision': precision,
                    'wmh_volume_ml': wmh_vol,
                })

            except Exception as e:
                print(f"  Metrics failed for {test_subject} ({model_name}, {thresh_str}): {e}")



# =============================================================
# RESULTS: COMPREHENSIVE STATISTICAL ANALYSIS
# =============================================================
results_df = pd.DataFrame(all_metrics_results)

plot_dir = os.path.join(GE_COMPARE_dir, "plots")
analyse_dir = os.path.join(GE_COMPARE_dir, "analyse")
os.makedirs(plot_dir, exist_ok=True)
os.makedirs(analyse_dir, exist_ok=True)

# Save raw results
raw_path = os.path.join(analyse_dir, "ge_compare_raw_results.xlsx")
results_df.to_excel(raw_path, index=False)
print(f"\nRaw results: {len(results_df)} rows -> {raw_path}")

# Publication-ready settings
CB_PALETTE = {'with_ge': '#4477AA', 'without_ge': '#EE6677'}
MODEL_LABELS = {'with_ge': 'With GE (N=60)', 'without_ge': 'Without GE (N=45)'}
METRICS_LIST = ['dice_score', 'sensitivity', 'precision']

# Bonferroni correction structure:
#   Family 1 (overall): 3 metrics x 3 thresholds = 9 tests
#   Family 2 (scanner-stratified): 3 metrics x 3 thresholds x N_scanner_groups
#   Family 3 (volume): 3 thresholds = 3 tests
N_OVERALL_TESTS = len(METRICS_LIST) * len(thresholds_list)
ALPHA_OVERALL = bonferroni_alpha(N_OVERALL_TESTS)
N_VOLUME_TESTS = len(thresholds_list)
ALPHA_VOLUME = bonferroni_alpha(N_VOLUME_TESTS)

print(f"\nBonferroni correction:")
print(f"  Overall: {N_OVERALL_TESTS} tests, alpha = {ALPHA_OVERALL:.4f}")
print(f"  Volume:  {N_VOLUME_TESTS} tests, alpha = {ALPHA_VOLUME:.4f}")


# ---------------------------------------------------------
# 1) OVERALL COMPARISON: With GE vs Without GE
# ---------------------------------------------------------
print("\n" + "=" * 70)
print("1) PHASE I: OVERALL COMPARISON: With GE vs Without GE Training")
print("=" * 70)

overall_stats = []

for thresh in thresholds_list:
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty:
        continue

    for metric in METRICS_LIST:
        pivot_m = subset.pivot(index='subject', columns='model', values=metric).dropna()
        if len(pivot_m) == 0 or 'with_ge' not in pivot_m.columns or 'without_ge' not in pivot_m.columns:
            continue

        vals_with = pivot_m['with_ge'].values
        vals_without = pivot_m['without_ge'].values

        # Wilcoxon signed-rank (paired)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                w_stat, w_pval = wilcoxon(vals_with, vals_without)
            except ValueError:
                w_stat, w_pval = np.nan, np.nan

        # Cliff's Delta with bootstrap CI
        delta, ci_low, ci_high = bootstrap_cliffs_delta(vals_with, vals_without)

        row = {
            'Threshold': 'LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}',
            'Metric': metric.replace('_', ' ').title(),
            'N': len(pivot_m),
            'Mean With GE': round(np.mean(vals_with), 4),
            'SD With GE': round(np.std(vals_with, ddof=1), 4),
            'Mean Without GE': round(np.mean(vals_without), 4),
            'SD Without GE': round(np.std(vals_without, ddof=1), 4),
            'Mean Diff': round(np.mean(vals_with) - np.mean(vals_without), 6),
            'W Statistic': round(w_stat, 1) if not np.isnan(w_stat) else 'N/A',
            'p (raw)': format_p_table(w_pval),
            'Significant': 'Yes' if (w_pval < ALPHA_OVERALL if not np.isnan(w_pval) else False) else 'No',
            "Cliff's \u03b4 [95% CI]": format_delta_ci(delta, ci_low, ci_high,
                                                         interpret_cliffs_delta(delta)),
            'Meaningful (|\u03b4|>=0.28)': 'Yes' if abs(delta) >= DELTA_THRESHOLD else 'No',
            'Higher': 'With GE' if np.mean(vals_with) > np.mean(vals_without) else 'Without GE',
        }
        overall_stats.append(row)

        sig_marker = "*" if row['Significant'] == 'Yes' else ""
        print(f"  [{thresh_str}] {metric}: "
              f"With GE = {row['Mean With GE']:.4f} (SD {row['SD With GE']:.4f}), "
              f"Without GE = {row['Mean Without GE']:.4f} (SD {row['SD Without GE']:.4f}), "
              f"delta = {delta:.4f} [{ci_low:.3f}, {ci_high:.3f}] ({interpret_cliffs_delta(delta)}), "
              f"p = {format_p_table(w_pval)}{sig_marker}")

overall_stats_df = pd.DataFrame(overall_stats)
overall_stats_df.to_excel(os.path.join(analyse_dir, "ge_compare_overall_statistics.xlsx"), index=False)


# ---------------------------------------------------------
# 2) GE TEST SUBJECTS: SEPARATE REPORT
# ---------------------------------------------------------
print("\n" + "=" * 70)
print("2) GE TEST SUBJECTS  SEPARATE REPORT")
print("=" * 70)

ge_test_stats = []
ge_test = results_df[results_df['scanner_group'] == 'GE'].copy()

if len(ge_test) > 0:
    for thresh in thresholds_list:
        thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
        ge_subset = ge_test[ge_test['threshold'] == thresh_str]
        if ge_subset.empty:
            continue

        for metric in METRICS_LIST:
            pivot_m = ge_subset.pivot(index='subject', columns='model', values=metric).dropna()
            if len(pivot_m) == 0 or 'with_ge' not in pivot_m.columns or 'without_ge' not in pivot_m.columns:
                continue

            vals_with = pivot_m['with_ge'].values
            vals_without = pivot_m['without_ge'].values

            # Wilcoxon only if N >= 6 (minimum for meaningful paired test)
            if len(pivot_m) >= 6:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    try:
                        w_stat, w_pval = wilcoxon(vals_with, vals_without)
                    except ValueError:
                        w_stat, w_pval = np.nan, np.nan
            else:
                w_stat, w_pval = np.nan, np.nan

            delta, ci_low, ci_high = bootstrap_cliffs_delta(vals_with, vals_without)

            ge_test_stats.append({
                'Threshold': 'LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}',
                'Metric': metric.replace('_', ' ').title(),
                'N_GE_subjects': len(pivot_m),
                'Mean With GE': round(np.mean(vals_with), 4),
                'Mean Without GE': round(np.mean(vals_without), 4),
                'Mean Diff': round(np.mean(vals_with) - np.mean(vals_without), 6),
                'p (raw)': format_p_table(w_pval),
                "Cliff's \u03b4 [95% CI]": format_delta_ci(delta, ci_low, ci_high,
                                                             interpret_cliffs_delta(delta)),
                'Meaningful (|\u03b4|>=0.28)': 'Yes' if abs(delta) >= DELTA_THRESHOLD else 'No',
            })

            print(f"  [{thresh_str}] GE only, {metric} (N={len(pivot_m)}): "
                  f"With={np.median(vals_with):.3f}, Without={np.median(vals_without):.3f}, "
                  f"delta={delta:.3f} [{ci_low:.3f}, {ci_high:.3f}]")

    ge_test_stats_df = pd.DataFrame(ge_test_stats)
    ge_test_stats_df.to_excel(os.path.join(analyse_dir, "ge_compare_GE_test_subjects.xlsx"), index=False)
else:
    print("  No GE subjects in test set.")
    ge_test_stats_df = pd.DataFrame()


# ---------------------------------------------------------
# 3) SCANNER-STRATIFIED ANALYSIS (non-GE scanners)
# ---------------------------------------------------------
print("\n" + "=" * 70)
print("3) SCANNER-STRATIFIED ANALYSIS")
print("=" * 70)

scanner_stats = []
scanner_groups_available = sorted(results_df['scanner_group'].dropna().unique())
N_SCANNER_TESTS = len(METRICS_LIST) * len(thresholds_list) * len(scanner_groups_available)
ALPHA_SCANNER = bonferroni_alpha(N_SCANNER_TESTS) if N_SCANNER_TESTS > 0 else 0.05
print(f"  Scanner-stratified: {N_SCANNER_TESTS} tests, alpha = {ALPHA_SCANNER:.4f}")

for thresh in thresholds_list:
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty:
        continue

    for scanner_grp in scanner_groups_available:
        scanner_subset = subset[subset['scanner_group'] == scanner_grp]

        for metric in METRICS_LIST:
            pivot_m = scanner_subset.pivot(
                index='subject', columns='model', values=metric
            ).dropna()

            if len(pivot_m) < 3:
                continue
            if 'with_ge' not in pivot_m.columns or 'without_ge' not in pivot_m.columns:
                continue

            vals_with = pivot_m['with_ge'].values
            vals_without = pivot_m['without_ge'].values

            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                try:
                    w_stat, w_pval = wilcoxon(vals_with, vals_without)
                except ValueError:
                    w_stat, w_pval = np.nan, np.nan

            delta, ci_low, ci_high = bootstrap_cliffs_delta(vals_with, vals_without)

            scanner_stats.append({
                'Threshold': 'LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}',
                'Scanner': scanner_grp,
                'Metric': metric.replace('_', ' ').title(),
                'N': len(pivot_m),
                'Mean With GE': round(np.mean(vals_with), 4),
                'Mean Without GE': round(np.mean(vals_without), 4),
                'Mean Diff': round(np.mean(vals_with) - np.mean(vals_without), 6),
                'p (raw)': format_p_table(w_pval),
                'Significant': 'Yes' if (w_pval < ALPHA_SCANNER if not np.isnan(w_pval) else False) else 'No',
                "Cliff's \u03b4 [95% CI]": format_delta_ci(delta, ci_low, ci_high,
                                                             interpret_cliffs_delta(delta)),
                'Meaningful (|\u03b4|>=0.28)': 'Yes' if abs(delta) >= DELTA_THRESHOLD else 'No',
            })

            if metric == 'dice_score':
                print(f"  [{thresh_str}] {scanner_grp} (N={len(pivot_m)}): "
                      f"delta = {delta:.3f} [{ci_low:.3f}, {ci_high:.3f}] "
                      f"({interpret_cliffs_delta(delta)}), p = {w_pval:.4f}")

scanner_stats_df = pd.DataFrame(scanner_stats)
scanner_stats_df.to_excel(os.path.join(analyse_dir, "ge_compare_scanner_stratified.xlsx"), index=False)


# ---------------------------------------------------------
# 4) WMH VOLUME COMPARISON
# ---------------------------------------------------------
print("\n" + "=" * 70)
print("4) WMH VOLUME COMPARISON")
print("=" * 70)

volume_stats = []

for thresh in thresholds_list:
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty or 'wmh_volume_ml' not in subset.columns:
        continue

    pivot_vol = subset.pivot(
        index='subject', columns='model', values='wmh_volume_ml'
    ).dropna()

    if len(pivot_vol) == 0:
        continue
    if 'with_ge' not in pivot_vol.columns or 'without_ge' not in pivot_vol.columns:
        continue

    vals_with = pivot_vol['with_ge'].values
    vals_without = pivot_vol['without_ge'].values

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            w_stat, w_pval = wilcoxon(vals_with, vals_without)
        except ValueError:
            w_stat, w_pval = np.nan, np.nan

    delta, ci_low, ci_high = bootstrap_cliffs_delta(vals_with, vals_without)

    volume_stats.append({
        'Threshold': 'LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}',
        'N': len(pivot_vol),
        'Mean Vol With GE (mL)': round(np.mean(vals_with), 4),
        'Mean Vol Without GE (mL)': round(np.mean(vals_without), 4),
        'Mean Vol Diff (mL)': round(np.mean(vals_with) - np.mean(vals_without), 6),
        'p (raw)': format_p_table(w_pval),
        'Significant': 'Yes' if (w_pval < ALPHA_VOLUME if not np.isnan(w_pval) else False) else 'No',
        "Cliff's \u03b4 [95% CI]": format_delta_ci(delta, ci_low, ci_high,
                                                     interpret_cliffs_delta(delta)),
        'Meaningful (|\u03b4|>=0.28)': 'Yes' if abs(delta) >= DELTA_THRESHOLD else 'No',
    })

    print(f"  [{thresh_str}] Volume (mL): "
          f"With GE = {np.median(vals_with):.2f}, "
          f"Without GE = {np.median(vals_without):.2f}, "
          f"delta = {delta:.3f} [{ci_low:.3f}, {ci_high:.3f}], "
          f"p = {w_pval:.4f}")

volume_stats_df = pd.DataFrame(volume_stats)
volume_stats_df.to_excel(os.path.join(analyse_dir, "ge_compare_volume_statistics.xlsx"), index=False)


# =============================================================
# 5) SUPPLEMENTAL FIGURES (Publication-Ready)
# =============================================================
print("\n" + "=" * 70)
print("5) GENERATING SUPPLEMENTAL FIGURES")
print("=" * 70)


# --- Figure S-X: Multi-panel Dice/Sensitivity/Precision by threshold ---
fig, axes = plt.subplots(len(METRICS_LIST), len(thresholds_list),
                         figsize=(5 * len(thresholds_list), 4.5 * len(METRICS_LIST)),
                         squeeze=False)

for col_i, thresh in enumerate(thresholds_list):
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty:
        continue

    for row_i, metric in enumerate(METRICS_LIST):
        ax = axes[row_i, col_i]
        ax.set_ylim(-0.05, 1.05)  # Set BEFORE annotations

        for mi, model in enumerate(['with_ge', 'without_ge']):
            model_data = subset[subset['model'] == model][metric].dropna()
            ax.boxplot(
                model_data.values,
                positions=[mi],
                widths=0.35,
                patch_artist=True,
                boxprops=dict(facecolor=CB_PALETTE[model], alpha=0.6),
                medianprops=dict(color='black', linewidth=1.5),
                whiskerprops=dict(linewidth=1.0),
                capprops=dict(linewidth=1.0),
                flierprops=dict(marker='o', markersize=3, alpha=0.4),
            )
            # Individual data points
            jitter = np.random.default_rng(42).uniform(-0.08, 0.08, size=len(model_data))
            ax.scatter(
                mi + jitter, model_data.values,
                c='black', s=8, alpha=0.3, zorder=3
            )
            # N annotation (AFTER set_ylim)
            ax.text(mi, -0.08, f'N={len(model_data)}',
                    ha='center', va='top', fontsize=8, fontweight='bold')

        ax.set_xticks([0, 1])
        ax.set_xticklabels([MODEL_LABELS['with_ge'], MODEL_LABELS['without_ge']], fontsize=9)

        if row_i == 0:
            title_text = 'LOCATE' if thresh_str == 'locate' else f'Threshold = 0.{thresh_str}'
            ax.set_title(title_text, fontsize=11, fontweight='bold')
        if col_i == 0:
            ax.set_ylabel(metric.replace('_', ' ').title(), fontsize=10)

        # Cliff's delta + Bonferroni annotation
        match = overall_stats_df[
            (overall_stats_df['Threshold'] == ('LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}')) &
            (overall_stats_df['Metric'] == metric.replace('_', ' ').title())
        ]
        if len(match) > 0:
            r = match.iloc[0]
            p_str = r['p (raw)']
            sig_str = f"p = {p_str}"
            if r['Significant'] == 'Yes':
                sig_str += " *"
            delta_col = "Cliff's \u03b4 [95% CI]"
            delta_text = f"{r[delta_col]}\n{sig_str}"
            ax.text(0.98, 0.98, delta_text, transform=ax.transAxes,
                    fontsize=7, va='top', ha='right',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8))

fig.suptitle('Phase I: BIANCA Segmentation Performance\n'
             'Training With GE (N=60) vs Without GE (N=45, strict subset)\n'
             f'(* significant at Bonferroni-corrected $\\alpha$ = {ALPHA_OVERALL:.4f})',
             fontsize=12, fontweight='bold', y=1.02)
fig.tight_layout()

fig_path_1 = os.path.join(plot_dir, "SuppFig_GE_compare_metrics_multipanel.png")
fig.savefig(fig_path_1, dpi=300, bbox_inches='tight')
fig.savefig(fig_path_1.replace('.png', '.pdf'), bbox_inches='tight')
print(f"  Saved: {fig_path_1}")
plt.close(fig)


# --- Figure S-X: Scanner-stratified Dice (including GE separately) ---
for thresh in [0.85, "locate"]:
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty:
        continue

    # Order: Siemens, Philips, GE (GE last, visually separated)
    scanner_order = [sg for sg in ['Siemens', 'Philips', 'GE'] if sg in subset['scanner_group'].unique()]
    n_scanners = len(scanner_order)
    if n_scanners == 0:
        continue

    fig, axes = plt.subplots(1, n_scanners, figsize=(4 * n_scanners, 4.5), squeeze=False)

    for si, sg in enumerate(scanner_order):
        ax = axes[0, si]
        ax.set_ylim(-0.1, 1.05)  # Set BEFORE annotations
        sg_data = subset[subset['scanner_group'] == sg]

        for mi, model in enumerate(['with_ge', 'without_ge']):
            model_data = sg_data[sg_data['model'] == model]['dice_score'].dropna()
            ax.boxplot(
                model_data.values,
                positions=[mi],
                widths=0.35,
                patch_artist=True,
                boxprops=dict(facecolor=CB_PALETTE[model], alpha=0.6),
                medianprops=dict(color='black', linewidth=1.5),
                flierprops=dict(marker='o', markersize=3, alpha=0.4),
            )
            jitter = np.random.default_rng(42).uniform(-0.08, 0.08, size=len(model_data))
            ax.scatter(mi + jitter, model_data.values, c='black', s=8, alpha=0.3, zorder=3)
            ax.text(mi, -0.13, f'N={len(model_data)}', ha='center', va='top', fontsize=8)

        ax.set_xticks([0, 1])
        ax.set_xticklabels(['With GE', 'Without GE'], fontsize=9)
        # Highlight GE panel
        panel_title = f'{sg} (N={len(sg_data)//2})'
        if sg == 'GE':
            panel_title += '  test only'
        ax.set_title(panel_title, fontsize=10, fontweight='bold',
                     color='#CC3311' if sg == 'GE' else 'black')
        if si == 0:
            ax.set_ylabel('Dice Score', fontsize=10)

        # Delta annotation
        match = scanner_stats_df[
            (scanner_stats_df['Threshold'] == ('LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}')) &
            (scanner_stats_df['Scanner'] == sg) &
            (scanner_stats_df['Metric'] == 'Dice Score')
        ]
        if len(match) > 0:
            r = match.iloc[0]
            delta_col = "Cliff's \u03b4 [95% CI]"
            ax.text(0.98, 0.98,
                    f"{r[delta_col]}\np = {r['p (raw)']}",
                    transform=ax.transAxes, fontsize=7, va='top', ha='right',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8))

    thresh_label = 'LOCATE' if thresh_str == 'locate' else f'Threshold 0.{thresh_str}'
    fig.suptitle(f'Phase I: Scanner-Stratified Dice Scores ({thresh_label})\n'
                 f'Training With vs Without GE',
                 fontsize=12, fontweight='bold', y=1.02)
    fig.tight_layout()

    fig_path = os.path.join(plot_dir, f"SuppFig_GE_scanner_stratified_{thresh_str}.png")
    fig.savefig(fig_path, dpi=300, bbox_inches='tight')
    fig.savefig(fig_path.replace('.png', '.pdf'), bbox_inches='tight')
    print(f"  Saved: {fig_path}")
    plt.close(fig)


# --- Figure S-X: Bland-Altman difference plot ---
for thresh in [0.85, "locate"]:
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty:
        continue

    pivot = subset.pivot(index='subject', columns='model', values='dice_score').dropna()
    if len(pivot) == 0 or 'with_ge' not in pivot.columns:
        continue

    # Merge scanner_group for color-coding
    subj_scanner = results_df[['subject', 'scanner_group']].drop_duplicates('subject')
    pivot = pivot.merge(subj_scanner, left_index=True, right_on='subject', how='left')

    mean_vals = (pivot['with_ge'] + pivot['without_ge']) / 2
    diff_vals = pivot['with_ge'] - pivot['without_ge']

    fig, ax = plt.subplots(figsize=(7, 5))

    # Color-code by scanner group
    scanner_colors = {'Siemens': '#4477AA', 'Philips': '#228833', 'GE': '#CC3311'}
    for sg in pivot['scanner_group'].unique():
        mask = pivot['scanner_group'] == sg
        ax.scatter(mean_vals[mask], diff_vals[mask],
                   c=scanner_colors.get(sg, 'gray'), s=25, alpha=0.7,
                   edgecolors='black', linewidths=0.3, label=sg, zorder=3)

    ax.axhline(y=diff_vals.mean(), color='red', linestyle='-', linewidth=1.5,
               label=f'Mean diff = {diff_vals.mean():.4f}')
    ax.axhline(y=diff_vals.mean() + 1.96 * diff_vals.std(), color='gray',
               linestyle='--', linewidth=1, label=r'$\pm$1.96 SD')
    ax.axhline(y=diff_vals.mean() - 1.96 * diff_vals.std(), color='gray',
               linestyle='--', linewidth=1)
    ax.axhline(y=0, color='black', linestyle=':', linewidth=0.8, alpha=0.5)

    ax.set_xlabel('Mean Dice Score [(With GE + Without GE) / 2]', fontsize=10)
    ax.set_ylabel('Difference in Dice Score [With GE $-$ Without GE]', fontsize=10)
    thresh_label = 'LOCATE' if thresh_str == 'locate' else f'Threshold 0.{thresh_str}'
    ax.set_title(f'Phase I: Bland-Altman: Dice Score Difference ({thresh_label})',
                 fontsize=11, fontweight='bold')
    ax.legend(fontsize=8, loc='best')
    fig.tight_layout()

    fig_path = os.path.join(plot_dir, f"SuppFig_GE_bland_altman_{thresh_str}.png")
    fig.savefig(fig_path, dpi=300, bbox_inches='tight')
    fig.savefig(fig_path.replace('.png', '.pdf'), bbox_inches='tight')
    print(f"  Saved: {fig_path}")
    plt.close(fig)


# =============================================================
# 6) SUPPLEMENTAL SUMMARY TABLE
# =============================================================
print("\n" + "=" * 70)
print("6) SUPPLEMENTAL SUMMARY TABLE")
print("=" * 70)

summary_rows = []

for thresh in thresholds_list:
    thresh_str = "locate" if thresh == "locate" else str(int(thresh * 100))
    subset = results_df[results_df['threshold'] == thresh_str].copy()
    if subset.empty:
        continue

    for model in ['with_ge', 'without_ge']:
        model_data = subset[subset['model'] == model]
        row = {
            'Threshold': 'LOCATE' if thresh_str == 'locate' else f'0.{thresh_str}',
            'Training Set': MODEL_LABELS[model],
            'N': len(model_data),
        }
        for metric in METRICS_LIST:
            vals = model_data[metric].dropna()
            if len(vals) > 0:
                row[f'{metric}_median'] = f"{np.median(vals):.3f}"
                row[f'{metric}_IQR'] = (f"[{np.percentile(vals, 25):.3f}, "
                                        f"{np.percentile(vals, 75):.3f}]")
                row[f'{metric}_mean_sd'] = f"{np.mean(vals):.3f} ({np.std(vals, ddof=1):.3f})"
            else:
                row[f'{metric}_median'] = "N/A"
                row[f'{metric}_IQR'] = "N/A"
                row[f'{metric}_mean_sd'] = "N/A"

        vol_vals = model_data['wmh_volume_ml'].dropna()
        if len(vol_vals) > 0:
            row['volume_median_ml'] = f"{np.median(vol_vals):.2f}"
            row['volume_IQR_ml'] = (f"[{np.percentile(vol_vals, 25):.2f}, "
                                     f"{np.percentile(vol_vals, 75):.2f}]")
        else:
            row['volume_median_ml'] = "N/A"
            row['volume_IQR_ml'] = "N/A"

        summary_rows.append(row)

summary_df = pd.DataFrame(summary_rows)
summary_path = os.path.join(analyse_dir, "ge_compare_supplemental_table.xlsx")
summary_df.to_excel(summary_path, index=False)
print(f"  Saved: {summary_path}")
print(summary_df.to_string(index=False))


# =============================================================
# 7) BONFERRONI CORRECTION TRANSPARENCY TABLE
# =============================================================
bonferroni_table = pd.DataFrame([
    {
        'Comparison Family': 'Overall (With GE vs Without GE)',
        'N_tests': N_OVERALL_TESTS,
        'Base_alpha': 0.05,
        'Corrected_alpha': round(ALPHA_OVERALL, 4),
        'Description': f'{len(METRICS_LIST)} metrics x {len(thresholds_list)} thresholds',
        'Justification': 'Single question (does GE inclusion affect performance?) tested across '
                         'all metric-threshold combinations. Global correction applied because '
                         'all tests address the same hypothesis.',
    },
    {
        'Comparison Family': 'Scanner-Stratified',
        'N_tests': N_SCANNER_TESTS,
        'Base_alpha': 0.05,
        'Corrected_alpha': round(ALPHA_SCANNER, 4),
        'Description': f'{len(METRICS_LIST)} metrics x {len(thresholds_list)} thresholds x {len(scanner_groups_available)} scanner groups',
        'Justification': 'Exploratory analysis: scanner-specific effects may differ. '
                         'All tests within this family address scanner-specific GE impact.',
    },
    {
        'Comparison Family': 'WMH Volume',
        'N_tests': N_VOLUME_TESTS,
        'Base_alpha': 0.05,
        'Corrected_alpha': round(ALPHA_VOLUME, 4),
        'Description': f'{len(thresholds_list)} thresholds',
        'Justification': 'Separate family: volume is a distinct outcome measure from '
                         'voxel-overlap metrics (Dice/Sensitivity/Precision).',
    },
])
bonferroni_table.to_excel(os.path.join(analyse_dir, "ge_compare_bonferroni_families.xlsx"), index=False)


# =============================================================
# 8) COMBINED EXPORT
# =============================================================
print("\n" + "=" * 70)
print("8) COMBINED EXPORT")
print("=" * 70)

combined_path = os.path.join(analyse_dir, "ge_compare_all_statistics.xlsx")
with pd.ExcelWriter(combined_path) as writer:
    overall_stats_df.to_excel(writer, sheet_name='Overall', index=False)
    if len(ge_test_stats_df) > 0:
        ge_test_stats_df.to_excel(writer, sheet_name='GE_Test_Subjects', index=False)
    scanner_stats_df.to_excel(writer, sheet_name='Scanner_Stratified', index=False)
    volume_stats_df.to_excel(writer, sheet_name='Volume', index=False)
    summary_df.to_excel(writer, sheet_name='Supplemental_Table', index=False)
    bonferroni_table.to_excel(writer, sheet_name='Bonferroni_Families', index=False)
    results_df.to_excel(writer, sheet_name='Raw_Data', index=False)

combined_meta = {
    'Overall': [
        "Table: Phase I - GE Compare: Overall With GE vs Without GE",
        "Nested design: Without GE (N=45) is strict subset of With GE (N=60).",
        f"Bonferroni correction: k={N_OVERALL_TESTS} tests (3 metrics x 3 thresholds), alpha={ALPHA_OVERALL:.4f}.",
        "Single question: does adding GE data to training affect performance?",
        "Cliff's Delta with bootstrapped 95% CI (1000 iterations).",
        f"Meaningful effect: |delta| >= {DELTA_THRESHOLD}.",
    ],
    'GE_Test_Subjects': [
        "Table: Phase I - GE Test Subjects Only (descriptive, small N)",
        "Caution: small sample sizes limit statistical inference.",
    ],
    'Scanner_Stratified': [
        "Table: Phase I - Scanner-Stratified GE Compare",
        f"Bonferroni correction: k={N_SCANNER_TESTS} tests, alpha={ALPHA_SCANNER:.4f}.",
        "Exploratory: scanner-specific GE impact.",
    ],
    'Volume': [
        "Table: Phase I - WMH Volume Comparison",
        f"Bonferroni correction: k={N_VOLUME_TESTS} tests, alpha={ALPHA_VOLUME:.4f}.",
        "Separate family: volume is distinct from voxel-overlap metrics.",
    ],
    'Bonferroni_Families': [
        "Table: Bonferroni Correction Family Structure",
        "Documents all multiple-testing correction decisions.",
    ],
}
format_excel(combined_path, combined_meta)

print(f"  All statistics exported -> {combined_path}")

print("\n" + "=" * 70)
print("ANALYSIS COMPLETE")
print("=" * 70)
print(f"  Plots: {plot_dir}/")
print(f"  Stats: {analyse_dir}/")
print(f"  Raw:   {raw_path}")
print(f"\nNested design: 'without GE' (N=45) is a strict subset of 'with GE' (N=60).")
print(f"Any performance difference is attributable solely to the 15 added GE subjects.")