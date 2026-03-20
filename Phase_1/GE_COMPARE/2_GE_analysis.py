#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GE Compare: Statistical Analysis (Paper-Ready Tables Only)
==========================================================

Revision context (NeuroImage: Clinical, Major Revision)
-------------------------------------------------------
This script addresses:
  (1) GE exclusion and robustness claims (R1 Comment 4; R5 #8)
  (2) Multiple comparisons transparency (R5 #5)

Technical design:
  - Nested training: Without GE (N=45) is strict subset of With GE (N=60)
  - Any performance difference attributable solely to 15 added GE subjects
  - Wilcoxon signed-rank (paired) with Bonferroni correction
  - Cliff's Delta with bootstrapped 95% CIs (1000 iterations)

Paper changes:
  Methods 2.3: Scanner exclusion rationale
  Supplemental: Tables S-X (this output)
  Limitations: GE generalizability acknowledged

Response to Reviewers:
  R1 Comment 4: Empirical GE comparison; "robustness" qualified
  R5 #8: Selection bias acknowledged; supplemental evidence

Output (single .xlsx, 4 sheets):
  1. Wilcoxon_Descriptive   combined inferential + descriptive stats
  2. Bland_Altman            agreement analysis
  3. Scanner_Stratified      per-scanner GE impact
  4. Bonferroni_Families     multiple testing transparency

Key findings (from actual results)
-----------------------------------
  Overall (N=70 test subjects):
    - All Cliff's Delta values are negligible (|delta| <= 0.047).
    - GE inclusion in training has no meaningful effect on non-GE
      segmentation performance.
    - Consistent pattern across all thresholds (B0.85, B0.90, B+L):
      Precision increases marginally with GE training (+0.014),
      Sensitivity decreases marginally (-0.005 to -0.009).
      Both shifts are negligible in magnitude.

  Bland-Altman:
    - Mean differences maximally 0.015 (Precision). LoAs are narrow.
    - No systematic bias for Dice (p = 0.097 / 0.741).
    - Statistically significant but negligible bias for Sensitivity
      and Precision -- reflects high statistical power from paired
      design (N=70), not meaningful effect size.

  Scanner-stratified:
    - Siemens and Philips: all effects negligible (|delta| < 0.10).
    - GE test subjects (N=5): medium-sized deltas (0.28-0.36) appear,
      but N=5 is too small for reliable inference. This supports the
      GE exclusion rationale: GE subjects benefit from GE-specific
      training data, but the sample is insufficient for generalization.

  Interpretation note for manuscript:
    - Several Wilcoxon tests reach statistical significance after
      Bonferroni correction (e.g., Sensitivity p < 0.001), but all
      corresponding Cliff's Delta values remain negligible. This is
      the classic "statistically significant but practically meaningless"
      scenario in paired designs with adequate sample size. The manuscript
      must frame this explicitly: statistical significance reflects
      the high power of paired comparisons at N=70, not the magnitude
      of the effect. Effect sizes (Cliff's Delta), not p-values, should
      drive the substantive conclusions.

References:
  Griffanti et al. (2016). BIANCA. NeuroImage: Clinical, 9, 235-242.
  Romano et al. (2006). Cliff's Delta for ordinal effect sizes.
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.stats import wilcoxon
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cliffs_delta import cliffs_delta as _cliffs_delta

warnings.filterwarnings('ignore')

# =============================================================
# CONFIGURATION
# =============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_RESULTS_PATH = os.path.join(SCRIPT_DIR, "analyse", "ge_compare_raw_results.xlsx")
ANALYSE_DIR = os.path.join(SCRIPT_DIR, "analyse")
os.makedirs(ANALYSE_DIR, exist_ok=True)

RANDOM_STATE = 42
N_BOOTSTRAP = 1000
DELTA_THRESHOLD = 0.28
METRICS_LIST = ['dice_score', 'sensitivity', 'precision']
MODEL_LABELS = {'with_ge': 'With GE (N=60)', 'without_ge': 'Without GE (N=45)'}


# =============================================================
# HELPER FUNCTIONS
# =============================================================
def bootstrap_cliffs_delta(x, y, n_boot=N_BOOTSTRAP, seed=RANDOM_STATE):
    rng = np.random.RandomState(seed)
    delta, _ = _cliffs_delta(x.tolist(), y.tolist())
    boot_deltas = []
    for _ in range(n_boot):
        d, _ = _cliffs_delta(
            rng.choice(x, size=len(x), replace=True).tolist(),
            rng.choice(y, size=len(y), replace=True).tolist()
        )
        boot_deltas.append(d)
    return delta, np.percentile(boot_deltas, 2.5), np.percentile(boot_deltas, 97.5)


def interpret_cliffs_delta(delta):
    abs_d = abs(delta)
    if abs_d < 0.147:
        return "negligible"
    elif abs_d < 0.28:
        return "small"
    elif abs_d < 0.43:
        return "medium"
    return "large"


def bonferroni_alpha(n_tests, base_alpha=0.05):
    return base_alpha / n_tests


def fmt_p(p):
    if pd.isna(p):
        return 'N/A'
    return '<0.001' if p < 0.001 else f"{p:.3f}"


def fmt_delta(delta, ci_lo, ci_hi):
    if pd.isna(delta):
        return 'N/A'
    mag = interpret_cliffs_delta(delta)
    return f"{delta:.4f} [{ci_lo:.3f}, {ci_hi:.3f}] ({mag})"


def format_excel(filepath, sheet_metadata=None):
    wb = load_workbook(filepath)
    h_fill = PatternFill('solid', fgColor='2F5496')
    h_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    m_font = Font(italic=True, size=9, name='Arial', color='555555')
    d_font = Font(size=9, name='Arial')
    brd = Border(bottom=Side(style='thin', color='D9D9D9'))
    ctr = Alignment(horizontal='center', vertical='center')
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for ws in wb.worksheets:
        n_meta = 0
        if sheet_metadata and ws.title in sheet_metadata:
            lines = sheet_metadata[ws.title]
            n_meta = len(lines)
            ws.insert_rows(1, n_meta)
            for i, line in enumerate(lines, 1):
                ws.cell(row=i, column=1, value=line).font = m_font
                ws.cell(row=i, column=1).alignment = lft
                for c in range(2, ws.max_column + 1):
                    ws.cell(row=i, column=c, value='')
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=ws.max_column)

        hr = n_meta + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=hr, column=col)
            cell.fill, cell.font, cell.alignment = h_fill, h_font, ctr

        for row in range(hr + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.font, c.alignment, c.border = d_font, ctr, brd
                if isinstance(c.value, float):
                    c.number_format = '0.0000'

        for col in range(1, ws.max_column + 1):
            mx = max(len(str(ws.cell(row=r, column=col).value or ''))
                     for r in range(hr, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 35)
    wb.save(filepath)


# =============================================================
# TABLE TO IMAGE (JPG)
# =============================================================

def table_to_image(df, filepath, col_widths=None):
    """Render a DataFrame as a publication-ready table image."""
    n_rows, n_cols = df.shape

    col_w = col_widths if col_widths else [max(len(str(c)), max(len(str(v)) for v in df[c])) * 0.11 + 0.1 for c in df.columns]
    fig_w = sum(col_w)
    row_h = 0.28
    fig_h = (n_rows + 1) * row_h

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    table = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc='center',
        loc='center',
    )
    table.auto_set_font_size(False)

    for j in range(n_cols):
        cell = table[0, j]
        cell.set_facecolor('#2F5496')
        cell.set_text_props(color='white', fontweight='bold',
                            fontsize=11, fontfamily='sans-serif')
        cell.set_edgecolor('#1a3a6e')
        cell.set_linewidth(1.5)
        cell.set_height(1.0 / (n_rows + 1))

    for i in range(1, n_rows + 1):
        for j in range(n_cols):
            cell = table[i, j]
            cell.set_facecolor('white')
            cell.set_edgecolor('black')
            cell.set_linewidth(0.8)
            cell.set_height(1.0 / (n_rows + 1))
            cell.set_text_props(fontsize=11, fontfamily='sans-serif')

    for j, w in enumerate(col_w):
        for i in range(n_rows + 1):
            table[i, j].set_width(w / fig_w)

    fig.patch.set_facecolor('#FEFEFE')
    fig.savefig(filepath, dpi=300, bbox_inches='tight', pad_inches=0,
                facecolor='#FEFEFE', edgecolor='none')
    plt.show()
    plt.close(fig)

    from PIL import Image, ImageChops
    img = Image.open(filepath).convert('RGB')
    bg = Image.new('RGB', img.size, (254, 254, 254))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        cropped = img.crop(bbox)
        arr = np.array(cropped)
        mask = (arr[:, :, 0] == 254) & (arr[:, :, 1] == 254) & (arr[:, :, 2] == 254)
        arr[mask] = [255, 255, 255]
        Image.fromarray(arr).save(filepath, quality=95)

    print(f"  Saved: {os.path.basename(filepath)}")


# =============================================================
# LOAD DATA
# =============================================================
print("=" * 60)
print("GE Compare: Paper-Ready Tables")
print("=" * 60)

df = pd.read_excel(RAW_RESULTS_PATH)
df['threshold'] = df['threshold'].astype(str)

THRESHOLDS = sorted(df['threshold'].unique(), key=lambda x: (x == 'locate', x))
scanner_groups = sorted(df['scanner_group'].dropna().unique())

N_OVERALL = len(METRICS_LIST) * len(THRESHOLDS)
ALPHA_OVERALL = bonferroni_alpha(N_OVERALL)
N_SCANNER = len(METRICS_LIST) * len(THRESHOLDS) * len(scanner_groups)
ALPHA_SCANNER = bonferroni_alpha(N_SCANNER) if N_SCANNER > 0 else 0.05

print(f"  {len(df)} rows, {df['subject'].nunique()} subjects")
print(f"  Bonferroni overall: k={N_OVERALL}, alpha={ALPHA_OVERALL:.4f}")
print(f"  Bonferroni scanner: k={N_SCANNER}, alpha={ALPHA_SCANNER:.4f}")


# =============================================================
# SHEET 1: WILCOXON + DESCRIPTIVE (COMBINED)
# =============================================================
print("\n--- Wilcoxon + Descriptive ---")
rows_wd = []

for ts in THRESHOLDS:
    sub = df[df['threshold'] == ts]
    if sub.empty:
        continue
    thresh_label = 'B+L' if ts == 'locate' else f'B0.{ts}'

    for metric in METRICS_LIST:
        piv = sub.pivot(index='subject', columns='model', values=metric).dropna()
        if len(piv) == 0 or 'with_ge' not in piv or 'without_ge' not in piv:
            continue

        vw = piv['with_ge'].values
        vwo = piv['without_ge'].values

        try:
            w_stat, w_p = wilcoxon(vw, vwo)
        except ValueError:
            w_stat, w_p = np.nan, np.nan

        delta, ci_lo, ci_hi = bootstrap_cliffs_delta(vw, vwo)

        rows_wd.append({
            'Threshold': thresh_label,
            'Metric': metric.replace('_', ' ').title(),
            'N': len(piv),
            'Mean With GE': round(np.mean(vw), 4),
            'SD With GE': round(np.std(vw, ddof=1), 4),
            'Median With GE': round(np.median(vw), 4),
            'Mean Without GE': round(np.mean(vwo), 4),
            'SD Without GE': round(np.std(vwo, ddof=1), 4),
            'Median Without GE': round(np.median(vwo), 4),
            'Mean Diff': round(np.mean(vw) - np.mean(vwo), 6),
            'W': round(w_stat, 1) if not np.isnan(w_stat) else 'N/A',
            'p': fmt_p(w_p),
            'Significant': 'Yes' if (not np.isnan(w_p) and w_p < ALPHA_OVERALL) else 'No',
            "Cliff's \u03b4 [95% CI]": fmt_delta(delta, ci_lo, ci_hi),
            'Meaningful (|\u03b4|\u22650.28)': 'Yes' if abs(delta) >= DELTA_THRESHOLD else 'No',
            'Higher': 'With GE' if np.mean(vw) > np.mean(vwo) else 'Without GE',
        })
        print(f"  [{ts}] {metric}: \u03b4={delta:.4f}, p={fmt_p(w_p)}")

df_wd = pd.DataFrame(rows_wd)


# =============================================================
# SHEET 2: BLAND-ALTMAN
# =============================================================
print("\n--- Bland-Altman ---")
rows_ba = []

for ts in THRESHOLDS:
    sub = df[df['threshold'] == ts]
    if sub.empty:
        continue
    thresh_label = 'B+L' if ts == 'locate' else f'B0.{ts}'

    for metric in METRICS_LIST:
        piv = sub.pivot(index='subject', columns='model', values=metric).dropna()
        if len(piv) == 0 or 'with_ge' not in piv or 'without_ge' not in piv:
            continue

        diffs = piv['with_ge'].values - piv['without_ge'].values
        md = np.mean(diffs)
        sd = np.std(diffs, ddof=1)

        try:
            _, p_bias = wilcoxon(diffs)
        except ValueError:
            p_bias = np.nan

        rows_ba.append({
            'Threshold': thresh_label,
            'Metric': metric.replace('_', ' ').title(),
            'N': len(diffs),
            'Mean Diff': round(md, 6),
            'SD Diff': round(sd, 6),
            'Median Diff': round(np.median(diffs), 6),
            'LoA Lower': round(md - 1.96 * sd, 4),
            'LoA Upper': round(md + 1.96 * sd, 4),
            'Min Diff': round(np.min(diffs), 4),
            'Max Diff': round(np.max(diffs), 4),
            'Bias p': fmt_p(p_bias),
            'Direction': 'With GE higher' if md > 0 else 'Without GE higher',
        })

df_ba = pd.DataFrame(rows_ba)


# =============================================================
# SHEET 3: SCANNER-STRATIFIED
# =============================================================
print("\n--- Scanner-Stratified ---")
rows_sc = []

for ts in THRESHOLDS:
    sub = df[df['threshold'] == ts]
    if sub.empty:
        continue
    thresh_label = 'B+L' if ts == 'locate' else f'B0.{ts}'

    for sg in scanner_groups:
        sg_sub = sub[sub['scanner_group'] == sg]
        for metric in METRICS_LIST:
            piv = sg_sub.pivot(index='subject', columns='model', values=metric).dropna()
            if len(piv) < 3 or 'with_ge' not in piv or 'without_ge' not in piv:
                continue

            vw = piv['with_ge'].values
            vwo = piv['without_ge'].values

            try:
                _, w_p = wilcoxon(vw, vwo)
            except ValueError:
                w_p = np.nan

            delta, ci_lo, ci_hi = bootstrap_cliffs_delta(vw, vwo)

            rows_sc.append({
                'Threshold': thresh_label,
                'Scanner': sg,
                'Metric': metric.replace('_', ' ').title(),
                'N': len(piv),
                'Mean With GE': f"{np.mean(vw):.4f}",
                'Mean Without GE': f"{np.mean(vwo):.4f}",
                'Mean Diff': f"{np.mean(vw) - np.mean(vwo):.6f}",
                "Cliff's \u03b4 [95% CI]": fmt_delta(delta, ci_lo, ci_hi),
            })

            if metric == 'dice_score':
                print(f"  [{ts}] {sg} (N={len(piv)}): \u03b4={delta:.3f}, p={fmt_p(w_p)}")

df_sc = pd.DataFrame(rows_sc)


# =============================================================
# SHEET 4: BONFERRONI FAMILIES
# =============================================================
df_bonf = pd.DataFrame([
    {
        'Comparison Family': 'Overall (With GE vs Without GE)',
        'N_tests': N_OVERALL,
        'Base_alpha': 0.05,
        'Corrected_alpha': round(ALPHA_OVERALL, 4),
        'Description': f'{len(METRICS_LIST)} metrics x {len(THRESHOLDS)} thresholds',
        'Justification': 'Single hypothesis tested across all metric-threshold combinations.',
    },
    {
        'Comparison Family': 'Scanner-Stratified',
        'N_tests': N_SCANNER,
        'Base_alpha': 0.05,
        'Corrected_alpha': round(ALPHA_SCANNER, 4),
        'Description': f'{len(METRICS_LIST)} metrics x {len(THRESHOLDS)} thresholds x {len(scanner_groups)} groups',
        'Justification': 'Exploratory: scanner-specific GE impact.',
    },
])


# =============================================================
# EXPORT: SINGLE COMBINED EXCEL
# =============================================================
out_path = os.path.join(ANALYSE_DIR, "ge_compare_paper_tables.xlsx")

with pd.ExcelWriter(out_path) as writer:
    df_wd.to_excel(writer, sheet_name='Wilcoxon_Descriptive', index=False)
    df_ba.to_excel(writer, sheet_name='Bland_Altman', index=False)
    df_sc.to_excel(writer, sheet_name='Scanner_Stratified', index=False)
    df_bonf.to_excel(writer, sheet_name='Bonferroni_Families', index=False)

format_excel(out_path, {
    'Wilcoxon_Descriptive': [
        "Table S-X: With GE (N=60) vs Without GE (N=45) Training Comparison",
        "Nested design: Without GE is strict subset of With GE.",
        f"Bonferroni: k={N_OVERALL}, corrected alpha={ALPHA_OVERALL:.4f}.",
        "Cliff's Delta with bootstrapped 95% CI (1000 iterations).",
        f"Meaningful effect threshold: |delta| >= {DELTA_THRESHOLD}.",
    ],
    'Bland_Altman': [
        "Table S-X: Bland-Altman Agreement (With GE vs Without GE)",
        "Difference = With GE - Without GE. Positive = With GE higher.",
        "95% LoA = mean diff +/- 1.96 x SD.",
    ],
    'Scanner_Stratified': [
        "Table S-X: Scanner-Stratified GE Impact",
        f"Bonferroni: k={N_SCANNER}, corrected alpha={ALPHA_SCANNER:.4f}.",
        "Exploratory analysis per scanner manufacturer.",
    ],
    'Bonferroni_Families': [
        "Table S-X: Bonferroni Correction Families",
        "Documents all multiple-testing correction decisions (R5 #5).",
    ],
})

print(f"\n{'=' * 60}")
print(f"DONE -> {out_path}")
print(f"  4 sheets: Wilcoxon_Descriptive, Bland_Altman,")
print(f"            Scanner_Stratified, Bonferroni_Families")
print(f"{'=' * 60}")

# Export tables as images
table_to_image(df_wd, os.path.join(ANALYSE_DIR, "ge_wilcoxon_descriptive.jpg"))
table_to_image(df_ba, os.path.join(ANALYSE_DIR, "ge_bland_altman.jpg"))
table_to_image(df_sc, os.path.join(ANALYSE_DIR, "ge_scanner_stratified.jpg"))
table_to_image(df_bonf, os.path.join(ANALYSE_DIR, "ge_bonferroni_families.jpg"))