#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase I: Threshold Comparison by Severity Level (Inpainted Condition)
=====================================================================

Revision context (NeuroImage: Clinical, Major Revision)
-------------------------------------------------------
Violin plots comparing three thresholds (B+L, B0.85, B0.90) for the
inpainted condition. Models are trained on the inpainted condition only;
segmentation is evaluated on the inpainted condition. Individual data points
are colored by WMH severity level (high, middle, low). All violin
plots share the same y-axis range for direct visual comparison.

Wilcoxon signed-rank pairwise comparisons with Bonferroni correction
shown as significance brackets.

Addresses: R1 Comment 1, R5 #4 (threshold selection justification).

Design: Models trained on inpainted condition, evaluated on inpainted condition.
n=89 subjects. Per-subject means averaged across 10 seeds x 5-fold stratified CV.

Key results
-----------
B+L achieved the highest Dice Score (0.567) and Sensitivity (0.687)
among the three thresholds, while B0.85 and B0.90
achieved higher Precision (0.637 and 0.674 respectively) at the cost
of substantially lower Sensitivity.

Violin plots (inpainted condition):
  Dice Score:  B+L (0.567) > B0.85 (0.545) > B0.90 (0.540).
               B+L vs B0.85: p=0.017*, delta=0.077 (negligible).
               B+L vs B0.90: p=0.035*, delta=0.097 (negligible).
               B0.85 vs B0.90: p=0.549 (ns).
  Sensitivity: B+L (0.687) >> B0.85 (0.549) >> B0.90 (0.510).
               All three pairwise comparisons p<0.001***.
               B+L vs B0.85: delta=0.414 (medium), meaningful.
               B+L vs B0.90: delta=0.523 (large), meaningful.
               B0.85 vs B0.90: delta=0.130 (negligible).
  Precision:   B+L (0.557) < B0.85 (0.637) < B0.90 (0.674).
               All three pairwise comparisons p<0.001***.
               B+L vs B0.85: delta=-0.213 (small).
               B+L vs B0.90: delta=-0.300 (small), meaningful.
               B0.85 vs B0.90: delta=-0.105 (negligible).

  Interpretation: B+L adaptive threshold achieves the best
  Sensitivity-Precision trade-off as measured by Dice Score. The
  fixed thresholds (B0.85, B0.90) sacrifice Sensitivity for Precision,
  producing a net loss in Dice. The Sensitivity advantage of B+L
  is medium-to-large in effect size (delta=0.41-0.52), representing
  a meaningful improvement in WMH detection. The Precision disadvantage
  is small (delta=-0.21 to -0.30), reflecting the expected trade-off
  of detecting more true positives alongside some additional false
  positives.

Bonferroni family structure: k=3 pairwise comparisons per metric
  (alpha_adj = 0.0167). Dice, Sensitivity, and Precision are treated
  as separate hypothesis families because each addresses a distinct
  aspect of segmentation performance. Global correction across all 9
  tests (k=9) was not applied to avoid overcorrection given the
  inherent correlation between voxel-overlap metrics.

Paper changes
-------------
  - Section 3.1 (Results): Threshold comparison figures.
  - Supplemental: Descriptive stats by threshold x severity.
  - Supplemental: Wilcoxon pairwise table with Bonferroni correction.

Outputs
-------
  dice_score_threshold_severity.png
  sensitivity_threshold_severity.png
  precision_threshold_severity.png
  severity_distribution.png
  threshold_comparison_results.xlsx
    -> Sheet 'Wilcoxon Pairwise': Bonferroni-corrected pairwise tests
    -> Sheet 'Descriptive': per-threshold x severity descriptive stats

@author: temuuleu
"""

import os
import glob
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import wilcoxon
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cliffs_delta import cliffs_delta as _cliffs_delta

warnings.filterwarnings('ignore')

# =============================================================
# CONFIG
# =============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
metrics_result_dir = os.path.join(SCRIPT_DIR, "bianca_result_dir")
output_dir = os.path.join(SCRIPT_DIR, "analysis", "voxel_threshold_severity")
os.makedirs(output_dir, exist_ok=True)

# Fixed condition
TRAIN_CONDITION = 'filled'
TEST_CONDITION = 'filled'

# Metadata file with ROI_Volume for severity histogram
METADATA_PATH = os.path.join(SCRIPT_DIR, "Phase_1", "LOCATE_SET", "bianca_pool_wihtouth_ge.xlsx")

THRESHOLDS = ['85', '90', 'locate']
THRESH_LABELS = {'85': 'B0.85', '90': 'B0.90', 'locate': 'B+L'}
THRESH_ORDER = ['locate', '85', '90']

METRICS = ['dice_score', 'sensitivity', 'precision']
METRIC_LABELS = {
    'dice_score': 'Dice Score',
    'sensitivity': 'Sensitivity',
    'precision': 'Precision',
}
METRIC_SHORT = {
    'dice_score': 'Dice_mean',
    'sensitivity': 'Sens_mean',
    'precision': 'Prec_mean',
}

SEVERITY_LEVELS = ['high', 'middle', 'low']
SEVERITY_COLORS = {
    'high': '#D62728',
    'middle': '#FF7F0E',
    'low': '#2CA02C',
}
SEVERITY_LABELS = {
    'high': 'High',
    'middle': 'Middle',
    'low': 'Low',
}

N_POSTHOC = 3  # 3 pairwise comparisons per metric family -> Bonferroni alpha=0.0167
DELTA_THRESHOLD = 0.28


# =============================================================
# HELPERS
# =============================================================

def cliffs_delta_wrapper(x, y):
    x, y = list(x), list(y)
    if len(x) == 0 or len(y) == 0:
        return 'N/A', 'N/A'
    d, size = _cliffs_delta(x, y)
    return round(d, 4), size


def bootstrap_cliffs_delta_ci(x, y, n_boot=1000, alpha=0.05, seed=42):
    rng = np.random.default_rng(seed)
    x, y = np.array(x), np.array(y)
    if len(x) < 2 or len(y) < 2:
        return np.nan, np.nan
    deltas = []
    for _ in range(n_boot):
        x_b = rng.choice(x, size=len(x), replace=True)
        y_b = rng.choice(y, size=len(y), replace=True)
        d, _ = _cliffs_delta(list(x_b), list(y_b))
        deltas.append(d)
    lo = np.percentile(deltas, 100 * alpha / 2)
    hi = np.percentile(deltas, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def bootstrap_ci(data, n_boot=1000, alpha=0.05, seed=42):
    """Bootstrap 95% CI for the mean."""
    rng = np.random.default_rng(seed)
    data = np.array(data)
    if len(data) < 2:
        return np.nan, np.nan
    means = [np.mean(rng.choice(data, size=len(data), replace=True))
             for _ in range(n_boot)]
    lo = np.percentile(means, 100 * alpha / 2)
    hi = np.percentile(means, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def format_p(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.4f}"


def format_p_table(p):
    """For Excel tables: readable string."""
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.3f}"


def sig_stars(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '***'
    if p < 0.01:
        return '**'
    if p < 0.05:
        return '*'
    return 'ns'


def format_delta_ci(delta, ci_lo, ci_hi, mag='N/A'):
    """Format Cliff's delta + 95% CI + interpretation into a single string."""
    if pd.isna(delta):
        return 'N/A'
    ci_lo_s = f"{ci_lo:.3f}" if not pd.isna(ci_lo) else 'N/A'
    ci_hi_s = f"{ci_hi:.3f}" if not pd.isna(ci_hi) else 'N/A'
    return f"{delta:.4f} [{ci_lo_s}, {ci_hi_s}] ({mag})"


# =============================================================
# LOAD + AGGREGATE
# =============================================================

def load_and_aggregate(df):
    """Filter pre-loaded DataFrame to inpainted condition, aggregate per subject x threshold."""
    df = df[(df['train_condition'] == TRAIN_CONDITION) &
            (df['test_condition'] == TEST_CONDITION)].copy()

    print(f"Filtered: {len(df)} rows | Thresholds: {sorted(df['threshold'].unique())}")

    group_cols = ['subject', 'threshold']
    extra_cols = [c for c in ['scanner', 'severity_level', 'lesion_type']
                  if c in df.columns]

    agg_dict = {m: 'mean' for m in METRICS if m in df.columns}
    for col in extra_cols:
        agg_dict[col] = 'first'

    df_agg = df.groupby(group_cols).agg(agg_dict).reset_index()
    print(f"Aggregated: {len(df_agg)} rows ({df_agg['subject'].nunique()} subjects)")

    if 'severity_level' in df_agg.columns:
        print(f"Severity distribution: {df_agg.groupby('severity_level')['subject'].nunique().to_dict()}")

    return df_agg


# =============================================================
# COMPUTE SHARED Y-AXIS FOR ALL VIOLIN PLOTS
# =============================================================

def compute_shared_violin_ylim(df_agg):
    """
    Compute shared y-axis limits across all violin metrics and thresholds.
    Includes headroom for significance brackets.
    """
    global_min = np.inf
    global_max = -np.inf
    for metric in METRICS:
        for thresh in THRESH_ORDER:
            vals = df_agg[df_agg['threshold'] == thresh][metric].dropna().values
            if len(vals) > 0:
                global_min = min(global_min, np.min(vals))
                global_max = max(global_max, np.max(vals))

    bracket_headroom = 0.04 + 3 * 0.06 + 0.06
    y_lo = max(-0.02, global_min - 0.05)
    y_hi = global_max + bracket_headroom

    return (y_lo, y_hi)


# =============================================================
# DESCRIPTIVE TABLE (threshold x severity)
# =============================================================

def descriptive_table(df_agg):
    """Descriptive stats per threshold x severity x metric."""
    rows = []
    for metric in METRICS:
        metric_label = METRIC_LABELS[metric]
        for thresh in THRESH_ORDER:
            t_label = THRESH_LABELS[thresh]
            sub_t = df_agg[df_agg['threshold'] == thresh]

            # Overall for this threshold
            vals_all = sub_t[metric].dropna().values
            ci_lo, ci_hi = bootstrap_ci(vals_all)
            rows.append({
                'Metric': metric_label,
                'Threshold': t_label,
                'Severity': 'All',
                'n': len(vals_all),
                'Mean': f"{np.mean(vals_all):.4f}" if len(vals_all) > 0 else 'N/A',
                'SD': f"{np.std(vals_all, ddof=1):.4f}" if len(vals_all) > 1 else 'N/A',
                'Median': f"{np.median(vals_all):.4f}" if len(vals_all) > 0 else 'N/A',
                'IQR': f"{np.percentile(vals_all, 75) - np.percentile(vals_all, 25):.4f}" if len(vals_all) > 0 else 'N/A',
                '95% CI_lo': f"{ci_lo:.4f}" if not np.isnan(ci_lo) else 'N/A',
                '95% CI_hi': f"{ci_hi:.4f}" if not np.isnan(ci_hi) else 'N/A',
            })

            # Per severity
            if 'severity_level' in sub_t.columns:
                for sev in SEVERITY_LEVELS:
                    vals = sub_t[sub_t['severity_level'] == sev][metric].dropna().values
                    ci_lo, ci_hi = bootstrap_ci(vals)
                    rows.append({
                        'Metric': metric_label,
                        'Threshold': t_label,
                        'Severity': SEVERITY_LABELS.get(sev, sev),
                        'n': len(vals),
                        'Mean': f"{np.mean(vals):.4f}" if len(vals) > 0 else 'N/A',
                        'SD': f"{np.std(vals, ddof=1):.4f}" if len(vals) > 1 else 'N/A',
                        'Median': f"{np.median(vals):.4f}" if len(vals) > 0 else 'N/A',
                        'IQR': f"{np.percentile(vals, 75) - np.percentile(vals, 25):.4f}" if len(vals) > 0 else 'N/A',
                        '95% CI_lo': f"{ci_lo:.4f}" if not np.isnan(ci_lo) else 'N/A',
                        '95% CI_hi': f"{ci_hi:.4f}" if not np.isnan(ci_hi) else 'N/A',
                    })

    return pd.DataFrame(rows)


# =============================================================
# WILCOXON PAIRWISE (3 thresholds)
# =============================================================

def compute_wilcoxon_pairwise(df_agg):
    """Wilcoxon signed-rank pairwise (Bonferroni-corrected) across 3 thresholds."""
    pw_rows = []

    for metric in METRICS:
        metric_label = METRIC_LABELS[metric]

        for t1, t2 in combinations(THRESH_ORDER, 2):
            tl1, tl2 = THRESH_LABELS[t1], THRESH_LABELS[t2]

            d1 = df_agg[df_agg['threshold'] == t1][['subject', metric]].set_index('subject')
            d2 = df_agg[df_agg['threshold'] == t2][['subject', metric]].set_index('subject')
            merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()

            if len(merged) < 2:
                continue

            v1 = merged[f'{metric}_1'].values
            v2 = merged[f'{metric}_2'].values

            try:
                stat, p_raw = wilcoxon(v1, v2, alternative='two-sided')
            except ValueError:
                stat, p_raw = np.nan, 1.0

            p_bonf = min(p_raw * N_POSTHOC, 1.0)
            delta_val, mag = cliffs_delta_wrapper(v1, v2)
            ci_lo, ci_hi = bootstrap_cliffs_delta_ci(v1, v2)
            delta_num = float(delta_val) if delta_val != 'N/A' else np.nan
            meaningful = abs(delta_num) >= DELTA_THRESHOLD if not np.isnan(delta_num) else False
            sig = p_bonf < 0.05

            if sig or meaningful:
                mean_diff = np.mean(v1) - np.mean(v2)
                higher = tl2 if mean_diff < 0 else tl1 if mean_diff > 0 else 'Equal'
            else:
                higher = 'None'

            pw_rows.append({
                'Metric': metric_label,
                'Comparison': f"{tl1} vs {tl2}",
                'W Statistic': f"{stat:.0f}" if not pd.isna(stat) else 'N/A',
                'p (Bonferroni)': format_p_table(p_bonf),
                "Cliff's \u03b4 [95% CI]": format_delta_ci(delta_num, ci_lo, ci_hi, mag),
                'Higher': higher,
            })

    return pd.DataFrame(pw_rows)


# =============================================================
# VIOLIN PLOT WITH SEVERITY-COLORED POINTS (shared y-axis)
# =============================================================

def plot_threshold_severity(df_agg, metric, ylim):
    """
    Violin plot: one violin per threshold, individual points
    colored by severity level. Uses shared y-axis limits.
    """
    metric_label = METRIC_LABELS[metric]
    has_severity = 'severity_level' in df_agg.columns

    fig, ax = plt.subplots(figsize=(10, 7))

    positions = [1, 2, 3]
    violin_data = []
    for thresh in THRESH_ORDER:
        vals = df_agg[df_agg['threshold'] == thresh][metric].dropna().values
        violin_data.append(vals)

    # Violin plots
    vp = ax.violinplot(violin_data, positions=positions, showmeans=False,
                       showmedians=False, showextrema=False, widths=0.7)
    for body in vp['bodies']:
        body.set_facecolor('#B0C4DE')
        body.set_edgecolor('#4C72B0')
        body.set_alpha(0.35)

    # Scatter points colored by severity + jitter
    rng = np.random.default_rng(42)
    for i, thresh in enumerate(THRESH_ORDER):
        sub = df_agg[df_agg['threshold'] == thresh]
        pos = positions[i]

        if has_severity:
            for sev in SEVERITY_LEVELS:
                mask = sub['severity_level'] == sev
                vals = sub.loc[mask, metric].dropna().values
                jitter = rng.uniform(-0.18, 0.18, size=len(vals))
                color = SEVERITY_COLORS.get(sev, '#888888')
                ax.scatter(pos + jitter, vals, c=color, s=35, alpha=0.75,
                           edgecolors='white', linewidths=0.4, zorder=3)
        else:
            vals = sub[metric].dropna().values
            jitter = rng.uniform(-0.18, 0.18, size=len(vals))
            ax.scatter(pos + jitter, vals, c='#4C72B0', s=35, alpha=0.75,
                       edgecolors='white', linewidths=0.4, zorder=3)

    # Mean, SD, CI per threshold
    annotation_texts = []
    for i, thresh in enumerate(THRESH_ORDER):
        vals = df_agg[df_agg['threshold'] == thresh][metric].dropna().values
        pos = positions[i]
        t_label = THRESH_LABELS[thresh]

        if len(vals) == 0:
            continue

        mean_val = np.mean(vals)
        sd_val = np.std(vals, ddof=1)
        ci_lo, ci_hi = bootstrap_ci(vals)

        # Mean dot
        ax.scatter([pos], [mean_val], c='black', s=90, zorder=5, marker='o')

        # SD bars
        ax.plot([pos, pos], [mean_val - sd_val, mean_val + sd_val],
                color='black', linewidth=2.5, zorder=4)
        ax.plot([pos - 0.1, pos + 0.1], [mean_val - sd_val, mean_val - sd_val],
                color='black', linewidth=2.5, zorder=4)
        ax.plot([pos - 0.1, pos + 0.1], [mean_val + sd_val, mean_val + sd_val],
                color='black', linewidth=2.5, zorder=4)

        # CI dotted lines
        ax.plot([pos - 0.08, pos + 0.08], [ci_lo, ci_lo],
                color='black', linewidth=1.8, linestyle=':', zorder=4)
        ax.plot([pos - 0.08, pos + 0.08], [ci_hi, ci_hi],
                color='black', linewidth=1.8, linestyle=':', zorder=4)

        annotation_texts.append(
            f"{t_label}\n"
            f"{METRIC_SHORT[metric]}={mean_val:.4f}\u00B1{sd_val:.4f}\n"
            f"CI=({ci_lo:.4f}, {ci_hi:.4f})"
        )

    ax.set_xticks(positions)
    ax.set_xticklabels(annotation_texts, fontsize=9, linespacing=1.4)
    ax.tick_params(axis='x', pad=8)

    ax.set_xlabel('Threshold Strategy', fontsize=12, labelpad=12)
    ax.set_ylabel(metric_label, fontsize=12)
    ax.set_title(f"Phase I: {metric_label} by Threshold (Inpainted Condition)",
                 fontsize=14, fontweight='bold', pad=15)

    # -- Wilcoxon significance brackets --
    bracket_pairs = list(combinations(range(len(THRESH_ORDER)), 2))
    sig_brackets = []
    for i1, i2 in bracket_pairs:
        t1, t2 = THRESH_ORDER[i1], THRESH_ORDER[i2]
        d1 = df_agg[df_agg['threshold'] == t1][['subject', metric]].set_index('subject')
        d2 = df_agg[df_agg['threshold'] == t2][['subject', metric]].set_index('subject')
        merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()
        if len(merged) < 2:
            continue
        v1 = merged[f'{metric}_1'].values
        v2 = merged[f'{metric}_2'].values
        try:
            _, p_pw = wilcoxon(v1, v2, alternative='two-sided')
        except ValueError:
            p_pw = 1.0
        p_adj = min(p_pw * N_POSTHOC, 1.0)

        label = f"p={format_p(p_adj)}"
        if p_adj < 0.05:
            label = sig_stars(p_adj)

        sig_brackets.append((positions[i1], positions[i2], label, p_adj))

    # Draw ALL brackets (significant or not)
    if sig_brackets:
        all_vals = np.concatenate([df_agg[df_agg['threshold'] == t][metric].dropna().values
                                   for t in THRESH_ORDER])
        y_data_max = np.max(all_vals)
        bracket_base = y_data_max + 0.04
        bracket_step = 0.06

        sig_brackets.sort(key=lambda b: b[1] - b[0])

        for b_idx, (x1, x2, label, p_adj) in enumerate(sig_brackets):
            y_bar = bracket_base + b_idx * bracket_step
            y_tip = y_bar - 0.015

            lw = 1.2 if p_adj < 0.05 else 0.8
            col = 'black' if p_adj < 0.05 else 'gray'

            ax.plot([x1, x1, x2, x2], [y_tip, y_bar, y_bar, y_tip],
                    color=col, linewidth=lw, zorder=6)
            ax.text((x1 + x2) / 2, y_bar + 0.005, label,
                    ha='center', va='bottom', fontsize=10,
                    fontweight='bold' if p_adj < 0.05 else 'normal',
                    color=col, zorder=6)

    # Apply shared y-axis
    ax.set_ylim(ylim)

    # Severity counts per threshold
    if has_severity:
        for i, thresh in enumerate(THRESH_ORDER):
            sub = df_agg[df_agg['threshold'] == thresh]
            pos = positions[i]
            counts = []
            for sev in SEVERITY_LEVELS:
                n_sev = (sub['severity_level'] == sev).sum()
                counts.append(f"{SEVERITY_LABELS[sev]}={n_sev}")
            count_text = "  ".join(counts)
            ax.text(pos, -0.01, count_text, ha='center', va='top',
                    fontsize=7, color='gray', style='italic',
                    transform=ax.get_xaxis_transform())

    # Legend
    legend_handles = []
    if has_severity:
        for sev in SEVERITY_LEVELS:
            legend_handles.append(mpatches.Patch(
                color=SEVERITY_COLORS[sev], label=f"{SEVERITY_LABELS[sev]} WMH"))
    legend_handles.append(plt.Line2D([0], [0], marker='o', color='black',
                                      linestyle='None', markersize=7, label='Mean'))
    legend_handles.append(plt.Line2D([0], [0], color='black', linewidth=2.5,
                                      label='\u00B1 SD'))
    legend_handles.append(plt.Line2D([0], [0], color='black', linewidth=1.8,
                                      linestyle=':', label='95% CI'))
    fig.legend(handles=legend_handles, loc='lower center',
               bbox_to_anchor=(0.5, -0.02), ncol=len(legend_handles),
               fontsize=9, framealpha=0.9, edgecolor='lightgray')

    fig.tight_layout(rect=[0, 0.05, 1, 1])
    fname = f"{metric}_threshold_severity"
    fig.savefig(os.path.join(output_dir, f"{fname}.png"), dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {fname}.png")


# =============================================================
# LOAD ROI VOLUMES FROM METADATA
# =============================================================

def load_roi_volumes():
    """
    Load ROI_Volume (ground-truth WMH volume in mL) and severity_level
    from pool metadata for the severity histogram.
    """
    if not os.path.isfile(METADATA_PATH):
        print(f"  WARNING: Metadata file not found: {METADATA_PATH}")
        return None

    meta = pd.read_excel(METADATA_PATH)
    required = ['subject', 'ROI_Volume', 'severity_level']
    missing = [c for c in required if c not in meta.columns]
    if missing:
        print(f"  WARNING: Missing columns in metadata: {missing}")
        return None

    cols = required + (['scanner'] if 'scanner' in meta.columns else [])
    meta = meta[cols].dropna(subset=['ROI_Volume', 'severity_level']).copy()
    meta['ROI_Volume'] = pd.to_numeric(meta['ROI_Volume'], errors='coerce')
    meta = meta.dropna(subset=['ROI_Volume'])
    print(f"  Loaded ROI_Volume for {len(meta)} subjects from metadata.")
    return meta


# =============================================================
# SEVERITY DISTRIBUTION PLOT
# =============================================================

def plot_severity_distribution(df_agg):
    """
    Three-panel figure explaining WMH severity stratification:
      A) Histogram of ground-truth WMH volume with severity-colored bars.
      B) Box plots of ROI_Volume per severity level.
      C) Severity by scanner (grouped bar).
      D) Overall counts bar.
    """
    sub = df_agg[df_agg['threshold'] == THRESH_ORDER[0]].copy()
    has_scanner = 'scanner' in sub.columns
    n_total = len(sub)

    roi_df = load_roi_volumes()

    has_roi = False
    if roi_df is not None:
        merged = sub[['subject', 'severity_level']].merge(
            roi_df[['subject', 'ROI_Volume']], on='subject', how='left')
        merged = merged.dropna(subset=['ROI_Volume'])
        if len(merged) > 0:
            has_roi = True
            sorted_vols = merged['ROI_Volume'].sort_values().values
            low_vals = merged[merged['severity_level'] == 'low']['ROI_Volume']
            mid_vals = merged[merged['severity_level'] == 'middle']['ROI_Volume']
            high_vals = merged[merged['severity_level'] == 'high']['ROI_Volume']

            cutoff_low_mid = (low_vals.max() + mid_vals.min()) / 2 if len(low_vals) > 0 and len(mid_vals) > 0 else sorted_vols[len(sorted_vols) // 3]
            cutoff_mid_high = (mid_vals.max() + high_vals.min()) / 2 if len(mid_vals) > 0 and len(high_vals) > 0 else sorted_vols[2 * len(sorted_vols) // 3]

    if has_roi and has_scanner:
        fig = plt.figure(figsize=(16, 10))
        gs = fig.add_gridspec(2, 2, height_ratios=[1.2, 1], hspace=0.35, wspace=0.3)
        ax_hist = fig.add_subplot(gs[0, 0])
        ax_box = fig.add_subplot(gs[0, 1])
        ax_scanner = fig.add_subplot(gs[1, 0])
        ax_counts = fig.add_subplot(gs[1, 1])
    elif has_roi:
        fig, axes = plt.subplots(1, 2, figsize=(14, 6))
        ax_hist, ax_box = axes
        ax_scanner = None
        ax_counts = None
    else:
        fig, ax_counts = plt.subplots(figsize=(6, 5))
        ax_hist = None
        ax_box = None
        ax_scanner = None

    # ─── PANEL A: HISTOGRAM ───
    if has_roi and ax_hist is not None:
        all_vols = merged['ROI_Volume'].values
        bin_edges = np.histogram_bin_edges(all_vols, bins='auto')
        if len(bin_edges) < 10:
            bin_edges = np.linspace(all_vols.min(), all_vols.max(), 16)
        elif len(bin_edges) > 30:
            bin_edges = np.linspace(all_vols.min(), all_vols.max(), 26)

        for sev in SEVERITY_LEVELS:
            sev_vols = merged[merged['severity_level'] == sev]['ROI_Volume'].values
            ax_hist.hist(sev_vols, bins=bin_edges, alpha=0.7,
                         color=SEVERITY_COLORS[sev],
                         label=f"{SEVERITY_LABELS[sev]} (n={len(sev_vols)})",
                         edgecolor='white', linewidth=0.8)

        ymax = ax_hist.get_ylim()[1]
        ax_hist.axvline(cutoff_low_mid, color='#333333', linestyle='--',
                        linewidth=2, zorder=5, label=f'Cutoff: {cutoff_low_mid:.1f} mL')
        ax_hist.axvline(cutoff_mid_high, color='#333333', linestyle='-.',
                        linewidth=2, zorder=5, label=f'Cutoff: {cutoff_mid_high:.1f} mL')

        ax_hist.axvspan(all_vols.min() - 1, cutoff_low_mid,
                        alpha=0.06, color=SEVERITY_COLORS['low'], zorder=0)
        ax_hist.axvspan(cutoff_low_mid, cutoff_mid_high,
                        alpha=0.06, color=SEVERITY_COLORS['middle'], zorder=0)
        ax_hist.axvspan(cutoff_mid_high, all_vols.max() + 1,
                        alpha=0.06, color=SEVERITY_COLORS['high'], zorder=0)

        ax_hist.set_xlabel('Ground-Truth WMH Volume (mL)', fontsize=11)
        ax_hist.set_ylabel('Number of Subjects', fontsize=11)
        ax_hist.set_title('(A) WMH Volume Distribution by Severity Tercile',
                          fontsize=12, fontweight='bold')
        ax_hist.legend(fontsize=8, loc='upper right', framealpha=0.9)
        ax_hist.spines['top'].set_visible(False)
        ax_hist.spines['right'].set_visible(False)

    # ─── PANEL B: BOX PLOTS ───
    if has_roi and ax_box is not None:
        rng = np.random.default_rng(42)
        positions_box = [1, 2, 3]
        box_data = [merged[merged['severity_level'] == sev]['ROI_Volume'].dropna().values
                     for sev in SEVERITY_LEVELS]

        bp = ax_box.boxplot(box_data, positions=positions_box, widths=0.5,
                            patch_artist=True, showfliers=False, zorder=2,
                            medianprops=dict(color='black', linewidth=2))
        for patch, sev in zip(bp['boxes'], SEVERITY_LEVELS):
            patch.set_facecolor(SEVERITY_COLORS[sev])
            patch.set_alpha(0.4)

        for i, sev in enumerate(SEVERITY_LEVELS):
            vals = merged[merged['severity_level'] == sev]['ROI_Volume'].dropna().values
            jitter = rng.uniform(-0.15, 0.15, size=len(vals))
            ax_box.scatter(positions_box[i] + jitter, vals,
                           c=SEVERITY_COLORS[sev], s=30, alpha=0.7,
                           edgecolors='white', linewidths=0.4, zorder=3)

        ax_box.set_xticks(positions_box)
        ax_box.set_xticklabels([f"{SEVERITY_LABELS[s]} WMH" for s in SEVERITY_LEVELS], fontsize=10)
        ax_box.set_ylabel('Ground-Truth WMH Volume (mL)', fontsize=11)
        ax_box.set_title('(B) WMH Volume by Severity Level', fontsize=12, fontweight='bold')
        ax_box.spines['top'].set_visible(False)
        ax_box.spines['right'].set_visible(False)

    # ─── PANEL C: SCANNER ───
    if has_scanner and ax_scanner is not None:
        scanners = sorted(sub['scanner'].unique())
        x = np.arange(len(scanners))
        width = 0.25

        for i, sev in enumerate(SEVERITY_LEVELS):
            counts = [len(sub[(sub['scanner'] == sc) & (sub['severity_level'] == sev)])
                      for sc in scanners]
            b = ax_scanner.bar(x + i * width, counts, width,
                               color=SEVERITY_COLORS[sev], edgecolor='white',
                               linewidth=1.2, label=SEVERITY_LABELS[sev])
            for bar, c in zip(b, counts):
                if c > 0:
                    ax_scanner.text(bar.get_x() + bar.get_width() / 2,
                                    bar.get_height() + 0.3,
                                    str(c), ha='center', va='bottom', fontsize=9)

        ax_scanner.set_xticks(x + width)
        ax_scanner.set_xticklabels(scanners, fontsize=10)
        ax_scanner.set_ylabel('Number of Subjects', fontsize=11)
        ax_scanner.set_title('(C) Severity by Scanner', fontsize=12, fontweight='bold')
        ax_scanner.legend(fontsize=9)
        ax_scanner.spines['top'].set_visible(False)
        ax_scanner.spines['right'].set_visible(False)

    # ─── PANEL D: COUNTS ───
    if ax_counts is not None:
        sev_counts = sub['severity_level'].value_counts().reindex(
            SEVERITY_LEVELS).fillna(0).astype(int)
        bars = ax_counts.bar(
            [SEVERITY_LABELS[s] for s in SEVERITY_LEVELS],
            [sev_counts[s] for s in SEVERITY_LEVELS],
            color=[SEVERITY_COLORS[s] for s in SEVERITY_LEVELS],
            edgecolor='white', linewidth=1.2, width=0.6)
        for bar, sev in zip(bars, SEVERITY_LEVELS):
            count = sev_counts[sev]
            pct = 100 * count / n_total
            ax_counts.text(bar.get_x() + bar.get_width() / 2,
                           bar.get_height() + 0.5,
                           f"n={count}\n({pct:.0f}%)",
                           ha='center', va='bottom', fontsize=10, fontweight='bold')
        ax_counts.set_ylabel('Number of Subjects', fontsize=11)
        ax_counts.set_title('(D) Severity Distribution', fontsize=12, fontweight='bold')
        ax_counts.set_ylim(0, max(sev_counts) * 1.3)
        ax_counts.spines['top'].set_visible(False)
        ax_counts.spines['right'].set_visible(False)

    fig.suptitle(
        f"Phase I: WMH Severity Stratification (n={n_total})",
        fontsize=14, fontweight='bold', y=1.01)
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, "severity_distribution.png"),
                dpi=300, bbox_inches='tight')
    plt.close(fig)
    print("  Saved: severity_distribution.png")


# =============================================================
# EXCEL FORMATTING
# =============================================================

def format_excel(filepath, sheet_metadata=None):
    wb = load_workbook(filepath)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    meta_font = Font(italic=True, size=9, name='Arial', color='555555')
    data_font = Font(size=9, name='Arial')
    border = Border(bottom=Side(style='thin', color='D9D9D9'))
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for ws in wb.worksheets:
        n_meta = 0
        if sheet_metadata and ws.title in sheet_metadata:
            meta_lines = sheet_metadata[ws.title]
            n_meta = len(meta_lines)
            ws.insert_rows(1, n_meta)
            for i, line in enumerate(meta_lines, start=1):
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = meta_font
                cell.alignment = left
                for c in range(2, ws.max_column + 1):
                    ws.cell(row=i, column=c, value='')
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=ws.max_column)

        header_row = n_meta + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'

        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                         for r in range(header_row, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 35)

    wb.save(filepath)


# =============================================================
# TABLE TO IMAGE (JPG)
# =============================================================

def table_to_image(df, filepath, col_widths=None):
    """Render a DataFrame as a publication-ready table image (Excel-like style)."""
    n_rows, n_cols = df.shape

    col_w = col_widths if col_widths else [max(len(str(c)), max(len(str(v)) for v in df[c])) * 0.14 + 0.15 for c in df.columns]
    fig_w = sum(col_w)
    row_h = 0.32
    fig_h = (n_rows + 1) * row_h

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    table = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc='center',
        loc='center',
    )
    table.auto_set_font_size(False)

    for j in range(n_cols):
        cell = table[0, j]
        cell.set_facecolor('#2F5496')
        cell.set_text_props(color='white', fontweight='bold',
                            fontsize=12, fontfamily='sans-serif')
        cell.set_edgecolor('#1a3a6e')
        cell.set_linewidth(1.5)
        cell.set_height(1.0 / (n_rows + 1))

    for i in range(1, n_rows + 1):
        for j in range(n_cols):
            cell = table[i, j]
            cell.set_facecolor('white')
            cell.set_edgecolor('black')
            cell.set_linewidth(0.8)
            cell.set_height(1.0 / (n_rows + 1))
            cell.set_text_props(fontsize=12, fontfamily='sans-serif')

    for j, w in enumerate(col_w):
        for i in range(n_rows + 1):
            table[i, j].set_width(w / fig_w)

    fig.patch.set_facecolor('#FEFEFE')
    fig.savefig(filepath, dpi=300, bbox_inches='tight', pad_inches=0,
                facecolor='#FEFEFE', edgecolor='none')
    plt.show()
    plt.close(fig)

    from PIL import Image, ImageChops
    img = Image.open(filepath).convert('RGB')
    bg = Image.new('RGB', img.size, (254, 254, 254))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        cropped = img.crop(bbox)
        arr = np.array(cropped)
        mask = (arr[:, :, 0] == 254) & (arr[:, :, 1] == 254) & (arr[:, :, 2] == 254)
        arr[mask] = [255, 255, 255]
        Image.fromarray(arr).save(filepath, quality=95)

    print(f"  Saved: {os.path.basename(filepath)}")


# =============================================================
# CONSOLE SUMMARY
# =============================================================

def print_summary(df_agg):
    """Print descriptive + pairwise summary to console for all metrics."""
    for metric in METRICS:
        metric_label = METRIC_LABELS[metric]
        print(f"\n--- Descriptive Summary: {metric_label} ---")
        for thresh in THRESH_ORDER:
            vals = df_agg[df_agg['threshold'] == thresh][metric].dropna().values
            ci_lo, ci_hi = bootstrap_ci(vals)
            print(f"  {THRESH_LABELS[thresh]:10s}: n={len(vals)}, "
                  f"Mean={np.mean(vals):.4f} +/- {np.std(vals, ddof=1):.4f}, "
                  f"Median={np.median(vals):.4f}, "
                  f"95% CI=[{ci_lo}, {ci_hi}]")

        print(f"\n--- Wilcoxon Pairwise: {metric_label} (Bonferroni k=3) ---")
        for t1, t2 in combinations(THRESH_ORDER, 2):
            d1 = df_agg[df_agg['threshold'] == t1][['subject', metric]].set_index('subject')
            d2 = df_agg[df_agg['threshold'] == t2][['subject', metric]].set_index('subject')
            merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()
            if len(merged) < 2:
                continue
            v1 = merged[f'{metric}_1'].values
            v2 = merged[f'{metric}_2'].values
            try:
                stat, p_pw = wilcoxon(v1, v2, alternative='two-sided')
            except ValueError:
                stat, p_pw = np.nan, 1.0
            p_adj = min(p_pw * N_POSTHOC, 1.0)
            delta, mag = cliffs_delta_wrapper(v1, v2)
            print(f"  {THRESH_LABELS[t1]} vs {THRESH_LABELS[t2]}: "
                  f"W={stat:.1f}, p_adj={format_p(p_adj)} {sig_stars(p_adj)}, "
                  f"Cliff's d={delta} ({mag})")


# =============================================================
# MAIN
# =============================================================

if __name__ == "__main__":
    print("=" * 70)
    print("PHASE I: THRESHOLD COMPARISON BY SEVERITY (INPAINTED CONDITION)")
    print(f"Trained on inpainted, evaluated on inpainted")
    print(f"Thresholds: {[THRESH_LABELS[t] for t in THRESH_ORDER]}")
    print(f"Metrics: {[METRIC_LABELS[m] for m in METRICS]}")
    print("=" * 70)

    # Load once
    files = sorted(glob.glob(os.path.join(
        metrics_result_dir, "bianca_metrics_seed_*.xlsx")))
    if not files:
        raise FileNotFoundError(f"No metric files in {metrics_result_dir}")

    dfs = [pd.read_excel(f) for f in files]
    df = pd.concat(dfs, ignore_index=True)
    df['threshold'] = df['threshold'].astype(str)

    df_agg = load_and_aggregate(df)

    # ── Shared y-axis for all violin plots ──
    violin_ylim = compute_shared_violin_ylim(df_agg)
    print(f"  Shared violin y-axis: [{violin_ylim[0]:.3f}, {violin_ylim[1]:.3f}]")

    # ── 1) Severity distribution ──
    print("\n--- Severity Distribution ---")
    plot_severity_distribution(df_agg)

    # ── 2) Violin plots per metric ──
    print("\n--- Violin Plots ---")
    for metric in METRICS:
        print(f"\n  Violin: {METRIC_LABELS[metric]}")
        plot_threshold_severity(df_agg, metric, violin_ylim)

    print_summary(df_agg)

    # ── 3) Tables ──
    desc_df = descriptive_table(df_agg)
    wilcoxon_df = compute_wilcoxon_pairwise(df_agg)

    # ── Save all tables in one Excel file ──
    xlsx_path = os.path.join(output_dir, "threshold_comparison_results.xlsx")

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        wilcoxon_df.to_excel(writer, sheet_name='Wilcoxon Pairwise', index=False)
        desc_df.to_excel(writer, sheet_name='Descriptive', index=False)

    meta = {
        'Wilcoxon Pairwise': [
            "Table: Phase I - Wilcoxon Signed-Rank Pairwise Tests (Threshold Comparison)",
            "Design: Models trained on inpainted condition, evaluated on inpainted condition.",
            f"Bonferroni correction: k={N_POSTHOC} pairwise comparisons per metric family (alpha=0.0167).",
            "Family structure: Dice, Sensitivity, Precision treated as separate hypothesis families,",
            "each addressing a distinct aspect of segmentation performance (overlap, detection, false positives).",
            "Global correction (k=9) not applied to avoid overcorrection given inherent metric correlation.",
            "Cliff's Delta with bootstrapped 95% CI (1000 iterations).",
            f"Meaningful effect: |delta| >= {DELTA_THRESHOLD}.",
            "Per-subject means across 10 seeds x 5-fold stratified CV.",
        ],
        'Descriptive': [
            "Table: Phase I - Descriptive Statistics by Threshold and WMH Severity Level",
            "Design: Models trained on inpainted condition, evaluated on inpainted condition.",
            "Per-subject means across 10 seeds x 5-fold stratified CV.",
            "95% CI: bootstrap confidence interval (1000 iterations).",
            "Severity: based on WMH burden stratification (high/middle/low).",
        ],
    }
    format_excel(xlsx_path, meta)
    print(f"\n  Saved combined table: {xlsx_path}")

    # Export tables as images
    wilcoxon_img = os.path.join(output_dir, "wilcoxon_thresholds.jpg")
    table_to_image(wilcoxon_df, wilcoxon_img)
    desc_img = os.path.join(output_dir, "descriptive_threshold_severity.jpg")
    table_to_image(desc_df, desc_img)

    print(f"\nAll outputs in: {output_dir}")