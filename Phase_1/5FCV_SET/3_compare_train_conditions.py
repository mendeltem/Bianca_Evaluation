#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Metrics by Training Condition + Bland-Altman (B+L only)
=======================================================

Revision context (NeuroImage: Clinical, Major Revision)
-------------------------------------------------------
Violin plots (one per metric: Dice Score, Sensitivity, Precision)
comparing performance across three training conditions (Non removed,
Removed, Inpainted) with the evaluation condition fixed to Inpainted.
Threshold: B+L. All violin plots share the same y-axis range
for direct visual comparison.

Additionally, Bland-Altman agreement plots are generated for the
B+L threshold, comparing all three training condition pairs:
  1. Non Removed vs Removed
  2. Non Removed vs Inpainted
  3. Removed vs Inpainted

Addresses: R1 Comment 1, R5 #9, R5 #4.

Design: Models trained on each condition separately, all evaluated on the
inpainted condition. Threshold: B+L. n=89 subjects.
Per-subject means averaged across 10 seeds x 5-fold stratified CV.

Key results
-----------
Training condition had no meaningful effect on segmentation accuracy.
All 9 Bonferroni-corrected Wilcoxon pairwise tests were non-significant
(smallest p_bonf = 0.104, Precision Non Removed vs Removed). All
Cliff's delta were negligible (max |delta| = 0.0043). This justifies
the use of a single training condition (Inpainted) for all subsequent
analyses.

Violin plots (B+L threshold):
  Dice Score:  Mean = 0.567 +/- 0.236 (all three conditions identical
               to 3rd decimal). All pairwise p_bonf >= 0.743.
  Sensitivity: Mean = 0.687 +/- 0.179. All pairwise p_bonf >= 0.128.
  Precision:   Mean = 0.557 +/- 0.287. All pairwise p_bonf >= 0.104.

Bland-Altman agreement (B+L threshold):
  95% LoA were extremely narrow across all comparisons:
    Non Removed vs Removed:   Dice [-0.003, 0.002], Sens [-0.006, 0.006],
                              Prec [-0.004, 0.003]
    Non Removed vs Inpainted: Dice [-0.002, 0.002], Sens [-0.003, 0.003],
                              Prec [-0.002, 0.002]
    Removed vs Inpainted:     Dice [-0.002, 0.002], Sens [-0.004, 0.004],
                              Prec [-0.003, 0.003]

  Removed vs Inpainted LoA were consistently narrower than Non Removed
  vs Removed LoA, confirming that the filling strategy (zero vs NAWM
  inpainting) has even less effect than the presence/absence of the
  stroke lesion in training data.

  3 of 9 uncorrected bias tests reached p < 0.05 (Sensitivity NR vs R
  p=0.043, Precision NR vs R p=0.035, Precision NR vs Inpainted p=0.048),
  but all with negligible Cliff's delta (max |delta| = 0.0043, all CIs
  encompassing zero). With 9 uncorrected tests, ~0.45 false positives
  are expected by chance alone. These marginal p-values reflect
  statistical noise, not meaningful systematic bias.

Bonferroni family structure: k=3 pairwise comparisons per metric
  (alpha_adj = 0.0167). Dice, Sensitivity, and Precision are treated
  as separate hypothesis families because each addresses a distinct
  aspect of segmentation performance. Global correction across all 9
  tests (k=9) was not applied to avoid overcorrection given the
  inherent correlation between voxel-overlap metrics.

Paper changes
-------------
  - Section 3.2 (Results): Violin + BA figures for training condition
    comparison, confirming negligible effect of training condition.
  - Supplemental: BA table with LoA, bias test, Cliff's delta [95% CI].
  - Supplemental: Wilcoxon pairwise table with Bonferroni correction.
  - Supplemental: Descriptive statistics per metric x condition.

Response to Reviewers
---------------------
  R1 Comment 1: "We compared all three training conditions (non_removed,
    removed, inpainted) with test condition fixed to inpainted. All 9
    Bonferroni-corrected pairwise Wilcoxon tests were non-significant
    (smallest p_bonf = 0.104). Bland-Altman analysis confirmed
    near-identical agreement across conditions (all Cliff's delta
    negligible, max |delta| = 0.004, 95% LoA within +/-0.006).
    Mean Dice scores were identical to the 3rd decimal place (0.567)
    across all three training conditions."

  R5 Comment 9: "Inpainting was tested as a third training condition.
    Removed vs Inpainted LoA (Dice: +/-0.002, Sensitivity: +/-0.004,
    Precision: +/-0.003) were consistently narrower than Non Removed vs
    Removed LoA (Dice: +/-0.003, Sensitivity: +/-0.006, Precision:
    +/-0.004), confirming the filling strategy has no meaningful effect
    on segmentation performance beyond the already negligible effect of
    lesion presence in training data."

  R5 Comment 4: "Bland-Altman plots provide practical agreement
    assessment beyond hypothesis testing. The extremely narrow 95% LoA
    (max +/-0.006) demonstrate that any differences between training
    conditions are negligible in magnitude and fall well below the
    measurement precision of WMH segmentation."

Outputs
-------
  dice_score_train_conditions_locate.png
  sensitivity_train_conditions_locate.png
  precision_train_conditions_locate.png
  BA_non_removed_vs_removed_locate.png
  BA_non_removed_vs_filled_locate.png
  BA_removed_vs_filled_locate.png
  locate_train_conditions_results.xlsx
    -> Sheet 'Wilcoxon Pairwise': Bonferroni-corrected pairwise tests
    -> Sheet 'Descriptive': per-condition descriptive stats
    -> Sheet 'Bland-Altman': BA agreement stats

@author: temuuleu
"""

import os
import glob
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import wilcoxon
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cliffs_delta import cliffs_delta as _cliffs_delta

warnings.filterwarnings('ignore')

# =============================================================
# CONFIG
# =============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
metrics_result_dir = os.path.join(SCRIPT_DIR, "bianca_result_dir")
output_dir = os.path.join(SCRIPT_DIR, "analysis", "dice_locate_train_conditions")
os.makedirs(output_dir, exist_ok=True)

# Fixed design
TEST_CONDITION = 'filled'   # inpainted test (fixed)
THRESHOLDS = ['locate']
THRESH_LABELS = {'85': 'B0.85', '90': 'B0.90', 'locate': 'B+L'}

# Three training conditions (x-axis)
COND_ORDER = ['non_removed', 'removed', 'filled']
COND_LABELS = {
    'non_removed': 'Non Removed',
    'removed': 'Removed',
    'filled': 'Inpainted',
}

# Violin plots: all three voxel-level metrics
VIOLIN_METRICS = ['dice_score', 'sensitivity', 'precision']
VIOLIN_METRIC_LABELS = {
    'dice_score': 'Dice Score',
    'sensitivity': 'Sensitivity',
    'precision': 'Precision',
}

# BA uses all three voxel-level metrics
BA_METRICS = ['dice_score', 'sensitivity', 'precision']
BA_METRIC_LABELS = {
    'dice_score': 'Dice Score',
    'sensitivity': 'Sensitivity',
    'precision': 'Precision',
}

# BA comparisons (all 3 pairs)
BA_COMPARISONS = [
    ('non_removed', 'removed'),
    ('non_removed', 'filled'),
    ('removed', 'filled'),
]

SEVERITY_LEVELS = ['high', 'middle', 'low']
SEVERITY_COLORS = {
    'high': '#D62728',
    'middle': '#FF7F0E',
    'low': '#2CA02C',
}
SEVERITY_LABELS = {
    'high': 'High',
    'middle': 'Middle',
    'low': 'Low',
}

# Lesion type styling (for BA plots)
LESION_COLORS = {
    'infarct': '#4C72B0',
    'infra': '#DD4444',
    'lacune': '#E48AC5',
    'mixed': '#55D6BE',
}

N_POSTHOC = 3  # 3 pairwise comparisons per metric family -> Bonferroni alpha=0.0167
DELTA_THRESHOLD = 0.28
OUTLIER_FACTOR = 2.5


# =============================================================
# HELPERS
# =============================================================

def cliffs_delta_wrapper(x, y):
    x, y = list(x), list(y)
    if len(x) == 0 or len(y) == 0:
        return 'N/A', 'N/A'
    d, size = _cliffs_delta(x, y)
    return round(d, 4), size


def bootstrap_cliffs_delta_ci(x, y, n_boot=1000, alpha=0.05, seed=42):
    rng = np.random.default_rng(seed)
    x, y = np.array(x), np.array(y)
    if len(x) < 2 or len(y) < 2:
        return np.nan, np.nan
    deltas = []
    for _ in range(n_boot):
        x_b = rng.choice(x, size=len(x), replace=True)
        y_b = rng.choice(y, size=len(y), replace=True)
        d, _ = _cliffs_delta(list(x_b), list(y_b))
        deltas.append(d)
    lo = np.percentile(deltas, 100 * alpha / 2)
    hi = np.percentile(deltas, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def bootstrap_ci(data, n_boot=1000, alpha=0.05, seed=42):
    rng = np.random.default_rng(seed)
    data = np.array(data)
    if len(data) < 2:
        return np.nan, np.nan
    means = [np.mean(rng.choice(data, size=len(data), replace=True))
             for _ in range(n_boot)]
    lo = np.percentile(means, 100 * alpha / 2)
    hi = np.percentile(means, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def format_p(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.4f}"


def format_p_table(p):
    """For Excel tables: readable string."""
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.3f}"


def sig_stars(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '***'
    if p < 0.01:
        return '**'
    if p < 0.05:
        return '*'
    return 'ns'


def format_delta_ci(delta, ci_lo, ci_hi, mag='N/A'):
    """Format Cliff's delta + 95% CI + interpretation into a single string."""
    if pd.isna(delta):
        return 'N/A'
    ci_lo_s = f"{ci_lo:.3f}" if not pd.isna(ci_lo) else 'N/A'
    ci_hi_s = f"{ci_hi:.3f}" if not pd.isna(ci_hi) else 'N/A'
    return f"{delta:.4f} [{ci_lo_s}, {ci_hi_s}] ({mag})"


# =============================================================
# LOAD + AGGREGATE
# =============================================================

def load_and_aggregate(df, threshold):
    """
    Filter pre-loaded DataFrame to inpainted evaluation condition + given threshold,
    aggregate per subject x train_condition.
    """
    df = df[(df['test_condition'] == TEST_CONDITION) &
            (df['threshold'] == threshold)].copy()

    print(f"Filtered: {len(df)} rows | Threshold: {threshold} | "
          f"Train conditions: {sorted(df['train_condition'].unique())}")

    group_cols = ['subject', 'train_condition']
    extra_cols = [c for c in ['scanner', 'severity_level', 'lesion_type']
                  if c in df.columns]

    all_metrics = list(set(VIOLIN_METRICS + BA_METRICS))
    agg_dict = {m: 'mean' for m in all_metrics if m in df.columns}
    for col in extra_cols:
        agg_dict[col] = 'first'

    df_agg = df.groupby(group_cols).agg(agg_dict).reset_index()
    print(f"Aggregated: {len(df_agg)} rows ({df_agg['subject'].nunique()} subjects)")

    if 'severity_level' in df_agg.columns:
        print(f"Severity: {df_agg.groupby('severity_level')['subject'].nunique().to_dict()}")

    return df_agg


# =============================================================
# COMPUTE SHARED Y-AXIS FOR ALL VIOLIN PLOTS
# =============================================================

def compute_shared_violin_ylim(df_agg):
    """
    Compute shared y-axis limits across all violin metrics and conditions.
    Includes headroom for significance brackets.
    """
    global_min = np.inf
    global_max = -np.inf
    for metric in VIOLIN_METRICS:
        for cond in COND_ORDER:
            vals = df_agg[df_agg['train_condition'] == cond][metric].dropna().values
            if len(vals) > 0:
                global_min = min(global_min, np.min(vals))
                global_max = max(global_max, np.max(vals))

    # Headroom for brackets (3 brackets * step + padding)
    bracket_headroom = 0.04 + 3 * 0.06 + 0.06
    y_lo = max(-0.02, global_min - 0.05)
    y_hi = global_max + bracket_headroom

    return (y_lo, y_hi)


# =============================================================
# WILCOXON PAIRWISE (for Excel table)
# =============================================================

def compute_wilcoxon_pairwise(df_agg):
    """
    Compute Wilcoxon signed-rank pairwise tests for all metrics
    with Bonferroni correction. Returns DataFrame for Excel.
    """
    rows = []
    for metric in VIOLIN_METRICS:
        metric_label = VIOLIN_METRIC_LABELS[metric]
        for c1, c2 in combinations(COND_ORDER, 2):
            label1, label2 = COND_LABELS[c1], COND_LABELS[c2]
            d1 = df_agg[df_agg['train_condition'] == c1][['subject', metric]].set_index('subject')
            d2 = df_agg[df_agg['train_condition'] == c2][['subject', metric]].set_index('subject')
            merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()

            if len(merged) < 2:
                continue

            v1 = merged[f'{metric}_1'].values
            v2 = merged[f'{metric}_2'].values

            try:
                stat, p_raw = wilcoxon(v1, v2, alternative='two-sided')
            except ValueError:
                stat, p_raw = np.nan, 1.0

            p_bonf = min(p_raw * N_POSTHOC, 1.0)
            delta, mag = cliffs_delta_wrapper(v1, v2)
            ci_lo, ci_hi = bootstrap_cliffs_delta_ci(v1, v2)
            meaningful = abs(delta) >= DELTA_THRESHOLD if not np.isnan(delta) else False
            sig = p_bonf < 0.05

            if sig or meaningful:
                mean_diff = np.mean(v1) - np.mean(v2)
                higher = label2 if mean_diff < 0 else label1 if mean_diff > 0 else 'Equal'
            else:
                higher = 'No meaningful difference'

            rows.append({
                'Metric': metric_label,
                'Comparison': f"{label1} vs {label2}",
                'W Statistic': f"{stat:.0f}" if not np.isnan(stat) else 'N/A',
                'p (Bonferroni)': format_p_table(p_bonf),
                "Cliff's \u03b4 [95% CI]": format_delta_ci(delta, ci_lo, ci_hi, mag),
                'Higher': higher,
            })

    return pd.DataFrame(rows)


# =============================================================
# DESCRIPTIVE STATS (for Excel table)
# =============================================================

def compute_descriptive(df_agg):
    """
    Compute descriptive statistics for all metrics x conditions.
    Returns DataFrame for Excel.
    """
    rows = []
    for metric in VIOLIN_METRICS:
        metric_label = VIOLIN_METRIC_LABELS[metric]
        for cond in COND_ORDER:
            vals = df_agg[df_agg['train_condition'] == cond][metric].dropna().values
            if len(vals) == 0:
                continue
            ci_lo, ci_hi = bootstrap_ci(vals)
            rows.append({
                'Metric': metric_label,
                'Training Condition': COND_LABELS[cond],
                'n': len(vals),
                'Mean': round(np.mean(vals), 4),
                'SD': round(np.std(vals, ddof=1), 4),
                'Median': round(np.median(vals), 4),
                'Min': round(np.min(vals), 4),
                'Max': round(np.max(vals), 4),
                'IQR_25': round(np.percentile(vals, 25), 4),
                'IQR_75': round(np.percentile(vals, 75), 4),
                '95% CI_lo': ci_lo,
                '95% CI_hi': ci_hi,
            })
    return pd.DataFrame(rows)


# =============================================================
# VIOLIN PLOT (generalized, with shared y-axis)
# =============================================================

def plot_metric_conditions(df_agg, threshold, thresh_label, metric, metric_label, ylim):
    """
    Violin plot: given metric across three training conditions,
    individual points colored by severity. Uses shared y-axis limits.
    """
    has_severity = 'severity_level' in df_agg.columns

    fig, ax = plt.subplots(figsize=(10, 7))

    positions = [1, 2, 3]
    violin_data = []
    for cond in COND_ORDER:
        vals = df_agg[df_agg['train_condition'] == cond][metric].dropna().values
        violin_data.append(vals)

    # Violins
    vp = ax.violinplot(violin_data, positions=positions, showmeans=False,
                       showmedians=False, showextrema=False, widths=0.7)
    for body in vp['bodies']:
        body.set_facecolor('#B0C4DE')
        body.set_edgecolor('#4C72B0')
        body.set_alpha(0.35)

    # Scatter points (severity-colored) + jitter
    rng = np.random.default_rng(42)
    for i, cond in enumerate(COND_ORDER):
        sub = df_agg[df_agg['train_condition'] == cond]
        pos = positions[i]

        if has_severity:
            for sev in SEVERITY_LEVELS:
                mask = sub['severity_level'] == sev
                vals = sub.loc[mask, metric].dropna().values
                jitter = rng.uniform(-0.18, 0.18, size=len(vals))
                color = SEVERITY_COLORS.get(sev, '#888888')
                ax.scatter(pos + jitter, vals, c=color, s=35, alpha=0.75,
                           edgecolors='white', linewidths=0.4, zorder=3)
        else:
            vals = sub[metric].dropna().values
            jitter = rng.uniform(-0.18, 0.18, size=len(vals))
            ax.scatter(pos + jitter, vals, c='#4C72B0', s=35, alpha=0.75,
                       edgecolors='white', linewidths=0.4, zorder=3)

    # Short labels for annotation
    METRIC_SHORT = {
        'dice_score': 'Dice_mean',
        'sensitivity': 'Sens_mean',
        'precision': 'Prec_mean',
    }

    # Mean, SD, CI annotations (show all 3 metrics per condition)
    annotation_texts = []
    for i, cond in enumerate(COND_ORDER):
        vals = df_agg[df_agg['train_condition'] == cond][metric].dropna().values
        pos = positions[i]

        if len(vals) == 0:
            continue

        mean_val = np.mean(vals)
        sd_val = np.std(vals, ddof=1)
        ci_lo, ci_hi = bootstrap_ci(vals)

        ax.scatter([pos], [mean_val], c='black', s=90, zorder=5, marker='o')

        ax.plot([pos, pos], [mean_val - sd_val, mean_val + sd_val],
                color='black', linewidth=2.5, zorder=4)
        ax.plot([pos - 0.1, pos + 0.1], [mean_val - sd_val, mean_val - sd_val],
                color='black', linewidth=2.5, zorder=4)
        ax.plot([pos - 0.1, pos + 0.1], [mean_val + sd_val, mean_val + sd_val],
                color='black', linewidth=2.5, zorder=4)

        ax.plot([pos - 0.08, pos + 0.08], [ci_lo, ci_lo],
                color='black', linewidth=1.8, linestyle=':', zorder=4)
        ax.plot([pos - 0.08, pos + 0.08], [ci_hi, ci_hi],
                color='black', linewidth=1.8, linestyle=':', zorder=4)

        annotation_texts.append(
            f"{COND_LABELS[cond]}\n"
            f"{METRIC_SHORT[metric]}={mean_val:.4f}\u00B1{sd_val:.4f}\n"
            f"CI=({ci_lo:.4f}, {ci_hi:.4f})"
        )

    ax.set_xticks(positions)
    ax.set_xticklabels(annotation_texts, fontsize=9, linespacing=1.4)
    ax.tick_params(axis='x', pad=8)

    ax.set_xlabel('Training Condition (Preprocessing)', fontsize=12, labelpad=12)
    ax.set_ylabel(metric_label, fontsize=12)
    ax.set_title(f"Phase I: {metric_label} by Training Condition ({thresh_label}, Test=Inpainted)",
                 fontsize=14, fontweight='bold', pad=15)

    # -- Wilcoxon significance brackets --
    bracket_pairs = list(combinations(range(len(COND_ORDER)), 2))
    sig_brackets = []
    for i1, i2 in bracket_pairs:
        c1, c2 = COND_ORDER[i1], COND_ORDER[i2]
        d1 = df_agg[df_agg['train_condition'] == c1][['subject', metric]].set_index('subject')
        d2 = df_agg[df_agg['train_condition'] == c2][['subject', metric]].set_index('subject')
        merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()
        if len(merged) < 2:
            continue
        v1 = merged[f'{metric}_1'].values
        v2 = merged[f'{metric}_2'].values
        try:
            _, p_pw = wilcoxon(v1, v2, alternative='two-sided')
        except ValueError:
            p_pw = 1.0
        p_adj = min(p_pw * N_POSTHOC, 1.0)
        delta, mag = cliffs_delta_wrapper(v1, v2)

        label = f"p={format_p(p_adj)}"
        if p_adj < 0.05:
            label = sig_stars(p_adj)

        sig_brackets.append((positions[i1], positions[i2], label, p_adj))

    # Draw ALL brackets
    if sig_brackets:
        all_vals = np.concatenate([
            df_agg[df_agg['train_condition'] == c][metric].dropna().values
            for c in COND_ORDER])
        y_data_max = np.max(all_vals)
        bracket_base = y_data_max + 0.04
        bracket_step = 0.06

        sig_brackets.sort(key=lambda b: b[1] - b[0])

        for b_idx, (x1, x2, label, p_adj) in enumerate(sig_brackets):
            y_bar = bracket_base + b_idx * bracket_step
            y_tip = y_bar - 0.015

            lw = 1.2 if p_adj < 0.05 else 0.8
            col = 'black' if p_adj < 0.05 else 'gray'

            ax.plot([x1, x1, x2, x2], [y_tip, y_bar, y_bar, y_tip],
                    color=col, linewidth=lw, zorder=6)
            ax.text((x1 + x2) / 2, y_bar + 0.005, label,
                    ha='center', va='bottom', fontsize=10,
                    fontweight='bold' if p_adj < 0.05 else 'normal',
                    color=col, zorder=6)

    # Apply shared y-axis
    ax.set_ylim(ylim)

    # Severity counts per condition
    if has_severity:
        for i, cond in enumerate(COND_ORDER):
            sub = df_agg[df_agg['train_condition'] == cond]
            pos = positions[i]
            counts = []
            for sev in SEVERITY_LEVELS:
                n_sev = (sub['severity_level'] == sev).sum()
                counts.append(f"{SEVERITY_LABELS[sev]}={n_sev}")
            count_text = "  ".join(counts)
            ax.text(pos, -0.01, count_text, ha='center', va='top',
                    fontsize=7, color='gray', style='italic',
                    transform=ax.get_xaxis_transform())

    # Legend
    legend_handles = []
    if has_severity:
        for sev in SEVERITY_LEVELS:
            legend_handles.append(mpatches.Patch(
                color=SEVERITY_COLORS[sev], label=f"{SEVERITY_LABELS[sev]} WMH"))
    legend_handles.append(plt.Line2D([0], [0], marker='o', color='black',
                                      linestyle='None', markersize=7, label='Mean'))
    legend_handles.append(plt.Line2D([0], [0], color='black', linewidth=2.5,
                                      label='\u00B1 SD'))
    legend_handles.append(plt.Line2D([0], [0], color='black', linewidth=1.8,
                                      linestyle=':', label='95% CI'))
    fig.legend(handles=legend_handles, loc='lower center',
               bbox_to_anchor=(0.5, -0.02), ncol=len(legend_handles),
               fontsize=9, framealpha=0.9, edgecolor='lightgray')

    fig.tight_layout(rect=[0, 0.05, 1, 1])
    fname = f"{metric}_train_conditions_{threshold}"
    fig.savefig(os.path.join(output_dir, f"{fname}.png"), dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {fname}.png")


# =============================================================
# BLAND-ALTMAN: COMPUTE STATS
# =============================================================

def compute_ba_stats(df_agg, cond1, cond2):
    """
    Compute Bland-Altman statistics for one comparison across
    dice_score, sensitivity, precision.
    """
    label1 = COND_LABELS[cond1]
    label2 = COND_LABELS[cond2]

    d1 = df_agg[df_agg['train_condition'] == cond1].set_index('subject')
    d2 = df_agg[df_agg['train_condition'] == cond2].set_index('subject')
    common = d1.index.intersection(d2.index)
    d1, d2 = d1.loc[common], d2.loc[common]

    has_lesion_type = 'lesion_type' in d1.columns
    lesion_types = d1['lesion_type'].values if has_lesion_type else np.array(['unknown'] * len(d1))

    panels = []
    for metric in BA_METRICS:
        metric_label = BA_METRIC_LABELS[metric]
        vals_1 = d1[metric].values
        vals_2 = d2[metric].values
        diffs = vals_1 - vals_2
        means = (vals_1 + vals_2) / 2.0

        mean_diff = np.mean(diffs)
        median_diff = np.median(diffs)
        sd_diff = np.std(diffs, ddof=1)
        loa_lower = mean_diff - 1.96 * sd_diff
        loa_upper = mean_diff + 1.96 * sd_diff

        outlier_lo = mean_diff - OUTLIER_FACTOR * sd_diff
        outlier_hi = mean_diff + OUTLIER_FACTOR * sd_diff
        is_outlier = (diffs < outlier_lo) | (diffs > outlier_hi)

        try:
            _, p_bias = wilcoxon(diffs, alternative='two-sided')
        except ValueError:
            p_bias = 1.0
        systematic = p_bias < 0.05

        delta, mag = cliffs_delta_wrapper(vals_1, vals_2)
        ci_lo, ci_hi = bootstrap_cliffs_delta_ci(vals_1, vals_2)
        meaningful = abs(delta) >= DELTA_THRESHOLD if not np.isnan(delta) else False

        if systematic or meaningful:
            higher = label2 if mean_diff < 0 else label1 if mean_diff > 0 else 'Equal'
        else:
            higher = 'No meaningful difference'

        panels.append({
            'Metric': metric_label,
            'n': len(diffs),
            'Mean Diff (bias)': round(mean_diff, 6),
            'Median Diff': round(median_diff, 6),
            'SD Diff': round(sd_diff, 6),
            '95% LoA Lower': round(loa_lower, 3),
            '95% LoA Upper': round(loa_upper, 3),
            'Systematic Bias': 'Yes' if systematic else 'No',
            'Bias p-value': format_p_table(p_bias),
            'Higher Condition': higher,
            "Cliff's \u03b4 [95% CI]": format_delta_ci(delta, ci_lo, ci_hi, mag),
            '_diffs': diffs,
            '_means': means,
            '_is_outlier': is_outlier,
            '_lesion_types': lesion_types,
            '_has_lesion_type': has_lesion_type,
            '_systematic': systematic,
            '_meaningful': meaningful,
            '_loa_lower': loa_lower,
            '_loa_upper': loa_upper,
            '_median_diff': median_diff,
            '_p_bias': p_bias,
            '_delta': delta,
            '_ci_lo': ci_lo,
            '_ci_hi': ci_hi,
            '_mag': mag,
            '_higher': higher,
        })

    return panels


def get_global_ylim(all_panels):
    """Symmetric y-axis limits across ALL BA panels."""
    global_max = 0
    for panels in all_panels:
        for p in panels:
            abs_max = np.max(np.abs(p['_diffs']))
            if abs_max > global_max:
                global_max = abs_max
    padding = global_max * 0.15
    ylim = global_max + padding
    return (-ylim, ylim)


# =============================================================
# BLAND-ALTMAN: PLOT
# =============================================================

def plot_ba_comparison(panels, cond1, cond2, ylim):
    """3-panel BA figure (dice, sensitivity, precision) with shared y-axis."""
    label1 = COND_LABELS[cond1]
    label2 = COND_LABELS[cond2]
    has_lesion_type = panels[0]['_has_lesion_type']

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    fig.suptitle(f'Phase I: Bland-Altman: {label1}  vs  {label2}  (B+L, evaluated on Inpainted)',
                 fontsize=14, fontweight='bold', y=1.02)

    for ax_idx, p in enumerate(panels):
        ax = axes[ax_idx]
        diffs = p['_diffs']
        means = p['_means']
        is_outlier = p['_is_outlier']
        lesion_types = p['_lesion_types']
        loa_lower = p['_loa_lower']
        loa_upper = p['_loa_upper']
        median_diff = p['_median_diff']
        metric_label = p['Metric']

        ax.set_ylim(ylim)

        for lt in sorted(set(lesion_types)):
            mask = (lesion_types == lt) & (~is_outlier)
            color = LESION_COLORS.get(lt, '#888888')
            ax.scatter(means[mask], diffs[mask],
                       c=color, marker='o', s=50, alpha=0.7,
                       edgecolors='white', linewidths=0.5,
                       label=lt if ax_idx == 0 else None)

        if np.any(is_outlier):
            for lt in sorted(set(lesion_types)):
                mask = (lesion_types == lt) & is_outlier
                if not np.any(mask):
                    continue
                color = LESION_COLORS.get(lt, '#888888')
                ax.scatter(means[mask], diffs[mask],
                           c=color, marker='x', s=100, linewidths=2,
                           label='Outlier' if ax_idx == 0 and lt == sorted(set(lesion_types))[0] else None)

        ax.axhline(median_diff, color='black', linestyle='--', linewidth=1, alpha=0.8)
        ax.axhline(loa_lower, color='green', linestyle='--', linewidth=1, alpha=0.7)
        ax.axhline(loa_upper, color='red', linestyle='--', linewidth=1, alpha=0.7)

        ax.set_title(f"{metric_label}\n95% LoA: {loa_lower:.3f} \u2013 {loa_upper:.3f}",
                     fontsize=11, fontweight='bold')
        ax.set_xlabel(f"Mean {metric_label}", fontsize=10)
        ax.set_ylabel("Difference" if ax_idx == 0 else "", fontsize=10)

    legend_handles = []
    if has_lesion_type:
        for lt in sorted(LESION_COLORS.keys()):
            legend_handles.append(plt.Line2D(
                [0], [0], marker='o', color='w',
                markerfacecolor=LESION_COLORS[lt], markersize=8, label=lt))
    legend_handles += [
        plt.Line2D([0], [0], color='green', linestyle='--', label='Lower LoA'),
        plt.Line2D([0], [0], color='red', linestyle='--', label='Upper LoA'),
        plt.Line2D([0], [0], color='black', linestyle='--', label='Median'),
        plt.Line2D([0], [0], marker='x', color='gray', linestyle='None',
                   markersize=8, label='Outlier'),
    ]
    fig.legend(handles=legend_handles, loc='lower center',
               bbox_to_anchor=(0.5, -0.05), ncol=min(len(legend_handles), 9),
               fontsize=9, framealpha=0.9, edgecolor='lightgray')

    fig.tight_layout()
    fname = f"BA_{cond1}_vs_{cond2}_locate"
    fig.savefig(os.path.join(output_dir, f"{fname}.png"), dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {fname}.png")


# =============================================================
# EXCEL FORMATTING
# =============================================================

def format_excel(filepath, sheet_metadata=None):
    wb = load_workbook(filepath)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    meta_font = Font(italic=True, size=9, name='Arial', color='555555')
    data_font = Font(size=9, name='Arial')
    border = Border(bottom=Side(style='thin', color='D9D9D9'))
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for ws in wb.worksheets:
        n_meta = 0
        if sheet_metadata and ws.title in sheet_metadata:
            meta_lines = sheet_metadata[ws.title]
            n_meta = len(meta_lines)
            ws.insert_rows(1, n_meta)
            for i, line in enumerate(meta_lines, start=1):
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = meta_font
                cell.alignment = left
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=ws.max_column)

        header_row = n_meta + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'

        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                         for r in range(header_row, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 35)

    wb.save(filepath)


# =============================================================
# TABLE TO IMAGE (JPG)
# =============================================================

def table_to_image(df, filepath, col_widths=None):
    """Render a DataFrame as a publication-ready table image (Excel-like style)."""
    n_rows, n_cols = df.shape

    col_w = col_widths if col_widths else [max(len(str(c)), max(len(str(v)) for v in df[c])) * 0.11 + 0.1 for c in df.columns]
    fig_w = sum(col_w)
    row_h = 0.28
    fig_h = (n_rows + 1) * row_h

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    table = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc='center',
        loc='center',
    )
    table.auto_set_font_size(False)

    for j in range(n_cols):
        cell = table[0, j]
        cell.set_facecolor('#2F5496')
        cell.set_text_props(color='white', fontweight='bold',
                            fontsize=11, fontfamily='sans-serif')
        cell.set_edgecolor('#1a3a6e')
        cell.set_linewidth(1.5)
        cell.set_height(1.0 / (n_rows + 1))

    for i in range(1, n_rows + 1):
        for j in range(n_cols):
            cell = table[i, j]
            cell.set_facecolor('white')
            cell.set_edgecolor('black')
            cell.set_linewidth(0.8)
            cell.set_height(1.0 / (n_rows + 1))
            cell.set_text_props(fontsize=11, fontfamily='sans-serif')

    for j, w in enumerate(col_w):
        for i in range(n_rows + 1):
            table[i, j].set_width(w / fig_w)

    fig.patch.set_facecolor('#FEFEFE')
    fig.savefig(filepath, dpi=300, bbox_inches='tight', pad_inches=0,
                facecolor='#FEFEFE', edgecolor='none')
    plt.show()
    plt.close(fig)

    from PIL import Image, ImageChops
    img = Image.open(filepath).convert('RGB')
    bg = Image.new('RGB', img.size, (254, 254, 254))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        cropped = img.crop(bbox)
        arr = np.array(cropped)
        mask = (arr[:, :, 0] == 254) & (arr[:, :, 1] == 254) & (arr[:, :, 2] == 254)
        arr[mask] = [255, 255, 255]
        Image.fromarray(arr).save(filepath, quality=95)

    print(f"  Saved: {os.path.basename(filepath)}")


# =============================================================
# CONSOLE SUMMARY
# =============================================================

def print_summary(df_agg, thresh_label):
    """Print descriptive + pairwise summary to console for all violin metrics."""
    for metric in VIOLIN_METRICS:
        metric_label = VIOLIN_METRIC_LABELS[metric]
        print(f"\n--- Descriptive Summary: {metric_label} ({thresh_label}) ---")
        for cond in COND_ORDER:
            vals = df_agg[df_agg['train_condition'] == cond][metric].dropna().values
            ci_lo, ci_hi = bootstrap_ci(vals)
            print(f"  {COND_LABELS[cond]:15s}: n={len(vals)}, "
                  f"Mean={np.mean(vals):.4f} +/- {np.std(vals, ddof=1):.4f}, "
                  f"Median={np.median(vals):.4f}, "
                  f"95% CI=[{ci_lo}, {ci_hi}]")

        print(f"\n--- Wilcoxon Pairwise: {metric_label} ({thresh_label}, Bonferroni k=3) ---")
        for c1, c2 in combinations(COND_ORDER, 2):
            d1 = df_agg[df_agg['train_condition'] == c1][['subject', metric]].set_index('subject')
            d2 = df_agg[df_agg['train_condition'] == c2][['subject', metric]].set_index('subject')
            merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()
            if len(merged) < 2:
                print(f"  {COND_LABELS[c1]} vs {COND_LABELS[c2]}: insufficient pairs")
                continue
            v1 = merged[f'{metric}_1'].values
            v2 = merged[f'{metric}_2'].values
            try:
                stat, p_pw = wilcoxon(v1, v2, alternative='two-sided')
            except ValueError:
                stat, p_pw = np.nan, 1.0
            p_adj = min(p_pw * N_POSTHOC, 1.0)
            delta, mag = cliffs_delta_wrapper(v1, v2)
            print(f"  {COND_LABELS[c1]} vs {COND_LABELS[c2]}: "
                  f"W={stat:.1f}, p_adj={format_p(p_adj)} {sig_stars(p_adj)}, "
                  f"Cliff's d={delta} ({mag})")


# =============================================================
# MAIN
# =============================================================

print("=" * 70)
print("PHASE I: METRICS BY TRAINING CONDITION (TEST=INPAINTED)")
print(f"Test: Inpainted | Thresholds: {[THRESH_LABELS[t] for t in THRESHOLDS]}")
print(f"Train conditions: {[COND_LABELS[c] for c in COND_ORDER]}")
print(f"Violin metrics: {[VIOLIN_METRIC_LABELS[m] for m in VIOLIN_METRICS]}")
print("=" * 70)

files = sorted(glob.glob(os.path.join(
    metrics_result_dir, "bianca_metrics_seed_*.xlsx")))
if not files:
    raise FileNotFoundError(f"No metric files in {metrics_result_dir}")

# Load once
dfs = [pd.read_excel(f) for f in files]
df = pd.concat(dfs, ignore_index=True)
df['threshold'] = df['threshold'].astype(str)

for threshold in THRESHOLDS:
    thresh_label = THRESH_LABELS[threshold]
    print(f"\n{'─' * 50}")
    print(f"  Threshold: {thresh_label}")
    print(f"{'─' * 50}")

    df_agg = load_and_aggregate(df, threshold)

    # ── Shared y-axis for all violin plots ──
    violin_ylim = compute_shared_violin_ylim(df_agg)
    print(f"  Shared violin y-axis: [{violin_ylim[0]:.3f}, {violin_ylim[1]:.3f}]")

    # ── Violin plots for each metric ──
    for metric in VIOLIN_METRICS:
        metric_label = VIOLIN_METRIC_LABELS[metric]
        print(f"\n  Violin: {metric_label}")
        plot_metric_conditions(df_agg, threshold, thresh_label,
                               metric, metric_label, violin_ylim)

    print_summary(df_agg, thresh_label)

    # ── Wilcoxon + Descriptive tables ──
    wilcoxon_df = compute_wilcoxon_pairwise(df_agg)
    descriptive_df = compute_descriptive(df_agg)

    # ── BA analysis: B+L threshold only ──
    if threshold == 'locate':
        print(f"\n{'─' * 50}")
        print(f"  BLAND-ALTMAN ANALYSIS (B+L, evaluated on Inpainted)")
        print(f"{'─' * 50}")

        all_ba_panels = []
        for cond1, cond2 in BA_COMPARISONS:
            panels = compute_ba_stats(df_agg, cond1, cond2)
            all_ba_panels.append((cond1, cond2, panels))

        ylim = get_global_ylim([p for _, _, p in all_ba_panels])
        print(f"  Global BA y-axis: [{ylim[0]:.4f}, {ylim[1]:.4f}]")

        all_ba_results = []
        for cond1, cond2, panels in all_ba_panels:
            label1, label2 = COND_LABELS[cond1], COND_LABELS[cond2]
            print(f"\n  BA: {label1} vs {label2}")

            plot_ba_comparison(panels, cond1, cond2, ylim)

            table_rows = [{k: v for k, v in p.items() if not k.startswith('_')}
                          for p in panels]
            ba_df = pd.DataFrame(table_rows)
            ba_df.insert(0, 'Comparison', f"{label1} vs {label2}")
            all_ba_results.append(ba_df)

            print(ba_df[['Metric', 'Mean Diff (bias)', '95% LoA Lower',
                         '95% LoA Upper', 'Systematic Bias', 'Higher Condition',
                         "Cliff's \u03b4 [95% CI]"]].to_string(index=False))

        combined_ba = pd.concat(all_ba_results, ignore_index=True)

        # ── Save all tables in one Excel file ──
        xlsx_path = os.path.join(output_dir, "locate_train_conditions_results.xlsx")

        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            wilcoxon_df.to_excel(writer, sheet_name='Wilcoxon Pairwise', index=False)
            descriptive_df.to_excel(writer, sheet_name='Descriptive', index=False)
            combined_ba.to_excel(writer, sheet_name='Bland-Altman', index=False)

        meta = {
            'Wilcoxon Pairwise': [
                "Table: Phase I - Wilcoxon Signed-Rank Pairwise Tests (Training Conditions)",
                f"Design: Models evaluated on {COND_LABELS[TEST_CONDITION]} condition, threshold B+L.",
                f"Bonferroni correction: k={N_POSTHOC} pairwise comparisons per metric family (alpha=0.0167).",
                "Family structure: Dice, Sensitivity, Precision treated as separate hypothesis families,",
                "each addressing a distinct aspect of segmentation performance (overlap, detection, false positives).",
                "Global correction (k=9) not applied to avoid overcorrection given inherent metric correlation.",
                "Cliff's Delta with bootstrapped 95% CI (1000 iterations).",
                f"Meaningful effect: |delta| >= {DELTA_THRESHOLD}.",
                "Per-subject means across 10 seeds x 5-fold stratified CV.",
            ],
            'Descriptive': [
                "Table: Phase I - Descriptive Statistics (Training Conditions)",
                f"Design: Models evaluated on {COND_LABELS[TEST_CONDITION]} condition, threshold B+L.",
                "95% CI via bootstrap (1000 iterations).",
                "Per-subject means across 10 seeds x 5-fold stratified CV.",
            ],
            'Bland-Altman': [
                "Table: Phase I - Bland-Altman Agreement (Training Conditions, B+L Threshold)",
                f"Design: Models evaluated on {COND_LABELS[TEST_CONDITION]} condition, threshold B+L.",
                "Difference = Condition_1 - Condition_2. Positive = Condition_1 higher.",
                "95% LoA = mean difference +/- 1.96 x SD of differences.",
                "Systematic bias: Wilcoxon signed-rank on differences (alpha=0.05, uncorrected).",
                "Note: 9 bias tests are performed without multiple-testing correction (standard BA practice).",
                "Marginal p-values (0.03-0.05) should be interpreted with caution; all Cliff's delta are negligible.",
                f"Higher Condition: only stated when bias is significant or |delta| >= {DELTA_THRESHOLD}.",
                "Y-axis: symmetric, shared across all 9 panels for direct comparison.",
                "Per-subject means across 10 seeds x 5-fold stratified CV.",
            ],
        }
        format_excel(xlsx_path, meta)
        print(f"\n  Saved combined table: {xlsx_path}")

        # Export tables as images
        wilcoxon_img = os.path.join(output_dir, "wilcoxon_train_conditions.jpg")
        table_to_image(wilcoxon_df, wilcoxon_img)
        desc_img = os.path.join(output_dir, "descriptive_train_conditions.jpg")
        table_to_image(descriptive_df, desc_img)
        ba_img = os.path.join(output_dir, "bland_altman_stats.jpg")
        table_to_image(combined_ba, ba_img)

print(f"\nAll outputs in: {output_dir}")