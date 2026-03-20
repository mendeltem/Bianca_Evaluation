#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cluster-Level Threshold Comparison by Severity (Inpainted Condition)
============================================================

Revision context (NeuroImage: Clinical, Major Revision)
-------------------------------------------------------
Cluster-level violin plots comparing three thresholds (B0.85, B0.90, B+L) for the
inpainted condition. Models are trained on the inpainted condition only;
segmentation is evaluated on the inpainted condition. Individual data points
are colored by WMH severity level (high, middle, low).

Outputs:
  1. Violin plot per metric (F1, Precision, Recall) with
     severity-colored scatter points, mean +/- SD, bootstrap 95% CI.
  2. Descriptive statistics table stratified by threshold x severity.
  3. Pairwise Wilcoxon signed-rank tests (Bonferroni-corrected, k=3).

Design: Models trained on inpainted condition, evaluated on inpainted condition.
Per-subject means averaged across 10 seeds x 5-fold stratified CV.

@author: temuuleu
"""

import os
import glob
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import wilcoxon
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cliffs_delta import cliffs_delta as _cliffs_delta

warnings.filterwarnings('ignore')

# =============================================================
# CONFIG
# =============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
metrics_result_dir = os.path.join(SCRIPT_DIR, "bianca_result_dir")
output_dir = os.path.join(SCRIPT_DIR, "analysis", "cluster_threshold_severity")
os.makedirs(output_dir, exist_ok=True)

# Fixed condition
TRAIN_CONDITION = 'filled'
TEST_CONDITION = 'filled'

THRESHOLDS = ['85', '90', 'locate']
THRESH_LABELS = {'85': 'B0.85', '90': 'B0.90', 'locate': 'B+L'}
THRESH_ORDER = ['locate', '85', '90']

METRICS = ['lesion_f1', 'lesion_precision', 'lesion_recall']
METRIC_LABELS = {
    'lesion_f1': 'Lesion-level F1',
    'lesion_precision': 'Lesion-level Precision',
    'lesion_recall': 'Lesion-level Recall',
}

SEVERITY_LEVELS = ['high', 'middle', 'low']
SEVERITY_COLORS = {
    'high': '#D62728',
    'middle': '#FF7F0E',
    'low': '#2CA02C',
}
SEVERITY_LABELS = {
    'high': 'High',
    'middle': 'Middle',
    'low': 'Low',
}

N_POSTHOC = 3


# =============================================================
# HELPERS
# =============================================================

def cliffs_delta_wrapper(x, y):
    x, y = list(x), list(y)
    if len(x) == 0 or len(y) == 0:
        return 'N/A', 'N/A'
    d, size = _cliffs_delta(x, y)
    return f"{d:.4f}", size


def bootstrap_ci(data, n_boot=1000, alpha=0.05, seed=42):
    """Bootstrap 95% CI for the mean."""
    rng = np.random.default_rng(seed)
    data = np.array(data)
    if len(data) < 2:
        return np.nan, np.nan
    means = [np.mean(rng.choice(data, size=len(data), replace=True))
             for _ in range(n_boot)]
    lo = np.percentile(means, 100 * alpha / 2)
    hi = np.percentile(means, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def format_p(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.4f}"


def sig_stars(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '***'
    if p < 0.01:
        return '**'
    if p < 0.05:
        return '*'
    return 'ns'


# =============================================================
# LOAD + AGGREGATE
# =============================================================

def load_and_aggregate():
    """Load, filter to inpainted condition, aggregate per subject x threshold."""
    files = sorted(glob.glob(os.path.join(
        metrics_result_dir, "bianca_metrics_seed_*.xlsx")))
    if not files:
        raise FileNotFoundError(f"No metric files in {metrics_result_dir}")

    dfs = [pd.read_excel(f) for f in files]
    df = pd.concat(dfs, ignore_index=True)
    df['threshold'] = df['threshold'].astype(str)

    # Filter: trained on inpainted, evaluated on inpainted
    df = df[(df['train_condition'] == TRAIN_CONDITION) &
            (df['test_condition'] == TEST_CONDITION)].copy()

    print(f"Filtered: {len(df)} rows | Thresholds: {sorted(df['threshold'].unique())}")

    # Aggregate per subject x threshold
    group_cols = ['subject', 'threshold']
    extra_cols = [c for c in ['scanner', 'severity_level', 'lesion_type']
                  if c in df.columns]

    agg_dict = {m: 'mean' for m in METRICS if m in df.columns}
    for col in extra_cols:
        agg_dict[col] = 'first'

    df_agg = df.groupby(group_cols).agg(agg_dict).reset_index()
    print(f"Aggregated: {len(df_agg)} rows ({df_agg['subject'].nunique()} subjects)")

    if 'severity_level' in df_agg.columns:
        print(f"Severity distribution: {df_agg.groupby('severity_level')['subject'].nunique().to_dict()}")

    return df_agg


# =============================================================
# DESCRIPTIVE TABLE (threshold x severity)
# =============================================================

def descriptive_table(df_agg):
    """Descriptive stats per threshold x severity x metric."""
    rows = []
    for metric in METRICS:
        metric_label = METRIC_LABELS[metric]
        for thresh in THRESH_ORDER:
            t_label = THRESH_LABELS[thresh]
            sub_t = df_agg[df_agg['threshold'] == thresh]

            # Overall for this threshold
            vals_all = sub_t[metric].dropna().values
            ci_lo, ci_hi = bootstrap_ci(vals_all)
            rows.append({
                'Metric': metric_label,
                'Threshold': t_label,
                'Severity': 'All',
                'n': len(vals_all),
                'Mean': f"{np.mean(vals_all):.4f}" if len(vals_all) > 0 else 'N/A',
                'SD': f"{np.std(vals_all, ddof=1):.4f}" if len(vals_all) > 1 else 'N/A',
                'Median': f"{np.median(vals_all):.4f}" if len(vals_all) > 0 else 'N/A',
                'IQR': f"{np.percentile(vals_all, 75) - np.percentile(vals_all, 25):.4f}" if len(vals_all) > 0 else 'N/A',
                '95% CI': f"[{ci_lo}, {ci_hi}]" if not np.isnan(ci_lo) else 'N/A',
            })

            # Per severity
            if 'severity_level' in sub_t.columns:
                for sev in SEVERITY_LEVELS:
                    vals = sub_t[sub_t['severity_level'] == sev][metric].dropna().values
                    ci_lo, ci_hi = bootstrap_ci(vals)
                    rows.append({
                        'Metric': metric_label,
                        'Threshold': t_label,
                        'Severity': SEVERITY_LABELS.get(sev, sev),
                        'n': len(vals),
                        'Mean': f"{np.mean(vals):.4f}" if len(vals) > 0 else 'N/A',
                        'SD': f"{np.std(vals, ddof=1):.4f}" if len(vals) > 1 else 'N/A',
                        'Median': f"{np.median(vals):.4f}" if len(vals) > 0 else 'N/A',
                        'IQR': f"{np.percentile(vals, 75) - np.percentile(vals, 25):.4f}" if len(vals) > 0 else 'N/A',
                        '95% CI': f"[{ci_lo}, {ci_hi}]" if not np.isnan(ci_lo) else 'N/A',
                    })

    return pd.DataFrame(rows)


# =============================================================
# WILCOXON PAIRWISE (3 thresholds)
# =============================================================

def analysis_wilcoxon_pairwise(df_agg):
    """Wilcoxon signed-rank pairwise (Bonferroni-corrected) across 3 thresholds."""
    pw_rows = []

    for metric in METRICS:
        metric_label = METRIC_LABELS[metric]

        for t1, t2 in combinations(THRESH_ORDER, 2):
            tl1, tl2 = THRESH_LABELS[t1], THRESH_LABELS[t2]

            d1 = df_agg[df_agg['threshold'] == t1][['subject', metric]].set_index('subject')
            d2 = df_agg[df_agg['threshold'] == t2][['subject', metric]].set_index('subject')
            merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()

            if len(merged) < 2:
                pw_rows.append({
                    'Metric': metric_label, 'Comparison': f"{tl1} vs {tl2}",
                    'W-stat': 'N/A', 'p-adj (Bonferroni)': 'N/A',
                    "Cliff's delta": 'N/A', 'Effect Size': 'N/A',
                })
                continue

            v1 = merged[f'{metric}_1'].values
            v2 = merged[f'{metric}_2'].values

            try:
                stat, p_pw = wilcoxon(v1, v2, alternative='two-sided')
            except ValueError:
                stat, p_pw = np.nan, 1.0

            p_adj = min(p_pw * N_POSTHOC, 1.0)
            delta, mag = cliffs_delta_wrapper(v1, v2)

            pw_rows.append({
                'Metric': metric_label,
                'Comparison': f"{tl1} vs {tl2}",
                'W-stat': f"{stat:.1f}" if not pd.isna(stat) else 'N/A',
                'p-adj (Bonferroni)': format_p(p_adj),
                "Cliff's delta": delta,
                'Effect Size': mag,
            })

    return pd.DataFrame(pw_rows)


# =============================================================
# VIOLIN PLOT WITH SEVERITY-COLORED POINTS
# =============================================================

def plot_threshold_severity(df_agg, metric):
    """
    Single violin plot: one violin per threshold, individual points
    colored by severity level.
    """
    metric_label = METRIC_LABELS[metric]
    has_severity = 'severity_level' in df_agg.columns

    fig, ax = plt.subplots(figsize=(10, 7))

    # Prepare data per threshold
    positions = [1, 2, 3]
    violin_data = []
    for thresh in THRESH_ORDER:
        vals = df_agg[df_agg['threshold'] == thresh][metric].dropna().values
        violin_data.append(vals)

    # Violin plots
    vp = ax.violinplot(violin_data, positions=positions, showmeans=False,
                       showmedians=False, showextrema=False, widths=0.7)
    for body in vp['bodies']:
        body.set_facecolor('#B0C4DE')
        body.set_edgecolor('#4C72B0')
        body.set_alpha(0.35)

    # Scatter points colored by severity + jitter
    rng = np.random.default_rng(42)
    for i, thresh in enumerate(THRESH_ORDER):
        sub = df_agg[df_agg['threshold'] == thresh]
        pos = positions[i]

        if has_severity:
            for sev in SEVERITY_LEVELS:
                mask = sub['severity_level'] == sev
                vals = sub.loc[mask, metric].dropna().values
                jitter = rng.uniform(-0.18, 0.18, size=len(vals))
                color = SEVERITY_COLORS.get(sev, '#888888')
                ax.scatter(pos + jitter, vals, c=color, s=35, alpha=0.75,
                           edgecolors='white', linewidths=0.4, zorder=3)
        else:
            vals = sub[metric].dropna().values
            jitter = rng.uniform(-0.18, 0.18, size=len(vals))
            ax.scatter(pos + jitter, vals, c='#4C72B0', s=35, alpha=0.75,
                       edgecolors='white', linewidths=0.4, zorder=3)

    # Mean, SD, CI per threshold
    annotation_texts = []
    for i, thresh in enumerate(THRESH_ORDER):
        vals = df_agg[df_agg['threshold'] == thresh][metric].dropna().values
        pos = positions[i]
        t_label = THRESH_LABELS[thresh]

        if len(vals) == 0:
            continue

        mean_val = np.mean(vals)
        sd_val = np.std(vals, ddof=1)
        ci_lo, ci_hi = bootstrap_ci(vals)

        # Mean dot
        ax.scatter([pos], [mean_val], c='black', s=90, zorder=5, marker='o')

        # SD bars (mean +/- SD)
        ax.plot([pos, pos], [mean_val - sd_val, mean_val + sd_val],
                color='black', linewidth=2.5, zorder=4)
        ax.plot([pos - 0.1, pos + 0.1], [mean_val - sd_val, mean_val - sd_val],
                color='black', linewidth=2.5, zorder=4)
        ax.plot([pos - 0.1, pos + 0.1], [mean_val + sd_val, mean_val + sd_val],
                color='black', linewidth=2.5, zorder=4)

        # CI as dotted lines
        ax.plot([pos - 0.08, pos + 0.08], [ci_lo, ci_lo],
                color='black', linewidth=1.8, linestyle=':', zorder=4)
        ax.plot([pos - 0.08, pos + 0.08], [ci_hi, ci_hi],
                color='black', linewidth=1.8, linestyle=':', zorder=4)

        # Collect annotation for x-axis label area
        annotation_texts.append(
            f"{t_label}\n"
            f"Mean={mean_val:.2f}\u00B1{sd_val:.2f}\n"
            f"CI=({ci_lo:.2f}, {ci_hi:.2f})"
        )

    # Custom x-axis labels with stats below
    ax.set_xticks(positions)
    ax.set_xticklabels(annotation_texts, fontsize=9, linespacing=1.4)
    ax.tick_params(axis='x', pad=8)

    ax.set_xlabel('Threshold Strategy', fontsize=12, labelpad=12)
    ax.set_ylabel(metric_label, fontsize=12)
    ax.set_title(f"{metric_label} by Threshold (Inpainted Condition)",
                 fontsize=14, fontweight='bold', pad=15)

    # -- Wilcoxon significance brackets --
    # Compute pairwise Wilcoxon between thresholds for this metric
    bracket_pairs = list(combinations(range(len(THRESH_ORDER)), 2))
    sig_brackets = []
    for i1, i2 in bracket_pairs:
        t1, t2 = THRESH_ORDER[i1], THRESH_ORDER[i2]
        d1 = df_agg[df_agg['threshold'] == t1][['subject', metric]].set_index('subject')
        d2 = df_agg[df_agg['threshold'] == t2][['subject', metric]].set_index('subject')
        merged = d1.join(d2, lsuffix='_1', rsuffix='_2', how='inner').dropna()
        if len(merged) < 2:
            continue
        v1 = merged[f'{metric}_1'].values
        v2 = merged[f'{metric}_2'].values
        try:
            _, p_pw = wilcoxon(v1, v2, alternative='two-sided')
        except ValueError:
            p_pw = 1.0
        p_adj = min(p_pw * N_POSTHOC, 1.0)
        if p_adj < 0.05:
            if p_adj < 0.001:
                stars = '***'
            elif p_adj < 0.01:
                stars = '**'
            else:
                stars = '*'
            sig_brackets.append((positions[i1], positions[i2], stars))

    # Draw brackets above the data
    if sig_brackets:
        # Find top of data
        all_vals = np.concatenate([df_agg[df_agg['threshold'] == t][metric].dropna().values
                                   for t in THRESH_ORDER])
        y_data_max = np.max(all_vals)
        bracket_base = y_data_max + 0.04
        bracket_step = 0.06

        # Sort brackets by span (narrowest first, drawn lowest)
        sig_brackets.sort(key=lambda b: b[1] - b[0])

        for b_idx, (x1, x2, stars) in enumerate(sig_brackets):
            y_bar = bracket_base + b_idx * bracket_step
            y_tip = y_bar - 0.015

            # Horizontal bar
            ax.plot([x1, x1, x2, x2], [y_tip, y_bar, y_bar, y_tip],
                    color='black', linewidth=1.2, zorder=6)
            # Stars
            ax.text((x1 + x2) / 2, y_bar + 0.005, stars,
                    ha='center', va='bottom', fontsize=11, fontweight='bold',
                    zorder=6)

        # Set ylim to accommodate brackets
        y_top = bracket_base + len(sig_brackets) * bracket_step + 0.06
        ax.set_ylim(-0.02, max(1.02, y_top))
    else:
        ax.set_ylim(-0.02, 1.02)

    # Severity count annotation per threshold (inside plot, bottom)
    if has_severity:
        for i, thresh in enumerate(THRESH_ORDER):
            sub = df_agg[df_agg['threshold'] == thresh]
            pos = positions[i]
            counts = []
            for sev in SEVERITY_LEVELS:
                n_sev = (sub['severity_level'] == sev).sum()
                counts.append(f"{SEVERITY_LABELS[sev]}={n_sev}")
            count_text = "  ".join(counts)
            ax.text(pos, -0.01, count_text, ha='center', va='top',
                    fontsize=7, color='gray', style='italic',
                    transform=ax.get_xaxis_transform())

    # Legend below plot
    legend_handles = []
    if has_severity:
        for sev in SEVERITY_LEVELS:
            legend_handles.append(mpatches.Patch(
                color=SEVERITY_COLORS[sev], label=f"{SEVERITY_LABELS[sev]} WMH"))
    legend_handles.append(plt.Line2D([0], [0], marker='o', color='black',
                                      linestyle='None', markersize=7, label='Mean'))
    legend_handles.append(plt.Line2D([0], [0], color='black', linewidth=2.5,
                                      label='\u00B1 SD'))
    legend_handles.append(plt.Line2D([0], [0], color='black', linewidth=1.8,
                                      linestyle=':', label='95% CI'))
    fig.legend(handles=legend_handles, loc='lower center',
               bbox_to_anchor=(0.5, -0.02), ncol=len(legend_handles),
               fontsize=9, framealpha=0.9, edgecolor='lightgray')

    fig.tight_layout(rect=[0, 0.05, 1, 1])
    fname = f"{metric}_threshold_severity"
    fig.savefig(os.path.join(output_dir, f"{fname}.png"), dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {fname}.png")


# =============================================================
# SEVERITY DISTRIBUTION PLOT
# =============================================================

def plot_severity_distribution(df_agg):
    """
    Bar chart showing severity level distribution, overall and by scanner.
    Uses one threshold slice (counts are identical across thresholds).
    """
    # Use one threshold to get unique subjects
    sub = df_agg[df_agg['threshold'] == THRESH_ORDER[0]].copy()
    has_scanner = 'scanner' in sub.columns

    if has_scanner:
        fig, axes = plt.subplots(1, 2, figsize=(12, 5), gridspec_kw={'width_ratios': [1, 1.8]})
        ax_overall, ax_scanner = axes
    else:
        fig, ax_overall = plt.subplots(figsize=(5, 5))

    # --- Overall severity counts ---
    sev_counts = sub['severity_level'].value_counts().reindex(SEVERITY_LEVELS).fillna(0).astype(int)
    bars = ax_overall.bar(
        [SEVERITY_LABELS[s] for s in SEVERITY_LEVELS],
        [sev_counts[s] for s in SEVERITY_LEVELS],
        color=[SEVERITY_COLORS[s] for s in SEVERITY_LEVELS],
        edgecolor='white', linewidth=1.2, width=0.6
    )
    # Count labels on bars
    for bar, sev in zip(bars, SEVERITY_LEVELS):
        count = sev_counts[sev]
        pct = 100 * count / len(sub)
        ax_overall.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                        f"n={count}\n({pct:.0f}%)",
                        ha='center', va='bottom', fontsize=10, fontweight='bold')

    ax_overall.set_ylabel('Number of Subjects', fontsize=11)
    ax_overall.set_title('Severity Distribution', fontsize=12, fontweight='bold')
    ax_overall.set_ylim(0, max(sev_counts) * 1.25)
    ax_overall.spines['top'].set_visible(False)
    ax_overall.spines['right'].set_visible(False)

    # --- By scanner (grouped bar) ---
    if has_scanner:
        scanners = sorted(sub['scanner'].unique())
        x = np.arange(len(scanners))
        width = 0.25

        for i, sev in enumerate(SEVERITY_LEVELS):
            counts = [len(sub[(sub['scanner'] == sc) & (sub['severity_level'] == sev)]) for sc in scanners]
            b = ax_scanner.bar(x + i * width, counts, width,
                               color=SEVERITY_COLORS[sev], edgecolor='white',
                               linewidth=1.2, label=SEVERITY_LABELS[sev])
            for bar, c in zip(b, counts):
                if c > 0:
                    ax_scanner.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                                    str(c), ha='center', va='bottom', fontsize=9)

        ax_scanner.set_xticks(x + width)
        ax_scanner.set_xticklabels(scanners, fontsize=10)
        ax_scanner.set_ylabel('Number of Subjects', fontsize=11)
        ax_scanner.set_title('Severity by Scanner', fontsize=12, fontweight='bold')
        ax_scanner.legend(fontsize=9)
        ax_scanner.spines['top'].set_visible(False)
        ax_scanner.spines['right'].set_visible(False)
        ax_scanner.set_ylim(0, ax_scanner.get_ylim()[1] * 1.2)

    fig.suptitle(f"WMH Severity Level Distribution (n={len(sub)})",
                 fontsize=14, fontweight='bold', y=1.02)
    fig.tight_layout()
    fig.savefig(os.path.join(output_dir, "severity_distribution.png"),
                dpi=300, bbox_inches='tight')
    plt.close(fig)
    print("  Saved: severity_distribution.png")


# =============================================================
# EXCEL FORMATTING
# =============================================================

def format_excel(filepath, sheet_metadata=None):
    wb = load_workbook(filepath)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    meta_font = Font(italic=True, size=9, name='Arial', color='555555')
    data_font = Font(size=9, name='Arial')
    border = Border(bottom=Side(style='thin', color='D9D9D9'))
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for ws in wb.worksheets:
        n_meta = 0
        if sheet_metadata and ws.title in sheet_metadata:
            meta_lines = sheet_metadata[ws.title]
            n_meta = len(meta_lines)
            ws.insert_rows(1, n_meta)
            for i, line in enumerate(meta_lines, start=1):
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = meta_font
                cell.alignment = left
                # Clear all cells in this row before merging
                for c in range(2, ws.max_column + 1):
                    ws.cell(row=i, column=c, value='')
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=ws.max_column)

        header_row = n_meta + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
                # Ensure decimal points (not commas) for numeric values
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'

        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                         for r in range(header_row, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 35)

    wb.save(filepath)


# =============================================================
# TABLE TO IMAGE (JPG)
# =============================================================

def table_to_image(df, filepath, col_widths=None):
    """Render a DataFrame as a publication-ready table image (Excel-like style)."""
    n_rows, n_cols = df.shape

    # Auto-calculate figure size (compact)
    col_w = col_widths if col_widths else [max(len(str(c)), max(len(str(v)) for v in df[c])) * 0.14 + 0.15 for c in df.columns]
    fig_w = sum(col_w)
    row_h = 0.32
    fig_h = (n_rows + 1) * row_h

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    table = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc='center',
        loc='center',
    )
    table.auto_set_font_size(False)

    # Style header - bold white on dark blue
    for j in range(n_cols):
        cell = table[0, j]
        cell.set_facecolor('#2F5496')
        cell.set_text_props(color='white', fontweight='bold',
                            fontsize=12, fontfamily='sans-serif')
        cell.set_edgecolor('#1a3a6e')
        cell.set_linewidth(1.5)
        cell.set_height(1.0 / (n_rows + 1))

    # Style data rows - white with solid black borders
    for i in range(1, n_rows + 1):
        for j in range(n_cols):
            cell = table[i, j]
            cell.set_facecolor('white')
            cell.set_edgecolor('black')
            cell.set_linewidth(0.8)
            cell.set_height(1.0 / (n_rows + 1))
            cell.set_text_props(fontsize=12, fontfamily='sans-serif')

    # Set column widths
    for j, w in enumerate(col_w):
        for i in range(n_rows + 1):
            table[i, j].set_width(w / fig_w)

    # Save with tiny colored border to enable clean cropping
    fig.patch.set_facecolor('#FEFEFE')
    fig.savefig(filepath, dpi=300, bbox_inches='tight', pad_inches=0,
                facecolor='#FEFEFE', edgecolor='none')
    plt.show()
    plt.close(fig)

    # Auto-crop the near-white border, then restore white background
    from PIL import Image, ImageChops
    img = Image.open(filepath).convert('RGB')
    bg = Image.new('RGB', img.size, (254, 254, 254))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        cropped = img.crop(bbox)
        arr = np.array(cropped)
        mask = (arr[:, :, 0] == 254) & (arr[:, :, 1] == 254) & (arr[:, :, 2] == 254)
        arr[mask] = [255, 255, 255]
        Image.fromarray(arr).save(filepath, quality=95)

    print(f"  Saved: {os.path.basename(filepath)}")


# =============================================================
# MAIN
# =============================================================

if __name__ == "__main__":
    print("=" * 70)
    print("CLUSTER-LEVEL THRESHOLD COMPARISON BY SEVERITY (INPAINTED CONDITION)")
    print(f"Train: Inpainted | Test: Inpainted")
    print(f"Thresholds: {[THRESH_LABELS[t] for t in THRESH_ORDER]}")
    print("=" * 70)

    df_agg = load_and_aggregate()

    # ── 1) Severity distribution ──
    print("\n--- Severity Distribution ---")
    plot_severity_distribution(df_agg)

    # ── 2) Violin plots per metric ──
    print("\n--- Violin Plots ---")
    for metric in METRICS:
        plot_threshold_severity(df_agg, metric)

    # ── 3) Descriptive table (threshold x severity) ──
    print("\n--- Descriptive Table ---")
    desc_df = descriptive_table(df_agg)

    desc_path = os.path.join(output_dir, "descriptive_threshold_severity.xlsx")
    with pd.ExcelWriter(desc_path, engine='openpyxl') as writer:
        desc_df.to_excel(writer, sheet_name='Descriptive', index=False)
    desc_meta = {
        'Descriptive': [
            "Table: Descriptive Statistics by Threshold and WMH Severity Level",
            "Design: Models trained on inpainted condition, evaluated on inpainted condition.",
            "Per-subject means across 10 seeds x 5-fold stratified CV.",
            "95% CI: bootstrap confidence interval (1000 iterations).",
            "Severity: based on WMH burden stratification (high/middle/low).",
        ],
    }
    format_excel(desc_path, desc_meta)
    print(f"  Saved: {desc_path}")

    # Export descriptive table as image
    desc_img = os.path.join(output_dir, "descriptive_threshold_severity.jpg")
    table_to_image(desc_df, desc_img)

    # Print Dice summary
    dice_desc = desc_df[desc_df['Metric'] == 'Lesion-level F1']
    print("\nLesion-level F1 summary:")
    print(dice_desc.to_string(index=False))

    # ── 4) Wilcoxon pairwise (3 thresholds) ──
    print("\n--- Wilcoxon Pairwise: 3 Thresholds ---")
    pw_df = analysis_wilcoxon_pairwise(df_agg)

    pw_path = os.path.join(output_dir, "wilcoxon_thresholds.xlsx")
    with pd.ExcelWriter(pw_path, engine='openpyxl') as writer:
        pw_df.to_excel(writer, sheet_name='Wilcoxon Pairwise', index=False)
    pw_meta = {
        'Wilcoxon Pairwise': [
            "Table: Wilcoxon Signed-Rank Pairwise Tests (Paired, Within-Subject)",
            "Bonferroni correction: 3 comparisons, corrected alpha = 0.0167.",
            "Effect size: Cliff's Delta (negligible/small/medium/large).",
            "Design: Models trained on inpainted condition, evaluated on inpainted condition.",
        ],
    }
    format_excel(pw_path, pw_meta)
    print(f"  Saved: {pw_path}")

    # Export Wilcoxon table as image
    pw_img = os.path.join(output_dir, "wilcoxon_thresholds.jpg")
    table_to_image(pw_df, pw_img)

    print("\nWilcoxon Pairwise:")
    print(pw_df.to_string(index=False))

    print(f"\nAll outputs in: {output_dir}")