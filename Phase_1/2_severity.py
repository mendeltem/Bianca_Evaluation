#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Severity Level Definitions (Phase I + Phase II-A)
=================================================
Generates Excel with WMH severity tercile definitions
for both analysis phases.

Phase I (n=89):  BeLOVE (n=49) + Challenge (n=40), excl. GE
Phase II-A (n=89): BeLOVE (n=59) + LOCATE pool (n=30)

Output: 2_severity_definitions.xlsx
"""

import os
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# --- Data sources ---
# Phase I: 5-fold CV pool (BeLOVE + Challenge, excl. GE, excl. LOCATE holdout)
PHASE1_XLSX = os.path.join(SCRIPT_DIR, "LOCATE_SET", "bianca_pool_wihtouth_ge.xlsx")

# Phase II-A: subjects with ground-truth WMH masks
PHASE2A_XLSX = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_DICE_ONLY.xlsx")

OUTPUT_XLSX = os.path.join(SCRIPT_DIR, "2_severity_definitions.xlsx")


# --- Load Phase I ---
print("Loading Phase I pool ...")
df_p1 = pd.read_excel(PHASE1_XLSX)
print(f"  Phase I: {len(df_p1)} subjects")

# --- Load Phase II-A ---
print("Loading Phase II-A ...")
df_p2a = pd.read_excel(PHASE2A_XLSX)
if "subject_with_mask" in df_p2a.columns:
    df_p2a = df_p2a[df_p2a["subject_with_mask"] == 1].copy()

# Deduplicate to one row per subject (take first occurrence)
if "subject" in df_p2a.columns:
    df_p2a = df_p2a.drop_duplicates(subset="subject", keep="first")
print(f"  Phase II-A: {len(df_p2a)} subjects")


# --- Styles ---
header_fill = PatternFill('solid', fgColor='2F5496')
header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
meta_font = Font(italic=True, size=9, name='Arial', color='555555')
data_font = Font(size=9, name='Arial')
bold_font = Font(bold=True, size=9, name='Arial')
border_thin = Border(bottom=Side(style='thin', color='D9D9D9'))
ctr = Alignment(horizontal='center', vertical='center')
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)


def write_meta(ws, meta_lines, max_col=10):
    for i, line in enumerate(meta_lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = meta_font
        cell.alignment = left_wrap
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=max_col)
    return len(meta_lines)


def write_header(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = ctr


def auto_width(ws, max_col, header_row, max_row):
    for col in range(1, max_col + 1):
        mx = max(len(str(ws.cell(row=r, column=col).value or ''))
                 for r in range(header_row, max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 35)


def write_severity_block(ws, start_row, df, cohort_label):
    """
    Write severity tercile stats for one cohort.
    Returns next available row.
    """
    if 'severity_level' not in df.columns or 'ROI_Volume' not in df.columns:
        ws.cell(row=start_row, column=1,
                value=f"{cohort_label}: severity_level or ROI_Volume not available").font = meta_font
        return start_row + 1

    sub = df[['severity_level', 'ROI_Volume']].dropna()
    all_v = sub['ROI_Volume']
    row = start_row

    # Cohort label row
    ws.cell(row=row, column=1, value=cohort_label).font = bold_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    # Overall
    ws.cell(row=row, column=1, value='Overall').font = bold_font
    ws.cell(row=row, column=2, value=len(sub))
    ws.cell(row=row, column=3, value=f"{all_v.min():.2f}{all_v.max():.2f}")
    ws.cell(row=row, column=4, value=round(all_v.mean(), 2))
    ws.cell(row=row, column=5, value=round(all_v.std(), 2))
    ws.cell(row=row, column=6, value=round(all_v.median(), 2))
    ws.cell(row=row, column=7, value=round(all_v.quantile(0.25), 2))
    ws.cell(row=row, column=8, value=round(all_v.quantile(0.75), 2))

    t33 = all_v.quantile(1/3)
    t67 = all_v.quantile(2/3)
    ws.cell(row=row, column=9, value=f"{t33:.2f} / {t67:.2f}")
    row += 1

    # Per severity level
    for sev in ['low', 'middle', 'high']:
        vals = sub[sub['severity_level'] == sev]['ROI_Volume']
        if len(vals) == 0:
            continue
        ws.cell(row=row, column=1, value=sev.capitalize()).font = data_font
        ws.cell(row=row, column=2, value=len(vals))
        ws.cell(row=row, column=3, value=f"{vals.min():.2f}{vals.max():.2f}")
        ws.cell(row=row, column=4, value=round(vals.mean(), 2))
        ws.cell(row=row, column=5, value=round(vals.std(), 2))
        ws.cell(row=row, column=6, value=round(vals.median(), 2))
        ws.cell(row=row, column=7, value=round(vals.quantile(0.25), 2))
        ws.cell(row=row, column=8, value=round(vals.quantile(0.75), 2))
        ws.cell(row=row, column=9, value='')
        row += 1

    # Format
    for r in range(start_row + 1, row):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.alignment = ctr
            cell.border = border_thin

    return row + 1  # blank row between cohorts


# =========================================================
# BUILD EXCEL
# =========================================================
wb = Workbook()

# --- SHEET 1: Severity Definitions ---
ws1 = wb.active
ws1.title = 'Severity Definitions'

meta1 = [
    "WMH severity level definitions based on ground-truth WMH volume terciles.",
    "Method: Tercile split (33rd/67th percentile) of ground-truth WMH volume (mL).",
    "ROI_Volume = total WMH volume from manual segmentation.",
    "Phase I: 5-fold CV pool (Cohort 1 + Challenge, excl. GE, excl. LOCATE holdout).",
    "Phase II-A: subjects with ground-truth WMH mask (Cohort 1 + part of Cohort 2).",
]
n_meta = write_meta(ws1, meta1, 9)

headers = ['Cohort / Severity', 'n', 'Range (mL)', 'Mean', 'SD', 'Median',
           'IQR Low', 'IQR High', 'Tercile Cutoffs (33rd / 67th)']
hr = n_meta + 1
write_header(ws1, hr, headers)

# Phase I
next_row = write_severity_block(ws1, hr + 1, df_p1,
    f"Phase I (n={len(df_p1)}): BeLOVE + Challenge, excl. GE")

# Phase II-A
next_row = write_severity_block(ws1, next_row, df_p2a,
    f"Phase II-A (n={len(df_p2a)}): BeLOVE + LOCATE pool")

auto_width(ws1, 9, hr, next_row)


# --- SHEET 2: Scanner × Severity (Phase I) ---
if 'scanner' in df_p1.columns and 'severity_level' in df_p1.columns:
    ws2 = wb.create_sheet('Scanner x Severity (Phase I)')
    ct = pd.crosstab(df_p1['scanner'], df_p1['severity_level'], margins=True)

    meta2 = [f"Scanner × WMH severity cross-tabulation (Phase I, n={len(df_p1)})."]
    n_meta2 = write_meta(ws2, meta2, ct.shape[1] + 1)
    hr2 = n_meta2 + 1
    write_header(ws2, hr2, ['Scanner'] + [str(c) for c in ct.columns])

    for r_idx, (idx_val, row_data) in enumerate(ct.iterrows(), hr2 + 1):
        ws2.cell(row=r_idx, column=1, value=str(idx_val)).font = data_font
        for c_idx, val in enumerate(row_data, 2):
            ws2.cell(row=r_idx, column=c_idx, value=int(val)).font = data_font
            ws2.cell(row=r_idx, column=c_idx).alignment = ctr
            ws2.cell(row=r_idx, column=c_idx).border = border_thin
    auto_width(ws2, ct.shape[1] + 1, hr2, hr2 + len(ct))


# --- SHEET 3: Scanner × Severity (Phase II-A) ---
if 'scanner' in df_p2a.columns and 'severity_level' in df_p2a.columns:
    ws3 = wb.create_sheet('Scanner x Severity (Phase II-A)')
    ct2 = pd.crosstab(df_p2a['scanner'], df_p2a['severity_level'], margins=True)

    meta3 = [f"Scanner × WMH severity cross-tabulation (Phase II-A, n={len(df_p2a)})."]
    n_meta3 = write_meta(ws3, meta3, ct2.shape[1] + 1)
    hr3 = n_meta3 + 1
    write_header(ws3, hr3, ['Scanner'] + [str(c) for c in ct2.columns])

    for r_idx, (idx_val, row_data) in enumerate(ct2.iterrows(), hr3 + 1):
        ws3.cell(row=r_idx, column=1, value=str(idx_val)).font = data_font
        for c_idx, val in enumerate(row_data, 2):
            ws3.cell(row=r_idx, column=c_idx, value=int(val)).font = data_font
            ws3.cell(row=r_idx, column=c_idx).alignment = ctr
            ws3.cell(row=r_idx, column=c_idx).border = border_thin
    auto_width(ws3, ct2.shape[1] + 1, hr3, hr3 + len(ct2))


# =========================================================
# SAVE
# =========================================================
wb.save(OUTPUT_XLSX)
print(f"\n✅ {OUTPUT_XLSX}")
print(f"   Sheet 1: Severity Definitions (Phase I + Phase II-A)")
print(f"   Sheet 2: Scanner × Severity (Phase I)")
print(f"   Sheet 3: Scanner × Severity (Phase II-A)")