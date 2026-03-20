#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cluster-Level Analysis of BIANCA WMH Segmentation
===================================================
@author: temuuleu

# =============================================================
# REVIEWER RESPONSE (R1, Comment 2; R5, #17)
# =============================================================
#
# R1 raised the concern that global Dice alone cannot distinguish
# genuine sensitivity gains from spatially distributed false positives.
# R5 (#17) noted that precision values were not contextualized.
#
# This script produces all cluster-level analyses for the manuscript:
#
# MAIN TEXT:
#   1. Voxel vs Cluster dissociation (Figure: Dice vs Cluster F1 scatter)
#   2. Cluster Precision vs Recall (Figure: bar chart)
#   3. R vs NR condition comparison (Figure: boxplot + Kruskal-Wallis)
#   4. Cliff's Delta for R vs NR on cluster-level
#
# SUPPLEMENTAL:
#   5. Post-hoc Wilcoxon with Bonferroni correction
#   6. Per-dataset breakdown (BeLOVE vs Challenge, Mann-Whitney U)
#   7. GT cluster count distribution
#   8. Summary tables (mean, SD, median per condition x threshold)
#
# NOT INCLUDED (not requested by reviewers):
#   - Per-scanner stratification (already covered in voxel-level analysis)
#   - Per-severity stratification (already covered in voxel-level analysis)
#   - Full 9x3 condition matrix (identical values, no added information)
#
# Methodology:
#   - Connected-component labeling: 26-connectivity
#   - Minimum cluster size: threshold- and dataset-specific (grid search)
#   - Overlap criterion: any-overlap (>=1 voxel), MICCAI 2017
#   - Filtering: prediction only; GT retained without filtering
#   - Evaluated: 10 seeds x 5-fold stratified CV x 89 subjects
# =============================================================

# =============================================================
# PAPER TEXT: METHODS
# =============================================================
#
# 2.X Lesion-level evaluation
#
# To complement voxel-level Dice coefficients, we computed lesion-level
# precision, recall, and F1 score using connected-component analysis.
# Binary segmentation masks were decomposed into spatially contiguous
# clusters using 26-connectivity (scipy.ndimage.label). Predicted
# clusters smaller than a threshold-specific minimum cluster size
# (determined via grid search; Supplemental Table X) were removed to
# exclude spurious components; ground truth masks were not filtered.
# Consistent with the MICCAI 2017 WMH Segmentation Challenge (Kuijf
# et al., 2019), a predicted cluster was classified as a true positive
# if it overlapped with any ground truth cluster by at least one voxel
# (any-overlap criterion). Lesion-level precision was defined as the
# proportion of predicted clusters classified as true positives, and
# lesion-level recall as the proportion of ground truth clusters
# detected. Lesion-level F1 was computed as the harmonic mean of
# precision and recall.
#
# To assess the effect of training preprocessing on cluster-level
# detection, Kruskal-Wallis H-tests were performed for each threshold
# and metric. Where the omnibus test was significant (alpha = 0.05),
# post-hoc pairwise Wilcoxon signed-rank tests were conducted with
# Bonferroni correction (corrected alpha = 0.05/3 = 0.0167). Effect
# sizes were quantified using Cliff's Delta, with |delta| >= 0.28 as
# the threshold for a meaningful difference.
# =============================================================

# =============================================================
# PAPER TEXT: RESULTS
# =============================================================
#
# 3.X Lesion-level evaluation
#
# Cluster-level F1 scores (0.855-0.877) substantially exceeded
# voxel-level Dice coefficients (0.540-0.567), indicating that BIANCA
# detected the majority of WMH clusters but delineated their boundaries
# imprecisely (Supplemental Figure X). Lesion-level precision was
# consistently high across all thresholding approaches (0.881-0.928),
# confirming that the majority of predicted clusters corresponded to
# genuine WMH regions. Lesion-level recall ranged from 0.849 to 0.905,
# with B+L achieving the highest detection rate.
#
# Kruskal-Wallis tests revealed no significant differences between
# training preprocessing conditions on any cluster-level metric
# (all p > 0.92; Supplemental Table X), confirming that lesion removal
# does not affect lesion detection performance. This finding, combined
# with the high and stable cluster-level precision (0.881-0.928),
# addresses the concern that higher WMH volumes after lesion removal
# might reflect spatially distributed false positives rather than
# genuine sensitivity gains.
# =============================================================
"""

import matplotlib
matplotlib.use('Agg')

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os
import glob
from scipy.stats import wilcoxon, mannwhitneyu, kruskal

try:
    from cliffs_delta import cliffs_delta
    HAS_CLIFFS = True
except ImportError:
    HAS_CLIFFS = False
    print("WARNING: cliffs_delta not installed. Install with: pip install cliffs_delta")

# =============================================================
# CONFIG
# =============================================================
SEED_DIR = "Phase_1/5FCV_SET/bianca_result_dir"
OUTPUT_DIR = "Phase_1/Cluster_analysis_results"
PLOT_DIR = os.path.join(OUTPUT_DIR, "plots")

# Labels for plots (LaTeX subscript) vs Excel (plain text)
TH_PLOT = {'85': r'B$_{0.85}$', '90': r'B$_{0.90}$', 'locate': 'B+L'}
TH_PLAIN = {'85': 'B_0.85', '90': 'B_0.90', 'locate': 'B+L'}
COND_LABELS = {
    'non_removed': 'Train Non Removed',
    'removed': 'Train Removed',
    'filled': 'Train Inpainted',
}

# All metrics: voxel-level + cluster-level
VOXEL_METRICS = ['dice_score', 'sensitivity', 'precision']
CLUSTER_METRICS = ['lesion_f1', 'lesion_precision', 'lesion_recall']
ALL_METRICS = VOXEL_METRICS + CLUSTER_METRICS

# Bootstrap CI parameters (consistent with manuscript: Efron & Tibshirani, 1993)
N_BOOTSTRAP = 1000
CI_ALPHA = 0.05  # 95% CI
RANDOM_SEED = 42

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(PLOT_DIR, exist_ok=True)


# =============================================================
# HELPER: Excel export with metadata
# =============================================================
def save_excel(filepath, dataframe, title, description):
    """Save DataFrame to Excel with descriptive metadata header."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    n_cols = len(dataframe.columns)
    merge_width = max(n_cols, 6)

    thin = Border(*(Side(style='thin', color='B0B0B0'),) * 4)

    # Title
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_width)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Description
    for line in description:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_width)
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name='Arial', size=9, italic=True, color='333333')
        c.fill = PatternFill('solid', fgColor='F2F2F2')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    row += 2  # empty separator

    # Headers
    for ci, col in enumerate(dataframe.columns, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font = Font(name='Arial', bold=True, size=10)
        c.fill = PatternFill('solid', fgColor='D6E4F0')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin

    # Data
    for _, data_row in dataframe.iterrows():
        row += 1
        for ci, val in enumerate(data_row, 1):
            c = ws.cell(row=row, column=ci)
            if isinstance(val, (np.bool_, bool)):
                c.value = str(val)
            elif isinstance(val, float) and not np.isnan(val):
                c.value = val
                c.number_format = '0.0000'
            else:
                c.value = val
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(horizontal='center')
            c.border = thin

    # Auto-width
    for ci in range(1, n_cols + 1):
        max_len = max(len(str(dataframe.columns[ci-1])),
                      dataframe.iloc[:, ci-1].astype(str).str.len().max() if len(dataframe) > 0 else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 35)

    wb.save(filepath)


# =============================================================
# HELPER: Bootstrap 95% confidence intervals
# =============================================================
def bootstrap_ci(values, n_boot=N_BOOTSTRAP, alpha=CI_ALPHA, seed=RANDOM_SEED):
    """
    Compute bootstrap 95% CI for the mean.

    Same-size resampling with replacement, 1000 iterations,
    consistent with manuscript methodology (Efron & Tibshirani, 1993).

    Returns (mean, ci_lower, ci_upper)
    """
    rng = np.random.default_rng(seed)
    values = np.asarray(values)
    n = len(values)
    boot_means = np.array([
        rng.choice(values, size=n, replace=True).mean()
        for _ in range(n_boot)
    ])
    ci_lo = np.percentile(boot_means, 100 * alpha / 2)
    ci_hi = np.percentile(boot_means, 100 * (1 - alpha / 2))
    return round(values.mean(), 4), round(ci_lo, 4), round(ci_hi, 4)


def fmt_ci(mean, lo, hi):
    """Format as 'mean [lo, hi]' for tables."""
    return f"{mean:.3f} [{lo:.3f}, {hi:.3f}]"


# =============================================================
# 1. LOAD DATA
# =============================================================
def load_data(seed_dir=SEED_DIR):
    files = sorted(glob.glob(os.path.join(seed_dir, "bianca_metrics_seed_*.xlsx")))
    if not files:
        files = sorted(glob.glob("bianca_metrics_seed_*.xlsx"))
    if not files:
        raise FileNotFoundError("No bianca_metrics_seed_*.xlsx files found.")
    df = pd.concat([pd.read_excel(f) for f in files], ignore_index=True)
    df['threshold'] = df['threshold'].astype(str)
    df['dataset'] = df['subject'].apply(
        lambda x: 'BeLOVE' if x.startswith('belove') else 'Challenge')
    print(f"Loaded {len(files)} seeds: {len(df)} rows, "
          f"{df['subject'].nunique()} subjects")
    return df


# =============================================================
# 2. VOXEL vs CLUSTER DISSOCIATION (Main Text Figure + Table)
# =============================================================
def voxel_vs_cluster(df):
    """
    Core analysis addressing R1 Comment 2:
    Cluster F1 >> Dice demonstrates that BIANCA detects most lesions
    but delineates boundaries imprecisely. High cluster precision
    confirms that detected regions are genuine WMH, not distributed FP.

    Reports both voxel-level (Dice, sensitivity, precision) and
    cluster-level (F1, precision, recall) with bootstrap 95% CIs.
    """
    filled = df[(df['train_condition'] == 'filled') & (df['test_condition'] == 'filled')]
    per_sub = filled.groupby(['subject', 'threshold', 'dataset']).agg(
        {m: 'mean' for m in ALL_METRICS}).reset_index()

    # Summary table with bootstrap CIs
    rows = []
    print("\n" + "=" * 70)
    print("VOXEL vs CLUSTER PERFORMANCE (train=filled, test=filled)")
    print("Bootstrap 95% CI (1000 iterations)")
    print("=" * 70)

    for th in ['85', '90', 'locate']:
        sub = per_sub[per_sub['threshold'] == th]
        row = {'threshold': TH_PLAIN[th]}

        # Voxel-level metrics with CIs
        for m, label in [('dice_score', 'Dice'),
                         ('sensitivity', 'Vox Sens'),
                         ('precision', 'Vox Prec')]:
            mean, lo, hi = bootstrap_ci(sub[m].values)
            row[f'{m}_mean'] = mean
            row[f'{m}_sd'] = round(sub[m].std(), 4)
            row[f'{m}_ci_lower'] = lo
            row[f'{m}_ci_upper'] = hi

        # Cluster-level metrics with CIs
        for m, label in [('lesion_f1', 'Clust F1'),
                         ('lesion_precision', 'Clust P'),
                         ('lesion_recall', 'Clust R')]:
            mean, lo, hi = bootstrap_ci(sub[m].values)
            row[f'{m}_mean'] = mean
            row[f'{m}_sd'] = round(sub[m].std(), 4)
            row[f'{m}_ci_lower'] = lo
            row[f'{m}_ci_upper'] = hi

        row['f1_minus_dice'] = round(
            row['lesion_f1_mean'] - row['dice_score_mean'], 4)
        rows.append(row)

        print(f"\n  {TH_PLOT[th]}:")
        print(f"    Voxel:   Dice={fmt_ci(row['dice_score_mean'], row['dice_score_ci_lower'], row['dice_score_ci_upper'])}  "
              f"Sens={fmt_ci(row['sensitivity_mean'], row['sensitivity_ci_lower'], row['sensitivity_ci_upper'])}  "
              f"Prec={fmt_ci(row['precision_mean'], row['precision_ci_lower'], row['precision_ci_upper'])}")
        print(f"    Cluster: F1={fmt_ci(row['lesion_f1_mean'], row['lesion_f1_ci_lower'], row['lesion_f1_ci_upper'])}  "
              f"P={fmt_ci(row['lesion_precision_mean'], row['lesion_precision_ci_lower'], row['lesion_precision_ci_upper'])}  "
              f"R={fmt_ci(row['lesion_recall_mean'], row['lesion_recall_ci_lower'], row['lesion_recall_ci_upper'])}")
        print(f"    Gap (cF1 - Dice) = {row['f1_minus_dice']:+.3f}")

    # Scatter plot: Dice vs Cluster F1
    colors = {'85': '#1f77b4', '90': '#ff7f0e', 'locate': '#2ca02c'}
    fig, ax = plt.subplots(figsize=(8, 7))
    for th in ['85', '90', 'locate']:
        sub = per_sub[per_sub['threshold'] == th]
        ax.scatter(sub['dice_score'], sub['lesion_f1'], alpha=0.5, s=30,
                   color=colors[th], label=TH_PLOT[th])
    ax.plot([0, 1], [0, 1], 'k--', alpha=0.3, lw=1)
    ax.set_xlabel('Voxel-level Dice', fontsize=12)
    ax.set_ylabel('Cluster-level F1', fontsize=12)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.2)
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    fig.tight_layout()
    fig.savefig(os.path.join(PLOT_DIR, "dice_vs_cluster_f1_scatter.png"), dpi=300)
    plt.close()
    print("  Saved: dice_vs_cluster_f1_scatter.png")

    return pd.DataFrame(rows)


# =============================================================
# 3. PRECISION vs RECALL: Voxel + Cluster (Main Text Figure)
# =============================================================
def precision_recall_bars(df):
    """
    Addresses R5 #17: Precision values contextualized.
    Shows voxel-level (Dice, Sens, Prec) vs cluster-level (F1, P, R)
    side by side, demonstrating that moderate voxel precision coexists
    with high cluster precision.
    """
    filled = df[(df['train_condition'] == 'filled') & (df['test_condition'] == 'filled')]

    thresholds = ['85', '90', 'locate']
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))

    # Panel A: Voxel-level
    vox_metrics = [('dice_score', 'Dice'), ('sensitivity', 'Sensitivity'), ('precision', 'Precision')]
    x = np.arange(len(thresholds))
    w = 0.25
    vox_colors = ['#5B9BD5', '#70AD47', '#ED7D31']
    for i, (m, label) in enumerate(vox_metrics):
        means = [filled[filled['threshold'] == t][m].mean() for t in thresholds]
        sds = [filled[filled['threshold'] == t][m].std() for t in thresholds]
        bars = axes[0].bar(x + i * w - w, means, w, yerr=sds, label=label,
                           color=vox_colors[i], capsize=3, alpha=0.85)
        for bar in bars:
            axes[0].text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.02,
                         f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=8)

    axes[0].set_xticks(x)
    axes[0].set_xticklabels([TH_PLOT[t] for t in thresholds], fontsize=11)
    axes[0].set_ylabel('Score', fontsize=11)
    axes[0].set_title('Voxel-level', fontsize=12, fontweight='bold')
    axes[0].legend(fontsize=9)
    axes[0].set_ylim(0, 1.15)
    axes[0].grid(True, alpha=0.2, axis='y')

    # Panel B: Cluster-level
    clu_metrics = [('lesion_f1', 'F1'), ('lesion_precision', 'Precision'), ('lesion_recall', 'Recall')]
    clu_colors = ['#5B9BD5', '#70AD47', '#ED7D31']
    for i, (m, label) in enumerate(clu_metrics):
        means = [filled[filled['threshold'] == t][m].mean() for t in thresholds]
        sds = [filled[filled['threshold'] == t][m].std() for t in thresholds]
        bars = axes[1].bar(x + i * w - w, means, w, yerr=sds, label=label,
                           color=clu_colors[i], capsize=3, alpha=0.85)
        for bar in bars:
            axes[1].text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.02,
                         f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=8)

    axes[1].set_xticks(x)
    axes[1].set_xticklabels([TH_PLOT[t] for t in thresholds], fontsize=11)
    axes[1].set_title('Cluster-level', fontsize=12, fontweight='bold')
    axes[1].legend(fontsize=9)
    axes[1].set_ylim(0, 1.15)
    axes[1].grid(True, alpha=0.2, axis='y')

    fig.tight_layout()
    fig.savefig(os.path.join(PLOT_DIR, "voxel_vs_cluster_bars.png"), dpi=300)
    plt.close()
    print("  Saved: voxel_vs_cluster_bars.png")


# =============================================================
# 4. CONDITION COMPARISON: Kruskal-Wallis + Wilcoxon + Cliff's Delta
#    (Main Text Results + Supplemental Tables)
# =============================================================
def condition_comparison(df):
    """
    Core analysis: Does training preprocessing affect cluster-level detection?

    1. Kruskal-Wallis H-test (omnibus, 3 training conditions)
    2. Post-hoc Wilcoxon signed-rank (Bonferroni alpha = 0.05/3)
    3. Cliff's Delta (|delta| >= 0.28 for meaningful difference)

    Test condition fixed to 'filled' (inpainted).
    Per-subject means averaged across seeds/folds for independence.
    """
    COMPARISONS = [
        ('non_removed', 'removed',  'Train Non Removed vs Train Removed'),
        ('non_removed', 'filled',   'Train Non Removed vs Train Inpainted'),
        ('removed',     'filled',   'Train Removed vs Train Inpainted'),
    ]
    BONF_ALPHA = 0.05 / len(COMPARISONS)
    train_conds = ['non_removed', 'removed', 'filled']

    omnibus_rows = []
    posthoc_rows = []

    print("\n" + "=" * 70)
    print("TRAIN CONDITION COMPARISON (cluster-level, test=filled)")
    print(f"Omnibus: Kruskal-Wallis | Post-hoc: Wilcoxon (Bonferroni alpha={BONF_ALPHA:.4f})")
    print("=" * 70)

    for th in ['85', '90', 'locate']:
        # Per-subject averages per condition
        cond_data = {}
        for tc in train_conds:
            cond_data[tc] = df[
                (df['train_condition'] == tc) &
                (df['test_condition'] == 'filled') &
                (df['threshold'] == th)
            ].groupby("subject")[ALL_METRICS].mean()

        print(f"\n  --- {TH_PLOT[th]} ---")

        for metric in ALL_METRICS:
            groups = [cond_data[tc][metric].values for tc in train_conds]
            try:
                h_stat, kw_p = kruskal(*groups)
            except ValueError:
                h_stat, kw_p = 0.0, 1.0

            omnibus_rows.append({
                'threshold': TH_PLAIN[th],
                'metric': metric,
                'groups': 'Non Removed, Removed, Inpainted',
                'H_statistic': round(h_stat, 4),
                'p_kruskal_wallis': round(kw_p, 4),
                'significant_alpha_0.05': kw_p < 0.05,
                'n_per_group': len(groups[0]),
            })

            if metric == 'lesion_f1':
                sig = "SIGNIFICANT" if kw_p < 0.05 else "n.s."
                print(f"    Kruskal-Wallis (F1): H={h_stat:.4f}, p={kw_p:.4f} ({sig})")

            # Post-hoc pairwise
            for cond_a, cond_b, label in COMPARISONS:
                merged = cond_data[cond_a].merge(
                    cond_data[cond_b], on='subject', suffixes=('_a', '_b'))
                vals_a = merged[f'{metric}_a'].values
                vals_b = merged[f'{metric}_b'].values

                try:
                    _, p_w = wilcoxon(vals_a, vals_b)
                except ValueError:
                    p_w = 1.0

                d_val, d_size = (np.nan, 'N/A')
                if HAS_CLIFFS:
                    d_val, d_size = cliffs_delta(vals_b, vals_a)

                # Bootstrap CIs for each condition mean
                m_a, lo_a, hi_a = bootstrap_ci(vals_a)
                m_b, lo_b, hi_b = bootstrap_ci(vals_b)
                diff_vals = vals_b - vals_a
                m_d, lo_d, hi_d = bootstrap_ci(diff_vals)

                posthoc_rows.append({
                    'threshold': TH_PLAIN[th],
                    'comparison': label,
                    'metric': metric,
                    'mean_a': m_a,
                    'ci_lower_a': lo_a,
                    'ci_upper_a': hi_a,
                    'mean_b': m_b,
                    'ci_lower_b': lo_b,
                    'ci_upper_b': hi_b,
                    'mean_diff': m_d,
                    'diff_ci_lower': lo_d,
                    'diff_ci_upper': hi_d,
                    'kruskal_wallis_p': round(kw_p, 4),
                    'omnibus_significant': kw_p < 0.05,
                    'p_wilcoxon': round(p_w, 4),
                    'bonferroni_alpha': round(BONF_ALPHA, 4),
                    'significant_bonferroni': (kw_p < 0.05) and (p_w < BONF_ALPHA),
                    'cliffs_delta': round(d_val, 4) if not np.isnan(d_val) else np.nan,
                    'effect_size_category': d_size,
                    'n_subjects': len(merged),
                })

        # Print post-hoc F1 summary
        for cond_a, cond_b, label in COMPARISONS:
            merged = cond_data[cond_a].merge(
                cond_data[cond_b], on='subject', suffixes=('_a', '_b'))
            f1_diff = merged['lesion_f1_b'].mean() - merged['lesion_f1_a'].mean()
            try:
                _, pw = wilcoxon(merged['lesion_f1_a'].values, merged['lesion_f1_b'].values)
            except ValueError:
                pw = 1.0
            d_str = ""
            if HAS_CLIFFS:
                d, s = cliffs_delta(merged['lesion_f1_b'].values, merged['lesion_f1_a'].values)
                d_str = f"  delta={d:.4f} ({s})"
            print(f"    {label}: dF1={f1_diff:+.4f}  p={pw:.4f}{d_str}")

    # Boxplot (Main Text Figure)
    filled_test = df[df['test_condition'] == 'filled']
    per_sub = filled_test.groupby(
        ['subject', 'train_condition', 'threshold']).agg({'lesion_f1': 'mean'}).reset_index()
    per_sub['train_label'] = per_sub['train_condition'].map(COND_LABELS)

    fig, axes = plt.subplots(1, 3, figsize=(15, 5), sharey=True)
    order = list(COND_LABELS.values())
    for i, th in enumerate(['85', '90', 'locate']):
        sub = per_sub[per_sub['threshold'] == th]
        sns.boxplot(data=sub, x='train_label', y='lesion_f1', order=order,
                    hue='train_label', hue_order=order, legend=False,
                    ax=axes[i], palette='Set2', width=0.6)
        axes[i].set_title(TH_PLOT[th], fontsize=12, fontweight='bold')
        axes[i].set_xlabel('')
        axes[i].set_ylim(0, 1.05)
        axes[i].grid(True, alpha=0.2, axis='y')
        axes[i].tick_params(axis='x', rotation=15)
    axes[0].set_ylabel('Cluster-level F1', fontsize=11)
    fig.tight_layout()
    fig.savefig(os.path.join(PLOT_DIR, "cluster_f1_condition_boxplot.png"), dpi=300)
    plt.close()
    print("  Saved: cluster_f1_condition_boxplot.png")

    return pd.DataFrame(omnibus_rows), pd.DataFrame(posthoc_rows)


# =============================================================
# 5. PER-DATASET: BeLOVE vs Challenge (Supplemental)
# =============================================================
def per_dataset(df):
    """
    Supplemental analysis: in-distribution (BeLOVE) vs
    out-of-distribution (Challenge) performance.
    Mann-Whitney U for unpaired comparison.
    """
    filled = df[(df['train_condition'] == 'filled') & (df['test_condition'] == 'filled')]

    # Descriptive
    desc_rows = []
    for ds in ['BeLOVE', 'Challenge']:
        dsdf = filled[filled['dataset'] == ds]
        n = dsdf['subject'].nunique()
        for th in ['85', '90', 'locate']:
            sub = dsdf[dsdf['threshold'] == th]
            per_sub = sub.groupby('subject')[ALL_METRICS].mean()
            row = {'dataset': ds, 'n': n, 'threshold': TH_PLAIN[th]}
            for m in ALL_METRICS:
                mean, lo, hi = bootstrap_ci(per_sub[m].values)
                row[f'{m}_mean'] = mean
                row[f'{m}_sd'] = round(per_sub[m].std(), 4)
                row[f'{m}_ci_lower'] = lo
                row[f'{m}_ci_upper'] = hi
            row['n_pred_mean'] = round(sub['n_pred_clusters'].mean(), 1)
            row['n_gt_mean'] = round(sub['n_gt_clusters'].mean(), 1)
            desc_rows.append(row)

    # Statistical comparison
    stat_rows = []
    print("\n" + "=" * 70)
    print("PER-DATASET (Mann-Whitney U: BeLOVE vs Challenge)")
    print("=" * 70)
    for th in ['85', '90', 'locate']:
        bel = filled[(filled['dataset'] == 'BeLOVE') & (filled['threshold'] == th)].groupby("subject")[ALL_METRICS].mean()
        cha = filled[(filled['dataset'] == 'Challenge') & (filled['threshold'] == th)].groupby("subject")[ALL_METRICS].mean()
        for m in ALL_METRICS:
            _, p = mannwhitneyu(bel[m].values, cha[m].values, alternative='two-sided')
            d_val, d_size = (np.nan, 'N/A')
            if HAS_CLIFFS:
                d_val, d_size = cliffs_delta(bel[m].values, cha[m].values)
            m_bel, lo_bel, hi_bel = bootstrap_ci(bel[m].values)
            m_cha, lo_cha, hi_cha = bootstrap_ci(cha[m].values)
            stat_rows.append({
                'threshold': TH_PLAIN[th], 'metric': m,
                'mean_belove': m_bel,
                'ci_belove': f'[{lo_bel:.4f}, {hi_bel:.4f}]',
                'mean_challenge': m_cha,
                'ci_challenge': f'[{lo_cha:.4f}, {hi_cha:.4f}]',
                'diff': round(m_bel - m_cha, 4),
                'p_mann_whitney': round(p, 4),
                'cliffs_delta': round(d_val, 4) if not np.isnan(d_val) else np.nan,
                'effect_size': d_size,
            })
            if m == 'lesion_f1':
                print(f"  {TH_PLAIN[th]} F1: BeLOVE={bel[m].mean():.4f} vs "
                      f"Challenge={cha[m].mean():.4f}  p={p:.4f}")

    return pd.DataFrame(desc_rows), pd.DataFrame(stat_rows)


# =============================================================
# 6. GT CLUSTER DISTRIBUTION (Supplemental)
# =============================================================
def gt_distribution(df):
    """Descriptive statistics of GT cluster counts (26-connectivity)."""
    filled = df[(df['train_condition'] == 'filled') &
                (df['test_condition'] == 'filled') &
                (df['threshold'] == '85')]
    gt = filled.groupby('subject')['n_gt_clusters'].mean()

    rows = []
    for label, data in [('All', gt),
                        ('BeLOVE', filled[filled['dataset'] == 'BeLOVE'].groupby('subject')['n_gt_clusters'].mean()),
                        ('Challenge', filled[filled['dataset'] == 'Challenge'].groupby('subject')['n_gt_clusters'].mean())]:
        rows.append({
            'group': label, 'n': len(data),
            'mean': round(data.mean(), 1), 'sd': round(data.std(), 1),
            'median': round(data.median(), 1),
            'min': int(data.min()), 'max': int(data.max()),
            'q25': int(data.quantile(0.25)), 'q75': int(data.quantile(0.75)),
        })

    print("\n" + "=" * 70)
    print("GT CLUSTER DISTRIBUTION (26-connectivity)")
    print("=" * 70)
    for r in rows:
        print(f"  {r['group']:>10s}: mean={r['mean']}, median={r['median']}, "
              f"range={r['min']}-{r['max']}")

    return pd.DataFrame(rows)


# =============================================================
# 7. SUMMARY TABLE (Supplemental)
# =============================================================
def summary_table(df):
    """Mean, SD, median per train x test x threshold (all 27 combos)."""
    metric_cols = ['dice_score', 'sensitivity', 'precision',
                   'lesion_f1', 'lesion_precision', 'lesion_recall',
                   'n_pred_clusters', 'n_gt_clusters']
    group_cols = ['train_condition', 'test_condition', 'threshold']

    summary = df.groupby(group_cols)[metric_cols].agg(['mean', 'std', 'median']).round(4)
    # Flatten MultiIndex columns
    flat = summary.reset_index()
    flat.columns = ['_'.join(str(c) for c in col).rstrip('_') for col in flat.columns]
    return flat


# =============================================================
# MAIN
# =============================================================
def main():
    print("=" * 70)
    print("CLUSTER-LEVEL ANALYSIS OF BIANCA WMH SEGMENTATION")
    print("=" * 70)

    df = load_data()

    # --- MAIN TEXT ---
    dissoc_df = voxel_vs_cluster(df)
    precision_recall_bars(df)
    omnibus_df, posthoc_df = condition_comparison(df)

    # --- SUPPLEMENTAL ---
    dataset_desc_df, dataset_stat_df = per_dataset(df)
    gt_df = gt_distribution(df)
    summary_df = summary_table(df)

    # =============================================================
    # SAVE
    # =============================================================
    print(f"\nSaving to {OUTPUT_DIR}/...")

    # Main Text tables
    save_excel(
        os.path.join(OUTPUT_DIR, "voxel_vs_cluster_dissociation.xlsx"),
        dissoc_df,
        "Voxel-Level vs Cluster-Level Performance Dissociation",
        ["Addresses R1 Comment 2: Cluster F1 substantially exceeds Dice, indicating that BIANCA detects most",
         "WMH clusters but delineates boundaries imprecisely. High cluster precision confirms detected regions",
         "are genuine WMH, not distributed false positives.",
         "Condition: train=filled, test=filled. Per-subject means across 10 seeds x 5-fold CV (n=89).",
         "Cluster parameters: 26-connectivity, any-overlap (MICCAI 2017), prediction-only MCS filtering."])

    save_excel(
        os.path.join(OUTPUT_DIR, "kruskal_wallis_train_condition.xlsx"),
        omnibus_df,
        "Kruskal-Wallis H-Test: Effect of Training Preprocessing on Cluster-Level Metrics",
        ["Question: Does training with Non Removed vs Removed vs Inpainted images affect lesion detection?",
         "Design: 3 groups (train conditions), test condition fixed to 'filled' (inpainted).",
         "Test: Kruskal-Wallis H-test (non-parametric omnibus, k=3 independent groups).",
         "Null hypothesis: All three training conditions produce equal cluster-level metrics.",
         "Significance: alpha = 0.05.",
         "Data: Per-subject means averaged across 10 seeds x 5-fold stratified CV (n=89 per group).",
         "If significant: see posthoc_wilcoxon_bonferroni.xlsx for pairwise comparisons."])

    save_excel(
        os.path.join(OUTPUT_DIR, "posthoc_wilcoxon_bonferroni.xlsx"),
        posthoc_df,
        "Post-hoc Wilcoxon Signed-Rank Tests with Bonferroni Correction",
        ["Pairwise comparisons of training conditions (paired by subject).",
         "Test: Wilcoxon signed-rank (non-parametric, paired).",
         "Correction: Bonferroni (3 comparisons, corrected alpha = 0.05/3 = 0.0167).",
         "Effect size: Cliff's Delta (|d| >= 0.28 = meaningful, per manuscript conventions).",
         "Column 'omnibus_significant': Kruskal-Wallis was significant for this metric.",
         "Column 'significant_bonferroni': True only if BOTH omnibus AND post-hoc significant.",
         "Note: Post-hoc tests interpretable only when omnibus is significant."])

    # Supplemental tables
    save_excel(
        os.path.join(OUTPUT_DIR, "per_dataset_descriptive.xlsx"),
        dataset_desc_df,
        "Cluster-Level Metrics by Dataset: BeLOVE vs WMH Segmentation Challenge",
        ["BeLOVE (n=58): in-distribution (included in cross-validated training).",
         "Challenge (n=31): out-of-distribution (external validation dataset).",
         "Condition: train=filled, test=filled."])

    save_excel(
        os.path.join(OUTPUT_DIR, "per_dataset_mann_whitney.xlsx"),
        dataset_stat_df,
        "Mann-Whitney U Test: BeLOVE vs Challenge Comparison",
        ["Test: Mann-Whitney U (non-parametric, unpaired, 2 groups).",
         "Significance: alpha = 0.05.",
         "Effect size: Cliff's Delta.",
         "Per-subject means (BeLOVE n=58, Challenge n=31)."])

    save_excel(
        os.path.join(OUTPUT_DIR, "gt_cluster_distribution.xlsx"),
        gt_df,
        "Ground Truth Cluster Count Distribution (26-Connectivity)",
        ["Descriptive statistics of the number of GT clusters per subject.",
         "26-connectivity used (consistent with all cluster-level analyses).",
         "No size filtering applied to ground truth (expert-delineated lesions)."])

    save_excel(
        os.path.join(OUTPUT_DIR, "summary_all_conditions.xlsx"),
        summary_df,
        "Full Summary: Mean, SD, Median per Train x Test x Threshold",
        ["All 3 train x 3 test x 3 threshold = 27 combinations.",
         "Statistics: mean, standard deviation, median for each metric.",
         "Averaged across 10 seeds x 5-fold stratified CV x 89 subjects."])

    # Raw data
    df.to_excel(os.path.join(OUTPUT_DIR, "all_seeds_raw.xlsx"), index=False)

    print(f"\nDone. Results: {OUTPUT_DIR}/  Plots: {PLOT_DIR}/")


if __name__ == "__main__":
    main()