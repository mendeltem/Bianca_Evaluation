#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase II-A: Correlation Analysis -- Stroke Lesion Volume vs WMH Volume Difference
=================================================================================
Trimmed AND untrimmed Spearman correlations with Bootstrap 95% CIs.

Revision context
----------------
  R5 Comment 15: Full untrimmed results must be provided alongside trimmed
                 for transparency.
  R5 Comment 12: Small subgroups (n<10) flagged; statistics reported with
                 caution note.
  R5 Comment 5:  Bonferroni correction per comparison family.
  R1 Comment 1 / R5 Comment 9: Zero-filling vs inpainting -- convergence
                 comparison isolates removal effect from filling strategy.

Paper changes
-------------
  Section 2.8 (Statistical Analysis): "To analyze the relationship
    between stroke lesion volume and WMH segmentation differences,
    Spearman rank correlations were computed between stroke lesion
    volume and signed WMH volume differences, stratified by lesion
    type (Spearman, 1904). Trimmed correlations excluding cases above
    the 90th percentile of stroke lesion volume are reported in the
    main text to minimize outlier influence (Wilcox, 2012; Pernet et
    al., 2012); full-sample values are provided in the Supplemental
    for transparency (Supplemental Table X). Bootstrap 95% confidence
    intervals (1000 iterations) were calculated on both samples
    (Efron & Tibshirani, 1993). Bonferroni correction was applied per
    pairwise comparison family (k=4 lesion types, corrected
    alpha=0.0125). Subgroups with n<10 are flagged and reported
    descriptively."

  Section 3.2.3 (Size Effects and Segmentation Performance): "Stroke
    lesion volume correlated with WMH volume differences for ischemic
    infarcts (trimmed Spearman correlation 0.74, p<0.001, 95% CI
    [0.39, 0.92], n=26; full: 0.78, p<0.001, 95% CI [0.54, 0.90],
    n=29), with differences ranging from near 0 mL in small lesions
    to 2.5 mL in large ones. Dice-based coloring showed removal
    improved segmentation accuracy in 48% of ischemic infarcts. Near-
    identical patterns were observed for non-removed vs inpainted
    (trimmed rho=0.81, p<0.001). Lacunes (trimmed rho=0.22, p=0.71),
    mixed lesions (trimmed rho=0.68, p=0.17, n=9), and infratentorial
    strokes (trimmed rho=0.00, p=1.00, n=4) showed non-significant
    correlations after Bonferroni correction; the latter two subgroups
    had insufficient sample sizes (n<10) and should be interpreted
    with caution. Convergence analysis (removed vs inpainted) showed
    no significant correlations for any lesion type (all p_bonf=1.0;
    infarcts trimmed rho=0.12), confirming that observed effects are
    attributable to lesion removal per se, independent of filling
    strategy."

  Figure X caption: "Scatter plots showing correlation between stroke
    lesion volume and WMH volume difference, stratified by lesion
    type (Phase II-A, n=89). Color coding: green=Dice improved after
    preprocessing, red=Dice worsened. Trimmed Spearman correlations
    (excluding top 10% of lesion volumes) were significant only for
    ischemic infarcts (rho=0.74, p<0.001). Full statistics in
    Supplemental Table X."

  Supplemental Table X: Trimmed + full Spearman correlations with
    bootstrap 95% CIs for all three pairwise comparisons, all lesion
    types.

Response to Reviewers
---------------------
  R5 Comment 15: "Thank you for raising this important transparency
    concern. We now report both trimmed (<=90th percentile of stroke
    lesion volume) and full-sample Spearman correlations for all
    lesion types across all three pairwise comparisons (Supplemental
    Table X). Trimmed correlations are reported in the main text to
    minimize outlier influence (Wilcox, 2012; Pernet et al., 2012),
    while full-sample values are provided alongside for complete
    transparency. The results are consistent: ischemic infarcts show
    significant size-dependent scaling in both analyses (trimmed
    rho=0.74, full rho=0.78, both p<0.001 after Bonferroni
    correction), while smaller subgroups (lacunes, mixed,
    infratentorial) show non-significant correlations in both
    (Section 3.2.3, Figure X, Supplemental Table X)."

  R5 Comment 12: "Thank you for this valid methodological concern.
    Subgroups with n<10 (infratentorial n=5, ICH n=1) are now
    explicitly flagged in tables and figures with a caution note.
    Mixed lesions (n=10) are reported with inferential statistics
    but noted as borderline. ICH (n=1) is excluded from all
    Bonferroni families and reported descriptively only. Non-
    significant correlations for small subgroups (mixed: trimmed
    rho=0.68, p=0.17; infratentorial: trimmed rho=0.00, p=1.00)
    are reported for completeness but not used to draw conclusions
    (Section 3.2.3)."

  R5 Comment 5: "Thank you for this important concern regarding
    multiple comparisons. Bonferroni correction is now applied per
    comparison family, with k=4 lesion types per pairwise comparison
    (corrected alpha=0.0125). The family structure is explicitly
    documented in Section 2.8 and in the Supplemental tables. ICH
    (n=1) is excluded from families. After correction, only ischemic
    infarcts remain significant (p<0.001 for both non-removed vs
    removed and non-removed vs inpainted comparisons)."

  R1 Comment 1 / R5 Comment 9 (convergence): "The convergence
    comparison (removed vs inpainted) provides direct evidence that
    the filling strategy has no measurable effect on the size-volume
    relationship. No significant correlations were observed for any
    lesion type (all p_bonf=1.0), and the Y-axis range collapsed
    from +/-2.5 mL to +/-0.4 mL. This isolates lesion removal as
    the active preprocessing step, independent of whether lesion
    voxels are zero-filled or inpainted with NAWM intensities
    (Section 3.2.3, Figure X)."


Key results
-----------
Non Removed vs Removed:
  Ischemic infarcts: trimmed rho=0.74, p<0.001 (Bonf), 95% CI
    [0.39, 0.92], n=26. Full: rho=0.78, p<0.001, CI [0.54, 0.90],
    n=29. Removal improved Dice in 48% of cases. Differences ranged
    from near 0 mL (small lesions) to 2.5 mL (large).
  Lacunes: trimmed rho=0.22, p=0.71 (ns), n=39. Full: 0.33, p=0.12.
  Mixed: trimmed rho=0.68, p=0.17 (ns), n=9. Full: 0.77, p=0.04
    (significant before Bonferroni, ns after). n=10, borderline.
  Infratentorial: trimmed rho=0.00, p=1.00 (ns), n=4. n<10.
  ICH: n=1, not computable.

Non Removed vs Inpainted:
  Near-identical pattern: infarcts trimmed rho=0.81, p<0.001.
  Confirms that inpainting and zero-filling produce equivalent
  correlations relative to non-removed baseline.

Removed vs Inpainted (Convergence):
  No significant correlations for any lesion type (all p_bonf=1.0).
  Infarcts trimmed rho=0.12, p=1.0. Y-axis range collapsed ~6-fold
  compared to NR comparisons. Filling strategy is irrelevant.


Design
------
  Three preprocessing conditions: non_removed, removed, filled (inpainted).
  Train = filled (fixed), threshold = 0.85.
  Per-subject means from 10 seeds x 5-fold stratified CV.

  Pairwise comparisons (3):
    1. Non Removed vs Removed
    2. Non Removed vs Inpainted
    3. Removed vs Inpainted  (convergence)

  X-axis: stroke lesion volume (mL)
  Y-axis: WMH volume difference (mL) between conditions (signed)

  Trimming: exclude subjects above 90th percentile of stroke lesion
  volume within each lesion type (Wilcox, 2012).

  Bootstrap 95% CIs (1000 iterations) on both trimmed and full samples
  (Efron & Tibshirani, 1993).

  Bonferroni correction: one family per pairwise comparison,
  k = n_lesion_types (4), corrected alpha = 0.05/4 = 0.0125.
  ICH (n=1) excluded from families, reported descriptively.

Statistical note
----------------
  Spearman rank correlation is reported throughout. No trend lines are
  plotted, as Spearman assumes monotonic but not necessarily linear
  relationships. The rho values in annotation boxes communicate
  correlation strength; scatter point coloring by Dice criterion
  (green = Dice improved, red = worsened) links volume differences
  to segmentation accuracy.

References
----------
  Efron, B., Tibshirani, R.J., 1993. An Introduction to the Bootstrap.
    Chapman and Hall, New York.
  Pernet, C.R., Wilcox, R., Rousselet, G.A., 2012. Robust Correlation
    Analyses. Front. Psychol. 3, 606.
  Wilcox, R.R., 2012. Introduction to Robust Estimation and Hypothesis
    Testing, 3rd ed. Academic Press, Amsterdam.

Outputs
-------
  correlation_results_{cond1}_vs_{cond2}.xlsx
    -> Sheet 'Publication': Trimmed + untrimmed side-by-side
    -> Sheet 'Detailed': All numerical values
  correlation_plot_{cond1}_vs_{cond2}.png
"""

import os
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
from scipy.stats import spearmanr
import matplotlib.pyplot as plt
import matplotlib.lines as mlines

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIG
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_DICE_ONLY.xlsx")
PLOT_DIR = os.path.join(SCRIPT_DIR, "plots", "6_CorrelationAnalysis")

TRIM_PERCENTILE = 90
N_BOOTSTRAP = 1000
BOOT_SEED = 42
MIN_N_RELIABLE = 10  # below this, flag as unreliable (R5 #12)

LESION_ORDER = ["infra", "lacune", "infarct", "mixed"]
LESION_ORDER_ICH = ["ICH"]
LESION_DISPLAY = {
    "infra":   "Infratentorial Strokes",
    "lacune":  "Lacunes",
    "infarct": "Ischemic Infarcts",
    "mixed":   "Mixed (Infarcts + Lacunes)",
    "ICH":     "Intracranial Hemorrhage",
}

COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
    "filled":      "Inpainted",
}

COMPARISONS = [
    ("non_removed", "removed"),
    ("non_removed", "filled"),
    ("removed",     "filled"),   # convergence
]

# Note: Scatter plots use Dice-criterion coloring (green/red/gray),
# not per-lesion-type colors.


# =============================================================================
# STATISTICS
# =============================================================================

def spearman_with_bootstrap_ci(x, y, n_boot=N_BOOTSTRAP, seed=BOOT_SEED):
    """
    Spearman correlation with bootstrap 95% CI.
    Returns dict with rho, p, ci_lower, ci_upper, n.
    """
    x, y = np.asarray(x, dtype=float), np.asarray(y, dtype=float)
    valid = ~(np.isnan(x) | np.isnan(y))
    x, y = x[valid], y[valid]
    n = len(x)

    if n < 4:
        return {"n": n, "rho": np.nan, "p": np.nan,
                "ci_lower": np.nan, "ci_upper": np.nan}

    rho, p = spearmanr(x, y)

    rng = np.random.default_rng(seed)
    boot_rhos = np.empty(n_boot)
    for i in range(n_boot):
        idx = rng.choice(n, size=n, replace=True)
        r, _ = spearmanr(x[idx], y[idx])
        boot_rhos[i] = r

    ci_lo = np.nanpercentile(boot_rhos, 2.5)
    ci_hi = np.nanpercentile(boot_rhos, 97.5)

    return {"n": n, "rho": rho, "p": p,
            "ci_lower": ci_lo, "ci_upper": ci_hi}


def compute_trimmed_and_full(x_all, y_all, trim_pct=TRIM_PERCENTILE):
    """
    Compute both full-sample and trimmed Spearman correlations.
    Trimming: exclude subjects above trim_pct percentile of x.
    Returns (full_stats, trimmed_stats, trim_info).
    """
    x_all, y_all = np.asarray(x_all, dtype=float), np.asarray(y_all, dtype=float)
    valid = ~(np.isnan(x_all) | np.isnan(y_all))
    x_clean, y_clean = x_all[valid], y_all[valid]

    # Full sample
    full_stats = spearman_with_bootstrap_ci(x_clean, y_clean)

    # Trimmed
    if len(x_clean) < 4:
        trimmed_stats = {"n": 0, "rho": np.nan, "p": np.nan,
                         "ci_lower": np.nan, "ci_upper": np.nan}
        trim_info = {"threshold_ml": np.nan, "n_excluded": 0}
    else:
        threshold = np.percentile(x_clean, trim_pct)
        mask = x_clean <= threshold
        x_trim, y_trim = x_clean[mask], y_clean[mask]
        trimmed_stats = spearman_with_bootstrap_ci(x_trim, y_trim)
        trim_info = {
            "threshold_ml": round(threshold, 2),
            "n_excluded": int((~mask).sum()),
        }

    return full_stats, trimmed_stats, trim_info


# =============================================================================
# ANALYSIS
# =============================================================================

def run_correlation_analysis(df, comparison, lesion_order):
    """
    For one pairwise comparison, compute correlations between stroke
    lesion volume and WMH volume difference, stratified by lesion type.

    Returns list of result dicts.
    """
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]

    vol_col1 = f"WMH_{cond1}_volume_ml"
    vol_col2 = f"WMH_{cond2}_volume_ml"
    lesion_vol_col = "infarct_volume_ml"

    if vol_col1 not in df.columns or vol_col2 not in df.columns:
        print(f"  ERROR: Missing columns {vol_col1} or {vol_col2}")
        return []

    results = []
    for lesion in lesion_order:
        ld = df[df["lesion_type"] == lesion].copy()
        n_subj = len(ld)
        if n_subj == 0:
            continue

        # Signed volume difference: cond2 - cond1
        # For NR vs R: positive = R yields higher WMH volume
        x = ld[lesion_vol_col].values
        y = (ld[vol_col2] - ld[vol_col1]).values

        small_n = n_subj < MIN_N_RELIABLE
        if small_n:
            print(f"    WARNING: {lesion} n={n_subj} < {MIN_N_RELIABLE}, "
                  f"interpret with caution (R5 #12)")

        full_stats, trim_stats, trim_info = compute_trimmed_and_full(x, y)

        results.append({
            "lesion_type": lesion,
            "comparison": f"{l1} vs {l2}",
            "n_full": full_stats["n"],
            "rho_full": full_stats["rho"],
            "p_full": full_stats["p"],
            "ci_lower_full": full_stats["ci_lower"],
            "ci_upper_full": full_stats["ci_upper"],
            "n_trimmed": trim_stats["n"],
            "rho_trimmed": trim_stats["rho"],
            "p_trimmed": trim_stats["p"],
            "ci_lower_trimmed": trim_stats["ci_lower"],
            "ci_upper_trimmed": trim_stats["ci_upper"],
            "trim_threshold_ml": trim_info["threshold_ml"],
            "n_excluded": trim_info["n_excluded"],
            "small_n": small_n,
        })

    return results


def apply_bonferroni(results):
    """
    Apply Bonferroni correction within one comparison family.
    k = number of lesion types in the family (excluding ICH).
    Corrects both full and trimmed p-values.
    """
    # Only correct non-ICH, non-small-n entries
    testable = [r for r in results if r["lesion_type"] != "ICH"]
    k = len(testable)
    if k == 0:
        return results

    alpha_adj = 0.05 / k

    for r in results:
        if r["lesion_type"] == "ICH":
            r["p_full_bonf"] = np.nan
            r["p_trimmed_bonf"] = np.nan
            r["k_bonf"] = k
            r["alpha_adj"] = alpha_adj
            r["sig_full"] = "N/A (n=1)"
            r["sig_trimmed"] = "N/A (n=1)"
        else:
            p_f = r["p_full"]
            p_t = r["p_trimmed"]
            r["p_full_bonf"] = min(p_f * k, 1.0) if not np.isnan(p_f) else np.nan
            r["p_trimmed_bonf"] = min(p_t * k, 1.0) if not np.isnan(p_t) else np.nan
            r["k_bonf"] = k
            r["alpha_adj"] = alpha_adj

            # Significance labels
            for suffix in ["full", "trimmed"]:
                p_bonf = r[f"p_{suffix}_bonf"]
                if np.isnan(p_bonf):
                    r[f"sig_{suffix}"] = "-"
                elif p_bonf < 0.001:
                    r[f"sig_{suffix}"] = "***"
                elif p_bonf < 0.01:
                    r[f"sig_{suffix}"] = "**"
                elif p_bonf < 0.05:
                    r[f"sig_{suffix}"] = "*"
                else:
                    r[f"sig_{suffix}"] = "ns"

    return results


# =============================================================================
# TABLE GENERATION
# =============================================================================

def fmt_p(p):
    """Format p-value for publication."""
    if np.isnan(p):
        return "-"
    if p < 0.001:
        return "<0.001"
    if p < 0.01:
        return f"{p:.3f}"
    return f"{p:.2f}"


def fmt_rho(rho):
    """Format correlation coefficient."""
    if np.isnan(rho):
        return "-"
    return f"{rho:.2f}"


def fmt_ci(lo, hi):
    """Format 95% CI."""
    if np.isnan(lo) or np.isnan(hi):
        return "-"
    return f"[{lo:.2f}, {hi:.2f}]"


def fmt_rho_ci(rho, ci_lo, ci_hi):
    """Compact ρ [CI] notation for publication tables."""
    if np.isnan(rho):
        return "-"
    ci = fmt_ci(ci_lo, ci_hi)
    return f"{rho:.2f} {ci}"


def generate_publication_table(all_comparison_results):
    """
    Single combined Supplemental table across all comparisons.
    Compact format: Comparison | Lesion Type | n_trim | ρ [CI] trim |
    p_bonf trim | n_full | ρ [CI] full | p_bonf full | Sig

    Addresses R5 #15 (transparency: trimmed + full side-by-side)
    and keeps the table to one page.
    """
    rows = []
    for comp_label, results in all_comparison_results:
        for r in results:
            display = LESION_DISPLAY.get(r["lesion_type"], r["lesion_type"])
            caution = " †" if r["small_n"] else ""

            row = {
                "Comparison": comp_label,
                "Lesion Type": display + caution,
                "n (trim)": r["n_trimmed"],
                "ρ [95% CI] (trim)": fmt_rho_ci(
                    r["rho_trimmed"], r["ci_lower_trimmed"], r["ci_upper_trimmed"]),
                "p_Bonf (trim)": fmt_p(r.get("p_trimmed_bonf", r["p_trimmed"])),
                "n (full)": r["n_full"],
                "ρ [95% CI] (full)": fmt_rho_ci(
                    r["rho_full"], r["ci_lower_full"], r["ci_upper_full"]),
                "p_Bonf (full)": fmt_p(r.get("p_full_bonf", r["p_full"])),
                "Sig": r.get("sig_trimmed", ""),
            }
            rows.append(row)

    return pd.DataFrame(rows)


def generate_detailed_table(results):
    """Detailed table with all raw numerical values."""
    rows = []
    for r in results:
        row = {
            "Lesion Type": r["lesion_type"],
            "Comparison": r["comparison"],
            "n_full": r["n_full"],
            "rho_full": round(r["rho_full"], 4) if not np.isnan(r["rho_full"]) else None,
            "p_full": r["p_full"],
            "p_full_bonf": r.get("p_full_bonf"),
            "CI_lower_full": round(r["ci_lower_full"], 4) if not np.isnan(r["ci_lower_full"]) else None,
            "CI_upper_full": round(r["ci_upper_full"], 4) if not np.isnan(r["ci_upper_full"]) else None,
            "n_trimmed": r["n_trimmed"],
            "rho_trimmed": round(r["rho_trimmed"], 4) if not np.isnan(r["rho_trimmed"]) else None,
            "p_trimmed": r["p_trimmed"],
            "p_trimmed_bonf": r.get("p_trimmed_bonf"),
            "CI_lower_trimmed": round(r["ci_lower_trimmed"], 4) if not np.isnan(r["ci_lower_trimmed"]) else None,
            "CI_upper_trimmed": round(r["ci_upper_trimmed"], 4) if not np.isnan(r["ci_upper_trimmed"]) else None,
            "trim_threshold_ml": r["trim_threshold_ml"],
            "n_excluded": r["n_excluded"],
            "k_bonf": r.get("k_bonf"),
            "alpha_adj": r.get("alpha_adj"),
            "small_n_flag": r["small_n"],
        }
        rows.append(row)
    return pd.DataFrame(rows)


# =============================================================================
# PLOTTING
# =============================================================================

def plot_correlation(df, comparison, results, save_path=None):
    """
    Scatter plots stratified by lesion type.

    Color coding by Dice criterion:
      Green = Dice improved (cond2 > cond1) → removal/inpainting beneficial
      Red   = Dice worsened (cond2 < cond1) → removal/inpainting detrimental
      Gray  = No change

    No trend line  Spearman ρ in annotation box communicates correlation
    strength without implying a parametric model.
    """
    cond1, cond2 = comparison
    l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
    vol_col1 = f"WMH_{cond1}_volume_ml"
    vol_col2 = f"WMH_{cond2}_volume_ml"
    dice_col1 = f"WMH_{cond1}_dice"
    dice_col2 = f"WMH_{cond2}_dice"
    lesion_vol_col = "infarct_volume_ml"

    # Dice criterion colors
    COLOR_BETTER = "#2ecc71"   # green: cond2 Dice > cond1 Dice
    COLOR_WORSE  = "#e74c3c"   # red:   cond2 Dice < cond1 Dice
    COLOR_EQUAL  = "#95a5a6"   # gray:  no change

    available = [lt for lt in LESION_ORDER if lt in df["lesion_type"].unique()]
    n_plots = len(available)
    if n_plots == 0:
        return None

    n_cols = 2
    n_rows = int(np.ceil(n_plots / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(14, 6 * n_rows))
    axes = np.atleast_1d(axes).flatten()

    # Compute Dice criterion per subject
    has_dice = dice_col1 in df.columns and dice_col2 in df.columns
    if has_dice:
        dice_diff = df[dice_col2] - df[dice_col1]
        df = df.copy()
        df["_dice_criterion"] = np.where(
            dice_diff > 1e-6, 1,
            np.where(dice_diff < -1e-6, -1, 0)
        )
    else:
        print(f"  WARNING: Dice columns not found ({dice_col1}, {dice_col2}), "
              f"using uniform color")

    res_lookup = {r["lesion_type"]: r for r in results}

    for idx, lesion in enumerate(available):
        ax = axes[idx]
        ld = df[df["lesion_type"] == lesion].copy()

        x = ld[lesion_vol_col].values
        y = (ld[vol_col2] - ld[vol_col1]).values

        valid = ~(np.isnan(x) | np.isnan(y))

        if has_dice:
            criterion = ld["_dice_criterion"].values

            for dice_val, color in [(1, COLOR_BETTER), (-1, COLOR_WORSE),
                                     (0, COLOR_EQUAL)]:
                mask = valid & (criterion == dice_val)
                if mask.sum() > 0:
                    ax.scatter(x[mask], y[mask], c=color, alpha=0.7, s=50,
                               edgecolors="white", linewidth=0.5, zorder=2)
        else:
            ax.scatter(x[valid], y[valid], c="#7f8c8d", alpha=0.7, s=50,
                       edgecolors="white", linewidth=0.5, zorder=2)

        # Zero reference line
        ax.axhline(y=0, color="gray", linestyle=":", linewidth=1, alpha=0.5)

        # Annotation box: trimmed + full
        r = res_lookup.get(lesion)
        if r:
            caution_note = "\n(n<10, interpret with caution) †" if r["small_n"] else ""

            # Count Dice categories for this subplot
            if has_dice:
                n_better = int((ld["_dice_criterion"] == 1).sum())
                n_worse = int((ld["_dice_criterion"] == -1).sum())
                n_equal = int((ld["_dice_criterion"] == 0).sum())
                n_total = n_better + n_worse + n_equal
                pct_better = round(100 * n_better / n_total) if n_total > 0 else 0
                dice_line = f"\n{l2} better: {n_better}/{n_total} ({pct_better}%)"
            else:
                dice_line = ""

            lines = [
                f"Trimmed: ρ = {fmt_rho(r['rho_trimmed'])}, "
                f"p = {fmt_p(r.get('p_trimmed_bonf', r['p_trimmed']))}, "
                f"n = {r['n_trimmed']}",
                f"95% CI: {fmt_ci(r['ci_lower_trimmed'], r['ci_upper_trimmed'])}",
                f"Full: ρ = {fmt_rho(r['rho_full'])}, "
                f"p = {fmt_p(r.get('p_full_bonf', r['p_full']))}, "
                f"n = {r['n_full']}",
                f"95% CI: {fmt_ci(r['ci_lower_full'], r['ci_upper_full'])}",
            ]
            stats_text = "\n".join(lines) + dice_line + caution_note

            ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, fontsize=8,
                    verticalalignment="top", horizontalalignment="left",
                    bbox=dict(boxstyle="round,pad=0.4", facecolor="white",
                              alpha=0.9, edgecolor="gray"))

        ax.set_xlabel("Stroke Lesion Volume (mL)", fontsize=11)
        ax.set_ylabel(f"WMH Volume Difference (mL)\n({l2} − {l1})", fontsize=11)
        ax.set_title(LESION_DISPLAY.get(lesion, lesion), fontsize=12, fontweight="bold")
        ax.grid(True, alpha=0.3)

    # Hide unused axes
    for idx in range(n_plots, len(axes)):
        axes[idx].set_visible(False)

    # Legend: Dice criterion + zero line
    legend_handles = [
        mlines.Line2D([], [], color=COLOR_BETTER, marker="o", linestyle="None",
                       markersize=8, label=f"{l2} Dice Better"),
        mlines.Line2D([], [], color=COLOR_WORSE, marker="o", linestyle="None",
                       markersize=8, label=f"{l2} Dice Worse"),
        mlines.Line2D([], [], color=COLOR_EQUAL, marker="o", linestyle="None",
                       markersize=8, label="No Change"),
        mlines.Line2D([], [], color="gray", linestyle=":", linewidth=1,
                       label="Zero difference"),
    ]

    fig.legend(handles=legend_handles, loc="lower center", ncol=4,
               bbox_to_anchor=(0.5, 0.02), frameon=True, fontsize=10)

    is_convergence = (cond1 == "removed" and cond2 == "filled")
    tag = " (Convergence)" if is_convergence else ""
    plt.suptitle(
        f"Correlation: Stroke Lesion Volume vs WMH Volume Difference\n"
        f"{l1} vs {l2}{tag}",
        fontsize=14, fontweight="bold", y=1.0)
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.12, top=0.92)

    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches="tight", facecolor="white")
        print(f"  Plot saved: {save_path}")

    plt.close(fig)
    return fig


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def save_results_excel(pub_table, detail_table, path, metadata_lines):
    """Save publication + detailed tables with metadata header."""
    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Publication table
        pub_table.to_excel(writer, sheet_name="Publication",
                           index=False, startrow=len(metadata_lines) + 1)
        ws = writer.sheets["Publication"]

        # Metadata header
        for i, line in enumerate(metadata_lines):
            ws.cell(row=i + 1, column=1, value=line).font = Font(
                italic=True, color="666666", size=9)
            ws.merge_cells(start_row=i + 1, start_column=1,
                           end_row=i + 1, end_column=len(pub_table.columns))

        # Header formatting
        header_row = len(metadata_lines) + 1
        for col_idx in range(1, len(pub_table.columns) + 1):
            cell = ws.cell(row=header_row + 1, column=col_idx)
            if cell.value:
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Highlight significant rows (Sig column = last column)
        green = PatternFill("solid", fgColor="C6EFCE")
        sig_col_idx = len(pub_table.columns)  # "Sig" is last column
        data_start = header_row + 2
        for row_idx in range(data_start, data_start + len(pub_table)):
            sig_val = ws.cell(row=row_idx, column=sig_col_idx).value
            if sig_val and sig_val.strip() in ("*", "**", "***"):
                for col_idx in range(1, len(pub_table.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green

        # Column widths (skip merged cells)
        from openpyxl.cell.cell import MergedCell
        for col in ws.columns:
            real_cells = [c for c in col if not isinstance(c, MergedCell)]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or "")) for c in real_cells), default=10)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 3, 25)

        # Detailed table
        detail_table.to_excel(writer, sheet_name="Detailed", index=False)

    print(f"  Excel saved: {path}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("  Correlation Analysis: Stroke Lesion Volume vs WMH Volume Difference")
    print("=" * 70)

    # Load data
    print(f"\nLoading: {XLSX_PATH}")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    if "subject_with_mask" in df.columns:
        df = df[df["subject_with_mask"] == 1].copy()
    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")

    print(f"Subjects: {len(df)}")
    print(f"Lesion types:\n{df['lesion_type'].value_counts()}")

    os.makedirs(PLOT_DIR, exist_ok=True)

    # Collect results across all comparisons for combined table
    all_comparison_results = []   # list of (label, results)
    all_detail_rows = []          # for Detailed sheet

    for cond1, cond2 in COMPARISONS:
        l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
        is_convergence = (cond1 == "removed" and cond2 == "filled")
        tag = "[Convergence] " if is_convergence else ""
        comp_label = f"{l1} vs {l2}" + (" (Convergence)" if is_convergence else "")

        print(f"\n{'='*60}")
        print(f"  {tag}{l1} vs {l2}")
        print(f"{'='*60}")

        # --- Main analysis (excl. ICH) ---
        results_main = run_correlation_analysis(df, (cond1, cond2), LESION_ORDER)
        results_main = apply_bonferroni(results_main)

        # --- ICH (separate, no Bonferroni) ---
        results_ich = run_correlation_analysis(df, (cond1, cond2), LESION_ORDER_ICH)
        for r in results_ich:
            r["p_full_bonf"] = np.nan
            r["p_trimmed_bonf"] = np.nan
            r["k_bonf"] = len(results_main)
            r["alpha_adj"] = 0.05 / max(len(results_main), 1)
            r["sig_full"] = "N/A (n=1)"
            r["sig_trimmed"] = "N/A (n=1)"

        all_results = results_main + results_ich
        all_comparison_results.append((comp_label, all_results))

        # --- Print summary ---
        print(f"\n  {'Lesion':<15} {'n_trim':>6} {'ρ_trim':>7} {'p_bonf':>10} "
              f"{'n_full':>6} {'ρ_full':>7} {'p_bonf':>10}")
        print(f"  {'-'*65}")
        for r in all_results:
            print(f"  {r['lesion_type']:<15} "
                  f"{r['n_trimmed']:>6} {fmt_rho(r['rho_trimmed']):>7} "
                  f"{fmt_p(r.get('p_trimmed_bonf', np.nan)):>10} "
                  f"{r['n_full']:>6} {fmt_rho(r['rho_full']):>7} "
                  f"{fmt_p(r.get('p_full_bonf', np.nan)):>10}")

        # Collect detailed rows
        detail_table = generate_detailed_table(all_results)
        all_detail_rows.append(detail_table)

        # --- Plot (still per comparison) ---
        plot_path = os.path.join(PLOT_DIR,
                                 f"correlation_plot_{cond1}_vs_{cond2}.png")
        plot_correlation(df, (cond1, cond2), all_results, save_path=plot_path)

    # --- Combined Supplemental Table (one Excel file) ---
    pub_table = generate_publication_table(all_comparison_results)
    combined_detail = pd.concat(all_detail_rows, ignore_index=True)

    k_tests = len(LESION_ORDER)  # 4 lesion types per family
    metadata = [
        "Supplemental Table X: Spearman Correlations  Stroke Lesion "
        "Volume vs WMH Volume Difference",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"N subjects (with stroke masks): {len(df)}",
        f"Bonferroni family: k={k_tests} lesion types per comparison, "
        f"corrected α = {0.05/k_tests:.4f}",
        f"Trimming: ≤{TRIM_PERCENTILE}th percentile of stroke lesion "
        f"volume (Wilcox, 2012).",
        f"Bootstrap 95% CIs: {N_BOOTSTRAP} iterations "
        f"(Efron & Tibshirani, 1993).",
        f"† = n<{MIN_N_RELIABLE}, interpret with caution.",
        "ICH (n=1) excluded from Bonferroni families, reported descriptively.",
    ]

    xlsx_path = os.path.join(PLOT_DIR, "correlation_results_combined.xlsx")
    save_results_excel(pub_table, combined_detail, xlsx_path, metadata)

    print("\nDone.")