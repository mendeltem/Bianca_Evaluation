#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase II-A: Bland-Altman Analysis  Test Condition Comparisons
==================================================================

Revision context: R1 Comment 1, R5 #4, R5 #9.

Design: train=inpainted (fixed, single model), threshold=0.85.
Test conditions: non_removed, removed, filled (n=89 subjects with GT).
Single run per subject (real-world application scenario, no repeated CV).

Comparisons (3):
  1. Non Removed vs Removed
  2. Non Removed vs Inpainted
  3. Removed vs Inpainted  (convergence)

Key results
-----------
Friedman omnibus (k=3 test conditions, paired):
  Dice not significant (Chi2=2.94, p=0.230). Sensitivity (Chi2=29.34,
  p<0.001) and Precision (Chi2=14.76, p<0.001) significant. The Friedman
  detects the opposing Sens/Prec shifts but not their net effect in Dice.

ICC(3,1)  agreement:
  Dice=0.9997, Sensitivity=0.9991, Precision=0.9993. All excellent
  (Koo & Li, 2016). >99.9% of variance is between-subject; <0.1% is
  attributable to test condition preprocessing. The three conditions
  are practically interchangeable.

TOST equivalence (margin +/- 0.05 points):
  All 9 comparisons: Equivalent = Yes (all p_TOST <0.001). Actual
  differences (max 0.004 Dice points) fall far within the margin.
  This provides positive evidence of equivalence, not merely absence
  of a detectable difference.

Wilcoxon signed-rank (paired, Bonferroni k=3, alpha_adj=0.0167):
  Dice Score:  No significant pairwise differences (all p_bonf >= 0.095).
  Sensitivity: NR vs R (p_bonf <0.001***), NR vs Inpainted (p_bonf <0.001***),
               R vs Inpainted (p_bonf = 0.293, ns). Both removal conditions
               produce marginally higher Sensitivity than Non Removed
               (delta = -0.024 and -0.022, both negligible).
  Precision:   NR vs R (p_bonf <0.001***), NR vs Inpainted (p_bonf <0.001***),
               R vs Inpainted (p_bonf = 1.000, ns). Non Removed produces
               marginally higher Precision (delta = 0.012 and 0.013, negligible).
  All Cliff's delta: negligible (max |delta| = 0.024, well below 0.28).

  Interpretation: The paired design (89 subjects) detects sub-0.005-point
  differences in Sensitivity and Precision. These show opposing directional
  shifts that cancel in Dice Score. Removing the stroke lesion slightly
  increases Sensitivity (BIANCA detects previously obscured WMH) while
  slightly decreasing Precision (more false positives from altered intensity
  normalization). All effects are negligible in magnitude.

Bland-Altman agreement:
  Non Removed vs Removed:   Dice LoA [-0.015, 0.017], Sens [-0.019, 0.013],
                            Prec [-0.023, 0.032]. Systematic bias in Sens
                            and Prec (p_bonf <0.001), not in Dice (p_bonf=1.0).
  Non Removed vs Inpainted: Dice LoA [-0.014, 0.017], Sens [-0.022, 0.016],
                            Prec [-0.023, 0.031]. Same pattern as above.
  Removed vs Inpainted:     Dice LoA [-0.002, 0.002], Sens [-0.006, 0.007],
                            Prec [-0.003, 0.003]. No systematic bias after
                            Bonferroni (all p_bonf >= 0.095). LoA ~10x
                            narrower than NR comparisons.

  Convergence conclusion: The near-identical agreement between Removed and
  Inpainted (LoA within +/-0.007 across all metrics, no significant bias)
  confirms that observed effects are attributable to lesion removal per se,
  independent of the filling strategy (zero-filling vs NAWM-based inpainting).

Bonferroni family structure:
  BA bias tests: k=3 per metric family (Dice, Sensitivity, Precision each
  form a separate family with 3 pairwise comparisons, alpha_adj=0.0167).
  Wilcoxon post-hoc: same structure (k=3 per metric, alpha_adj=0.0167).

Paper changes
-------------
  Section 3.2 (Results): BA figures for test condition comparison.
  Supplemental: BA table, Friedman omnibus, Wilcoxon post-hoc, ICC, TOST.

Response to Reviewers
---------------------
  R1 Comment 1: "TOST equivalence testing formally demonstrated that all
    pairwise differences between test conditions fall within +/-0.05 points
    (all p <0.001). ICC(3,1) exceeded 0.999 for all metrics, confirming
    that test condition preprocessing does not meaningfully affect BIANCA
    segmentation performance. Bland-Altman convergence analysis (Removed
    vs Inpainted, 95% LoA for Dice: [-0.002, 0.002]) further confirms
    that zero-filling and inpainting produce equivalent results."

  R5 Comment 9: "The opposing directional pattern in Sensitivity (Removed >
    Non Removed, delta = -0.024) and Precision (Non Removed > Removed,
    delta = 0.012) is consistent with the mechanistic explanation: removing
    the stroke lesion allows BIANCA to detect previously obscured WMH
    (higher Sensitivity) while introducing additional false positives
    (lower Precision). These effects cancel in Dice Score and are
    negligible in magnitude (ICC > 0.999 across all metrics)."

  R5 Comment 4: "Bland-Altman 95% LoA confirm that the maximum agreement
    range across all test condition comparisons is [-0.023, 0.032]
    (Precision, NR vs R). Convergence LoA are within +/-0.007 for all
    metrics. TOST equivalence was demonstrated for all 9 comparisons
    (p <0.001, margin +/-0.05), providing positive evidence that the
    filling strategy has no meaningful effect on segmentation performance."

Outputs
-------
  BA_{cond1}_vs_{cond2}.png           3-panel BA plots (shared y-axis)
  bland_altman_test_conditions.xlsx   combined BA table (9 rows)
  omnibus_test_conditions.xlsx        Friedman + Wilcoxon + ICC + TOST

Statistical methods
-------------------
  Friedman test: Non-parametric omnibus for repeated measures (k=3).
    Replaces Kruskal-Wallis which assumes independent groups.
  ICC(3,1): Two-way mixed, single measures, consistency (Koo & Li, 2016).
    Quantifies agreement  values near 1.0 = conditions interchangeable.
  TOST: Two One-Sided Wilcoxon tests with margin +/- 0.05 points.
    Positive evidence of equivalence (evidence of absence), not just
    failure to reject H0 (absence of evidence).

@author: temuuleu
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import wilcoxon, friedmanchisquare
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cliffs_delta import cliffs_delta as _cliffs_delta

warnings.filterwarnings('ignore')

# =============================================================
# GLOBAL PLOT STYLE  improved readability
# =============================================================
plt.rcParams.update({
    'font.size': 19,
    'axes.titlesize': 20,
    'axes.labelsize': 15,
    'xtick.labelsize': 15,
    'ytick.labelsize': 15,
    'legend.fontsize': 15,
    'figure.titlesize': 20,
})

# =============================================================
# CONFIG
# =============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_DICE_ONLY.xlsx")
output_dir = os.path.join(SCRIPT_DIR, "analysis", "bland_altman")
os.makedirs(output_dir, exist_ok=True)

# Design parameters
TRAIN_CONDITION = 'filled'   # fixed: trained on inpainted only
THRESHOLD = '85'

# Comparisons to run (test condition pairs)
COMPARISONS = [
    ('non_removed', 'removed'),
    ('non_removed', 'filled'),
    ('removed', 'filled'),          # convergence
]

COND_LABELS = {
    'non_removed': 'Non Removed',
    'removed': 'Removed',
    'filled': 'Inpainted',
}

# Voxel-level metrics only (matching your existing plots)
BA_METRICS = ['dice_score', 'sensitivity', 'precision']
BA_METRIC_LABELS = {
    'dice_score': 'Dice Score',
    'sensitivity': 'Sensitivity',
    'precision': 'Precision',
}

# Lesion type styling
LESION_COLORS = {
    'infarct': '#4C72B0',
    'infra': '#DD4444',
    'lacune': '#E48AC5',
    'mixed': '#55D6BE',
}

# Cliff's Delta threshold
DELTA_THRESHOLD = 0.28

# Outlier: beyond mean +/- 2.5*SD of differences
OUTLIER_FACTOR = 2.5

# Bonferroni for BA bias tests: k=3 per metric family
# (each metric has 3 comparisons: NR vs R, NR vs F, R vs F)
N_BA_BONF = 3


# =============================================================
# HELPERS
# =============================================================

def cliffs_delta_wrapper(x, y):
    x, y = list(x), list(y)
    if len(x) == 0 or len(y) == 0:
        return np.nan, 'N/A'
    d, size = _cliffs_delta(x, y)
    return round(d, 4), size


def bootstrap_cliffs_delta_ci(x, y, n_boot=1000, alpha=0.05, seed=42):
    rng = np.random.default_rng(seed)
    x, y = np.array(x), np.array(y)
    if len(x) < 2 or len(y) < 2:
        return np.nan, np.nan
    deltas = []
    for _ in range(n_boot):
        x_b = rng.choice(x, size=len(x), replace=True)
        y_b = rng.choice(y, size=len(y), replace=True)
        d, _ = _cliffs_delta(list(x_b), list(y_b))
        deltas.append(d)
    lo = np.percentile(deltas, 100 * alpha / 2)
    hi = np.percentile(deltas, 100 * (1 - alpha / 2))
    return round(lo, 4), round(hi, 4)


def format_p(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '<0.001'
    return f"{p:.3f}"


# =============================================================
# LOAD + AGGREGATE
# =============================================================

def load_and_aggregate():
    """
    Load wide-format LOCATE_Results_Metrics_DICE_ONLY.xlsx and melt
    into long format: one row per subject x test_condition.

    Wide columns: WMH_{condition}_{metric_suffix}
    where condition = non_removed | removed | filled
    and metric_suffix = dice | sens | prec | lesion_f1 | lesion_prec | lesion_rec
    """
    if not os.path.exists(XLSX_PATH):
        raise FileNotFoundError(f"File not found: {XLSX_PATH}")

    df = pd.read_excel(XLSX_PATH)
    print(f"Loaded: {len(df)} rows from {os.path.basename(XLSX_PATH)}")

    # Filter: only subjects with ground-truth mask
    if 'subject_with_mask' in df.columns:
        df = df[df['subject_with_mask'] == 1].copy()
        print(f"Filtered (subject_with_mask=1): {len(df)} rows")

    # Replace ICB -> ICH
    if 'lesion_type' in df.columns:
        df['lesion_type'] = df['lesion_type'].replace('ICB', 'ICH')

    # Wide-to-long: metric suffix mapping
    SUFFIX_TO_METRIC = {
        'dice': 'dice_score',
        'sens': 'sensitivity',
        'prec': 'precision',
    }

    # Extra columns to carry over
    extra_cols = [c for c in ['scanner', 'severity_level', 'lesion_type']
                  if c in df.columns]

    # Identify subject column
    subj_col = 'subject' if 'subject' in df.columns else df.columns[0]

    rows = []
    for cond in ['non_removed', 'removed', 'filled']:
        for _, row in df.iterrows():
            entry = {
                'subject': row[subj_col],
                'test_condition': cond,
            }
            for suffix, metric_name in SUFFIX_TO_METRIC.items():
                col = f"WMH_{cond}_{suffix}"
                entry[metric_name] = row.get(col, np.nan)
            for ec in extra_cols:
                entry[ec] = row.get(ec, np.nan)
            rows.append(entry)

    df_agg = pd.DataFrame(rows)
    print(f"Melted: {len(df_agg)} rows ({df_agg['subject'].nunique()} subjects x "
          f"{df_agg['test_condition'].nunique()} conditions)")
    return df_agg


# =============================================================
# BLAND-ALTMAN: COMPUTE STATS (pass 1)
# =============================================================

def compute_ba_stats(df_agg, cond1, cond2):
    """
    Compute Bland-Altman statistics for one comparison (3 metrics).
    Returns list of dicts with all stats + raw arrays for plotting.
    """
    label1 = COND_LABELS[cond1]
    label2 = COND_LABELS[cond2]

    d1 = df_agg[df_agg['test_condition'] == cond1].set_index('subject')
    d2 = df_agg[df_agg['test_condition'] == cond2].set_index('subject')
    common = d1.index.intersection(d2.index)
    d1, d2 = d1.loc[common], d2.loc[common]

    has_lesion_type = 'lesion_type' in d1.columns
    if has_lesion_type:
        lesion_types = d1['lesion_type'].fillna('unknown').values
    else:
        lesion_types = np.array(['unknown'] * len(d1))

    panels = []
    for metric in BA_METRICS:
        metric_label = BA_METRIC_LABELS[metric]
        vals_1 = d1[metric].values
        vals_2 = d2[metric].values
        diffs = vals_1 - vals_2
        means = (vals_1 + vals_2) / 2.0

        mean_diff = np.mean(diffs)
        median_diff = np.median(diffs)
        sd_diff = np.std(diffs, ddof=1)
        loa_lower = mean_diff - 1.96 * sd_diff
        loa_upper = mean_diff + 1.96 * sd_diff

        outlier_lo = mean_diff - OUTLIER_FACTOR * sd_diff
        outlier_hi = mean_diff + OUTLIER_FACTOR * sd_diff
        is_outlier = (diffs < outlier_lo) | (diffs > outlier_hi)

        try:
            _, p_bias = wilcoxon(diffs, alternative='two-sided')
        except ValueError:
            p_bias = 1.0
        p_bias_bonf = min(p_bias * N_BA_BONF, 1.0)
        systematic = p_bias_bonf < 0.05

        delta, mag = cliffs_delta_wrapper(vals_1, vals_2)
        ci_lo, ci_hi = bootstrap_cliffs_delta_ci(vals_1, vals_2)
        meaningful = abs(delta) >= DELTA_THRESHOLD if not np.isnan(delta) else False

        # Direction only stated when effect is significant or meaningful
        if systematic or meaningful:
            higher = label2 if mean_diff < 0 else label1 if mean_diff > 0 else 'Equal'
        else:
            higher = 'No meaningful difference'

        panels.append({
            # Stats for table
            'Metric': metric_label,
            'n': len(diffs),
            'Mean Diff (bias)': round(mean_diff, 6),
            'Median Diff': round(median_diff, 6),
            'SD Diff': round(sd_diff, 6),
            '95% LoA Lower': round(loa_lower, 3),
            '95% LoA Upper': round(loa_upper, 3),
            'Systematic Bias': 'Yes' if systematic else 'No',
            'Bias p (raw)': format_p(p_bias),
            'Bias p (Bonferroni)': format_p(p_bias_bonf),
            'Higher Condition': higher,
            "Cliff's delta": delta,
            'Effect Size': mag,
            'delta CI_lo': ci_lo,
            'delta CI_hi': ci_hi,
            # Raw data for plotting
            '_diffs': diffs,
            '_means': means,
            '_is_outlier': is_outlier,
            '_lesion_types': lesion_types,
            '_has_lesion_type': has_lesion_type,
            '_systematic': systematic,
            '_meaningful': meaningful,
            '_loa_lower': loa_lower,
            '_loa_upper': loa_upper,
            '_median_diff': median_diff,
            '_p_bias': p_bias,
            '_p_bias_bonf': p_bias_bonf,
            '_delta': delta,
            '_ci_lo': ci_lo,
            '_ci_hi': ci_hi,
            '_mag': mag,
            '_higher': higher,
        })

    return panels


def get_global_ylim(all_panels):
    """
    Find symmetric y-axis limits across ALL panels (all comparisons,
    all metrics). Uses the max absolute difference value + 10% padding.
    """
    global_max = 0
    for panels in all_panels:
        for p in panels:
            diffs = p['_diffs']
            abs_max = np.max(np.abs(diffs))
            if abs_max > global_max:
                global_max = abs_max
    padding = global_max * 0.15
    ylim = global_max + padding
    return (-ylim, ylim)


# =============================================================
# BLAND-ALTMAN: PLOT (pass 2, with shared y-axis)
# =============================================================

def plot_ba_comparison(panels, cond1, cond2, ylim):
    """
    Plot one 3-panel Bland-Altman figure with fixed y-axis limits.
    """
    label1 = COND_LABELS[cond1]
    label2 = COND_LABELS[cond2]
    has_lesion_type = panels[0]['_has_lesion_type']

    fig, axes = plt.subplots(1, 3, figsize=(21, 6))
    fig.suptitle(f'Phase II-A: Bland-Altman: {label1}  vs  {label2}',
                 fontsize=16, fontweight='bold', y=1.02)

    for ax_idx, p in enumerate(panels):
        ax = axes[ax_idx]
        diffs = p['_diffs']
        means = p['_means']
        is_outlier = p['_is_outlier']
        lesion_types = p['_lesion_types']
        loa_lower = p['_loa_lower']
        loa_upper = p['_loa_upper']
        median_diff = p['_median_diff']
        systematic = p['_systematic']
        meaningful = p['_meaningful']
        metric_label = p['Metric']

        # Set shared y-axis
        ax.set_ylim(ylim)

        # Scatter by lesion type (non-outliers)
        for lt in sorted(set(lesion_types)):
            mask = (lesion_types == lt) & (~is_outlier)
            color = LESION_COLORS.get(lt, '#888888')
            ax.scatter(means[mask], diffs[mask],
                       c=color, marker='o', s=60, alpha=0.7,
                       edgecolors='white', linewidths=0.5,
                       label=lt if ax_idx == 0 else None)

        # Outliers
        if np.any(is_outlier):
            for lt in sorted(set(lesion_types)):
                mask = (lesion_types == lt) & is_outlier
                if not np.any(mask):
                    continue
                color = LESION_COLORS.get(lt, '#888888')
                ax.scatter(means[mask], diffs[mask],
                           c=color, marker='x', s=120, linewidths=2.5,
                           label='Outlier' if ax_idx == 0 and lt == sorted(set(lesion_types))[0] else None)

        # Reference lines
        ax.axhline(median_diff, color='black', linestyle='--', linewidth=1.2, alpha=0.8)
        ax.axhline(loa_lower, color='green', linestyle='--', linewidth=1.2, alpha=0.7)
        ax.axhline(loa_upper, color='red', linestyle='--', linewidth=1.2, alpha=0.7)

        # LoA annotations (at left edge of x-axis)
        x_left = ax.get_xlim()[0] if ax.get_xlim()[0] != 0 else 0.01
        ax.text(x_left, loa_upper, f' {loa_upper:.2f}', color='red',
                fontsize=11, fontweight='bold', va='bottom', ha='left')
        ax.text(x_left, loa_lower, f' {loa_lower:.2f}', color='green',
                fontsize=11, fontweight='bold', va='top', ha='left')
        ax.text(x_left, median_diff, f' {median_diff:.2f}', color='black',
                fontsize=11, fontweight='bold', va='bottom', ha='left')

        # Title
        ax.set_title(f"{metric_label}\n95% LoA: [{loa_lower:.3f}, {loa_upper:.3f}]",
                     fontsize=14, fontweight='bold')
        ax.set_xlabel(f"Mean {metric_label}", fontsize=13)
        ax.set_ylabel("Difference", fontsize=13)

        # Tick label size
        ax.tick_params(axis='both', labelsize=11)

    # Legend
    if has_lesion_type:
        legend_handles = []
        for lt in sorted(LESION_COLORS.keys()):
            legend_handles.append(mpatches.Patch(color=LESION_COLORS[lt], label=lt))
        legend_handles.append(plt.Line2D([0], [0], color='green', linestyle='--', label='Lower LoA'))
        legend_handles.append(plt.Line2D([0], [0], color='red', linestyle='--', label='Upper LoA'))
        legend_handles.append(plt.Line2D([0], [0], color='black', linestyle='--', label='Median'))
        legend_handles.append(plt.Line2D([0], [0], color='gray', marker='x',
                                          linestyle='None', markersize=10, label='Outlier'))
        fig.legend(handles=legend_handles, loc='lower center',
                   ncol=len(legend_handles), fontsize=11,
                   bbox_to_anchor=(0.5, -0.06))

    fig.tight_layout()

    fname = f"BA_{cond1}_vs_{cond2}"
    fig.savefig(os.path.join(output_dir, f"{fname}.png"), dpi=300, bbox_inches='tight')
    plt.show()
    plt.close(fig)
    print(f"  Saved: {fname}.png")


# =============================================================
# FRIEDMAN + WILCOXON POST-HOC + ICC + TOST
# =============================================================

TEST_CONDITIONS = ['non_removed', 'removed', 'filled']
N_POSTHOC = 3  # 3 pairwise comparisons -> Bonferroni alpha = 0.0167

# TOST equivalence margin: 0.05 points on metric scale
# Rationale: smallest difference that could plausibly affect downstream
# analyses. Conservative choice  typical Dice measurement noise in
# WMH segmentation exceeds 0.05 points.
TOST_MARGIN = 0.05


def sig_stars(p):
    if pd.isna(p):
        return 'N/A'
    if p < 0.001:
        return '***'
    if p < 0.01:
        return '**'
    if p < 0.05:
        return '*'
    return 'ns'


def compute_icc(df_wide, metric_cols):
    """
    Compute ICC(3,1)  two-way mixed, single measures, consistency.
    Appropriate for: same subjects rated under fixed conditions.

    Parameters
    ----------
    df_wide : DataFrame with one row per subject, columns = conditions.
    metric_cols : list of column names (one per condition).

    Returns
    -------
    icc : float, ICC(3,1) value.
    """
    data = df_wide[metric_cols].dropna().values
    n, k = data.shape  # n subjects, k conditions

    # Grand mean
    grand_mean = np.mean(data)

    # Mean squares
    row_means = np.mean(data, axis=1)
    col_means = np.mean(data, axis=0)

    SS_rows = k * np.sum((row_means - grand_mean) ** 2)
    SS_cols = n * np.sum((col_means - grand_mean) ** 2)
    SS_total = np.sum((data - grand_mean) ** 2)
    SS_error = SS_total - SS_rows - SS_cols

    MS_rows = SS_rows / (n - 1)
    MS_error = SS_error / ((n - 1) * (k - 1))

    # ICC(3,1): (MS_rows - MS_error) / (MS_rows + (k-1)*MS_error)
    denom = MS_rows + (k - 1) * MS_error
    if denom == 0:
        return np.nan
    icc = (MS_rows - MS_error) / denom
    return round(icc, 6)


def interpret_icc(icc):
    """Koo & Li (2016) benchmarks for ICC interpretation."""
    if pd.isna(icc):
        return 'N/A'
    if icc >= 0.90:
        return 'excellent'
    if icc >= 0.75:
        return 'good'
    if icc >= 0.50:
        return 'moderate'
    return 'poor'


def tost_paired_wilcoxon(x, y, margin):
    """
    Non-parametric TOST using paired Wilcoxon signed-rank tests.

    Tests whether the median difference between x and y falls within
    [-margin, +margin]. Equivalence is declared if BOTH one-sided
    tests are significant (p < 0.05).

    Returns
    -------
    p_upper : p-value for H0: diff >= +margin (test: diff - margin < 0)
    p_lower : p-value for H0: diff <= -margin (test: diff + margin > 0)
    p_tost  : max(p_upper, p_lower)  overall TOST p-value
    equivalent : bool
    """
    diff = np.array(x) - np.array(y)

    # Test 1: H0: median(diff) >= +margin → one-sided less
    shifted_upper = diff - margin
    try:
        _, p_upper = wilcoxon(shifted_upper, alternative='less')
    except ValueError:
        p_upper = 1.0

    # Test 2: H0: median(diff) <= -margin → one-sided greater
    shifted_lower = diff + margin
    try:
        _, p_lower = wilcoxon(shifted_lower, alternative='greater')
    except ValueError:
        p_lower = 1.0

    p_tost = max(p_upper, p_lower)
    equivalent = p_tost < 0.05

    return p_upper, p_lower, p_tost, equivalent


def analysis_omnibus(df_agg):
    """
    Friedman omnibus (paired, k=3 test conditions) + Wilcoxon
    signed-rank post-hoc (Bonferroni) + ICC(3,1) + TOST equivalence.

    Returns dict of DataFrames: {sheet_name: df}
    """
    friedman_rows = []
    pw_rows = []
    desc_rows = []
    icc_rows = []
    tost_rows = []

    for metric in BA_METRICS:
        metric_label = BA_METRIC_LABELS[metric]

        # ── Pivot to wide: subject x condition ──
        pivot = df_agg.pivot(index='subject', columns='test_condition',
                             values=metric).dropna()

        # Descriptive per condition
        groups = {}
        for cond in TEST_CONDITIONS:
            if cond not in pivot.columns:
                continue
            vals = pivot[cond].values
            groups[cond] = vals
            n = len(vals)
            desc_rows.append({
                'Metric': metric_label,
                'Test Condition': COND_LABELS[cond],
                'n': n,
                'Mean': round(np.mean(vals), 4),
                'SD': round(np.std(vals, ddof=1), 4),
                'Median': round(np.median(vals), 4),
                'IQR': round(np.percentile(vals, 75) - np.percentile(vals, 25), 4),
            })

        # ── Friedman omnibus (correct for paired data) ──
        valid = [groups[c] for c in TEST_CONDITIONS if c in groups]
        if len(valid) == 3 and all(len(v) > 0 for v in valid):
            chi2, p_friedman = friedmanchisquare(*valid)
            chi2 = round(chi2, 4)
        else:
            chi2, p_friedman = np.nan, np.nan

        friedman_rows.append({
            'Metric': metric_label,
            'Chi2': chi2,
            'p-value': format_p(p_friedman),
            'Significant (alpha=0.05)': sig_stars(p_friedman),
        })

        # ── ICC(3,1) ──
        icc_val = compute_icc(pivot, [c for c in TEST_CONDITIONS if c in pivot.columns])
        icc_rows.append({
            'Metric': metric_label,
            'ICC(3,1)': icc_val,
            'Interpretation': interpret_icc(icc_val),
            'n': len(pivot),
        })

        # ── Wilcoxon post-hoc + TOST per pair ──
        for c1, c2 in combinations(TEST_CONDITIONS, 2):
            label1, label2 = COND_LABELS[c1], COND_LABELS[c2]

            if c1 not in pivot.columns or c2 not in pivot.columns:
                continue

            v1 = pivot[c1].values
            v2 = pivot[c2].values

            # Wilcoxon
            try:
                stat, p_pw = wilcoxon(v1, v2, alternative='two-sided')
            except ValueError:
                stat, p_pw = np.nan, 1.0

            p_adj = min(p_pw * N_POSTHOC, 1.0)
            delta, mag = cliffs_delta_wrapper(v1, v2)
            ci_lo, ci_hi = bootstrap_cliffs_delta_ci(v1, v2)
            meaningful = 'Yes' if abs(delta) >= DELTA_THRESHOLD else 'No'

            pw_rows.append({
                'Metric': metric_label,
                'Comparison': f"{label1} vs {label2}",
                'n_paired': len(v1),
                'W-stat': round(stat, 2) if not pd.isna(stat) else np.nan,
                'p-value': format_p(p_pw),
                'p-adj (Bonferroni)': format_p(p_adj),
                'Significant (alpha=0.0167)': sig_stars(p_adj),
                "Cliff's delta": delta,
                'Effect Size': mag,
                'delta CI_lo': ci_lo, 'delta CI_hi': ci_hi,
                'Meaningful (|delta|>=0.28)': meaningful,
            })

            # TOST
            p_upper, p_lower, p_tost, equivalent = tost_paired_wilcoxon(
                v1, v2, TOST_MARGIN)
            mean_diff = np.mean(v1 - v2)
            tost_rows.append({
                'Metric': metric_label,
                'Comparison': f"{label1} vs {label2}",
                'n': len(v1),
                'Mean Diff': round(mean_diff, 6),
                f'Margin (+/- {TOST_MARGIN})': TOST_MARGIN,
                'p (upper)': format_p(p_upper),
                'p (lower)': format_p(p_lower),
                'p (TOST)': format_p(p_tost),
                'Equivalent': 'Yes' if equivalent else 'No',
            })

    sheets = {
        'Descriptive': pd.DataFrame(desc_rows),
        'Friedman': pd.DataFrame(friedman_rows),
        'Wilcoxon Post-hoc': pd.DataFrame(pw_rows),
        'ICC': pd.DataFrame(icc_rows),
        'TOST Equivalence': pd.DataFrame(tost_rows),
    }
    return sheets


# =============================================================
# EXCEL FORMATTING
# =============================================================

def format_excel(filepath, sheet_metadata=None):
    wb = load_workbook(filepath)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    meta_font = Font(italic=True, size=9, name='Arial', color='555555')
    data_font = Font(size=9, name='Arial')
    sig_fill = PatternFill('solid', fgColor='E2EFDA')
    meaningful_fill = PatternFill('solid', fgColor='FCE4D6')
    border = Border(bottom=Side(style='thin', color='D9D9D9'))
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for ws in wb.worksheets:
        n_meta = 0
        if sheet_metadata and ws.title in sheet_metadata:
            meta_lines = sheet_metadata[ws.title]
            n_meta = len(meta_lines)
            ws.insert_rows(1, n_meta)
            for i, line in enumerate(meta_lines, start=1):
                cell = ws.cell(row=i, column=1, value=line)
                cell.font = meta_font
                cell.alignment = left
                ws.merge_cells(start_row=i, start_column=1,
                               end_row=i, end_column=ws.max_column)

        header_row = n_meta + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Find highlight columns
        sig_col = None
        meaningful_col = None
        equivalent_col = None
        systematic_col = None
        for c in range(1, ws.max_column + 1):
            hval = str(ws.cell(row=header_row, column=c).value or '')
            if 'Significant' in hval:
                sig_col = c
            if 'Meaningful' in hval:
                meaningful_col = c
            if hval == 'Equivalent':
                equivalent_col = c
            if 'Systematic' in hval:
                systematic_col = c

        equiv_fill = PatternFill('solid', fgColor='D6EAF8')  # light blue

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = center
                cell.border = border

            # Green for significant (Wilcoxon/Friedman)
            if sig_col:
                sig_val = ws.cell(row=row, column=sig_col).value
                if sig_val and sig_val not in ('ns', 'N/A'):
                    for c2 in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c2).fill = sig_fill

            # Green for systematic bias (BA table)
            if systematic_col:
                sys_val = ws.cell(row=row, column=systematic_col).value
                if sys_val == 'Yes':
                    for c2 in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c2).fill = sig_fill

            # Orange for meaningful effect
            if meaningful_col:
                m_val = ws.cell(row=row, column=meaningful_col).value
                if m_val == 'Yes':
                    for c2 in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c2).fill = meaningful_fill

            # Blue for equivalence proven
            if equivalent_col:
                eq_val = ws.cell(row=row, column=equivalent_col).value
                if eq_val == 'Yes':
                    for c2 in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c2).fill = equiv_fill

        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                         for r in range(header_row, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 35)

    wb.save(filepath)


# =============================================================
# MAIN
# =============================================================

if __name__ == "__main__":
    print("=" * 70)
    print("PHASE II-A: BLAND-ALTMAN  TEST CONDITION COMPARISONS")
    print(f"Train: {COND_LABELS[TRAIN_CONDITION]} (fixed) | Threshold: 0.{THRESHOLD}")
    print(f"Test conditions: {[COND_LABELS[c] for c in TEST_CONDITIONS]}")
    print(f"Comparisons: {len(COMPARISONS)} (incl. convergence: Removed vs Inpainted)")
    print("=" * 70)

    df_agg = load_and_aggregate()
    

    # === PASS 1: Compute all stats ===
    all_panels = []  # list of (cond1, cond2, panels)
    for cond1, cond2 in COMPARISONS:
        panels = compute_ba_stats(df_agg, cond1, cond2)
        all_panels.append((cond1, cond2, panels))

    # === Find global symmetric y-axis across all 9 panels ===
    ylim = get_global_ylim([panels for _, _, panels in all_panels])
    print(f"\nGlobal y-axis limits: [{ylim[0]:.4f}, {ylim[1]:.4f}]")

    # === PASS 2: Plot with shared y-axis ===
    all_results = []
    for cond1, cond2, panels in all_panels:
        label1, label2 = COND_LABELS[cond1], COND_LABELS[cond2]
        print(f"\n--- {label1} vs {label2} ---")

        plot_ba_comparison(panels, cond1, cond2, ylim)

        # Build table (strip internal keys starting with '_')
        table_rows = [{k: v for k, v in p.items() if not k.startswith('_')} for p in panels]
        ba_df = pd.DataFrame(table_rows)
        ba_df.insert(0, 'Comparison', f"{label1} vs {label2}")
        all_results.append(ba_df)

        print(ba_df[['Metric', 'Mean Diff (bias)', '95% LoA Lower',
                      '95% LoA Upper', 'Systematic Bias', 'Bias p (Bonferroni)',
                      'Higher Condition', "Cliff's delta"]].to_string(index=False))

    # Save combined table
    combined = pd.concat(all_results, ignore_index=True)
    xlsx_path = os.path.join(output_dir, "bland_altman_test_conditions.xlsx")

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        combined.to_excel(writer, sheet_name='Bland-Altman', index=False)

    meta = {
        'Bland-Altman': [
            "Table: Phase II-A  Bland-Altman Agreement (Test Condition Comparisons)",
            f"Design: train={COND_LABELS[TRAIN_CONDITION]}, threshold=0.{THRESHOLD}",
            "Comparisons: NR vs R, NR vs Inpainted, R vs Inpainted (convergence).",
            "Difference = Condition_1 - Condition_2. Positive = Condition_1 higher.",
            "95% LoA = mean difference +/- 1.96 x SD of differences.",
            f"Systematic bias: Wilcoxon signed-rank, Bonferroni-corrected (k={N_BA_BONF} per metric family, alpha_adj={0.05/N_BA_BONF:.4f}).",
            "Family structure: Dice, Sensitivity, Precision each form a separate family with 3 comparisons.",
            f"Higher Condition: only stated when bias is significant (corrected) or |delta| >= {DELTA_THRESHOLD}.",
            "Y-axis: symmetric, shared across all 9 panels for direct comparison.",
            "Single run per subject (real-world application scenario, no repeated CV).",
        ],
    }
    format_excel(xlsx_path, meta)
    print(f"\n  Saved table: {xlsx_path}")

    # === Friedman + Wilcoxon + ICC + TOST (3 test conditions) ===
    print("\n" + "=" * 70)
    print("FRIEDMAN + ICC + TOST: 3 TEST CONDITIONS")
    print("=" * 70)

    omnibus_sheets = analysis_omnibus(df_agg)

    omnibus_path = os.path.join(output_dir, "omnibus_test_conditions.xlsx")
    with pd.ExcelWriter(omnibus_path, engine='openpyxl') as writer:
        for sheet_name, sheet_df in omnibus_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    omnibus_meta = {
        'Descriptive': [
            "Table: Phase II-A  Descriptive Statistics by Test Condition",
            f"Design: train={COND_LABELS[TRAIN_CONDITION]}, threshold=0.{THRESHOLD}",
            "Single run per subject (real-world application scenario, no repeated CV).",
        ],
        'Friedman': [
            "Table: Phase II-A  Friedman Test (Omnibus, k=3 Test Conditions, Paired)",
            "Groups: Non Removed, Removed, Inpainted (within-subject).",
            "Null hypothesis: all three test conditions produce equal distributions.",
            "Note: Friedman is the correct non-parametric omnibus for repeated measures.",
            "Replaces Kruskal-Wallis (which assumes independent groups).",
        ],
        'Wilcoxon Post-hoc': [
            "Table: Phase II-A  Wilcoxon Signed-Rank Post-hoc Tests (Paired, Within-Subject)",
            "Bonferroni correction: 3 comparisons per metric, corrected alpha = 0.0167.",
            "Effect size: Cliff's Delta with bootstrap 95% CI (1000 iterations).",
            f"Meaningful effect threshold: |delta| >= {DELTA_THRESHOLD} (Hess & Kromrey, 2004).",
        ],
        'ICC': [
            "Table: Phase II-A  Intraclass Correlation Coefficient ICC(3,1)",
            "Model: two-way mixed, single measures, consistency.",
            "Benchmarks (Koo & Li, 2016): <0.50 poor, 0.50-0.75 moderate,",
            "  0.75-0.90 good, >=0.90 excellent.",
            "Interpretation: ICC quantifies how interchangeable the three",
            "  test conditions are  values near 1.0 indicate that",
            "  between-subject variance dominates over between-condition variance.",
        ],
        'TOST Equivalence': [
            "Table: Phase II-A  TOST Equivalence Testing (Paired, Non-Parametric)",
            f"Equivalence margin: +/- {TOST_MARGIN} points on metric scale.",
            "Method: Two One-Sided Wilcoxon signed-rank tests.",
            "  p(upper): H0: diff >= +margin (one-sided less).",
            "  p(lower): H0: diff <= -margin (one-sided greater).",
            "  p(TOST) = max(p_upper, p_lower). Equivalent if p(TOST) < 0.05.",
            "Interpretation: Unlike standard NHST (absence of evidence),",
            "  TOST provides positive evidence of equivalence (evidence of absence).",
            f"Margin rationale: {TOST_MARGIN} points is a conservative threshold;",
            "  typical WMH segmentation measurement noise exceeds this value.",
        ],
    }
    format_excel(omnibus_path, omnibus_meta)
    print(f"  Saved: {omnibus_path}")

    print("\nFriedman results:")
    print(omnibus_sheets['Friedman'].to_string(index=False))
    print("\nICC results:")
    print(omnibus_sheets['ICC'].to_string(index=False))
    print("\nTOST Equivalence results:")
    print(omnibus_sheets['TOST Equivalence'].to_string(index=False))
    print("\nWilcoxon post-hoc results:")
    print(omnibus_sheets['Wilcoxon Post-hoc'].to_string(index=False))

    print(f"\nAll outputs in: {output_dir}")