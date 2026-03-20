#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Descriptive Statistics & Demographics  BIANCA WMH Revision
=============================================================
Three preprocessing conditions: Non Removed | Removed | Inpainted
Data source: LOCATE_Results_Metrics_DICE_ONLY.xlsx

Generates:
  1. Lesion volume summary (overall, by lesion type, by scanner)
  2. WMH volume statistics per condition (Total, Deep, Periventricular)
  3. Demographics table (sex, age)
  4. Lesion size distribution categories
  5. Boxplots by scanner and lesion type


Revision context

  Section 2.1 (Study Population): demographics, lesion characteristics
  Table 1: Cohort demographics and lesion volumes


Paper changes

  Section 2.1: Updated demographics for Phase 2a cohort
  Table 1: Three-condition volume summary
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

XLSX_PATH = "RESULTS/LOCATE_Results_Metrics_DICE_ONLY.xlsx"
PLOT_DIR = "./plots/4_Descriptive_Statistics"

LESION_ORDER = ["lacune", "infarct", "infra", "mixed", "ICH"]
SCANNER_ORDER = ["Prisma_fit", "Tim Trio", "Philips"]

LESION_DISPLAY = {
    "lacune":  "Lacunes",
    "infarct": "Infarcts",
    "infra":   "Infratentorial strokes",
    "mixed":   "Mixed (infarcts+lacunes)",
    "ICH":     "Intracranial hemorrhage",
}

COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
    "filled":      "Inpainted",
}

# WMH compartments
COMPARTMENTS = [
    ("WMH",     "Total"),
    ("deepWMH", "Deep"),
    ("perWMH",  "Periventricular"),
]


# ─────────────────────────────────────────────
# LESION SIZE SUMMARY
# ─────────────────────────────────────────────

def create_lesion_size_summary(df):
    """Comprehensive lesion volume statistics."""

    print("=" * 80)
    print("OVERALL LESION VOLUME STATISTICS")
    print("=" * 80)

    vol = df["infarct_volume_ml"].dropna()
    overall = {
        "N": len(vol),
        "Median (mL)": vol.median(),
        "Mean (mL)": vol.mean(),
        "Min (mL)": vol.min(),
        "Max (mL)": vol.max(),
        "Q1 (mL)": vol.quantile(0.25),
        "Q3 (mL)": vol.quantile(0.75),
        "IQR (mL)": vol.quantile(0.75) - vol.quantile(0.25),
        "Std Dev": vol.std(),
    }
    for k, v in overall.items():
        print(f"  {k}: {v:.2f}" if isinstance(v, float) else f"  {k}: {v}")

    # By lesion type
    print("\n" + "=" * 80)
    print("LESION VOLUME BY LESION TYPE")
    print("=" * 80)

    lesion_rows = []
    for les in LESION_ORDER:
        ld = df[df["lesion_type"] == les]["infarct_volume_ml"].dropna()
        if len(ld) == 0:
            continue
        row = {
            "Lesion Type": LESION_DISPLAY.get(les, les),
            "N": len(ld),
            "Median (mL)": round(ld.median(), 2),
            "Mean (mL)": round(ld.mean(), 2),
            "Min (mL)": round(ld.min(), 2),
            "Max (mL)": round(ld.max(), 2),
            "Q1 (mL)": round(ld.quantile(0.25), 2),
            "Q3 (mL)": round(ld.quantile(0.75), 2),
            "IQR (mL)": round(ld.quantile(0.75) - ld.quantile(0.25), 2),
        }
        lesion_rows.append(row)
        print(f"\n  {les.upper()} (n={len(ld)}): Median={row['Median (mL)']} mL, "
              f"Range={row['Min (mL)']}-{row['Max (mL)']} mL")

    lesion_df = pd.DataFrame(lesion_rows)

    # By scanner
    print("\n" + "=" * 80)
    print("LESION VOLUME BY SCANNER")
    print("=" * 80)

    scanner_rows = []
    for sc in SCANNER_ORDER:
        sd = df[df["scanner"] == sc]["infarct_volume_ml"].dropna()
        if len(sd) == 0:
            continue
        row = {
            "Scanner": sc,
            "N": len(sd),
            "Median (mL)": round(sd.median(), 2),
            "Mean (mL)": round(sd.mean(), 2),
            "Min (mL)": round(sd.min(), 2),
            "Max (mL)": round(sd.max(), 2),
            "Q1 (mL)": round(sd.quantile(0.25), 2),
            "Q3 (mL)": round(sd.quantile(0.75), 2),
        }
        scanner_rows.append(row)
        print(f"\n  {sc} (n={len(sd)}): Median={row['Median (mL)']} mL")

    scanner_df = pd.DataFrame(scanner_rows)

    # Size distribution
    print("\n" + "=" * 80)
    print("LESION SIZE DISTRIBUTION")
    print("=" * 80)

    n_total = len(vol)
    bins = [
        ("<2 mL", vol < 2),
        ("2-7.5 mL", (vol >= 2) & (vol < 7.5)),
        (">=7.5 mL", vol >= 7.5),
    ]
    dist_rows = []
    for label, mask in bins:
        count = mask.sum()
        pct = 100 * count / n_total if n_total > 0 else 0
        dist_rows.append({"Category": label, "N": count, "Percentage": round(pct, 1)})
        print(f"  {label}: {count} ({pct:.1f}%)")

    dist_df = pd.DataFrame(dist_rows)

    return {
        "overall": overall,
        "by_lesion_type": lesion_df,
        "by_scanner": scanner_df,
        "distribution": dist_df,
    }


# ─────────────────────────────────────────────
# WMH VOLUME STATISTICS PER CONDITION
# ─────────────────────────────────────────────

def create_wmh_volume_stats(df):
    """WMH volume descriptive stats for each condition and compartment."""

    print("\n" + "=" * 80)
    print("WMH VOLUME STATISTICS BY CONDITION")
    print("=" * 80)

    rows = []
    for cond_key, cond_label in COND_LABELS.items():
        for prefix, comp_name in COMPARTMENTS:
            col = f"{prefix}_{cond_key}_volume_ml"
            if col not in df.columns:
                continue
            v = df[col].dropna()
            if len(v) == 0:
                continue

            row = {
                "Condition": cond_label,
                "Compartment": comp_name,
                "N": len(v),
                "Median (mL)": round(v.median(), 2),
                "Mean (mL)": round(v.mean(), 2),
                "Q1 (mL)": round(v.quantile(0.25), 2),
                "Q3 (mL)": round(v.quantile(0.75), 2),
                "Min (mL)": round(v.min(), 2),
                "Max (mL)": round(v.max(), 2),
                "Std Dev": round(v.std(), 2),
            }
            rows.append(row)

    wmh_df = pd.DataFrame(rows)
    print(wmh_df.to_string(index=False))
    return wmh_df


# ─────────────────────────────────────────────
# DEMOGRAPHICS
# ─────────────────────────────────────────────

def create_demographics_table(df):
    """Sex and age summary, overall and by lesion type."""

    print("\n" + "=" * 80)
    print("DEMOGRAPHICS")
    print("=" * 80)

    rows = []

    # Overall
    if "sex" in df.columns and "age" in df.columns:
        n_total = len(df)
        n_male = (df["sex"] == "male").sum() if df["sex"].dtype == object else (df["sex"] == 1).sum()
        n_female = n_total - n_male
        age = df["age"].dropna()

        rows.append({
            "Group": "Overall",
            "N": n_total,
            "Male": n_male,
            "Female": n_female,
            "Male %": round(100 * n_male / n_total, 1) if n_total > 0 else 0,
            "Age Median": round(age.median(), 1),
            "Age Mean": round(age.mean(), 1),
            "Age Min": round(age.min(), 1),
            "Age Max": round(age.max(), 1),
            "Age Q1": round(age.quantile(0.25), 1),
            "Age Q3": round(age.quantile(0.75), 1),
        })

        print(f"\n  Overall (n={n_total}): {n_male}M / {n_female}F, "
              f"Age median={age.median():.1f} ({age.min():.0f}-{age.max():.0f})")

    # By lesion type
    for les in LESION_ORDER:
        ld = df[df["lesion_type"] == les]
        if len(ld) == 0:
            continue
        n = len(ld)
        n_m = (ld["sex"] == "male").sum() if ld["sex"].dtype == object else (ld["sex"] == 1).sum()
        n_f = n - n_m
        age_l = ld["age"].dropna()

        rows.append({
            "Group": LESION_DISPLAY.get(les, les),
            "N": n,
            "Male": n_m,
            "Female": n_f,
            "Male %": round(100 * n_m / n, 1) if n > 0 else 0,
            "Age Median": round(age_l.median(), 1) if len(age_l) > 0 else np.nan,
            "Age Mean": round(age_l.mean(), 1) if len(age_l) > 0 else np.nan,
            "Age Min": round(age_l.min(), 1) if len(age_l) > 0 else np.nan,
            "Age Max": round(age_l.max(), 1) if len(age_l) > 0 else np.nan,
            "Age Q1": round(age_l.quantile(0.25), 1) if len(age_l) > 0 else np.nan,
            "Age Q3": round(age_l.quantile(0.75), 1) if len(age_l) > 0 else np.nan,
        })

        print(f"  {les} (n={n}): {n_m}M / {n_f}F, "
              f"Age median={age_l.median():.1f}" if len(age_l) > 0 else "")

    # By scanner
    for sc in SCANNER_ORDER:
        sd = df[df["scanner"] == sc]
        if len(sd) == 0:
            continue
        n = len(sd)
        n_m = (sd["sex"] == "male").sum() if sd["sex"].dtype == object else (sd["sex"] == 1).sum()
        n_f = n - n_m
        age_s = sd["age"].dropna()

        rows.append({
            "Group": f"Scanner: {sc}",
            "N": n,
            "Male": n_m,
            "Female": n_f,
            "Male %": round(100 * n_m / n, 1) if n > 0 else 0,
            "Age Median": round(age_s.median(), 1) if len(age_s) > 0 else np.nan,
            "Age Mean": round(age_s.mean(), 1) if len(age_s) > 0 else np.nan,
            "Age Min": round(age_s.min(), 1) if len(age_s) > 0 else np.nan,
            "Age Max": round(age_s.max(), 1) if len(age_s) > 0 else np.nan,
            "Age Q1": round(age_s.quantile(0.25), 1) if len(age_s) > 0 else np.nan,
            "Age Q3": round(age_s.quantile(0.75), 1) if len(age_s) > 0 else np.nan,
        })

    demo_df = pd.DataFrame(rows)
    return demo_df


# ─────────────────────────────────────────────
# BOXPLOTS
# ─────────────────────────────────────────────

def create_scanner_boxplots(df, plot_dir):
    """WMH volume boxplots by scanner for each condition."""

    palette = {
        "Total": "#2ca02c",
        "Periventricular": "#1f77b4",
        "Deep": "#ff7f0e",
    }

    for cond_key, cond_label in COND_LABELS.items():
        plot_data = []
        for prefix, comp_name in COMPARTMENTS:
            col = f"{prefix}_{cond_key}_volume_ml"
            if col not in df.columns:
                continue
            tmp = df[["scanner"]].copy()
            tmp["volume"] = df[col]
            tmp["compartment"] = comp_name
            plot_data.append(tmp)

        if not plot_data:
            continue

        melted = pd.concat(plot_data, ignore_index=True).dropna()

        fig, ax = plt.subplots(figsize=(12, 7))
        sns.boxplot(x="scanner", y="volume", hue="compartment",
                    data=melted, palette=palette, showfliers=True,
                    order=SCANNER_ORDER, ax=ax)

        ax.set_xlabel("Scanner Type", fontsize=12, fontweight="bold")
        ax.set_ylabel("WMH Volume (mL)", fontsize=12, fontweight="bold")
        ax.set_title(f"WMH Volume by Scanner  {cond_label}",
                     fontsize=14, fontweight="bold")
        ax.legend(title="Compartment", loc="upper right")
        ax.grid(True, alpha=0.3, axis="y")
        plt.tight_layout()

        path = os.path.join(plot_dir, f"boxplot_scanner_{cond_key}.png")
        fig.savefig(path, dpi=300, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        print(f"  Scanner boxplot saved: {path}")


def create_lesion_boxplots(df, plot_dir):
    """WMH volume boxplots by lesion type for each condition."""

    palette = {
        "Total": "#2ca02c",
        "Periventricular": "#1f77b4",
        "Deep": "#ff7f0e",
    }

    for cond_key, cond_label in COND_LABELS.items():
        plot_data = []
        for prefix, comp_name in COMPARTMENTS:
            col = f"{prefix}_{cond_key}_volume_ml"
            if col not in df.columns:
                continue
            tmp = df[["lesion_type"]].copy()
            tmp["volume"] = df[col]
            tmp["compartment"] = comp_name
            plot_data.append(tmp)

        if not plot_data:
            continue

        melted = pd.concat(plot_data, ignore_index=True).dropna()

        fig, ax = plt.subplots(figsize=(12, 7))
        sns.boxplot(x="lesion_type", y="volume", hue="compartment",
                    data=melted, palette=palette, showfliers=True,
                    order=LESION_ORDER, ax=ax)

        ax.set_xlabel("Lesion Type", fontsize=12, fontweight="bold")
        ax.set_ylabel("WMH Volume (mL)", fontsize=12, fontweight="bold")
        ax.set_title(f"WMH Volume by Lesion Type  {cond_label}",
                     fontsize=14, fontweight="bold")
        ax.legend(title="Compartment", loc="upper right")
        ax.grid(True, alpha=0.3, axis="y")
        plt.tight_layout()

        path = os.path.join(plot_dir, f"boxplot_lesion_{cond_key}.png")
        fig.savefig(path, dpi=300, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        print(f"  Lesion boxplot saved: {path}")


# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────

def save_all_to_excel(stats_results, wmh_stats, demo_df, filepath):
    """Save all tables to multi-sheet Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        stats_results["by_lesion_type"].to_excel(writer, sheet_name="Lesion_Volume_by_Type", index=False)
        stats_results["by_scanner"].to_excel(writer, sheet_name="Lesion_Volume_by_Scanner", index=False)
        stats_results["distribution"].to_excel(writer, sheet_name="Size_Distribution", index=False)
        wmh_stats.to_excel(writer, sheet_name="WMH_Volume_by_Condition", index=False)
        demo_df.to_excel(writer, sheet_name="Demographics", index=False)

    # Format
    wb = openpyxl.load_workbook(filepath)
    thin = Border(left=Side("thin"), right=Side("thin"),
                  top=Side("thin"), bottom=Side("thin"))
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal="center")
                if cell.row == 1:
                    cell.font = Font(bold=True, size=11)
    wb.save(filepath)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading data …")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    if "subject_with_mask" in df.columns:
        df = df[df["subject_with_mask"] == 1].copy()
    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")

    print(f"Subjects: {len(df)}")
    print(f"Lesion types:\n{df['lesion_type'].value_counts()}")
    if "scanner" in df.columns:
        print(f"Scanners:\n{df['scanner'].value_counts()}")

    os.makedirs(PLOT_DIR, exist_ok=True)

    # 1. Lesion size summary
    stats_results = create_lesion_size_summary(df)

    # 2. WMH volume stats per condition
    wmh_stats = create_wmh_volume_stats(df)

    # 3. Demographics
    demo_df = create_demographics_table(df)
    print("\nDemographics Table:")
    print(demo_df.to_string(index=False))

    # 4. Boxplots
    print("\nGenerating boxplots …")
    create_scanner_boxplots(df, PLOT_DIR)
    create_lesion_boxplots(df, PLOT_DIR)

    # 5. Save everything
    excel_path = os.path.join(PLOT_DIR, "descriptive_statistics_all.xlsx")
    save_all_to_excel(stats_results, wmh_stats, demo_df, excel_path)
    print(f"\nAll tables saved: {excel_path}")

    print("\nDone.")