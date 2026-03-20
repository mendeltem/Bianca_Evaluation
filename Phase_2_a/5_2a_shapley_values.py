#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SHAP Feature Importance Analysis -- BIANCA WMH Revision
========================================================
Three preprocessing conditions: Non Removed | Removed | Inpainted

For each pairwise comparison, a Random Forest regression model is trained
to predict the absolute WMH volume difference from clinical and imaging
features. SHAP (SHapley Additive exPlanations) values quantify each
feature's contribution to the prediction.


MODEL TRAINING METHODOLOGY (for Response to Reviewers)
------------------------------------------------------

Aim:
  To identify which clinical and imaging factors most strongly influence
  the magnitude of WMH volume differences between preprocessing conditions.

Target variable:
  Absolute WMH volume difference |Cond1 - Cond2| (in mL) for each
  pairwise comparison (non_removed vs removed, non_removed vs inpainted).

Predictor features (7 variables):
  1. Wahlund score (ARWMC)    ordinal; total WMH severity rating
  2. Sex                      binary; 0=female, 1=male
  3. Age                      continuous; years
  4. Stroke lesion volume     continuous; mL (infarct_volume_ml)
  5. Brain volume             continuous; mL (brain_volume_ml)
  6. Scanner type             integer-coded: Siemens (Prisma fit, Tim Trio) = 1,
                               Philips = 2. Grouping reflects vendor (both Siemens
                               scanners are 3T with comparable FLAIR contrast).
  7. Lesion type              integer-coded: infra=1, lacune=2, infarct=3,
                               mixed=4, ICH=5. Tree-based models are invariant
                               to monotonic transformations of categorical
                               encodings; integer coding does not impose an
                               ordinal assumption on the model.

Model:
  sklearn.ensemble.RandomForestRegressor
    - n_estimators = 100
    - max_depth = 5
    - random_state = 42
    - oob_score = True (for unbiased R² estimation)
  Features were z-standardized (StandardScaler) prior to training.
  The model was trained on the full dataset (no train/test split) because
  the goal is explanatory (feature importance via SHAP), not predictive
  generalization. This is consistent with the SHAP framework where the
  model serves as a surrogate to decompose feature contributions
  (Lundberg & Lee, 2017; Molnar, 2022). Out-of-bag (OOB) R² is reported
  alongside in-sample R² to provide a less optimistic performance estimate.

SHAP computation:
  - shap.TreeExplainer (exact, polynomial-time for tree-based models)
  - Mean absolute SHAP values quantify global feature importance
  - 95% confidence intervals computed via bootstrap (1000 iterations),
    consistent with all other analyses in this manuscript
  - Percentage influence = (mean |SHAP_i|) / sum(mean |SHAP_j|) * 100

Interpretation:
  SHAP values indicate how much each feature shifts the predicted volume
  difference from the baseline (mean prediction). A high SHAP value for
  "stroke lesion volume" means that lesion size is a strong predictor of
  how much the WMH segmentation result changes between conditions.

References:
  - Lundberg, S.M., Lee, S.-I., 2017. A Unified Approach to Interpreting
    Model Predictions. NeurIPS 30.
  - Molnar, C., 2022. Interpretable Machine Learning, 2nd ed.
    christophm.github.io/interpretable-ml-book


Revision context
----------------
  R1 Comment 1 / R5 Comment 9: zero-filling vs inpainting
  R5 Comment 4: which factors drive volume differences
  R5 Comment 10: absolute model performance (R², MAE, OOB) required


Paper changes
-------------
  Section 2.10: SHAP methodology description (see METHODS below)
  Section 3.3: Feature importance results (see RESULTS below)
  Section 4.x: Mechanistic interpretation (see DISCUSSION below)
  Figure X: SHAP percentage importance bar plots with bootstrap 95% CI


METHODS (Section 2.10)
---------------------
To identify factors associated with WMH volume differences between
preprocessing conditions, a Random Forest regression model (100 trees,
max depth 5; scikit-learn) was trained on z-standardized features to
predict the absolute WMH volume difference for each pairwise comparison.
Features included stroke lesion volume, Wahlund score (ARWMC), age, sex,
brain volume, scanner type (Siemens vs Philips), and lesion type
(integer-coded; tree-based models are invariant to monotonic
transformations of categorical encodings). The model was trained on the
full dataset without train-test split, as the goal was explanatory
decomposition of feature contributions rather than predictive
generalization (Molnar, 2022). Out-of-bag R² and mean absolute error
(MAE) are reported alongside in-sample R² to characterize surrogate
model fit. Feature importance was quantified using SHAP values
(TreeExplainer; Lundberg and Lee, 2017), reported as percentage of
total mean absolute SHAP with bootstrap 95% confidence intervals
(1000 iterations).


RESULTS (Section 3.3)
---------------------
Stroke lesion volume accounted for the largest share of SHAP importance
in predicting WMH volume differences between non_removed and removed
conditions (76.3%, 95% CI [61.2, 94.6], mean |SHAP| = 0.482;
in-sample R² = 0.895, OOB R² = 0.357, MAE = 0.133 mL), followed by
Wahlund score (9.1%) and age (8.7%). The same ranking was observed for
non_removed vs inpainted (77.9%, 95% CI [61.7, 96.9]; in-sample
R² = 0.904, OOB R² = 0.386, MAE = 0.130 mL), with nearly identical
feature importance profiles. Scanner type contributed ≤0.1% across
both comparisons.
--> See Figure X (SHAP importance plots).


DISCUSSION (Section 4.x)
------------------------
The SHAP analysis indicated that stroke lesion volume was the
highest-ranked predictor of WMH volume differences between preprocessing
conditions. This is consistent with the mechanistic explanation: larger
stroke lesions cause greater distortion of the FLAIR intensity histogram
during normalization, affecting BIANCA's k-NN classifier performance on
subtle WMH. The virtually identical feature importance profiles between
both comparisons (non_removed vs removed and non_removed vs inpainted)
suggest that the choice of filling strategy has minimal influence on
which factors are associated with the volume difference; rather, it is
the act of lesion removal itself that accounts for the observed effects.


RESPONSE TO REVIEWERS
---------------------

R5 Comment 4 (factors driving volume differences):
  "Thank you for raising this point. SHAP analysis identified stroke
  lesion volume as the highest-ranked predictor of WMH volume differences
  (76.3% and 77.9% for non_removed vs removed and non_removed vs
  inpainted, respectively; in-sample R² = 0.895 and 0.904;
  OOB R² = 0.357 and 0.386; MAE = 0.133 and 0.130 mL). Feature
  importance rankings were virtually identical across both comparisons,
  suggesting that the filling strategy does not meaningfully alter which
  factors are associated with the differences. See Section 2.10,
  Section 3.3, and Figure X."

R5 Comment 10 (absolute model performance):
  "Thank you for this important point. We now report in-sample R²,
  out-of-bag R², and MAE for the Random Forest surrogate model. We have
  replaced the term 'dominant determinant' with 'highest-ranked predictor'
  throughout the manuscript to avoid overstating relative importance
  without absolute context. See Section 2.10 and Section 3.3."

R1 Comment 1 / R5 Comment 9 (zero-filling vs inpainting):
  "The SHAP feature importance profiles were virtually identical between
  non_removed vs removed and non_removed vs inpainted (R² = 0.895 vs
  0.904; stroke lesion volume contribution: 76.3% vs 77.9%). This
  provides complementary evidence that zero-filling and inpainting
  produce comparable results, as both comparisons are associated with
  the same factors in similar proportions."
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import shap
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error
from scipy.stats import sem

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "RESULTS", "LOCATE_Results_Metrics_DICE_ONLY.xlsx")
PLOT_DIR = os.path.join(SCRIPT_DIR, "plots", "5_SHAP_Analysis")

FEATURE_COLS = ["ARWMC", "sex", "age", "infarct_volume",
                "brain_volume", "scanner", "lesion_type"]

FEATURE_DISPLAY = {
    "ARWMC": "Wahlund (ARWMC)",
    "sex": "Sex",
    "age": "Age",
    "infarct_volume": "Stroke Lesion Volume",
    "brain_volume": "Brain Volume",
    "scanner": "Scanner Type",
    "lesion_type": "Lesion Type",
}

COND_LABELS = {
    "non_removed": "Non Removed",
    "removed":     "Removed",
    "filled":      "Inpainted",
}

COMPARISONS = [
    ("non_removed", "removed"),
    ("non_removed", "filled"),
]

# Integer encodings for categorical variables
SCANNER_MAPPING = {"Prisma_fit": 1, "Tim Trio": 1, "Philips": 2}
LESION_TYPE_MAPPING = {"infra": 1, "lacune": 2, "infarct": 3, "mixed": 4, "ICH": 5}
SEX_MAPPING = {"Women": 0, "Men": 1, "female": 0, "male": 1, "F": 0, "M": 1}

# Model parameters (documented for reviewers)
RF_PARAMS = {
    "n_estimators": 100,
    "max_depth": 5,
    "random_state": 42,
    "oob_score": True,
    "n_jobs": -1,
}

# Bootstrap parameters (consistent with all other analyses)
N_BOOTSTRAP = 1000
BOOTSTRAP_SEED = 42
BOOTSTRAP_ALPHA = 0.05


# ─────────────────────────────────────────────
# DATA PREPARATION
# ─────────────────────────────────────────────

def prepare_features(df):
    """
    Prepare feature matrix with integer encodings.

    Non-destructive: creates new columns (*_encoded) and maps them to
    the FEATURE_COLS names on a copy, preserving the original DataFrame.

    Encoding rationale (for reviewers):
      - Scanner: Siemens scanners (Prisma fit, Tim Trio) grouped as 1,
        Philips as 2. Both Siemens scanners are 3T with comparable FLAIR
        contrast; grouping reflects vendor differences relevant to
        intensity characteristics.
      - Lesion type: integer-coded (infra=1, lacune=2, infarct=3,
        mixed=4, ICH=5). Tree-based models are invariant to monotonic
        transformations; this coding does not impose ordinal assumptions.
      - Sex: binary (0=female, 1=male).
    """
    df_feat = df.copy()

    # Map categoricals to integers (new columns to avoid overwriting)
    df_feat["scanner_encoded"] = df_feat["scanner"].map(SCANNER_MAPPING)
    df_feat["lesion_type_encoded"] = df_feat["lesion_type"].map(LESION_TYPE_MAPPING)
    df_feat["sex_encoded"] = df_feat["sex"].map(SEX_MAPPING)

    # Map to FEATURE_COLS names for model input
    df_feat["ARWMC"] = df_feat["Wahlund"]
    df_feat["sex"] = df_feat["sex_encoded"]
    df_feat["infarct_volume"] = df_feat["infarct_volume_ml"]
    df_feat["brain_volume"] = df_feat["brain_volume_ml"]
    df_feat["scanner"] = df_feat["scanner_encoded"]
    df_feat["lesion_type"] = df_feat["lesion_type_encoded"]

    # Validate all feature columns exist
    missing = [c for c in FEATURE_COLS if c not in df_feat.columns]
    if missing:
        raise ValueError(f"Missing feature columns after preparation: {missing}")

    # Drop rows with NaN in features
    n_before = len(df_feat)
    df_clean = df_feat.dropna(subset=FEATURE_COLS)
    n_dropped = n_before - len(df_clean)
    if n_dropped > 0:
        print(f"    Dropped {n_dropped} rows with NaN in features")

    return df_clean


# ─────────────────────────────────────────────
# BOOTSTRAP CI FOR SHAP
# ─────────────────────────────────────────────

def bootstrap_shap_ci(abs_shap_values, n_boot=N_BOOTSTRAP,
                       alpha=BOOTSTRAP_ALPHA, seed=BOOTSTRAP_SEED):
    """
    Bootstrap 95% CI for mean |SHAP| per feature.

    Consistent with bootstrap CI methodology used throughout the
    manuscript (1000 iterations, percentile method).

    Parameters
    ----------
    abs_shap_values : np.ndarray, shape (n_subjects, n_features)
    n_boot : int
    alpha : float
    seed : int

    Returns
    -------
    dict : {feature_index: (ci_lo, ci_hi)}
    """
    rng = np.random.default_rng(seed)
    n_subjects, n_features = abs_shap_values.shape
    cis = {}

    for j in range(n_features):
        vals = abs_shap_values[:, j]
        boot_means = np.empty(n_boot)
        for b in range(n_boot):
            idx = rng.choice(n_subjects, size=n_subjects, replace=True)
            boot_means[b] = vals[idx].mean()
        ci_lo = np.percentile(boot_means, 100 * alpha / 2)
        ci_hi = np.percentile(boot_means, 100 * (1 - alpha / 2))
        cis[j] = (ci_lo, ci_hi)

    return cis


def calculate_shap_statistics(shap_values, feature_names):
    """
    Compute mean |SHAP|, bootstrap 95% CI, and percentage per feature.
    """
    abs_shap = np.abs(shap_values)
    boot_cis = bootstrap_shap_ci(abs_shap)

    stats = {}
    total_mean = abs_shap.mean(axis=0).sum()

    for i, feat in enumerate(feature_names):
        vals = abs_shap[:, i]
        mean_val = vals.mean()
        ci_lo, ci_hi = boot_cis[i]
        stats[feat] = {
            "mean": mean_val,
            "std": vals.std(),
            "ci_lo": ci_lo,
            "ci_hi": ci_hi,
            "pct": 100 * mean_val / total_mean,
        }
    return stats


# ─────────────────────────────────────────────
# MODEL TRAINING
# ─────────────────────────────────────────────

def train_and_explain(df_clean, target_col, comparison_label):
    """
    Train RandomForest and compute SHAP values.

    Reports in-sample R², OOB R², and MAE as requested by R5 #10.
    """
    print(f"\n  Training model for: {comparison_label}")

    X = df_clean[FEATURE_COLS].copy()
    y = df_clean[target_col].copy()

    # Drop NaN in target
    valid = y.notna()
    X, y = X[valid], y[valid]
    print(f"    Features: {X.shape}, Target: {y.shape}")

    # Standardize
    scaler = StandardScaler()
    X_scaled = pd.DataFrame(scaler.fit_transform(X),
                            columns=FEATURE_COLS, index=X.index)

    # Train
    model = RandomForestRegressor(**RF_PARAMS)
    model.fit(X_scaled, y)

    r2_insample = model.score(X_scaled, y)
    r2_oob = model.oob_score_
    y_pred = model.predict(X_scaled)
    mae = mean_absolute_error(y, y_pred)

    print(f"    R² (in-sample) = {r2_insample:.4f}")
    print(f"    R² (OOB)       = {r2_oob:.4f}")
    print(f"    MAE            = {mae:.4f} mL")

    # SHAP
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X_scaled)
    stats = calculate_shap_statistics(shap_values, FEATURE_COLS)

    return {
        "model": model,
        "explainer": explainer,
        "shap_values": shap_values,
        "X_scaled": X_scaled,
        "X_raw": X,
        "y": y,
        "r2_insample": r2_insample,
        "r2_oob": r2_oob,
        "mae": mae,
        "stats": stats,
    }


# ─────────────────────────────────────────────
# PLOT: PERCENTAGE IMPORTANCE WITH BOOTSTRAP CI
# ─────────────────────────────────────────────

def plot_percentage_importance(stats, comparison_label, r2_insample,
                                r2_oob, mae, save_path, dpi=300):
    """
    Single horizontal bar plot: % importance with bootstrap 95% CI.
    Colorblind-friendly palette (viridis).
    """
    features = list(stats.keys())
    pcts = np.array([stats[f]["pct"] for f in features])
    means = np.array([stats[f]["mean"] for f in features])
    total_mean = means.sum()

    # Bootstrap CI in percentage space
    ci_lo_pcts = np.array([100 * stats[f]["ci_lo"] / total_mean for f in features])
    ci_hi_pcts = np.array([100 * stats[f]["ci_hi"] / total_mean for f in features])

    # Sort descending
    idx = np.argsort(pcts)[::-1]
    features_s = [FEATURE_DISPLAY.get(features[i], features[i]) for i in idx]
    pcts_s = pcts[idx]
    ci_lo_s = ci_lo_pcts[idx]
    ci_hi_s = ci_hi_pcts[idx]

    # Asymmetric error bars
    err_lo = pcts_s - ci_lo_s
    err_hi = ci_hi_s - pcts_s
    err_lo = np.clip(err_lo, 0, None)
    err_hi = np.clip(err_hi, 0, None)

    # Colorblind-friendly: viridis
    colors = plt.cm.viridis(np.linspace(0.25, 0.85, len(features_s)))

    fig, ax = plt.subplots(figsize=(10, 5))
    y_pos = np.arange(len(features_s))

    ax.barh(y_pos, pcts_s, xerr=[err_lo, err_hi], color=colors,
            capsize=4, error_kw={"linewidth": 1.5, "elinewidth": 1.5,
                                  "color": "0.3"})

    for i, (p, lo, hi) in enumerate(zip(pcts_s, ci_lo_s, ci_hi_s)):
        ax.text(hi + 0.8, i, f"{p:.1f}%  [{lo:.1f}, {hi:.1f}]",
                va="center", fontsize=9)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(features_s, fontsize=10)
    ax.set_xlabel("Percentage of Total |SHAP| Importance (%)", fontsize=11)
    ax.set_title(
        f"SHAP Feature Importance -- {comparison_label}\n"
        f"R² = {r2_insample:.3f} (in-sample), "
        f"R²_OOB = {r2_oob:.3f}, MAE = {mae:.3f} mL",
        fontsize=12)
    ax.invert_yaxis()
    ax.set_xlim(0, ci_hi_pcts.max() + 12)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    plt.tight_layout()
    fig.savefig(save_path, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"    Plot saved: {save_path}")


# ─────────────────────────────────────────────
# RESULTS TABLE
# ─────────────────────────────────────────────

def create_results_table(stats, comparison_label, r2_insample, r2_oob, mae):
    """Publication-ready SHAP importance table with model performance."""
    features = list(stats.keys())
    pcts = np.array([stats[f]["pct"] for f in features])
    idx = np.argsort(pcts)[::-1]

    rows = []
    for i in idx:
        f = features[i]
        s = stats[f]
        rows.append({
            "Feature": FEATURE_DISPLAY.get(f, f),
            "Comparison": comparison_label,
            "Mean |SHAP|": round(s["mean"], 4),
            "Std": round(s["std"], 4),
            "95% CI Lower": round(s["ci_lo"], 4),
            "95% CI Upper": round(s["ci_hi"], 4),
            "% Importance": round(s["pct"], 1),
            "R² (in-sample)": round(r2_insample, 4),
            "R² (OOB)": round(r2_oob, 4),
            "MAE (mL)": round(mae, 4),
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────

def save_all_tables(all_tables, filepath):
    """Save all comparison tables to multi-sheet Excel with formatting."""
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for name, tbl in all_tables.items():
            tbl.to_excel(writer, sheet_name=name[:31], index=False)

    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    thin = Border(left=Side("thin"), right=Side("thin"),
                  top=Side("thin"), bottom=Side("thin"))
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal="center")
                if cell.row == 1:
                    cell.font = Font(bold=True, size=11)
    wb.save(filepath)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("SHAP Feature Importance Analysis")
    print("=" * 60)

    print("\nLoading data ...")
    try:
        df = pd.read_excel(XLSX_PATH)
    except FileNotFoundError:
        raise SystemExit(f"File not found: {XLSX_PATH}")

    # Standard filters (consistent with all scripts)
    if "subject_with_mask" in df.columns:
        df = df[df["subject_with_mask"] == 1].copy()
    if "lesion_type" in df.columns:
        df["lesion_type"] = df["lesion_type"].replace("ICB", "ICH")

    print(f"Subjects with GT mask: {len(df)}")

    # Prepare features
    df_feat = prepare_features(df)
    print(f"After feature preparation: {len(df_feat)} subjects")

    os.makedirs(PLOT_DIR, exist_ok=True)

    all_tables = {}
    all_results = {}

    for cond1, cond2 in COMPARISONS:
        l1, l2 = COND_LABELS[cond1], COND_LABELS[cond2]
        comp_label = f"{l1} vs {l2}"
        comp_key = f"{cond1}_vs_{cond2}"

        # Target: absolute WMH volume difference
        vol1 = f"WMH_{cond1}_volume_ml"
        vol2 = f"WMH_{cond2}_volume_ml"
        target_col = f"WMH_abs_diff_{comp_key}"
        df_feat[target_col] = (df_feat[vol1] - df_feat[vol2]).abs()

        # Train & explain
        res = train_and_explain(df_feat, target_col, comp_label)
        all_results[comp_key] = res

        # Single plot: percentage with bootstrap CI
        plot_percentage_importance(
            res["stats"], comp_label,
            res["r2_insample"], res["r2_oob"], res["mae"],
            os.path.join(PLOT_DIR, f"shap_importance_{comp_key}.png"))

        # Table
        tbl = create_results_table(res["stats"], comp_label,
                                   res["r2_insample"], res["r2_oob"],
                                   res["mae"])
        all_tables[comp_key] = tbl

        # Print summary
        print(f"\n    Feature Importance -- {comp_label}")
        print(f"    R²={res['r2_insample']:.4f} (in-sample), "
              f"R²_OOB={res['r2_oob']:.4f}, MAE={res['mae']:.4f} mL")
        print(f"    {'Feature':<25} {'Mean |SHAP|':<12} "
              f"{'95% CI':<20} {'%':>6}")
        print(f"    {'-' * 65}")
        for _, row in tbl.iterrows():
            print(f"    {row['Feature']:<25} {row['Mean |SHAP|']:<12.4f} "
                  f"[{row['95% CI Lower']:.4f}, {row['95% CI Upper']:.4f}]"
                  f"  {row['% Importance']:>5.1f}%")

    # Save all tables
    excel_path = os.path.join(PLOT_DIR, "shap_feature_importance_all.xlsx")
    save_all_tables(all_tables, excel_path)
    print(f"\n  All tables saved: {excel_path}")

    # Summary comparison
    print("\n" + "=" * 60)
    print("COMPARISON SUMMARY")
    print("=" * 60)
    for comp_key, res in all_results.items():
        print(f"\n  {comp_key}:")
        print(f"    R² in-sample = {res['r2_insample']:.4f}")
        print(f"    R² OOB       = {res['r2_oob']:.4f}")
        print(f"    MAE          = {res['mae']:.4f} mL")
        top_feat = max(res["stats"], key=lambda f: res["stats"][f]["pct"])
        print(f"    Highest-ranked: {FEATURE_DISPLAY[top_feat]} "
              f"({res['stats'][top_feat]['pct']:.1f}%)")

    print("\nDone.")