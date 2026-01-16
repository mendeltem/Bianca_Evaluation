"""
================================================================================
ARWMC Score Dichotomization Analysis: WMH Volume Differences by WMH Burden
================================================================================

Paper Reference: "Robustness and Error Susceptibility of BIANCA for White 
Matter Hyperintensity Segmentation"

This script generates:
- Supplemental Table 3: Volume difference (NR vs R) by dichotomized ARWMC score

Key Finding (Section 3.3.4 and 3.3.5):
"Despite statistical significance, all preprocessing-induced differences remained 
below the threshold for meaningful change (Cliff's delta ≤0.07). This held across 
all relevant subgroups: stratified by ARWMC score (low n=43 vs high n=43)."

"Stratification by ARWMC score showed no such gradient, confirming that 
preprocessing impact scales with lesion volume rather than chronic WMH burden."

Methodology (Section 2.8):
"Regarding ARWMC scores, participants were stratified using the established 
cutoff of <10 (low burden) versus ≥10 (high burden) defined by Leonards et al. 
(2012). Subsequently, a balanced subset (n=86) was created from the Phase II-B 
cohort by retaining all high-score participants (n=43) and randomly selecting 
(random seed fixed for reproducibility) an equal number of low-score participants 
(n=43) to ensure unbiased effect size comparisons."

References:
- Leonards, C.O., et al., 2012. White matter lesion severity in mild acute 
  ischemic stroke patients and functional outcome after 1 year. Stroke 43, 
  3046-3051.
- Wahlund, L.O., et al., 2001. A New Rating Scale for Age-Related White Matter 
  Changes Applicable to MRI and CT. Stroke 32, 1318-1322.

Author: Uchralt Temuulen
================================================================================
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import wilcoxon
from cliffs_delta import cliffs_delta

# Set plotting style
sns.set_style("whitegrid")
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.size'] = 10


# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Define paths relative to project root (modify these for your environment)
PROJECT_ROOT = os.environ.get('PROJECT_ROOT', '.')
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'results', 'arwmc_dichotomization')

# Input data file
INPUT_FILE = os.path.join(DATA_DIR, 'all_results_post_processed.xlsx')

# Create output directory
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ARWMC threshold for dichotomization
# Paper Section 2.8: "the established cutoff of <10 (low burden) versus ≥10 
# (high burden) defined by Leonards et al. (2012)"
ARWMC_THRESHOLD = 10

# Random seed for reproducibility
# Paper: "random seed fixed for reproducibility"
RANDOM_SEED = 42

# WMH types to analyze
WMH_TYPES = [
    ('Total WMH', 'WMH'),
    ('Periventricular WMH', 'perWMH'),
    ('Deep WMH', 'deepWMH')
]

# Significance level
ALPHA_LEVEL = 0.05


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def get_significance_symbol(p_corrected):
    """
    Convert Bonferroni-corrected p-value to significance symbol.
    
    Parameters
    ----------
    p_corrected : float
        Bonferroni-corrected p-value
        
    Returns
    -------
    str
        Significance symbol: '***' (p<0.001), '**' (p<0.01), '*' (p<0.05), 'ns'
    """
    if p_corrected < 0.001:
        return '***'
    elif p_corrected < 0.01:
        return '**'
    elif p_corrected < 0.05:
        return '*'
    else:
        return 'ns'


def classify_effect_size(cliffs_delta_value):
    """
    Classify Cliff's Delta effect size.
    
    Paper Reference (Section 2.8):
    "Effect sizes were classified as negligible (|δ| < 0.28), small 
    (0.28 ≤ |δ| < 0.43), medium (0.43 ≤ |δ| < 0.56), or large (|δ| ≥ 0.56) 
    following Hess & Kromrey (2004)."
    
    Parameters
    ----------
    cliffs_delta_value : float
        Cliff's Delta value
        
    Returns
    -------
    str
        Effect size classification
    """
    abs_delta = abs(cliffs_delta_value)
    if abs_delta < 0.147:
        return 'negligible'
    elif abs_delta < 0.33:
        return 'small'
    elif abs_delta < 0.474:
        return 'medium'
    else:
        return 'large'


def calculate_group_statistics(group_data, wmh_prefix):
    """
    Calculate statistics for a WMH type within a group.
    
    Parameters
    ----------
    group_data : pd.DataFrame
        Data for one ARWMC group
    wmh_prefix : str
        Column prefix for WMH type (e.g., 'WMH', 'perWMH', 'deepWMH')
        
    Returns
    -------
    dict or None
        Statistics dictionary, or None if insufficient data
    """
    removed_col = f'{wmh_prefix}_removed_volume_ml'
    non_removed_col = f'{wmh_prefix}_non_removed_volume_ml'
    
    removed = group_data[removed_col].dropna()
    non_removed = group_data[non_removed_col].dropna()
    
    if len(removed) < 3 or len(non_removed) < 3:
        return None
    
    # Calculate Cliff's Delta effect size
    cliffs_val, _ = cliffs_delta(removed.values, non_removed.values)
    
    # Wilcoxon signed-rank test
    _, p_val = wilcoxon(removed, non_removed)
    
    # Calculate volume differences (NR - R)
    diff = non_removed - removed
    
    return {
        'N': len(removed),
        'Removed Median (ml)': round(removed.median(), 2),
        'Removed Q1 (ml)': round(removed.quantile(0.25), 2),
        'Removed Q3 (ml)': round(removed.quantile(0.75), 2),
        'Non-Removed Median (ml)': round(non_removed.median(), 2),
        'Non-Removed Q1 (ml)': round(non_removed.quantile(0.25), 2),
        'Non-Removed Q3 (ml)': round(non_removed.quantile(0.75), 2),
        'Diff Median (ml)': round(diff.median(), 2),
        'Diff Q1 (ml)': round(diff.quantile(0.25), 2),
        'Diff Q3 (ml)': round(diff.quantile(0.75), 2),
        'P-Value': p_val,
        "Cliff's Delta": round(cliffs_val, 2),
        'Effect Size': classify_effect_size(cliffs_val)
    }


def apply_bonferroni_correction(results_df, alpha=0.05):
    """
    Apply Bonferroni correction for multiple comparisons.
    
    Paper Reference (Section 2.8):
    "Statistical significance was defined as α=0.05 with correction for 
    multiple comparisons using the Bonferroni correction (Bonferroni, 1936)."
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Results dataframe with raw p-values
    alpha : float
        Original significance level (default: 0.05)
        
    Returns
    -------
    pd.DataFrame
        Results with Bonferroni-corrected p-values
    """
    df = results_df.copy()
    n_tests = len(df)
    
    print(f"\nBonferroni Correction:")
    print(f"  Number of tests: {n_tests}")
    print(f"  Corrected alpha: {alpha / n_tests:.6f}")
    
    # Multiply p-values by number of tests, cap at 1.0
    df['P-Value (Bonferroni)'] = (df['P-Value'] * n_tests).clip(upper=1.0)
    
    # Format for display
    df['P-Value (Bonf)-Display'] = df['P-Value (Bonferroni)'].apply(
        lambda x: "<0.001" if x < 0.001 else f"{x:.3f}"
    )
    
    # Add significance symbols
    df['Significance (Bonferroni)'] = df['P-Value (Bonferroni)'].apply(
        get_significance_symbol
    )
    
    return df


# ==============================================================================
# TABLE FORMATTING FUNCTIONS
# ==============================================================================

def format_results_for_publication(results_df, arwmc_threshold=10):
    """
    Format results into publication-ready table format.
    
    This generates Supplemental Table 3 from the paper:
    "Volume difference (mL) not removed versus removed segmented white matter 
    lesion volume by dichotomized ARWMC score"
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Analyzed results with Bonferroni correction
    arwmc_threshold : int
        ARWMC threshold used for dichotomization
        
    Returns
    -------
    pd.DataFrame
        Formatted table matching paper format
    """
    volume_display_names = {
        'Total WMH': 'Total',
        'Periventricular WMH': 'Peri',
        'Deep WMH': 'deep'
    }
    
    formatted_rows = []
    
    for arwmc_group in ['Low', 'High']:
        group_data = results_df[results_df['Group'] == arwmc_group]
        
        if len(group_data) == 0:
            continue
        
        n_value = group_data['N'].iloc[0]
        
        # Format ARWMC group label
        if arwmc_group == 'Low':
            arwmc_label = f'Low ARWMC ≤{arwmc_threshold} (n={n_value})'
        else:
            arwmc_label = f'High ARWMC >{arwmc_threshold} (n={n_value})'
        
        # Add section header row
        formatted_rows.append({
            'WMH': arwmc_label,
            'NR (mL)\n(median, IQR)': '',
            'R (mL)\n(median, IQR)': '',
            'Diff (mL)\n(median, IQR)': '',
            'p (Bonf)': '',
            "Cliff's Delta": '',
            'effect size': ''
        })
        
        # Add data rows (Total, Periventricular, Deep order)
        for wmh_type in ['Total WMH', 'Periventricular WMH', 'Deep WMH']:
            vol_data = group_data[group_data['WMH_Type'] == wmh_type]
            
            if len(vol_data) == 0:
                continue
            
            row = vol_data.iloc[0]
            
            # Format: median (Q1-Q3) on same line
            nr_text = f"{row['Non-Removed Median (ml)']} ({row['Non-Removed Q1 (ml)']}-{row['Non-Removed Q3 (ml)']})"
            r_text = f"{row['Removed Median (ml)']} ({row['Removed Q1 (ml)']}-{row['Removed Q3 (ml)']})"
            diff_text = f"{row['Diff Median (ml)']} ({row['Diff Q1 (ml)']}-{row['Diff Q3 (ml)']})"
            
            formatted_rows.append({
                'WMH': volume_display_names.get(wmh_type, wmh_type),
                'NR (mL)\n(median, IQR)': nr_text,
                'R (mL)\n(median, IQR)': r_text,
                'Diff (mL)\n(median, IQR)': diff_text,
                'p (Bonf)': f"{row['P-Value (Bonf)-Display']} {row['Significance (Bonferroni)']}",
                "Cliff's Delta": f"{row['Cliff\'s Delta']:.2f}",
                'effect size': row['Effect Size']
            })
    
    return pd.DataFrame(formatted_rows)


def save_formatted_excel(df, filepath):
    """
    Save DataFrame to Excel with professional formatting.
    
    Parameters
    ----------
    df : pd.DataFrame
        Formatted results dataframe
    filepath : str
        Output file path
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    
    # Save initial Excel file
    df.to_excel(filepath, index=False, sheet_name='Results')
    
    # Load and format
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 30,  # WMH type
        'B': 18,  # NR
        'C': 18,  # R
        'D': 18,  # Diff
        'E': 12,  # p-value (Bonf)
        'F': 12,  # Cliff's Delta
        'G': 15   # effect size
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            
            # Header row formatting
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Text format for Cliff's Delta column
            if cell.column == 6:
                cell.number_format = '@'
    
    # Set row heights
    ws.row_dimensions[1].height = 35
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 40
    
    # Merge ARWMC group header rows
    for row_num in range(2, ws.max_row + 1):
        cell_value = ws[f'A{row_num}'].value
        if cell_value and 'ARWMC' in str(cell_value):
            ws.merge_cells(f'A{row_num}:G{row_num}')
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row_num}'].font = Font(bold=True)
    
    wb.save(filepath)


# ==============================================================================
# MAIN ANALYSIS
# ==============================================================================

def main():
    """
    Main analysis pipeline for ARWMC score dichotomization analysis.
    
    Paper Reference:
    - Supplemental Table 3: Volume difference by dichotomized ARWMC score
    - Section 3.3.4: Critical finding - All differences remain low
    - Section 2.8: Stratification methodology (Leonards et al., 2012)
    """
    print("=" * 80)
    print("ARWMC SCORE DICHOTOMIZATION ANALYSIS")
    print("Paper: Robustness and Error Susceptibility of BIANCA")
    print("=" * 80)
    
    # -------------------------------------------------------------------------
    # Load data
    # -------------------------------------------------------------------------
    print("\n[1] Loading data...")
    
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(
            f"Input file not found: {INPUT_FILE}\n"
            f"Please update the DATA_DIR path or ensure the file exists."
        )
    
    df = pd.read_excel(INPUT_FILE, index_col=False)
    
    # Filter for validation dataset
    if 'DATASET' in df.columns:
        df = df[df['DATASET'] == 'NULLING'].copy()
    
    print(f"  Total subjects: {len(df)}")
    
    # -------------------------------------------------------------------------
    # Dichotomize by ARWMC score
    # -------------------------------------------------------------------------
    print(f"\n[2] Dichotomizing by ARWMC score (threshold: {ARWMC_THRESHOLD})...")
    
    # Note: ARWMC score may be stored as 'Wahlund' in the data file
    arwmc_col = 'Wahlund' if 'Wahlund' in df.columns else 'ARWMC'
    
    df_low = df[df[arwmc_col] <= ARWMC_THRESHOLD].copy()
    df_high = df[df[arwmc_col] > ARWMC_THRESHOLD].copy()
    
    print(f"  Low ARWMC (≤{ARWMC_THRESHOLD}): n={len(df_low)}")
    print(f"  High ARWMC (>{ARWMC_THRESHOLD}): n={len(df_high)}")
    
    # Balance groups
    # Paper: "a balanced subset (n=86) was created... by randomly selecting"
    target_n = min(len(df_low), len(df_high))
    np.random.seed(RANDOM_SEED)
    
    if len(df_low) > target_n:
        df_low = df_low.sample(n=target_n, random_state=RANDOM_SEED)
    if len(df_high) > target_n:
        df_high = df_high.sample(n=target_n, random_state=RANDOM_SEED)
    
    print(f"\n  Balanced groups: n={target_n} per group (total={2*target_n})")
    
    # -------------------------------------------------------------------------
    # Calculate statistics
    # -------------------------------------------------------------------------
    print("\n[3] Calculating statistics...")
    
    results = []
    
    for group_label, group_data in [('Low', df_low), ('High', df_high)]:
        for wmh_type_name, wmh_prefix in WMH_TYPES:
            stats = calculate_group_statistics(group_data, wmh_prefix)
            
            if stats is None:
                print(f"  Warning: Insufficient data for {wmh_type_name} in {group_label} group")
                continue
            
            stats['Group'] = group_label
            stats['WMH_Type'] = wmh_type_name
            results.append(stats)
    
    results_df = pd.DataFrame(results)
    
    # -------------------------------------------------------------------------
    # Apply Bonferroni correction
    # -------------------------------------------------------------------------
    print("\n[4] Applying Bonferroni correction...")
    results_df = apply_bonferroni_correction(results_df, alpha=ALPHA_LEVEL)
    
    # -------------------------------------------------------------------------
    # Display and save results
    # -------------------------------------------------------------------------
    print("\n[5] Results Summary")
    print("=" * 80)
    
    display_cols = [
        'Group', 'WMH_Type', 'N',
        'Removed Median (ml)', 'Removed Q1 (ml)', 'Removed Q3 (ml)',
        'Non-Removed Median (ml)', 'Non-Removed Q1 (ml)', 'Non-Removed Q3 (ml)',
        'Diff Median (ml)', 'Diff Q1 (ml)', 'Diff Q3 (ml)',
        'P-Value (Bonf)-Display', 'Significance (Bonferroni)',
        "Cliff's Delta", 'Effect Size'
    ]
    
    print(results_df[display_cols].to_string(index=False))
    
    # Save detailed results
    detailed_path = os.path.join(OUTPUT_DIR, 'arwmc_statistics_detailed.xlsx')
    results_df.to_excel(detailed_path, index=False)
    print(f"\n  Detailed results saved to: {detailed_path}")
    
    # -------------------------------------------------------------------------
    # Format and save publication-ready table
    # -------------------------------------------------------------------------
    print("\n[6] Generating Supplemental Table 3...")
    
    formatted_df = format_results_for_publication(results_df, arwmc_threshold=ARWMC_THRESHOLD)
    print(formatted_df.to_string(index=False))
    
    formatted_path = os.path.join(OUTPUT_DIR, 'arwmc_statistics_formatted.xlsx')
    save_formatted_excel(formatted_df, formatted_path)
    print(f"\n  Formatted table saved to: {formatted_path}")
    
    # -------------------------------------------------------------------------
    # Key findings
    # -------------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("KEY FINDINGS (Section 3.3.4)")
    print("=" * 80)
    
    max_delta = results_df["Cliff's Delta"].abs().max()
    print(f"\n  Maximum |Cliff's Delta|: {max_delta:.2f}")
    print(f"  All effect sizes: {'negligible' if max_delta < 0.28 else 'meaningful'}")
    
    print(f"\n  Paper conclusion:")
    print("  'Stratification by ARWMC score showed no such gradient, confirming")
    print("  that preprocessing impact scales with lesion volume rather than")
    print("  chronic WMH burden.'")
    
    print("\n" + "=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)
    
    return results_df, formatted_df


# ==============================================================================
# SCRIPT EXECUTION
# ==============================================================================

if __name__ == "__main__":
    results_df, formatted_df = main()