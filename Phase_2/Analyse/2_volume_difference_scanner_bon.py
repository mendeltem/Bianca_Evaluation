"""
================================================================================
WMH Volume Difference Analysis by Scanner Type: R vs NR Conditions
================================================================================

Paper Reference: "Robustness and Error Susceptibility of BIANCA for White 
Matter Hyperintensity Segmentation"

This script performs the statistical analysis for:
- Table 2: Volume difference (NR vs R) by scanner type
- Section 3.3.6: Critical discovery - Scanner emerges as dominant determinant
- Supplemental Figure 2: SHAP analysis showing scanner importance

Key Finding (Section 3.3.6):
"When stratified by scanner type, preprocessing effects varied across systems,
with volume differences being larger for Philips (median 0.08 mL) compared to 
Siemens scanners (Prisma_fit: 0.04 mL; Tim Trio: 0.02 mL)."

Statistical Methods (Section 2.8):
- Wilcoxon signed-rank test for paired comparisons (non-normal distributions)
- Cliff's Delta for non-parametric effect size estimation
- Bonferroni correction for multiple comparisons

Author: Uchralt Temuulen

================================================================================
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from cliffs_delta import cliffs_delta

# Optional: Import custom plotting setup if available
try:
    from setup import setup_plot_styling
    setup_plot_styling()
except ImportError:
    print("Note: Custom plot styling not available, using defaults")


# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Define paths relative to project root (modify these for your environment)
PROJECT_ROOT = os.environ.get('PROJECT_ROOT', '.')
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'results', 'scanner_analysis')

# Input data file
INPUT_FILE = os.path.join(DATA_DIR, 'all_results_post_processed.xlsx')

# Create output directories
OUTPUT_DIR_FULL = os.path.join(OUTPUT_DIR, 'full_cohort')
OUTPUT_DIR_MASKED = os.path.join(OUTPUT_DIR, 'masked_subjects')
os.makedirs(OUTPUT_DIR_FULL, exist_ok=True)
os.makedirs(OUTPUT_DIR_MASKED, exist_ok=True)

# Analysis parameters
# Paper Section 2.2: Scanner distribution in Phase II-B
# - Philips Ingenia: n=36 (17.1%)
# - Siemens Tim Trio: n=51 (24.2%)
# - Siemens Prisma_fit: n=124 (58.8%)
SCANNER_ORDER = ['Prisma_fit', 'Tim Trio', 'Philips']
ALPHA_LEVEL = 0.05

# Figure parameters
DPI = 300
FIG_WIDTH = 12
FIG_HEIGHT = 9

# Color palette for visualizations
VOLUME_TYPE_COLORS = {
    'Total WMH': '#2ca02c',
    'Periventricular WMH': '#1f77b4',
    'Deep WMH': '#ff7f0e'
}


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def calculate_volume_differences(df):
    """
    Calculate WMH volume differences between non-removed and removed conditions.
    
    Paper Reference (Section 2.4):
    Volume differences represent the effect of lesion removal preprocessing
    on WMH segmentation output.
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with WMH volume columns
        
    Returns
    -------
    pd.DataFrame
        Dataframe with added volume difference columns
    """
    df = df.copy()
    
    # Calculate signed volume differences (NR - R)
    df['WMH_volume_diff'] = (
        df['WMH_non_removed_volume_ml'] - df['WMH_removed_volume_ml']
    )
    df['deepWMH_volume_diff'] = (
        df['deepWMH_non_removed_volume_ml'] - df['deepWMH_removed_volume_ml']
    )
    df['perWMH_volume_diff'] = (
        df['perWMH_non_removed_volume_ml'] - df['perWMH_removed_volume_ml']
    )
    
    # Calculate absolute differences for magnitude analysis
    df['WMH_volume_diff_abs'] = df['WMH_volume_diff'].abs()
    df['deepWMH_volume_diff_abs'] = df['deepWMH_volume_diff'].abs()
    df['perWMH_volume_diff_abs'] = df['perWMH_volume_diff'].abs()
    
    return df


def get_significance_symbol(p_corrected):
    """
    Convert Bonferroni-corrected p-value to significance symbol.
    
    Parameters
    ----------
    p_corrected : float
        Bonferroni-corrected p-value
        
    Returns
    -------
    str
        Significance symbol: '***' (p<0.001), '**' (p<0.01), '*' (p<0.05), 'ns'
    """
    if p_corrected < 0.001:
        return '***'
    elif p_corrected < 0.01:
        return '**'
    elif p_corrected < 0.05:
        return '*'
    else:
        return 'ns'


def get_volume_columns(volume_type):
    """
    Get column names for a given volume type.
    
    Parameters
    ----------
    volume_type : str
        One of 'Total', 'Periventricular', or 'Deep'
        
    Returns
    -------
    tuple
        (removed_column, non_removed_column, diff_column)
    """
    column_mapping = {
        'Total': (
            'WMH_removed_volume_ml',
            'WMH_non_removed_volume_ml',
            'WMH_volume_diff_abs'
        ),
        'Periventricular': (
            'perWMH_removed_volume_ml',
            'perWMH_non_removed_volume_ml',
            'perWMH_volume_diff_abs'
        ),
        'Deep': (
            'deepWMH_removed_volume_ml',
            'deepWMH_non_removed_volume_ml',
            'deepWMH_volume_diff_abs'
        )
    }
    return column_mapping[volume_type]


def analyze_scanner_group(scanner_data, scanner_name):
    """
    Perform statistical analysis for a single scanner type.
    
    Paper Reference (Section 3.3.6):
    "When stratified by scanner type, preprocessing effects varied across 
    systems, with volume differences being larger for Philips (median 0.08 mL) 
    compared to Siemens scanners (Prisma_fit: 0.04 mL; Tim Trio: 0.02 mL)."
    
    Parameters
    ----------
    scanner_data : pd.DataFrame
        Subset of data for one scanner type
    scanner_name : str
        Name of the scanner being analyzed
        
    Returns
    -------
    list
        List of dictionaries containing results for each volume type
    """
    results = []
    volume_types = ['Total', 'Periventricular', 'Deep']
    
    for volume_type in volume_types:
        removed_col, non_removed_col, diff_col = get_volume_columns(volume_type)
        
        removed_values = scanner_data[removed_col].dropna()
        non_removed_values = scanner_data[non_removed_col].dropna()
        diff_values = scanner_data[diff_col].dropna()
        
        if len(diff_values) == 0:
            print(f"  Warning: No valid data for {volume_type} in {scanner_name}")
            continue
        
        # Calculate Cliff's Delta effect size
        # Paper Section 2.8: "Effect sizes were calculated using Cliff's Delta"
        delta, effect_category = cliffs_delta(
            removed_values.values,
            non_removed_values.values
        )
        
        # Perform Wilcoxon signed-rank test (paired, non-parametric)
        # Paper Section 2.8: "Wilcoxon signed-rank test due to the paired 
        # nature of the comparisons and non-normal distribution"
        wilcoxon_result = stats.wilcoxon(removed_values, non_removed_values)
        
        results.append({
            'Scanner': scanner_name,
            'Volume Type': volume_type,
            'N': len(diff_values),
            # Removed condition statistics
            'Removed Median (ml)': round(removed_values.median(), 2),
            'Removed Q1 (ml)': round(removed_values.quantile(0.25), 2),
            'Removed Q3 (ml)': round(removed_values.quantile(0.75), 2),
            # Non-removed condition statistics
            'Non-Removed Median (ml)': round(non_removed_values.median(), 2),
            'Non-Removed Q1 (ml)': round(non_removed_values.quantile(0.25), 2),
            'Non-Removed Q3 (ml)': round(non_removed_values.quantile(0.75), 2),
            # Volume difference statistics
            'Diff Median (ml)': round(diff_values.median(), 2),
            'Diff Q1 (ml)': round(diff_values.quantile(0.25), 2),
            'Diff Q3 (ml)': round(diff_values.quantile(0.75), 2),
            # Statistical test results
            'W-Statistic': round(wilcoxon_result.statistic, 2),
            'P-Value': wilcoxon_result.pvalue,
            "Cliff's Delta": round(delta, 2),
            'Effect Size': effect_category
        })
    
    return results


def apply_bonferroni_correction(results_df, alpha=0.05):
    """
    Apply Bonferroni correction for multiple comparisons.
    
    Paper Reference (Section 2.8):
    "Statistical significance was defined as α=0.05 with correction for 
    multiple comparisons using the Bonferroni correction (Bonferroni, 1936)."
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Results dataframe with raw p-values
    alpha : float
        Original significance level (default: 0.05)
        
    Returns
    -------
    pd.DataFrame
        Results with Bonferroni-corrected p-values (raw p-values removed)
    """
    df = results_df.copy()
    n_tests = len(df)
    
    print(f"\nBonferroni Correction:")
    print(f"  Number of tests: {n_tests}")
    print(f"  Corrected alpha: {alpha / n_tests:.6f}")
    
    # Multiply p-values by number of tests, cap at 1.0
    df['P-Value (Bonferroni)'] = (df['P-Value'] * n_tests).clip(upper=1.0)
    
    # Format for display
    df['P-Value (Bonf)-Display'] = df['P-Value (Bonferroni)'].apply(
        lambda x: "<0.001" if x < 0.001 else f"{x:.3f}"
    )
    
    # Add significance symbols
    df['Significance (Bonferroni)'] = df['P-Value (Bonferroni)'].apply(
        get_significance_symbol
    )
    
    # Remove uncorrected p-value (only report Bonferroni-corrected)
    df = df.drop(columns=['P-Value'])
    
    return df


def format_results_for_publication(results_df, group_by='Scanner'):
    """
    Format results into publication-ready table format.
    
    This generates Table 2 from the paper:
    "Volume difference (mL) not removed (NR) versus removed (R) segmented 
    white matter hyperintensity volume by scanner type"
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Analyzed results with Bonferroni correction
    group_by : str
        Grouping variable ('Scanner' or 'Lesion Type')
        
    Returns
    -------
    pd.DataFrame
        Formatted table matching paper Table 2 format
    """
    # Mapping for volume type display names
    volume_display_names = {
        'Total': 'Total',
        'Periventricular': 'Peri',
        'Deep': 'deep'
    }
    
    # Determine group order based on grouping variable
    if group_by == 'Scanner':
        group_order = SCANNER_ORDER
        group_col = 'Scanner'
    else:
        group_order = ['infra', 'lacune', 'infarct', 'mixed', 'ICH']
        group_col = 'Lesion Type'
    
    formatted_rows = []
    
    for group in group_order:
        group_data = results_df[results_df[group_col] == group]
        
        if len(group_data) == 0:
            continue
        
        n_value = group_data['N'].iloc[0]
        
        # Add section header row
        formatted_rows.append({
            'WMH': f'{group} n={n_value}',
            'NR (mL)\n(median, IQR)': '',
            'R (mL)\n(median, IQR)': '',
            'Diff (mL)\n(median, IQR)': '',
            'p (Bonf)': '',
            "Cliff's Delta": '',
            'effect size': ''
        })
        
        # Add data rows (Total, Periventricular, Deep order)
        for vol_type in ['Total', 'Periventricular', 'Deep']:
            vol_data = group_data[group_data['Volume Type'] == vol_type]
            
            if len(vol_data) == 0:
                continue
            
            row = vol_data.iloc[0]
            
            # Format: median (Q1-Q3) on same line
            nr_text = f"{row['Non-Removed Median (ml)']} ({row['Non-Removed Q1 (ml)']}-{row['Non-Removed Q3 (ml)']})"
            r_text = f"{row['Removed Median (ml)']} ({row['Removed Q1 (ml)']}-{row['Removed Q3 (ml)']})"
            diff_text = f"{row['Diff Median (ml)']} ({row['Diff Q1 (ml)']}-{row['Diff Q3 (ml)']})"
            
            formatted_rows.append({
                'WMH': volume_display_names.get(vol_type, vol_type),
                'NR (mL)\n(median, IQR)': nr_text,
                'R (mL)\n(median, IQR)': r_text,
                'Diff (mL)\n(median, IQR)': diff_text,
                'p (Bonf)': f"{row['P-Value (Bonf)-Display']} {row['Significance (Bonferroni)']}",
                "Cliff's Delta": f"{row['Cliff\'s Delta']:.2f}",
                'effect size': row['Effect Size']
            })
    
    return pd.DataFrame(formatted_rows)


def save_formatted_excel(df, filepath, group_by='Scanner'):
    """
    Save DataFrame to Excel with professional formatting.
    
    Parameters
    ----------
    df : pd.DataFrame
        Formatted results dataframe
    filepath : str
        Output file path
    group_by : str
        Grouping variable for identifying header rows
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    
    # Save initial Excel file
    df.to_excel(filepath, index=False, sheet_name='Results')
    
    # Load and format
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 30,  # WMH type / Scanner
        'B': 18,  # NR
        'C': 18,  # R
        'D': 18,  # Diff
        'E': 12,  # p-value (Bonf)
        'F': 12,  # Cliff's Delta
        'G': 15   # effect size
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            
            # Header row formatting
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Text format for Cliff's Delta column
            if cell.column == 6:
                cell.number_format = '@'
    
    # Set row heights
    ws.row_dimensions[1].height = 35
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 40
    
    # Identify header row keywords based on grouping
    if group_by == 'Scanner':
        keywords = ['prisma', 'tim trio', 'philips']
    else:
        keywords = ['strokes', 'lacunes', 'infarcts', 'mixed', 'hemorrhage']
    
    # Merge header rows
    for row_num in range(2, ws.max_row + 1):
        cell_value = ws[f'A{row_num}'].value
        if cell_value and any(kw in str(cell_value).lower() for kw in keywords):
            ws.merge_cells(f'A{row_num}:G{row_num}')
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row_num}'].font = Font(bold=True)
    
    wb.save(filepath)


def create_boxplot_visualization(df, output_path, title_suffix=''):
    """
    Create boxplot visualization of volume differences by scanner type.
    
    Paper Reference (Table 2):
    Visualizes the distribution of preprocessing effects across scanner types.
    
    Parameters
    ----------
    df : pd.DataFrame
        Data with volume difference columns
    output_path : str
        Path to save the figure
    title_suffix : str
        Additional text for plot title (e.g., '- Masked Data Only')
    """
    # Prepare data for plotting
    volume_type_mapping = {
        'WMH_volume_diff_abs': 'Total WMH',
        'perWMH_volume_diff_abs': 'Periventricular WMH',
        'deepWMH_volume_diff_abs': 'Deep WMH'
    }
    
    melted_df = df.melt(
        id_vars=['scanner'],
        value_vars=['WMH_volume_diff_abs', 'perWMH_volume_diff_abs', 'deepWMH_volume_diff_abs'],
        var_name='volume_type',
        value_name='absolute_difference'
    )
    
    melted_df['volume_type'] = melted_df['volume_type'].map(volume_type_mapping)
    melted_df = melted_df.dropna()
    
    print(f"  Data points for plotting: {len(melted_df)}")
    
    # Create figure
    plt.figure(figsize=(FIG_WIDTH, FIG_HEIGHT), dpi=DPI)
    plt.rcParams['font.family'] = 'monospace'
    
    ax = sns.boxplot(
        x='scanner',
        y='absolute_difference',
        hue='volume_type',
        data=melted_df,
        palette=VOLUME_TYPE_COLORS,
        showfliers=True,
        order=SCANNER_ORDER
    )
    
    # Labels and title
    plt.xlabel('Scanner Type', fontsize=12, fontweight='bold')
    plt.ylabel('Absolute Volume Difference\n|Non-Removed - Removed| (mL)', fontsize=12, fontweight='bold')
    plt.title(
        f'WMH Volume Differences by Scanner Type{title_suffix}\n(Absolute Volume Differences in mL)',
        fontsize=14,
        fontweight='bold',
        pad=20
    )
    
    # Formatting
    ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right')
    plt.legend(
        title="Volume Type",
        loc='center left',
        bbox_to_anchor=(1, 0.5),
        title_fontsize=11,
        fontsize=10
    )
    plt.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    
    # Save
    plt.savefig(output_path, dpi=DPI, bbox_inches='tight', facecolor='white')
    print(f"  Plot saved to: {output_path}")
    plt.close()


# ==============================================================================
# MAIN ANALYSIS FUNCTIONS
# ==============================================================================

def run_scanner_analysis(df, output_dir, data_label='Full Cohort'):
    """
    Run complete scanner-stratified analysis pipeline.
    
    Paper Reference (Section 3.3.6):
    "SHAP analysis revealed the influence of scanner heterogeneity increased 
    31-fold (0.002 → 0.062), rising from 0.8% to 16.1% relative importance 
    from Phase II-A to Phase II-B as Philips scanner representation increased 
    from 9.3% to 17.1%."
    
    Parameters
    ----------
    df : pd.DataFrame
        Input data with volume columns
    output_dir : str
        Directory to save outputs
    data_label : str
        Label for print statements (e.g., 'Full Cohort' or 'Masked Subjects')
        
    Returns
    -------
    pd.DataFrame
        Results dataframe with statistics
    """
    print(f"\n{'=' * 80}")
    print(f"SCANNER ANALYSIS - {data_label.upper()}")
    print(f"{'=' * 80}")
    
    print(f"\nScanner distribution:")
    print(df['scanner'].value_counts().to_string())
    
    # Collect results for all scanners
    all_results = []
    
    for scanner in SCANNER_ORDER:
        scanner_data = df[df['scanner'] == scanner]
        
        if len(scanner_data) == 0:
            print(f"  Warning: No data for scanner '{scanner}'")
            continue
        
        print(f"\n  Processing: {scanner} (n={len(scanner_data)})")
        results = analyze_scanner_group(scanner_data, scanner)
        all_results.extend(results)
    
    if len(all_results) == 0:
        print("  No results generated.")
        return None
    
    results_df = pd.DataFrame(all_results)
    
    # Apply Bonferroni correction
    results_df = apply_bonferroni_correction(results_df, alpha=ALPHA_LEVEL)
    
    # Define display columns (Bonferroni-corrected p-values only)
    display_cols = [
        'Scanner', 'Volume Type', 'N',
        'Removed Median (ml)', 'Removed Q1 (ml)', 'Removed Q3 (ml)',
        'Non-Removed Median (ml)', 'Non-Removed Q1 (ml)', 'Non-Removed Q3 (ml)',
        'Diff Median (ml)', 'Diff Q1 (ml)', 'Diff Q3 (ml)',
        'W-Statistic',
        'P-Value (Bonf)-Display', 'Significance (Bonferroni)',
        "Cliff's Delta", 'Effect Size'
    ]
    
    # Print results
    print(f"\n{'=' * 80}")
    print(f"RESULTS - {data_label.upper()}")
    print(f"{'=' * 80}")
    print(results_df[display_cols].to_string(index=False))
    
    # Save detailed results
    detailed_path = os.path.join(output_dir, 'scanner_volume_difference_detailed.xlsx')
    results_df[display_cols].to_excel(detailed_path, index=False)
    print(f"\n  Detailed results saved to: {detailed_path}")
    
    # Format and save publication-ready table
    formatted_df = format_results_for_publication(results_df, group_by='Scanner')
    formatted_path = os.path.join(output_dir, 'scanner_volume_difference_formatted.xlsx')
    save_formatted_excel(formatted_df, formatted_path, group_by='Scanner')
    print(f"  Formatted table saved to: {formatted_path}")
    
    # Create visualization
    plot_path = os.path.join(output_dir, 'scanner_volume_difference_boxplot.png')
    title_suffix = '' if data_label == 'Full Cohort' else f' - {data_label}'
    create_boxplot_visualization(df, plot_path, title_suffix)
    
    return results_df


def main():
    """
    Main analysis pipeline for scanner-stratified WMH volume difference analysis.
    
    Paper Reference:
    - Table 2: Volume difference by scanner type
    - Section 3.3.6: Scanner emerges as dominant determinant
    - Phase II-B (n=211): Philips n=36 (17.1%), Tim Trio n=51 (24.2%), 
      Prisma_fit n=124 (58.8%)
    """
    print("=" * 80)
    print("WMH VOLUME DIFFERENCE ANALYSIS BY SCANNER TYPE")
    print("Paper: Robustness and Error Susceptibility of BIANCA")
    print("=" * 80)
    
    # -------------------------------------------------------------------------
    # Load and prepare data
    # -------------------------------------------------------------------------
    print("\n[1] Loading data...")
    
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(
            f"Input file not found: {INPUT_FILE}\n"
            f"Please update the DATA_DIR path or ensure the file exists."
        )
    
    df = pd.read_excel(INPUT_FILE, index_col=False)
    
    print(f"  Total subjects loaded: {len(df)}")
    
    # Calculate volume differences if not present
    if 'WMH_volume_diff_abs' not in df.columns:
        df = calculate_volume_differences(df)
        print("  Volume differences calculated.")
    
    print(f"\n  Scanner distribution (all data):")
    print(df['scanner'].value_counts().to_string())
    
    # -------------------------------------------------------------------------
    # Analysis 1: Full cohort
    # -------------------------------------------------------------------------
    print("\n[2] Running analysis on full cohort...")
    results_full = run_scanner_analysis(df, OUTPUT_DIR_FULL, 'Full Cohort')
    
    # -------------------------------------------------------------------------
    # Analysis 2: Masked subjects only (supplementary)
    # -------------------------------------------------------------------------
    print("\n[3] Running analysis on masked subjects only...")
    
    if 'subject_with_mask' in df.columns:
        df_masked = df[df['subject_with_mask'] == 1].copy()
        print(f"  Masked subjects: {len(df_masked)}")
        
        if len(df_masked) > 0:
            results_masked = run_scanner_analysis(df_masked, OUTPUT_DIR_MASKED, 'Masked Subjects')
        else:
            print("  No masked subjects found.")
            results_masked = None
    else:
        print("  Column 'subject_with_mask' not found. Skipping masked analysis.")
        results_masked = None
    
    # -------------------------------------------------------------------------
    # Summary
    # -------------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("ANALYSIS SUMMARY")
    print("=" * 80)
    
    print(f"\nKey Finding (Section 3.3.6):")
    print("  Scanner type emerged as the second most important factor")
    print("  after lesion volume in determining segmentation variability.")
    print(f"\n  Volume differences by scanner (from Table 2):")
    print("    - Philips: median 0.08 mL")
    print("    - Siemens Prisma_fit: median 0.04 mL")
    print("    - Siemens Tim Trio: median 0.02 mL")
    
    print(f"\nStatistical approach:")
    print(f"  - Test: Wilcoxon signed-rank (paired, non-parametric)")
    print(f"  - Effect size: Cliff's Delta")
    print(f"  - Multiple comparison correction: Bonferroni")
    
    print(f"\nOutput locations:")
    print(f"  Full cohort: {OUTPUT_DIR_FULL}")
    print(f"  Masked subjects: {OUTPUT_DIR_MASKED}")
    
    print("\n" + "=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)
    
    return results_full, results_masked


# ==============================================================================
# SCRIPT EXECUTION
# ==============================================================================

if __name__ == "__main__":
    results_full, results_masked = main()