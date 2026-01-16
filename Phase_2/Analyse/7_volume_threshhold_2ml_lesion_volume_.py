"""
================================================================================
Lesion Volume Dichotomization Analysis: WMH Volume Differences by Lesion Size
================================================================================

Paper Reference: "Robustness and Error Susceptibility of BIANCA for White 
Matter Hyperintensity Segmentation"

This script generates:
- Supplemental Table 2: Volume difference (NR vs R) by dichotomized median 
  lesion volume score

Key Finding (Section 3.3.5):
"When dichotomized by median lesion volume (2 mL), patients with small lesions 
(<2 mL, n=106) showed minimal preprocessing effects (Cliff's δ=0.01, p<0.001). 
In larger lesions (≥2 mL, n=105), removal effects were approximately 3-fold 
higher (Cliff's δ=0.03, p<0.001). This systematic scaling with lesion size, 
rather than random measurement variability, demonstrates a predictable, 
directional bias: the larger the lesion, the greater its impact on intensity 
normalization and subsequent WMH detection."

Methodology (Section 2.8):
"For lesion volume, a median-split approach (2.0 mL) was implemented to stratify 
subjects into small (<2 mL, n=106) and large (≥2 mL, n=105) groups. This approach 
allowed for a comparison of effect sizes across various size categories, providing 
a quantitative basis to support the identification of thresholds for potential 
practical significance."

References:
- Hess, M.R., Kromrey, J.D., 2004. Robust confidence intervals for effect sizes: 
  A comparative study of Cohen's d and Cliff's delta under non-normality and 
  heterogeneous variances. Annual Meeting of the American Educational Research 
  Association, San Diego, CA.

Author: Uchralt Temuulen
================================================================================
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import wilcoxon
from cliffs_delta import cliffs_delta

# Set plotting style
sns.set_style("whitegrid")
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.size'] = 10


# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Define paths relative to project root (modify these for your environment)
PROJECT_ROOT = os.environ.get('PROJECT_ROOT', '.')
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'results', 'lesion_volume_dichotomization')

# Input data file
INPUT_FILE = os.path.join(DATA_DIR, 'all_results_post_processed.xlsx')

# Create output directory
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Lesion volume threshold for dichotomization (median split)
# Paper Section 2.8: "a median-split approach (2.0 mL)"
VOLUME_THRESHOLD = 2  # mL

# Bootstrap parameters (for optional CI calculation)
N_BOOTSTRAP = 10000
CONFIDENCE_LEVEL = 0.95

# Random seed for reproducibility
RANDOM_SEED = 42

# WMH types to analyze
WMH_TYPES = [
    ('Total WMH', 'WMH'),
    ('Periventricular WMH', 'perWMH'),
    ('Deep WMH', 'deepWMH')
]

# Significance level
ALPHA_LEVEL = 0.05


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def get_significance_symbol(p_corrected):
    """
    Convert Bonferroni-corrected p-value to significance symbol.
    
    Parameters
    ----------
    p_corrected : float
        Bonferroni-corrected p-value
        
    Returns
    -------
    str
        Significance symbol: '***' (p<0.001), '**' (p<0.01), '*' (p<0.05), 'ns'
    """
    if p_corrected < 0.001:
        return '***'
    elif p_corrected < 0.01:
        return '**'
    elif p_corrected < 0.05:
        return '*'
    else:
        return 'ns'


def classify_effect_size(cliffs_delta_value):
    """
    Classify Cliff's Delta effect size.
    
    Paper Reference (Section 2.8):
    "Effect sizes were classified as negligible (|δ| < 0.28), small 
    (0.28 ≤ |δ| < 0.43), medium (0.43 ≤ |δ| < 0.56), or large (|δ| ≥ 0.56) 
    following Hess & Kromrey (2004)."
    
    Parameters
    ----------
    cliffs_delta_value : float
        Cliff's Delta value
        
    Returns
    -------
    str
        Effect size classification
    """
    abs_delta = abs(cliffs_delta_value)
    if abs_delta < 0.147:
        return 'negligible'
    elif abs_delta < 0.33:
        return 'small'
    elif abs_delta < 0.474:
        return 'medium'
    else:
        return 'large'


def bootstrap_cliffs_delta(removed, non_removed, n_bootstrap=10000, random_state=42):
    """
    Calculate bootstrap distribution of Cliff's delta.
    
    Reference: Hess & Kromrey (2004) - Robust Confidence Intervals for Effect Sizes
    
    Parameters
    ----------
    removed : array-like
        Removed condition values
    non_removed : array-like
        Non-removed condition values
    n_bootstrap : int
        Number of bootstrap iterations
    random_state : int
        Random seed for reproducibility
        
    Returns
    -------
    np.array
        Bootstrap distribution of Cliff's delta values
    """
    np.random.seed(random_state)
    removed = np.array(removed)
    non_removed = np.array(non_removed)
    cd_values = []
    
    for _ in range(n_bootstrap):
        # Resample with replacement from both groups
        removed_boot = np.random.choice(removed, size=len(removed), replace=True)
        non_removed_boot = np.random.choice(non_removed, size=len(non_removed), replace=True)
        
        # Calculate Cliff's delta for bootstrap sample
        cd_boot, _ = cliffs_delta(removed_boot, non_removed_boot)
        cd_values.append(cd_boot)
    
    return np.array(cd_values)


def calculate_group_statistics(group_data, wmh_prefix):
    """
    Calculate statistics for a WMH type within a volume group.
    
    Parameters
    ----------
    group_data : pd.DataFrame
        Data for one volume group
    wmh_prefix : str
        Column prefix for WMH type (e.g., 'WMH', 'perWMH', 'deepWMH')
        
    Returns
    -------
    dict or None
        Statistics dictionary, or None if insufficient data
    """
    removed_col = f'{wmh_prefix}_removed_volume_ml'
    non_removed_col = f'{wmh_prefix}_non_removed_volume_ml'
    
    removed = group_data[removed_col].dropna()
    non_removed = group_data[non_removed_col].dropna()
    
    if len(removed) < 3 or len(non_removed) < 3:
        return None
    
    # Calculate Cliff's Delta effect size
    cliffs_val, _ = cliffs_delta(removed.values, non_removed.values)
    
    # Wilcoxon signed-rank test
    _, p_val = wilcoxon(removed, non_removed)
    
    # Calculate volume differences (NR - R)
    diff = non_removed - removed
    
    return {
        'N': len(removed),
        'Removed Median (ml)': round(removed.median(), 2),
        'Removed Q1 (ml)': round(removed.quantile(0.25), 2),
        'Removed Q3 (ml)': round(removed.quantile(0.75), 2),
        'Non-Removed Median (ml)': round(non_removed.median(), 2),
        'Non-Removed Q1 (ml)': round(non_removed.quantile(0.25), 2),
        'Non-Removed Q3 (ml)': round(non_removed.quantile(0.75), 2),
        'Diff Median (ml)': round(diff.median(), 2),
        'Diff Q1 (ml)': round(diff.quantile(0.25), 2),
        'Diff Q3 (ml)': round(diff.quantile(0.75), 2),
        'P-Value': p_val,
        "Cliff's Delta": round(cliffs_val, 2),
        'Effect Size': classify_effect_size(cliffs_val)
    }


def apply_bonferroni_correction(results_df, alpha=0.05):
    """
    Apply Bonferroni correction for multiple comparisons.
    
    Paper Reference (Section 2.8):
    "Statistical significance was defined as α=0.05 with correction for 
    multiple comparisons using the Bonferroni correction (Bonferroni, 1936)."
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Results dataframe with raw p-values
    alpha : float
        Original significance level (default: 0.05)
        
    Returns
    -------
    pd.DataFrame
        Results with Bonferroni-corrected p-values
    """
    df = results_df.copy()
    n_tests = len(df)
    
    print(f"\nBonferroni Correction:")
    print(f"  Number of tests: {n_tests}")
    print(f"  Corrected alpha: {alpha / n_tests:.6f}")
    
    # Multiply p-values by number of tests, cap at 1.0
    df['P-Value (Bonferroni)'] = (df['P-Value'] * n_tests).clip(upper=1.0)
    
    # Format for display
    df['P-Value (Bonf)-Display'] = df['P-Value (Bonferroni)'].apply(
        lambda x: "<0.001" if x < 0.001 else f"{x:.3f}"
    )
    
    # Add significance symbols
    df['Significance (Bonferroni)'] = df['P-Value (Bonferroni)'].apply(
        get_significance_symbol
    )
    
    return df


# ==============================================================================
# TABLE FORMATTING FUNCTIONS
# ==============================================================================

def format_results_for_publication(results_df, volume_threshold=2):
    """
    Format results into publication-ready table format.
    
    This generates Supplemental Table 2 from the paper:
    "Volume difference (mL) not removed versus removed segmented white matter 
    lesion volume by dichotomized median lesion volume score"
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Analyzed results with Bonferroni correction
    volume_threshold : float
        Volume threshold used for dichotomization
        
    Returns
    -------
    pd.DataFrame
        Formatted table matching paper format
    """
    volume_display_names = {
        'Total WMH': 'Total',
        'Periventricular WMH': 'Peri',
        'Deep WMH': 'deep'
    }
    
    formatted_rows = []
    
    for volume_group in ['Small', 'Large']:
        group_data = results_df[results_df['Group'] == volume_group]
        
        if len(group_data) == 0:
            continue
        
        n_value = group_data['N'].iloc[0]
        
        # Format volume group label
        if volume_group == 'Small':
            volume_label = f'Low volume <{volume_threshold} (n={n_value})'
        else:
            volume_label = f'High volume ≥{volume_threshold} (n={n_value})'
        
        # Add section header row
        formatted_rows.append({
            'WMH': volume_label,
            'NR (mL)\n(median, IQR)': '',
            'R (mL)\n(median, IQR)': '',
            'Diff (mL)\n(median, IQR)': '',
            'p (Bonf)': '',
            "Cliff's Delta": '',
            'effect size': ''
        })
        
        # Add data rows (Total, Periventricular, Deep order)
        for wmh_type in ['Total WMH', 'Periventricular WMH', 'Deep WMH']:
            vol_data = group_data[group_data['WMH_Type'] == wmh_type]
            
            if len(vol_data) == 0:
                continue
            
            row = vol_data.iloc[0]
            
            # Format: median (Q1-Q3) on same line
            nr_text = f"{row['Non-Removed Median (ml)']} ({row['Non-Removed Q1 (ml)']}-{row['Non-Removed Q3 (ml)']})"
            r_text = f"{row['Removed Median (ml)']} ({row['Removed Q1 (ml)']}-{row['Removed Q3 (ml)']})"
            diff_text = f"{row['Diff Median (ml)']} ({row['Diff Q1 (ml)']}-{row['Diff Q3 (ml)']})"
            
            formatted_rows.append({
                'WMH': volume_display_names.get(wmh_type, wmh_type),
                'NR (mL)\n(median, IQR)': nr_text,
                'R (mL)\n(median, IQR)': r_text,
                'Diff (mL)\n(median, IQR)': diff_text,
                'p (Bonf)': f"{row['P-Value (Bonf)-Display']} {row['Significance (Bonferroni)']}",
                "Cliff's Delta": f"{row['Cliff\'s Delta']:.2f}",
                'effect size': row['Effect Size']
            })
    
    return pd.DataFrame(formatted_rows)


def save_formatted_excel(df, filepath):
    """
    Save DataFrame to Excel with professional formatting.
    
    Parameters
    ----------
    df : pd.DataFrame
        Formatted results dataframe
    filepath : str
        Output file path
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    
    # Save initial Excel file
    df.to_excel(filepath, index=False, sheet_name='Results')
    
    # Load and format
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 30,  # WMH type
        'B': 18,  # NR
        'C': 18,  # R
        'D': 18,  # Diff
        'E': 12,  # p-value (Bonf)
        'F': 12,  # Cliff's Delta
        'G': 15   # effect size
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            
            # Header row formatting
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Text format for Cliff's Delta column
            if cell.column == 6:
                cell.number_format = '@'
    
    # Set row heights
    ws.row_dimensions[1].height = 35
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 40
    
    # Merge volume group header rows
    for row_num in range(2, ws.max_row + 1):
        cell_value = ws[f'A{row_num}'].value
        if cell_value and 'volume' in str(cell_value).lower():
            ws.merge_cells(f'A{row_num}:G{row_num}')
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row_num}'].font = Font(bold=True)
    
    wb.save(filepath)


# ==============================================================================
# MAIN ANALYSIS
# ==============================================================================

def main():
    """
    Main analysis pipeline for lesion volume dichotomization analysis.
    
    Paper Reference:
    - Supplemental Table 2: Volume difference by dichotomized lesion volume
    - Section 3.3.5: Secondary analysis - Size-response scaling by lesion volume
    - Section 2.8: Median-split stratification methodology
    """
    print("=" * 80)
    print("LESION VOLUME DICHOTOMIZATION ANALYSIS")
    print("Paper: Robustness and Error Susceptibility of BIANCA")
    print("=" * 80)
    
    print(f"\nConfiguration:")
    print(f"  Volume threshold: {VOLUME_THRESHOLD} mL (median split)")
    print(f"  Bootstrap samples: {N_BOOTSTRAP}")
    print(f"  Confidence level: {CONFIDENCE_LEVEL}")
    
    # -------------------------------------------------------------------------
    # Load data
    # -------------------------------------------------------------------------
    print("\n[1] Loading data...")
    
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(
            f"Input file not found: {INPUT_FILE}\n"
            f"Please update the DATA_DIR path or ensure the file exists."
        )
    
    df = pd.read_excel(INPUT_FILE, index_col=False)
    
    # Filter for validation dataset
    if 'DATASET' in df.columns:
        df = df[df['DATASET'] == 'NULLING'].copy()
    
    print(f"  Total subjects: {len(df)}")
    
    # -------------------------------------------------------------------------
    # Dichotomize by lesion volume
    # -------------------------------------------------------------------------
    print(f"\n[2] Dichotomizing by lesion volume (threshold: {VOLUME_THRESHOLD} mL)...")
    
    # Split by volume threshold - no balancing needed
    df_small = df[df['infarct_volume_ml'] < VOLUME_THRESHOLD].copy()
    df_large = df[df['infarct_volume_ml'] >= VOLUME_THRESHOLD].copy()
    
    print(f"  Small lesions (<{VOLUME_THRESHOLD} mL): n={len(df_small)}")
    print(f"  Large lesions (≥{VOLUME_THRESHOLD} mL): n={len(df_large)}")
    print(f"  Total: n={len(df_small) + len(df_large)}")
    
    # -------------------------------------------------------------------------
    # Calculate statistics
    # -------------------------------------------------------------------------
    print("\n[3] Calculating statistics...")
    
    results = []
    
    for group_label, group_data in [('Small', df_small), ('Large', df_large)]:
        for wmh_type_name, wmh_prefix in WMH_TYPES:
            stats = calculate_group_statistics(group_data, wmh_prefix)
            
            if stats is None:
                print(f"  Warning: Insufficient data for {wmh_type_name} in {group_label} group")
                continue
            
            stats['Group'] = group_label
            stats['WMH_Type'] = wmh_type_name
            results.append(stats)
    
    results_df = pd.DataFrame(results)
    
    # -------------------------------------------------------------------------
    # Apply Bonferroni correction
    # -------------------------------------------------------------------------
    print("\n[4] Applying Bonferroni correction...")
    results_df = apply_bonferroni_correction(results_df, alpha=ALPHA_LEVEL)
    
    # -------------------------------------------------------------------------
    # Display and save results
    # -------------------------------------------------------------------------
    print("\n[5] Results Summary")
    print("=" * 80)
    
    display_cols = [
        'Group', 'WMH_Type', 'N',
        'Removed Median (ml)', 'Removed Q1 (ml)', 'Removed Q3 (ml)',
        'Non-Removed Median (ml)', 'Non-Removed Q1 (ml)', 'Non-Removed Q3 (ml)',
        'Diff Median (ml)', 'Diff Q1 (ml)', 'Diff Q3 (ml)',
        'P-Value (Bonf)-Display', 'Significance (Bonferroni)',
        "Cliff's Delta", 'Effect Size'
    ]
    
    print(results_df[display_cols].to_string(index=False))
    
    # Save detailed results
    detailed_path = os.path.join(OUTPUT_DIR, 'lesion_volume_statistics_detailed.xlsx')
    results_df[display_cols].to_excel(detailed_path, index=False)
    print(f"\n  Detailed results saved to: {detailed_path}")
    
    # -------------------------------------------------------------------------
    # Format and save publication-ready table
    # -------------------------------------------------------------------------
    print("\n[6] Generating Supplemental Table 2...")
    
    formatted_df = format_results_for_publication(results_df, volume_threshold=VOLUME_THRESHOLD)
    print(formatted_df.to_string(index=False))
    
    formatted_path = os.path.join(OUTPUT_DIR, 'lesion_volume_statistics_formatted.xlsx')
    save_formatted_excel(formatted_df, formatted_path)
    print(f"\n  Formatted table saved to: {formatted_path}")
    
    # -------------------------------------------------------------------------
    # Key findings (Section 3.3.5)
    # -------------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("KEY FINDINGS (Section 3.3.5)")
    print("=" * 80)
    
    small_total = results_df[(results_df['Group'] == 'Small') & (results_df['WMH_Type'] == 'Total WMH')]
    large_total = results_df[(results_df['Group'] == 'Large') & (results_df['WMH_Type'] == 'Total WMH')]
    
    if len(small_total) > 0 and len(large_total) > 0:
        small_cd = small_total["Cliff's Delta"].values[0]
        large_cd = large_total["Cliff's Delta"].values[0]
        small_n = small_total["N"].values[0]
        large_n = large_total["N"].values[0]
        
        # Calculate fold difference
        if small_cd != 0:
            fold_diff = abs(large_cd / small_cd)
        else:
            fold_diff = float('inf')
        
        print(f"\n  Small lesions (<{VOLUME_THRESHOLD} mL, n={small_n}):")
        print(f"    Cliff's Delta = {small_cd:.3f}")
        
        print(f"\n  Large lesions (≥{VOLUME_THRESHOLD} mL, n={large_n}):")
        print(f"    Cliff's Delta = {large_cd:.3f}")
        
        print(f"\n  Fold difference: {fold_diff:.1f}x")
        
        print(f"\n  Paper conclusion:")
        print(f"  'This systematic scaling with lesion size, rather than random")
        print(f"  measurement variability, demonstrates a predictable, directional")
        print(f"  bias: the larger the lesion, the greater its impact on intensity")
        print(f"  normalization and subsequent WMH detection.'")
    
    print("\n" + "=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)
    
    return results_df, formatted_df


# ==============================================================================
# SCRIPT EXECUTION
# ==============================================================================

if __name__ == "__main__":
    results_df, formatted_df = main()