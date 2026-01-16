"""
================================================================================
WMH Volume Difference Analysis: Lesion Removed (R) vs Non-Removed (NR)
================================================================================

Paper Reference: "Robustness and Error Susceptibility of BIANCA for White 
Matter Hyperintensity Segmentation"

This script performs the statistical analysis for:
- Table 1: Comparison of WMH segmentation volumes (NR vs R) by lesion type
- Section 3.3.3: Main findings - Lesion removal effects vary by lesion type
- Section 3.3.4: Critical finding - All differences remain low (Cliff's δ ≤0.07)

Statistical Methods (Section 2.8):
- Wilcoxon signed-rank test for paired comparisons (non-normal distributions)
- Cliff's Delta for non-parametric effect size estimation
- Bonferroni correction for multiple comparisons

Author: Uchralt Temuulen
================================================================================
"""

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from cliffs_delta import cliffs_delta

# Optional: Import custom plotting setup if available
try:
    from setup import setup_plot_styling
    setup_plot_styling()
except ImportError:
    print("Note: Custom plot styling not available, using defaults")

# ==============================================================================
# CONFIGURATION
# ==============================================================================

# Define paths relative to project root (modify these for your environment)
PROJECT_ROOT = os.environ.get('PROJECT_ROOT', '.')
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'results', 'volume_difference_analysis')

# Input data file
INPUT_FILE = os.path.join(DATA_DIR, 'all_results_post_processed.xlsx')

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Output file paths
RESULTS_DETAILED_PATH = os.path.join(OUTPUT_DIR, 'lesion_volume_difference_detailed.xlsx')
RESULTS_FORMATTED_PATH = os.path.join(OUTPUT_DIR, 'lesion_volume_difference_formatted.xlsx')
FIGURE_PATH = os.path.join(OUTPUT_DIR, 'lesion_volume_difference.png')

# Analysis parameters
LESION_TYPE_ORDER = ['infra', 'lacune', 'infarct', 'mixed', 'ICH']
ALPHA_LEVEL = 0.05  # Significance level before Bonferroni correction

# Figure parameters
DPI = 1000
FIG_WIDTH = 12
FIG_HEIGHT = 9


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def calculate_volume_differences(df):
    """
    Calculate WMH volume differences between non-removed and removed conditions.
    
    Paper Reference (Section 2.4):
    "We compared two preprocessing strategies: (1) R (removed): lesions were 
    masked from FLAIR images prior to BIANCA training and segmentation; 
    (2) NR (not removed): BIANCA segmentation was performed on original FLAIR 
    images, with lesion voxels subtracted post-hoc."
    
    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with WMH volume columns
        
    Returns
    -------
    pd.DataFrame
        Dataframe with added volume difference columns
    """
    df = df.copy()
    
    # Calculate signed volume differences (NR - R)
    # Positive values indicate NR > R (non-removed yields higher volume)
    df['WMH_volume_diff'] = (
        df['WMH_non_removed_volume_ml'] - df['WMH_removed_volume_ml']
    )
    df['deepWMH_volume_diff'] = (
        df['deepWMH_non_removed_volume_ml'] - df['deepWMH_removed_volume_ml']
    )
    df['perWMH_volume_diff'] = (
        df['perWMH_non_removed_volume_ml'] - df['perWMH_removed_volume_ml']
    )
    
    # Calculate absolute volume differences for magnitude analysis
    df['WMH_volume_diff_abs'] = df['WMH_volume_diff'].abs()
    df['deepWMH_volume_diff_abs'] = df['deepWMH_volume_diff'].abs()
    df['perWMH_volume_diff_abs'] = df['perWMH_volume_diff'].abs()
    
    return df


def get_significance_symbol(p_corrected):
    """
    Convert Bonferroni-corrected p-value to significance symbol.
    
    Parameters
    ----------
    p_corrected : float
        Bonferroni-corrected p-value
        
    Returns
    -------
    str
        Significance symbol: '***' (p<0.001), '**' (p<0.01), '*' (p<0.05), 'ns'
    """
    if p_corrected < 0.001:
        return '***'
    elif p_corrected < 0.01:
        return '**'
    elif p_corrected < 0.05:
        return '*'
    else:
        return 'ns'


def analyze_lesion_type(lesion_data, lesion_type):
    """
    Perform statistical analysis for a single lesion type.
    
    Paper Reference (Section 2.8 - Statistical Analysis):
    "Primary outcome comparisons and effect size differences between 'NR' and 
    'R' conditions were analyzed using the Wilcoxon signed-rank test due to 
    the paired nature of the comparisons and non-normal distribution of volume 
    differences. Effect sizes were calculated using Cliff's Delta."
    
    Parameters
    ----------
    lesion_data : pd.DataFrame
        Subset of data for one lesion type
    lesion_type : str
        Name of the lesion type being analyzed
        
    Returns
    -------
    list
        List of dictionaries containing results for each volume type
    """
    results = []
    
    # Define volume types and their corresponding column names
    volume_configs = [
        {
            'name': 'Total',
            'removed_col': 'WMH_removed_volume_ml',
            'non_removed_col': 'WMH_non_removed_volume_ml',
            'diff_col': 'WMH_volume_diff_abs'
        },
        {
            'name': 'Periventricular',
            'removed_col': 'perWMH_removed_volume_ml',
            'non_removed_col': 'perWMH_non_removed_volume_ml',
            'diff_col': 'perWMH_volume_diff_abs'
        },
        {
            'name': 'Deep',
            'removed_col': 'deepWMH_removed_volume_ml',
            'non_removed_col': 'deepWMH_non_removed_volume_ml',
            'diff_col': 'deepWMH_volume_diff_abs'
        }
    ]
    
    for config in volume_configs:
        removed_values = lesion_data[config['removed_col']].dropna()
        non_removed_values = lesion_data[config['non_removed_col']].dropna()
        diff_values = lesion_data[config['diff_col']].dropna()
        
        if len(diff_values) == 0:
            print(f"  Warning: No valid data for {config['name']} in {lesion_type}")
            continue
        
        # Calculate Cliff's Delta effect size
        # Comparing removed vs non-removed distributions
        delta, effect_category = cliffs_delta(
            removed_values.values, 
            non_removed_values.values
        )
        
        # Perform Wilcoxon signed-rank test (paired, non-parametric)
        wilcoxon_result = stats.wilcoxon(removed_values, non_removed_values)
        
        # Calculate descriptive statistics
        results.append({
            'Lesion Type': lesion_type,
            'Volume Type': config['name'],
            'N': len(diff_values),
            # Removed condition statistics
            'Removed Median (ml)': round(removed_values.median(), 2),
            'Removed Q1 (ml)': round(removed_values.quantile(0.25), 2),
            'Removed Q3 (ml)': round(removed_values.quantile(0.75), 2),
            # Non-removed condition statistics
            'Non-Removed Median (ml)': round(non_removed_values.median(), 2),
            'Non-Removed Q1 (ml)': round(non_removed_values.quantile(0.25), 2),
            'Non-Removed Q3 (ml)': round(non_removed_values.quantile(0.75), 2),
            # Volume difference statistics
            'Diff Median (ml)': round(diff_values.median(), 2),
            'Diff Q1 (ml)': round(diff_values.quantile(0.25), 2),
            'Diff Q3 (ml)': round(diff_values.quantile(0.75), 2),
            # Statistical test results
            'W-Statistic': round(wilcoxon_result.statistic, 2),
            'P-Value': wilcoxon_result.pvalue,
            "Cliff's Delta": round(delta, 2),
            'Effect Size': effect_category
        })
    
    return results


def apply_bonferroni_correction(results_df, alpha=0.05):
    """
    Apply Bonferroni correction for multiple comparisons.
    
    Paper Reference (Section 2.8):
    "Statistical significance was defined as α=0.05 with correction for 
    multiple comparisons using the Bonferroni correction (Bonferroni, 1936)."
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Results dataframe with raw p-values
    alpha : float
        Original significance level (default: 0.05)
        
    Returns
    -------
    pd.DataFrame
        Results with Bonferroni-corrected p-values
    """
    df = results_df.copy()
    n_tests = len(df)
    
    print(f"\nBonferroni Correction:")
    print(f"  Number of tests: {n_tests}")
    print(f"  Corrected alpha: {alpha / n_tests:.6f}")
    
    # Multiply p-values by number of tests, cap at 1.0
    df['P-Value (Bonferroni)'] = (df['P-Value'] * n_tests).clip(upper=1.0)
    
    # Format for display
    df['P-Value (Bonf)-Display'] = df['P-Value (Bonferroni)'].apply(
        lambda x: "<0.001" if x < 0.001 else f"{x:.3f}"
    )
    
    # Add significance symbols
    df['Significance (Bonferroni)'] = df['P-Value (Bonferroni)'].apply(
        get_significance_symbol
    )
    
    return df


def format_results_for_publication(results_df):
    """
    Format results into publication-ready table format.
    
    This generates Table 1 from the paper:
    "Comparison of WMH segmentation volumes lesions not removed (NR) and 
    removed (R)"
    
    Parameters
    ----------
    results_df : pd.DataFrame
        Analyzed results with Bonferroni correction
        
    Returns
    -------
    pd.DataFrame
        Formatted table matching paper Table 1 format
    """
    # Mapping for display names
    lesion_display_names = {
        'infra': 'infratentorial strokes',
        'lacune': 'lacunes',
        'infarct': 'infarcts',
        'mixed': 'mixed (infarcts+lacunes)',
        'ICH': 'intracranial hemorrhage'
    }
    
    volume_display_names = {
        'Total': 'Total',
        'Periventricular': 'Peri',
        'Deep': 'deep'
    }
    
    formatted_rows = []
    
    for lesion in LESION_TYPE_ORDER:
        lesion_data = results_df[results_df['Lesion Type'] == lesion]
        
        if len(lesion_data) == 0:
            continue
        
        n_value = lesion_data['N'].iloc[0]
        lesion_display = lesion_display_names.get(lesion, lesion)
        
        # Add section header row
        formatted_rows.append({
            'WMH': f'{lesion_display} n={n_value}',
            'NR (mL)\n(median, IQR)': '',
            'R (mL)\n(median, IQR)': '',
            'Diff (mL)\n(median, IQR)': '',
            'p (Bonf)': '',
            "Cliff's Delta": '',
            'effect size': ''
        })
        
        # Add data rows (Total, Periventricular, Deep order)
        for vol_type in ['Total', 'Periventricular', 'Deep']:
            vol_data = lesion_data[lesion_data['Volume Type'] == vol_type]
            
            if len(vol_data) == 0:
                continue
            
            row = vol_data.iloc[0]
            
            # Format: median on first line, (Q1-Q3) on second line
            nr_text = f"{row['Non-Removed Median (ml)']}\n({row['Non-Removed Q1 (ml)']}-{row['Non-Removed Q3 (ml)']})"
            r_text = f"{row['Removed Median (ml)']}\n({row['Removed Q1 (ml)']}-{row['Removed Q3 (ml)']})"
            diff_text = f"{row['Diff Median (ml)']}\n({row['Diff Q1 (ml)']}-{row['Diff Q3 (ml)']})"
            
            formatted_rows.append({
                'WMH': volume_display_names.get(vol_type, vol_type),
                'NR (mL)\n(median, IQR)': nr_text,
                'R (mL)\n(median, IQR)': r_text,
                'Diff (mL)\n(median, IQR)': diff_text,
                'p (Bonf)': f"{row['P-Value (Bonf)-Display']} {row['Significance (Bonferroni)']}",
                "Cliff's Delta": f"{row['Cliff\'s Delta']:.2f}",
                'effect size': row['Effect Size']
            })
    
    return pd.DataFrame(formatted_rows)


def save_formatted_excel(df, filepath):
    """
    Save DataFrame to Excel with professional formatting.
    
    Parameters
    ----------
    df : pd.DataFrame
        Formatted results dataframe
    filepath : str
        Output file path
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    
    # Save initial Excel file
    df.to_excel(filepath, index=False, sheet_name='Results')
    
    # Load and format
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 30,  # WMH type
        'B': 18,  # NR
        'C': 18,  # R
        'D': 18,  # Diff
        'E': 12,  # p-value
        'F': 12,  # Cliff's Delta
        'G': 15   # effect size
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            
            # Header row formatting
            if cell.row == 1:
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Text format for Cliff's Delta column
            if cell.column == 6:
                cell.number_format = '@'
    
    # Set row heights
    ws.row_dimensions[1].height = 35
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 40
    
    # Merge lesion type header rows
    lesion_keywords = ['strokes', 'lacunes', 'infarcts', 'mixed', 'hemorrhage']
    for row_num in range(2, ws.max_row + 1):
        cell_value = ws[f'A{row_num}'].value
        if cell_value and any(kw in str(cell_value).lower() for kw in lesion_keywords):
            ws.merge_cells(f'A{row_num}:G{row_num}')
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row_num}'].font = Font(bold=True)
    
    wb.save(filepath)


# ==============================================================================
# MAIN ANALYSIS
# ==============================================================================

def main():
    """
    Main analysis pipeline for WMH volume difference comparison.
    
    Paper Reference:
    - Phase II-B (n=211): Large-scale real-world validation
    - Table 1: Comparison of WMH segmentation volumes by lesion type
    """
    
    print("=" * 80)
    print("WMH VOLUME DIFFERENCE ANALYSIS: R vs NR CONDITIONS")
    print("Paper: Robustness and Error Susceptibility of BIANCA")
    print("=" * 80)
    
    # -------------------------------------------------------------------------
    # Load and prepare data
    # -------------------------------------------------------------------------
    print("\n[1] Loading data...")
    
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(
            f"Input file not found: {INPUT_FILE}\n"
            f"Please update the DATA_DIR path or ensure the file exists."
        )
    
    df = pd.read_excel(INPUT_FILE, index_col=False)
    
    # Filter for validation dataset (Phase II-B)
    # Paper: "Phase II-B included only BeLOVE participants who were not used 
    # for algorithm optimization in Phase I"
    df = df[df["DATASET"] == "NULLING"].copy()
    
    print(f"  Total subjects: {len(df)}")
    print(f"  Lesion type distribution:")
    print(df['lesion_type'].value_counts().to_string())
    
    # Calculate volume differences
    df = calculate_volume_differences(df)
    
    # -------------------------------------------------------------------------
    # Statistical analysis by lesion type
    # -------------------------------------------------------------------------
    print("\n[2] Performing statistical analysis by lesion type...")
    
    all_results = []
    
    for lesion in LESION_TYPE_ORDER:
        lesion_data = df[df["lesion_type"] == lesion]
        
        if len(lesion_data) == 0:
            print(f"  Warning: No data for lesion type '{lesion}'")
            continue
        
        print(f"  Processing: {lesion} (n={len(lesion_data)})")
        results = analyze_lesion_type(lesion_data, lesion)
        all_results.extend(results)
    
    results_df = pd.DataFrame(all_results)
    
    # -------------------------------------------------------------------------
    # Apply Bonferroni correction
    # -------------------------------------------------------------------------
    print("\n[3] Applying Bonferroni correction for multiple comparisons...")
    results_df = apply_bonferroni_correction(results_df, alpha=ALPHA_LEVEL)
    
    # -------------------------------------------------------------------------
    # Display and save results
    # -------------------------------------------------------------------------
    print("\n[4] Results Summary")
    print("=" * 80)
    
    display_columns = [
        "Lesion Type", "Volume Type", "N",
        "Removed Median (ml)", "Removed Q1 (ml)", "Removed Q3 (ml)",
        "Non-Removed Median (ml)", "Non-Removed Q1 (ml)", "Non-Removed Q3 (ml)",
        "Diff Median (ml)", "Diff Q1 (ml)", "Diff Q3 (ml)",
        "W-Statistic", "P-Value (Bonf)-Display", "Significance (Bonferroni)",
        "Cliff's Delta", "Effect Size"
    ]
    
    print(results_df[display_columns].to_string(index=False))
    
    # Save detailed results
    results_df[display_columns].to_excel(RESULTS_DETAILED_PATH, index=False)
    print(f"\n  Detailed results saved to: {RESULTS_DETAILED_PATH}")
    
    # -------------------------------------------------------------------------
    # Format and save publication-ready table
    # -------------------------------------------------------------------------
    print("\n[5] Generating publication-ready table (Table 1)...")
    
    formatted_df = format_results_for_publication(results_df)
    print(formatted_df.to_string(index=False))
    
    save_formatted_excel(formatted_df, RESULTS_FORMATTED_PATH)
    print(f"\n  Formatted table saved to: {RESULTS_FORMATTED_PATH}")
    
    # -------------------------------------------------------------------------
    # Key findings summary
    # -------------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("KEY FINDINGS (Section 3.3.4)")
    print("=" * 80)
    
    max_delta = results_df["Cliff's Delta"].abs().max()
    print(f"  Maximum |Cliff's Delta|: {max_delta:.2f}")
    print(f"  Interpretation: All effect sizes are {'negligible' if max_delta < 0.28 else 'meaningful'}")
    print(f"  Paper conclusion: 'All preprocessing-induced differences remained")
    print(f"    below the threshold for meaningful change (Cliff's delta ≤0.07)'")
    
    return results_df, formatted_df


# ==============================================================================
# SUPPLEMENTARY ANALYSIS: MASKED SUBJECTS ONLY
# ==============================================================================

def analyze_masked_subjects(df_original):
    """
    Repeat analysis for subjects with masks only.
    
    This supplementary analysis examines whether findings hold when restricted
    to subjects who had manual lesion masks available.
    
    Parameters
    ----------
    df_original : pd.DataFrame
        Original dataframe (before filtering)
        
    Returns
    -------
    pd.DataFrame
        Results for masked subjects only
    """
    print("\n" + "=" * 80)
    print("SUPPLEMENTARY ANALYSIS: MASKED SUBJECTS ONLY")
    print("=" * 80)
    
    # Filter for subjects with masks
    df_masked = df_original[df_original["subject_with_mask"] == 1].copy()
    
    print(f"  Total masked subjects: {len(df_masked)}")
    print(f"  Lesion type distribution:")
    print(df_masked['lesion_type'].value_counts().to_string())
    
    if len(df_masked) == 0:
        print("  No masked subjects found. Skipping supplementary analysis.")
        return None
    
    # Calculate volume differences
    df_masked = calculate_volume_differences(df_masked)
    
    # Perform analysis
    all_results = []
    
    for lesion in LESION_TYPE_ORDER:
        lesion_data = df_masked[df_masked["lesion_type"] == lesion]
        
        if len(lesion_data) == 0:
            continue
        
        print(f"  Processing (masked): {lesion} (n={len(lesion_data)})")
        results = analyze_lesion_type(lesion_data, lesion)
        all_results.extend(results)
    
    if len(all_results) == 0:
        print("  No results generated for masked subjects.")
        return None
    
    results_df = pd.DataFrame(all_results)
    results_df = apply_bonferroni_correction(results_df, alpha=ALPHA_LEVEL)
    
    return results_df


# ==============================================================================
# SCRIPT EXECUTION
# ==============================================================================

if __name__ == "__main__":
    # Run main analysis
    results_df, formatted_df = main()
    
    # Optionally run supplementary analysis on masked subjects
    # Uncomment the following lines if needed:
    # 
    # df_full = pd.read_excel(INPUT_FILE, index_col=False)
    # df_full = df_full[df_full["DATASET"] == "NULLING"].copy()
    # results_masked = analyze_masked_subjects(df_full)
    
    print("\n" + "=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)